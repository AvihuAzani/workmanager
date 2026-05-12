"""
cellcom_malai.py — שליפת מלאי ציוד (inventory) מ-API של סלקום
==============================================================

תיאור:
------
סקריפט Python ששולף את רשימת הציוד המוקצה לטכנאי מ-API הפנימי של סלקום,
כולל סיריאלים, כמויות, ופרטי ציוד.

הפונקציונליות:
--------------
- שולף inventory אישי של הטכנאי לפי טוקן
- מציג רשימת פריטי ציוד עם מספרים סיריאליים
- הפלט משמש בסיס לטאב "מחסן" ב-chat_server.py

שימוש:
------
  python cellcom_malai.py

הערה:
-----
  בגרסה המלאה, המלאי נשלף אוטומטית מ-/api/inventory ב-chat_server.py
  ומאוחסן ב-inventory_history.json לניטור שינויים לאורך זמן.
"""

import requests
import os
import json
from datetime import date
import urllib3
urllib3.disable_warnings()

# ============================================================
# הגדרות
# ============================================================
PHONE_DEVICE_ID = "FF210C6E-5313-4961-846D-229DF3FAC0FC"
DEVICE_MODEL    = "ios_Apple_iPhone 16_SysVer_26.3.1_appVer_23.03.26.1P"
OUTPUT_DIR      = r"C:\Users\Avihu\Documents\cellcom_reports"
TOKEN_FILE      = os.path.join(OUTPUT_DIR, "token.txt")

BASE_HEADERS = {
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "he-IL,he;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Content-Type": "application/json",
    "format": "application/json",
    "phonedeviceid": PHONE_DEVICE_ID,
    "devicemodel": DEVICE_MODEL,
    "User-Agent": "HomeTechAppClient/23.03.26 CFNetwork/3860.400.51 Darwin/25.3.0",
}

# ============================================================
# טוקן
# ============================================================
def get_token():
    if not os.path.exists(TOKEN_FILE):
        print(f"לא נמצא קובץ טוקן. הרץ קודם את cellcom_fetcher.py")
        exit(1)
    with open(TOKEN_FILE, "r") as f:
        token = f.read().strip()
    if not token:
        print("קובץ הטוקן ריק.")
        exit(1)
    print("טוקן נטען בהצלחה.")
    return token

# ============================================================
# שליפת מלאי
# ============================================================
def fetch_inventory(token):
    url = "https://tech-api.cellcom.co.il/api/technician/authorize/Inventory/GetMyInventory"
    headers = {**BASE_HEADERS, "Authorization": f"Bearer {token}"}
    body = {
        "DeviceModel": DEVICE_MODEL,
        "IsRefresh": True,
        "PhoneDeviceId": PHONE_DEVICE_ID
    }
    try:
        response = requests.post(url, headers=headers, json=body, verify=False, timeout=15)
        if response.status_code == 401:
            print("הטוקן פג — הרץ קודם את cellcom_fetcher.py")
            exit(1)
        data = response.json()
        if data.get("Header", {}).get("ReturnCode") != "00":
            print(f"שגיאה: {data.get('Header', {}).get('ReturnCodeMessage')}")
            return []
        return data["Body"]["result"]["myInventoryCategories"]
    except Exception as e:
        print(f"שגיאה: {e}")
        return []

# ============================================================
# ייצוא לאקסל
# ============================================================
def export_to_excel(categories):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        os.system("pip install openpyxl")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    today = date.today()
    filename = f"malaí_{today.strftime('%Y-%m-%d')}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    BLUE        = "4472C4"
    GREEN       = "70AD47"
    RED         = "FF0000"
    WHITE       = "FFFFFF"
    HEADER_FNT  = Font(color=WHITE, bold=True, size=11)
    CENTER      = Alignment(horizontal="center", vertical="center")
    RIGHT       = Alignment(horizontal="right", vertical="center")

    # צבע לכל קטגוריה
    CAT_COLORS = {
        "מלאי אינטרנט תקין": "4472C4",
        "מלאי TV תקין":       "70AD47",
        "מלאי חסום":          "FF4444",
    }

    # שיט סיכום
    ws_sum = wb.create_sheet("סיכום מלאי")
    ws_sum.sheet_view.rightToLeft = True

    sum_headers = ["קטגוריה", "פריט", "כמות", "מק\"ט"]
    sum_col_widths = [20, 35, 10, 12]

    for col, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=col, value=h)
        cell.fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
        cell.font = HEADER_FNT
        cell.alignment = CENTER
    ws_sum.row_dimensions[1].height = 22

    sum_row = 2
    total_items = 0

    for cat in categories:
        cat_title = cat.get("title", "")
        cat_color = CAT_COLORS.get(cat_title, "4472C4")
        cat_fill  = PatternFill(start_color=cat_color, end_color=cat_color, fill_type="solid")

        # שיט לכל קטגוריה
        ws = wb.create_sheet(cat_title[:31])
        ws.sheet_view.rightToLeft = True

        headers = ["פריט", "כמות", "מק\"ט", "מספרי סריאל"]
        col_widths = [35, 10, 12, 80]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = PatternFill(start_color=cat_color, end_color=cat_color, fill_type="solid")
            cell.font = HEADER_FNT
            cell.alignment = CENTER
        ws.row_dimensions[1].height = 22

        # כותרת קטגוריה
        ws.cell(row=1, column=1).value = f"{cat_title} — סה\"כ: {cat.get('inventorySum', 0)}"

        items = cat.get("inventoryItems", [])
        for row_idx, item in enumerate(items, 2):
            fill_color = "EEF4FF" if row_idx % 2 == 0 else "FFFFFF"
            row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            serials = item.get("inventorySerialItems") or []
            serials_str = ", ".join(serials) if serials else ""

            row_data = [
                item.get("title", ""),
                item.get("amount", 0),
                item.get("catalog", ""),
                serials_str,
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = row_fill
                cell.alignment = RIGHT

        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # סיכום בתחתית השיט
        summary_row = len(items) + 3
        ws.cell(row=summary_row, column=1, value=f"סה\"כ פריטים: {cat.get('inventorySum', 0)}").font = Font(bold=True, size=12)

        # הוספה לשיט סיכום
        for item in items:
            row_data = [
                cat_title,
                item.get("title", ""),
                item.get("amount", 0),
                item.get("catalog", ""),
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws_sum.cell(row=sum_row, column=col, value=value)
                if col == 1:
                    cell.fill = cat_fill
                    cell.font = Font(color=WHITE, bold=True)
                else:
                    fill_color = "EEF4FF" if sum_row % 2 == 0 else "FFFFFF"
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.alignment = RIGHT
            sum_row += 1
            total_items += item.get("amount", 0)

    # סיכום כללי בשיט סיכום
    for col, width in enumerate(sum_col_widths, 1):
        ws_sum.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws_sum.cell(row=sum_row + 2, column=1, value=f"סה\"כ פריטים במלאי: {total_items}").font = Font(bold=True, size=12)

    wb.save(filepath)
    print(f"אקסל נשמר: {filepath}")
    return filepath

# ============================================================
# הרצה ראשית
# ============================================================
if __name__ == "__main__":
    token = get_token()
    print("שולף מלאי...")
    categories = fetch_inventory(token)
    if categories:
        total = sum(c.get("inventorySum", 0) for c in categories)
        print(f"נמצאו {len(categories)} קטגוריות, סה\"כ {total} פריטים")
        excel_path = export_to_excel(categories)
        print("הושלם!")
        os.startfile(excel_path)
    else:
        print("לא נמצא מלאי.")
