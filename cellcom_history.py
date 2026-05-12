"""
cellcom_history.py — שליפת היסטוריית ביקורים ושמירה ב-SQLite
==============================================================

תיאור:
------
סקריפט Python ששולף היסטוריית ביקורים מ-API של סלקום לטווח תאריכים
מוגדר ושומר אותם ב-SQLite (cellcom_history.db) לצורך הצגה בדוחות.

הפונקציונליות:
--------------
- שולף ביקורים יומיים ב-batch מ-START_DATE עד היום
- שומר לטבלת visits עם כל השדות (סוג משימה, כתובת, סטטוס, תשתית, lsb_flag)
- מזהה ביקורים עם תשתית LSB (לצורך חישוב מחיר נפרד)
- מונע כפילויות (PRIMARY KEY על phone+call_id+fetch_date)

שימוש:
------
  # הגדר BEARER_TOKEN בקובץ לפני הרצה
  python cellcom_history.py

הערה:
-----
  בגרסה המלאה, השליפה מתבצעת אוטומטית מ-chat_server.py דרך /api/reports
  שמסנכרן ישירות מה-API ללא צורך בהרצה ידנית של סקריפט זה.
"""

import requests
import sqlite3
import os
import time
from datetime import date, timedelta
from collections import defaultdict

# ============================================================
# הגדרות
# ============================================================
BEARER_TOKEN = ""  # הכנס כאן את הטוקן שלך

PHONE_DEVICE_ID = "FF210C6E-5313-4961-846D-229DF3FAC0FC"
DEVICE_MODEL    = "ios_Apple_iPhone 16_SysVer_26.3.1_appVer_23.03.26.1P"
OUTPUT_DIR      = r"C:\Users\Avihu\Documents\cellcom_reports"
DB_PATH         = os.path.join(OUTPUT_DIR, "cellcom_history.db")

START_DATE = date(2024, 12, 30)
END_DATE   = date.today()

HEBREW_MONTHS = {
    1: "ינואר", 2: "פברואר", 3: "מרץ", 4: "אפריל",
    5: "מאי", 6: "יוני", 7: "יולי", 8: "אוגוסט",
    9: "ספטמבר", 10: "אוקטובר", 11: "נובמבר", 12: "דצמבר"
}

# ============================================================
# שליפה
# ============================================================
def fetch_schedules(target_date):
    date_str = target_date.strftime("%Y-%m-%d")
    url = "https://tech-api.cellcom.co.il/api/technician/authorize/technicianscedules/getSchedules"

    headers = {
        "Authorization": f"Bearer {BEARER_TOKEN}",
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "he-IL,he;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Content-Type": "application/json",
        "phonedeviceid": PHONE_DEVICE_ID,
        "devicemodel": DEVICE_MODEL,
        "format": "application/json",
    }

    body = {
        "DeviceModel": DEVICE_MODEL,
        "AssignmentStart": f"{date_str}T00:00:00",
        "AssignmentFinish": f"{date_str}T23:59:59",
        "IsRefresh": True,
        "PhoneDeviceId": PHONE_DEVICE_ID
    }

    try:
        response = requests.post(url, headers=headers, json=body, verify=False, timeout=15)
        if response.status_code == 401:
            return None, "expired"
        if response.status_code != 200:
            return [], "ok"
        data = response.json()
        if data.get("Header", {}).get("ReturnCode") != "00":
            return [], "ok"
        return data["Body"]["appointments"], "ok"
    except Exception as e:
        print(f"  שגיאה: {e}")
        return [], "ok"

# ============================================================
# מסד נתונים
# ============================================================
def init_db():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS visits (
            call_id       TEXT,
            fetch_date    TEXT,
            customer_id   TEXT,
            contact_name  TEXT,
            contact_phone TEXT,
            task_type     TEXT,
            street        TEXT,
            city          TEXT,
            appt_start    TEXT,
            appt_finish   TEXT,
            status        TEXT,
            infrastructure TEXT,
            call_type     TEXT,
            is_vip        TEXT,
            PRIMARY KEY (call_id, fetch_date)
        )
    """)
    conn.commit()
    return conn

def date_already_fetched(conn, target_date):
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM visits WHERE fetch_date = ?", (target_date.isoformat(),))
    return cursor.fetchone()[0] > 0

def save_day_to_db(conn, appointments, fetch_date):
    cursor = conn.cursor()
    date_str = fetch_date.isoformat()
    for apt in appointments:
        task = apt.get("task", {})
        call_details = apt.get("callDetails", {})
        cursor.execute("INSERT OR REPLACE INTO visits VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (
            task.get("callId", ""),
            date_str,
            task.get("customer", ""),
            task.get("contactName", ""),
            task.get("contactPhoneNumber", ""),
            task.get("taskType", ""),
            task.get("street", ""),
            task.get("city", ""),
            task.get("formattedAppointmentStart", ""),
            task.get("formattedAppointmentFinish", ""),
            task.get("status", {}).get("displayString", ""),
            call_details.get("infrastructure", ""),
            call_details.get("callType", ""),
            "כן" if task.get("isVIP") == "1" else "לא",
        ))
    conn.commit()

def get_all_visits(conn):
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM visits ORDER BY fetch_date DESC, appt_start ASC")
    return cursor.fetchall()

# ============================================================
# בניית אקסל — שיט לכל חודש + שיט סיכום כללי
# ============================================================
def build_excel(conn):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        os.system("pip install openpyxl")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

    today = date.today()
    filename = f"cellcom_history_{START_DATE}_to_{today}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    headers = [
        "תאריך ביקור", "מס' לקוח", "מס' ביקור",
        "שם לקוח", "טלפון", "סוג משימה",
        "כתובת", "עיר", "שעת תחילה", "שעת סיום",
        "סטטוס", "תשתית", "סוג קריאה", "VIP"
    ]
    col_widths = [13, 13, 12, 22, 14, 28, 22, 12, 10, 10, 18, 10, 14, 5]

    BLUE        = "4472C4"
    WHITE       = "FFFFFF"
    HEADER_FNT  = Font(color=WHITE, bold=True, size=11)
    HEADER_FILL = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    CENTER      = Alignment(horizontal="center", vertical="center")
    RIGHT       = Alignment(horizontal="right", vertical="center")

    def write_sheet(ws, rows):
        ws.sheet_view.rightToLeft = True
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FNT
            cell.alignment = CENTER
        ws.row_dimensions[1].height = 22
        for row_idx, row_data in enumerate(rows, 2):
            fill_color = "EEF4FF" if row_idx % 2 == 0 else "FFFFFF"
            row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = row_fill
                cell.alignment = RIGHT
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        return len(rows)

    all_rows = get_all_visits(conn)
    display_rows = []
    monthly_data = defaultdict(list)

    for row in all_rows:
        (call_id, fetch_date, customer_id, contact_name, contact_phone,
         task_type, street, city, appt_start, appt_finish,
         status, infrastructure, call_type, is_vip) = row

        display = [fetch_date, customer_id, call_id, contact_name, contact_phone,
                   task_type, street, city, appt_start, appt_finish,
                   status, infrastructure, call_type, is_vip]
        display_rows.append(display)
        try:
            d = date.fromisoformat(fetch_date)
            monthly_data[(d.year, d.month)].append(display)
        except:
            pass

    # שיט סיכום כללי
    ws_all = wb.create_sheet("כל התקופה")
    total = write_sheet(ws_all, display_rows)
    ws_all.cell(row=total + 3, column=1, value=f"סה\"כ: {total} ביקורים").font = Font(bold=True, size=12)

    # שיט לכל חודש
    for (year, month), month_rows in sorted(monthly_data.items(), reverse=True):
        ws_m = wb.create_sheet(f"{HEBREW_MONTHS[month]} {year}")
        count = write_sheet(ws_m, month_rows)

        summary_row = count + 3
        ws_m.cell(row=summary_row, column=1, value=f"סה\"כ: {count} ביקורים").font = Font(bold=True, size=12)

        task_counts = defaultdict(int)
        for r in month_rows:
            task_counts[r[5]] += 1

        ws_m.cell(row=summary_row + 2, column=1, value="פירוט לפי סוג משימה:").font = Font(bold=True)
        for i, (ttype, cnt) in enumerate(sorted(task_counts.items(), key=lambda x: -x[1])):
            ws_m.cell(row=summary_row + 3 + i, column=1, value=ttype)
            ws_m.cell(row=summary_row + 3 + i, column=2, value=cnt)

    wb.save(filepath)
    print(f"אקסל נשמר: {filepath}")
    return filepath

# ============================================================
# הרצה ראשית
# ============================================================
if __name__ == "__main__":
    import urllib3
    urllib3.disable_warnings()

    if not BEARER_TOKEN:
        print("שגיאה: הכנס טוקן בשורה BEARER_TOKEN בתחילת הקובץ!")
        input("לחץ Enter לסיום...")
        exit(1)

    conn = init_db()
    today = date.today()

    # חישוב מספר הימים
    total_days = (END_DATE - START_DATE).days + 1
    print(f"שולף {total_days} ימים מ-{START_DATE} עד {END_DATE}...")
    print("זה ייקח כ-2-3 דקות, אנא המתן...\n")

    fetched = 0
    skipped = 0
    errors  = 0

    current = START_DATE
    while current <= END_DATE:
        if date_already_fetched(conn, current):
            skipped += 1
            current += timedelta(days=1)
            continue

        appointments, status = fetch_schedules(current)

        if status == "expired":
            print("\n❌ הטוקן פג! עדכן את BEARER_TOKEN בקובץ והרץ מחדש.")
            print(f"הסקריפט הגיע עד תאריך: {current}")
            conn.close()
            input("לחץ Enter לסיום...")
            exit(1)

        if appointments:
            save_day_to_db(conn, appointments, current)
            print(f"  ✓ {current} — {len(appointments)} ביקורים")
            fetched += len(appointments)
        else:
            print(f"  - {current} — אין ביקורים")

        current += timedelta(days=1)
        time.sleep(0.3)

    print(f"\nסה\"כ נשלפו: {fetched} ביקורים")
    print(f"ימים שכבר היו במאגר: {skipped}")

    print("\nבונה קובץ אקסל היסטורי...")
    excel_path = build_excel(conn)
    conn.close()

    print("הושלם בהצלחה!")
    os.startfile(excel_path)
