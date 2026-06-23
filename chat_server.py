"""
chat_server.py — שרת Flask מרכזי לאפליקציית WorkManager לטכנאי סלקום
======================================================================

תיאור כללי:
-----------
אפליקציית Web מלאה (Single Page Application) שרצה בתוך Docker Container
של HomeAssistant (או כ-standalone server) ומאפשרת לטכנאי סלקום:

1. בנק פקעות     — שליפה בזמן אמת מ-API של סלקום, הצגת כרטיסי משימה
                   עם פרטים מלאים (כתובת, לקוח, ציוד, תשתית),
                   מיון לפי מיקום ג'יאוגרפי, הצגה על מפה (Leaflet/OpenStreetMap)

2. מחסן           — ניהול מלאי ציוד: הוספה, הפחתה, היסטוריה,
                   דוח ציוד לפי תאריך, החזרות עם תמונות

3. דוחות          — היסטוריית ביקורים ב-SQLite, סיכום לפי סוג משימה,
                   חישוב הכנסות (PPC/LSB/חריגה), ייצוא Excel עם סימון אדום
                   לביקורים שלא הושלמו, הצגה על מפה

4. מחירון         — הגדרת מחיר לכל סוג משימה (כולל PPC/LSB לתשתית סיבים),
                   שמירה ב-JSON לכל משתמש

5. מנהל           — ניהול משתמשים, הענקת הרשאות עם תפוגה לפי חודשים,
                   לוח ניהול מתקפל

6. צ'אט פנימי     — מערכת הודעות פנימית בין משתמשים ומנהל

מבנה האימות:
------------
- OTP בSMS עם Redis/localStorage
- X-Token / X-Phone headers בכל בקשת API
- Fallback לquery params (לצורך ייצוא Excel ב-window.location.href)
- הרשאות per-feature עם תוקף תאריך

קבצי נתונים:
------------
- cellcom_history.db      — SQLite עם היסטוריית ביקורים, משתמשים, החזרות, הודעות
- saved_tokens.json       — טוקני API של משתמשים
- prices/<phone>.json     — מחירונים per-user
- prices/overrides_<phone>.json — חריגות מחיר per-card
- return_photos/          — תמונות החזרת ציוד
- qr_payment.png          — קוד QR לתשלום

הרצה:
-----
  בתוך HomeAssistant Docker:
    pip3 install flask requests openpyxl && python3 chat_server.py

  לוקאלית עם Cloudflare Tunnel:
    python run_local.py

  Production (Railway/Heroku):
    gunicorn chat_server:app (ראה Procfile)

תלויות:
-------
  flask, requests, urllib3, openpyxl, sqlite3 (builtin)
  Frontend: Leaflet.js (מפות), Tesseract.js (OCR)
"""

import os, re, json, logging, requests, urllib3, sqlite3
from datetime import datetime, date, timedelta
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from flask import Flask, request, jsonify, session, Response, stream_with_context, send_file

urllib3.disable_warnings()
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "cellcom-secret-2026")

TOKENS_FILE         = os.path.join(os.path.dirname(__file__), "saved_tokens.json")
INVENTORY_HIST_FILE = os.path.join(os.path.dirname(__file__), "inventory_history.json")
DB_PATH             = os.path.join(os.path.dirname(__file__), "cellcom_history.db")
PRICES_DIR          = os.path.join(os.path.dirname(__file__), "prices")
RETURNS_PHOTOS_DIR  = os.path.join(os.path.dirname(__file__), "return_photos")
ARCHIVE_DIR         = os.path.join(os.path.dirname(__file__), "reports_archive")
os.makedirs(ARCHIVE_DIR, exist_ok=True)
QR_PAYMENT_FILE     = os.path.join(os.path.dirname(__file__), "qr_payment.png")
ADMIN_PHONE         = "00526845629"  # ← מספר אדמין (לוגין מקומי, ללא Cellcom)
ADMIN_EMPLOYEE_ID   = "104427"       # ← מספר עובד אדמין (מחליף OTP)
_ADMIN_LOCAL_TOKEN  = "CELLCOM-ADMIN-LOCAL-2026-X9"   # טוקן קבוע לחשבון אדמין
TRIAL_DAYS          = 7
SUB_DAYS            = 30

# ============================================================
# ניהול טוקני Cellcom (saved_tokens.json)
# ============================================================
def _load_tokens():
    if os.path.exists(TOKENS_FILE):
        try:
            with open(TOKENS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def _save_tokens(tokens):
    with open(TOKENS_FILE, "w", encoding="utf-8") as f:
        json.dump(tokens, f, ensure_ascii=False, indent=2)

def save_token_for_phone(phone, token):
    """שומר טוקן Cellcom לפלאפון — תוקף עד סוף היום"""
    tokens = _load_tokens()
    tokens[phone] = {"token": token, "date": date.today().isoformat()}
    _save_tokens(tokens)

def get_valid_token(phone):
    """מחזיר טוקן תקף לפלאפון (מהיום בלבד) או None"""
    tokens = _load_tokens()
    entry = tokens.get(phone)
    if not entry:
        return None
    if isinstance(entry, str):          # פורמט ישן — טוקן בלבד
        return entry
    if entry.get("date") == date.today().isoformat():
        return entry.get("token")
    return None

# הרשאות ברירת מחדל למשתמש רגיל
DEFAULT_PERMISSIONS = {
    "premium":           False,   # גישה לכל הלשוניות (מלבד פקעות שתמיד גלויה)
    "full_card":         False,   # כרטיס פקעה מלא
    "inventory":         False,   # לשונית מחסן
    "equipment_report":  False,   # תת-לשונית דוח ציוד
    "reports":           False,   # לשונית דוחות
    "prices":            False,   # לשונית מחירון
}

def get_auth():
    """קורא טוקן ופלאפון מה-header / query-param (client-side storage) — fallback לסשן (OTP flow)."""
    token = (request.headers.get('X-Token', '').strip()
             or request.args.get('token', '').strip()
             or session.get('token', ''))
    phone = (request.headers.get('X-Phone', '').strip()
             or request.args.get('phone', '').strip()
             or session.get('phone', ''))
    return (token or None), phone

# ============================================================
# דוחות — SQLite
# ============================================================
def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS visits (
            phone TEXT DEFAULT '',
            call_id TEXT, fetch_date TEXT, customer_id TEXT,
            contact_name TEXT, contact_phone TEXT, task_type TEXT,
            street TEXT, city TEXT, appt_start TEXT, appt_finish TEXT,
            status TEXT, infrastructure TEXT, call_type TEXT, is_vip TEXT,
            lsb_flag INTEGER DEFAULT 0,
            PRIMARY KEY (phone, call_id, fetch_date)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS returns (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            phone TEXT,
            serial TEXT,
            description TEXT,
            return_date TEXT,
            photo_filename TEXT,
            created_at TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            phone TEXT PRIMARY KEY,
            name TEXT,
            employee_id TEXT,
            registered_at TEXT,
            trial_expires TEXT,
            subscription_expires TEXT,
            is_admin INTEGER DEFAULT 0,
            permissions TEXT DEFAULT '{}'
        )
    """)
    try:
        conn.execute("ALTER TABLE returns ADD COLUMN phone TEXT DEFAULT ''")
    except:
        pass
    try:
        conn.execute("ALTER TABLE returns ADD COLUMN quantity INTEGER DEFAULT 1")
    except:
        pass
    # migration — הוסף עמודות אם לא קיימות
    try:
        conn.execute("ALTER TABLE visits ADD COLUMN lsb_flag INTEGER DEFAULT 0")
    except:
        pass
    try:
        conn.execute("ALTER TABLE visits ADD COLUMN ban TEXT DEFAULT ''")
    except:
        pass
    try:
        conn.execute("ALTER TABLE visits ADD COLUMN phone TEXT DEFAULT ''")
    except:
        pass
    try:
        conn.execute("ALTER TABLE visits ADD COLUMN system_source TEXT DEFAULT 'JET'")
    except:
        pass
    try:
        conn.execute("ALTER TABLE visits ADD COLUMN user_id TEXT DEFAULT ''")
    except:
        pass
    conn.execute("""
        CREATE TABLE IF NOT EXISTS messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            from_phone TEXT NOT NULL,
            to_phone TEXT NOT NULL,
            body TEXT NOT NULL,
            created_at TEXT NOT NULL,
            is_read INTEGER DEFAULT 0
        )
    """)
    try:
        conn.execute("ALTER TABLE users ADD COLUMN permissions TEXT DEFAULT '{}'")
    except:
        pass
    try:
        conn.execute("ALTER TABLE users ADD COLUMN premium_expires TEXT DEFAULT NULL")
    except:
        pass
    conn.commit()
    return conn

def date_already_fetched(conn, target_date, phone=""):
    cur = conn.execute(
        "SELECT COUNT(*) FROM visits WHERE fetch_date=? AND phone=?",
        (target_date.isoformat(), phone)
    )
    return cur.fetchone()[0] > 0

def save_day_to_db(conn, appointments, fetch_date, phone=""):
    date_str = fetch_date.isoformat()
    # הוסף עמודות חדשות אם לא קיימות
    for col_def in [
        "ALTER TABLE visits ADD COLUMN system_source TEXT DEFAULT 'JET'",
        "ALTER TABLE visits ADD COLUMN user_id TEXT DEFAULT ''",
    ]:
        try: conn.execute(col_def)
        except: pass
    for apt in appointments:
        task = apt.get("task", {})
        cd   = apt.get("callDetails", {})
        # בדוק אם LSB מופיע בכל שדות האפוינטמנט
        raw_text = json.dumps(apt, ensure_ascii=False).upper()
        lsb_flag = 1 if "LSB" in raw_text else 0
        cust_id = task.get("customer","") or task.get("customerId","") or task.get("customerNumber","")
        # BAN — נסה שדות רבים כולל billingAccountNumber, accountNumber וכד'
        ban_val = (task.get("ban","") or task.get("BAN","")
                   or task.get("billingAccountNumber","") or task.get("accountNumber","")
                   or task.get("jetBan","") or task.get("jetAccountNumber","")
                   or task.get("accountId","") or task.get("billingAccount","")
                   or "")
        # ✅ BAN מתוך actions[] — אם אין ב-task, חפש בפעולות שבוצעו בפקע
        if not ban_val or ban_val == cust_id:
            for action in (apt.get("actions") or []):
                a_ban = (action.get("ban","") or action.get("Ban","") or "").strip()
                # רק BAN נומרי (לא PI-prefix, לא ריק)
                if a_ban and a_ban.isdigit() and len(a_ban) >= 6:
                    ban_val = a_ban
                    break
                # גם בתוך jsonParams (JSON string) — חפש "Ban":"XXXXXX"
                if not ban_val or ban_val == cust_id:
                    jp = action.get("jsonParams","") or ""
                    if jp and isinstance(jp, str):
                        m = re.search(r'"Ban"\s*:\s*"(\d{6,})"', jp)
                        if m:
                            ban_val = m.group(1)
                            break
        if not ban_val:
            ban_val = cust_id
        # user_id — השדה שה-API מצפה לו ב-GetTechnicianVisitsHistory
        user_id_val = (task.get("user_id","") or task.get("userId","")
                       or task.get("siebel_id","") or task.get("siebelId","") or cust_id)
        # system_source — JET או SIEBEL (קובע את userType ב-API)
        src_val = (cd.get("sourceSystem","") or task.get("system_source","")
                   or task.get("sourceSystem","") or task.get("systemSource","") or "JET")
        conn.execute(
            "INSERT OR REPLACE INTO visits(phone,call_id,fetch_date,customer_id,contact_name,"
            "contact_phone,task_type,street,city,appt_start,appt_finish,status,"
            "infrastructure,call_type,is_vip,lsb_flag,ban,system_source,user_id) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                phone,
                task.get("callId",""), date_str,
                cust_id, task.get("contactName",""),
                task.get("contactPhoneNumber",""), task.get("taskType",""),
                task.get("street",""), task.get("city",""),
                task.get("formattedAppointmentStart",""), task.get("formattedAppointmentFinish",""),
                task.get("status",{}).get("displayString","") if isinstance(task.get("status"),dict) else task.get("status",""),
                cd.get("infrastructure",""), cd.get("callType",""),
                "כן" if task.get("isVIP")=="1" else "לא",
                lsb_flag, ban_val, src_val, user_id_val,
            )
        )
    conn.commit()

def get_visits_for_range(conn, start_date, end_date, phone=""):
    cur = conn.execute(
        "SELECT phone,call_id,fetch_date,customer_id,contact_name,contact_phone,task_type,"
        "street,city,appt_start,appt_finish,status,infrastructure,call_type,is_vip,lsb_flag,"
        "COALESCE(ban,''),COALESCE(system_source,'JET'),COALESCE(user_id,'') "
        "FROM visits WHERE fetch_date>=? AND fetch_date<=? AND phone=? "
        "ORDER BY fetch_date DESC, appt_start ASC",
        (start_date.isoformat(), end_date.isoformat(), phone)
    )
    cols = ["phone","call_id","fetch_date","customer_id","contact_name","contact_phone",
            "task_type","street","city","appt_start","appt_finish",
            "status","infrastructure","call_type","is_vip","lsb_flag","ban",
            "system_source","user_id"]
    return [dict(zip(cols, row)) for row in cur.fetchall()]

def fetch_schedules_api(token, target_date):
    date_str = target_date.strftime("%Y-%m-%d")
    try:
        headers = {**BASE_HEADERS, "Authorization": f"Bearer {token}"}
        r = requests.post(
            "https://tech-api.cellcom.co.il/api/technician/authorize/technicianscedules/getSchedules",
            headers=headers,
            json={"DeviceModel": DEVICE_MODEL, "IsRefresh": True, "PhoneDeviceId": PHONE_DEVICE_ID,
                  "AssignmentStart": f"{date_str}T00:00:00",
                  "AssignmentFinish": f"{date_str}T23:59:59"},
            verify=False, timeout=40
        )
        data = r.json()
    except Exception:
        return []
    if not data or data.get("Header", {}).get("ReturnCode") != "00":
        return []
    return data.get("Body", {}).get("appointments", [])

# ============================================================
# מחירון
# ============================================================
def _prices_file(phone):
    os.makedirs(PRICES_DIR, exist_ok=True)
    safe = "".join(c if c.isalnum() else "_" for c in (phone or "default"))
    return os.path.join(PRICES_DIR, f"{safe}.json")

def _load_prices_raw(phone=""):
    """טוען את קובץ המחירון הגולמי (dict ישן או list חדש)."""
    path = _prices_file(phone)
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    if phone and phone != ADMIN_PHONE:
        admin_path = _prices_file(ADMIN_PHONE)
        if os.path.exists(admin_path):
            try:
                with open(admin_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
    return {}

def _normalize_prices_history(raw):
    """ממיר פורמט ישן (dict) לרשימת גרסאות [{effective_from, prices}]."""
    if isinstance(raw, list):
        return raw
    # פורמט ישן — dict אחד ללא תאריך
    return [{"effective_from": "2000-01-01", "prices": raw}]

def load_prices(phone=""):
    """מחזיר את המחירון העדכני ביותר (dict)."""
    raw = _load_prices_raw(phone)
    history = _normalize_prices_history(raw)
    if not history:
        return {}
    latest = max(history, key=lambda e: e["effective_from"])
    return latest["prices"]

def load_prices_for_date(phone, date_str):
    """מחזיר את המחירון שהיה בתוקף בתאריך נתון (YYYY-MM-DD)."""
    raw = _load_prices_raw(phone)
    history = _normalize_prices_history(raw)
    # כל הגרסאות שתאריך תחילתן <= תאריך הביקור
    valid = [e for e in history if e["effective_from"] <= date_str]
    if not valid:
        # אם אין גרסה קודמת — השתמש בהקדומה ביותר
        valid = history
    if not valid:
        return {}
    return max(valid, key=lambda e: e["effective_from"])["prices"]

def load_prices_history(phone=""):
    """מחזיר את כל היסטוריית המחירונים מסודרת לפי תאריך."""
    raw = _load_prices_raw(phone)
    history = _normalize_prices_history(raw)
    return sorted(history, key=lambda e: e["effective_from"])

def save_prices(prices, phone="", effective_from=None):
    """שומר גרסת מחירון חדשה עם תאריך תחילה."""
    from datetime import date as _date
    if effective_from is None:
        effective_from = _date.today().strftime("%Y-%m-%d")
    raw = _load_prices_raw(phone)
    history = _normalize_prices_history(raw)
    # החלף אם כבר קיימת גרסה לאותו תאריך, אחרת הוסף
    for entry in history:
        if entry["effective_from"] == effective_from:
            entry["prices"] = prices
            break
    else:
        history.append({"effective_from": effective_from, "prices": prices})
    history.sort(key=lambda e: e["effective_from"])
    path = _prices_file(phone)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)

def get_all_task_types():
    try:
        conn = sqlite3.connect(DB_PATH)
        cur = conn.execute("SELECT DISTINCT task_type FROM visits WHERE task_type != '' ORDER BY task_type")
        types = [row[0] for row in cur.fetchall()]
        conn.close()
        return types
    except:
        return []

# ============================================================
# מנויים / הרשמה
# ============================================================
def db_get_user(phone):
    conn = sqlite3.connect(DB_PATH)
    row = conn.execute(
        "SELECT phone,name,employee_id,registered_at,trial_expires,subscription_expires,is_admin,permissions,premium_expires "
        "FROM users WHERE phone=?", (phone,)
    ).fetchone()
    conn.close()
    if not row: return None
    u = dict(zip(['phone','name','employee_id','registered_at','trial_expires','subscription_expires','is_admin','permissions','premium_expires'], row))
    # parse permissions JSON
    try:
        u['permissions'] = json.loads(u['permissions'] or '{}')
    except:
        u['permissions'] = {}
    return u

def get_effective_permissions(phone):
    """מחזיר הרשאות אפקטיביות — מנהל מקבל הכל, משתמש מקבל merge של ברירת מחדל + שמור + פרמיום תאריך"""
    if phone == ADMIN_PHONE:
        return {k: True for k in DEFAULT_PERMISSIONS}
    user = db_get_user(phone)
    if not user: return dict(DEFAULT_PERMISSIONS)
    if user.get('is_admin'):
        return {k: True for k in DEFAULT_PERMISSIONS}
    saved = user.get('permissions', {})
    merged = dict(DEFAULT_PERMISSIONS)
    today = date.today()
    for k, v in saved.items():
        if k not in DEFAULT_PERMISSIONS:
            continue
        if v is True:                      # הרשאה קבועה
            merged[k] = True
        elif v is False or v is None:      # כבוי
            merged[k] = False
        elif isinstance(v, str):           # תאריך ISO
            try:
                merged[k] = date.fromisoformat(v) >= today
            except:
                merged[k] = False
    # פרמיום לפי תאריך תפוגה (grant-premium מהצ'אט)
    prem_exp = user.get('premium_expires', '')
    if prem_exp:
        try:
            if date.fromisoformat(prem_exp) >= today:
                merged['premium'] = True
        except:
            pass
    return merged

def db_register_user(phone, name, employee_id):
    today = date.today()
    trial_exp = (today + timedelta(days=TRIAL_DAYS)).isoformat()
    premium_exp = (today + timedelta(days=14)).isoformat()   # פרמיום אוטומטי 14 יום
    is_adm = 1 if phone == ADMIN_PHONE else 0
    conn = sqlite3.connect(DB_PATH)
    conn.execute(
        "INSERT OR IGNORE INTO users (phone,name,employee_id,registered_at,trial_expires,is_admin,premium_expires) VALUES (?,?,?,?,?,?,?)",
        (phone, name, employee_id, today.isoformat(), trial_exp, is_adm, premium_exp)
    )
    conn.commit(); conn.close()

def db_extend_subscription(phone, days=None):
    if days is None: days = SUB_DAYS
    today = date.today()
    conn = sqlite3.connect(DB_PATH)
    row = conn.execute("SELECT subscription_expires FROM users WHERE phone=?", (phone,)).fetchone()
    if row and row[0]:
        try: base = max(date.fromisoformat(row[0]), today)
        except: base = today
    else:
        base = today
    new_exp = (base + timedelta(days=days)).isoformat()
    conn.execute("UPDATE users SET subscription_expires=? WHERE phone=?", (new_exp, phone))
    conn.commit(); conn.close()
    return new_exp

def get_sub_status(phone):
    """Returns (status_str, expires_date_str)
       status: 'active' | 'not_registered'
    """
    if phone == ADMIN_PHONE:
        return 'active', None
    user = db_get_user(phone)
    if not user:
        return 'not_registered', None
    return 'active', None

# ============================================================
# מלאי — היסטוריה
# ============================================================
def load_inventory_history():
    if os.path.exists(INVENTORY_HIST_FILE):
        try:
            with open(INVENTORY_HIST_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"last_updated": "", "serials": {}}

def save_inventory_history(hist):
    try:
        with open(INVENTORY_HIST_FILE, "w", encoding="utf-8") as f:
            json.dump(hist, f, ensure_ascii=False, indent=2)
    except Exception as e:
        import logging; logging.error(f"save_inventory_history failed: {e}")

def process_inventory_snapshot(categories):
    """
    קורא מלאי חדש, משווה עם ההיסטוריה.
    לכל סריאל: אם כבר קיים → שמור תאריך ישן, אם חדש → רשום היום.
    מחזיר רשימת פריטים מועשרת עם first_seen.
    """
    today = date.today().isoformat()
    hist = load_inventory_history()
    serials_hist = hist.get("serials", {})

    result = []
    for cat in categories:
        cat_title = cat.get("title", "")
        # קביעת סטטוס לפי שם הקטגוריה
        if "תקין" in cat_title:
            status = "תקין"
        elif "חסום" in cat_title or "לא תקין" in cat_title:
            status = "חסום"
        else:
            status = cat_title

        for item in cat.get("inventoryItems", []):
            desc    = item.get("title", "")
            catalog = item.get("catalog", "")
            serials = item.get("inventorySerialItems") or []
            amount  = item.get("amount", 0)

            if serials:
                for sn in serials:
                    if sn in serials_hist:
                        first_seen = serials_hist[sn]["first_seen"]
                    else:
                        first_seen = today
                    # עדכן היסטוריה
                    serials_hist[sn] = {
                        "first_seen": first_seen,
                        "description": desc,
                        "catalog": catalog,
                        "category": cat_title,
                        "status": status,
                    }
                    result.append({
                        "serial": sn,
                        "description": desc,
                        "catalog": catalog,
                        "category": cat_title,
                        "status": status,
                        "first_seen": first_seen,
                        "days_held": (date.today() - date.fromisoformat(first_seen)).days,
                    })
            else:
                # פריט ללא סריאל — מעקב לפי מק"ט+תיאור
                key = f"NO_SN__{catalog}__{desc}"
                if key in serials_hist:
                    first_seen = serials_hist[key]["first_seen"]
                else:
                    first_seen = today
                serials_hist[key] = {
                    "first_seen": first_seen,
                    "description": desc,
                    "catalog": catalog,
                    "category": cat_title,
                    "status": status,
                }
                result.append({
                    "serial": "",
                    "description": desc,
                    "catalog": catalog,
                    "category": cat_title,
                    "status": status,
                    "first_seen": first_seen,
                    "days_held": (date.today() - date.fromisoformat(first_seen)).days,
                    "amount": amount,
                })

    hist["last_updated"] = today
    hist["serials"] = serials_hist
    save_inventory_history(hist)
    return result

# ============================================================
# הגדרות
# ============================================================
PHONE_DEVICE_ID = "FF210C6E-5313-4961-846D-229DF3FAC0FC"
DEVICE_MODEL    = "ios_Apple_iPhone 16_SysVer_26.5_appVer_09.05.26.1P"
CLIENT_ID       = "354193a2-8d29-11ea-bc55-0242ac130004"
CLIENT_SECRET   = "354193a2-8d29-11ea-bc55-0242ac130003"

BASE_HEADERS = {
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "he-IL,he;q=0.9",
    "Content-Type": "application/json",
    "format": "application/json",
    "phonedeviceid": PHONE_DEVICE_ID,
    "devicemodel": DEVICE_MODEL,
    "User-Agent": "HomeTechAppClient/09.05.26 CFNetwork/3860.600.12 Darwin/25.5.0",
}

INFRA_MAP = {
    "FB": "סיבים IBC", "BF": "סיבים בזק", "BN": "נחושת בזק",
    "HO": "HOT", "NV": "סיבים NV", "IB": "IBC",
}

# ============================================================
# Cellcom API
# ============================================================
def cellcom_post(url, body, token=None):
    headers = {**BASE_HEADERS}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    else:
        headers["Authorization"] = "default"
    try:
        r = requests.post(url, headers=headers, json=body, verify=False, timeout=15)
        return r.json()
    except Exception:
        return {}

def login_step1(phone, employee_id):
    data = cellcom_post("https://tech-api.cellcom.co.il/api/technician/loginStep1", {
        "DeviceModel": DEVICE_MODEL, "PhoneNumber": phone, "EmployeeId": employee_id,
        "ClientId": CLIENT_ID, "LoginType": "EMPLOYEEPHONE", "Scope": "USERNAME",
        "SId": "ios_Apple_iPhone 16_SysVer_26.3.1", "PhoneDeviceId": PHONE_DEVICE_ID,
    })
    body = data.get("Body", {})
    if not body.get("isSuccess"):
        raise Exception("שגיאה בשליחת SMS")
    return body["ticketId"]

def login_step2(phone, employee_id, otp, ticket_id):
    data = cellcom_post("https://tech-api.cellcom.co.il/api/technician/loginStep2", {
        "PhoneNumber": phone, "EmployeeId": employee_id,
        "ClientId": CLIENT_ID, "ClientSecret": CLIENT_SECRET,
        "LoginType": "EMPLOYEEPHONE", "Scope": "USERNAME",
        "SId": "ios_Apple_iPhone 16_SysVer_26.3.1",
        "OtpCode": otp, "OtpGuid": ticket_id, "PhoneDeviceId": PHONE_DEVICE_ID,
    })
    rc = data.get("Header", {}).get("ReturnCode")
    if rc != "0":
        raise Exception(data.get("Header", {}).get("ReturnCodeMessage", "קוד שגוי"))
    return data["Body"]["access_token"]

def fetch_inventory(token):
    data = cellcom_post(
        "https://tech-api.cellcom.co.il/api/technician/authorize/Inventory/GetMyInventory",
        {"DeviceModel": DEVICE_MODEL, "IsRefresh": True, "PhoneDeviceId": PHONE_DEVICE_ID},
        token
    )
    if data.get("Header", {}).get("ReturnCode") != "00":
        return []
    return data.get("Body", {}).get("result", {}).get("myInventoryCategories", [])

_tasks_cache = {}   # token -> (timestamp, tasks)
TASKS_CACHE_TTL = 120  # שניות

def fetch_tasks(token):
    import time as _time
    now = _time.time()
    if token in _tasks_cache:
        ts, cached = _tasks_cache[token]
        if now - ts < TASKS_CACHE_TTL:
            return cached
    # getPotentialTasks לוקח זמן — timeout ארוך יותר מהרגיל
    headers = {**BASE_HEADERS, "Authorization": f"Bearer {token}"}
    r = requests.post(
        "https://tech-api.cellcom.co.il/api/technician/authorize/callActivities/getPotentialTasks",
        headers=headers,
        json={"DeviceModel": DEVICE_MODEL, "PhoneDeviceId": PHONE_DEVICE_ID},
        verify=False, timeout=50
    )
    try:
        data = r.json()
    except Exception:
        data = {}
    if data.get("Header", {}).get("ReturnCode") != "00":
        return []
    rd = data.get("Body", {}).get("ResponseData", "{}")
    if isinstance(rd, str):
        try:
            rd = json.loads(rd)
        except Exception:
            rd = {}
    if not isinstance(rd, dict):
        rd = {}
    tasks_obj = rd.get("Tasks") or {}
    tasks = tasks_obj.get("Task", []) if isinstance(tasks_obj, dict) else []
    if not isinstance(tasks, list):
        tasks = [tasks] if tasks else []
    _tasks_cache[token] = (_time.time(), tasks)
    return tasks

def fetch_call_details(token, call_id, customer_id, source):
    try:
        headers = {**BASE_HEADERS, "Authorization": f"Bearer {token}"}
        r = requests.post(
            "https://tech-api.cellcom.co.il/api/technician/authorize/callDetails/getCallDetails",
            headers=headers,
            json={"DeviceModel": DEVICE_MODEL, "CallId": call_id, "CrmCustomerId": customer_id,
                  "SourceSystem": source, "IsRefresh": False, "PhoneDeviceId": PHONE_DEVICE_ID},
            verify=False, timeout=10
        )
        data = r.json()
    except Exception:
        return {"_error": "timeout"}
    rc = data.get("Header", {}).get("ReturnCode", "")
    if rc != "00":
        return {"_error": data.get("Header", {}).get("ReturnCodeMessage", f"RC={rc}")}
    return data.get("Body", {})

def _find_ibc(obj):
    """מחפש ibcBuildingUrl בכל עומק במבנה JSON"""
    if isinstance(obj, dict):
        if "ibcBuildingUrl" in obj and obj["ibcBuildingUrl"]:
            return obj["ibcBuildingUrl"]
        for v in obj.values():
            r = _find_ibc(v)
            if r: return r
    elif isinstance(obj, list):
        for item in obj:
            r = _find_ibc(item)
            if r: return r
    return ""

# שמות שדות BAN אפשריים בתשובת getCallDetails
_BAN_FIELDS = [
    "ban","BAN","billingAccountNumber","accountNumber","jetBan",
    "jetAccountNumber","accountId","billingAccount","customerId",
    "customer","customerNumber","siebelId","siebel_id",
]

def _extract_numeric_ban(obj, depth=0):
    """מחפש שדה BAN נומרי (לא PI-prefix) בכל עומק במבנה JSON.
    מחזיר (field_name, value) או None."""
    if depth > 10:
        return None
    if isinstance(obj, dict):
        for k in _BAN_FIELDS:
            v = obj.get(k, "")
            if v and isinstance(v, str):
                vs = v.strip()
                if vs and vs.isdigit() and len(vs) >= 6:
                    return (k, vs)
        # חיפוש רקורסיבי בתתי-מילונים
        for k, v in obj.items():
            result = _extract_numeric_ban(v, depth+1)
            if result:
                return result
    elif isinstance(obj, list):
        for item in obj:
            result = _extract_numeric_ban(item, depth+1)
            if result:
                return result
    return None

def parse_equipment(body):
    infra_code = infra_name = technology = ibc_url = ""
    line_code = internet_user = router_model = router_serial = ""
    planned, existing, tv = [], [], []
    spm = body.get("supplyProductServicesModel", {})
    technology = spm.get("generalInfo", {}).get("technologyType", "")
    ibc_url = _find_ibc(body)
    for service in spm.get("services", []):
        if service.get("serviceType") == "INTERNET":
            infra_model = service.get("infrastructureModel", {}) or {}
            ic = infra_model.get("infraComOperator", "") or ""
            infra_code = ic
            infra_name = INFRA_MAP.get(ic, ic)
            line_code = infra_model.get("lineCode", "") or ""
            internet_user = service.get("internetUser", "") or ""
            eq_list = service.get("equipmentList", []) or []
            if eq_list:
                first = eq_list[0]
                router_model  = first.get("productDescription", "") or ""
                router_serial = first.get("serialNumber", "") or ""
        elif service.get("serviceType") == "TV":
            for kit in service.get("equipmentList", []):
                for item in kit.get("serialItems", []):
                    if item.get("productStatus") == "Installed":
                        d = item.get("productDescription", "")
                        s = item.get("serialNumber", "")
                        if d: tv.append(f"{d} ({s})" if s else d)
    for item in spm.get("allVisitSerialEquipment", []):
        status = item.get("productStatus", "")
        desc = item.get("productDescription", "") or item.get("originProductDescription", "")
        actions = item.get("availableEquipmentActions", [])
        if "Install" in actions or status in ["SupplyProcess", "Add"]:
            if desc: planned.append(desc)
        elif status == "Installed" and desc:
            existing.append(desc)
    return {
        "infra_name": infra_name, "technology": technology,
        "planned": " | ".join(planned),
        "existing": " | ".join(existing),
        "tv": " | ".join(tv),
        "ibc_url": ibc_url,
        "line_code": line_code,
        "internet_user": internet_user,
        "router_model": router_model,
        "router_serial": router_serial,
    }

def fetch_visit_history(token, ban, user_id, user_type="JET", customer_id=None):
    """שולף היסטוריית ביקורים — מנסה כמה קומבינציות אוטומטית.

    מבנה נכון ל-PI-prefix (לפי traffic שנלכד מהאפליקציה):
      ban=<customer_id נומרי>  userId=<PI-value>  userType=JET
    """
    ban_str  = str(ban         or "").strip()
    uid_str  = str(user_id     or "").strip()
    cid_str  = str(customer_id or "").strip()
    # אם ban ריק — נשתמש ב-customer_id או user_id
    if not ban_str:
        ban_str = cid_str or uid_str
    HIST_URL = ("https://tech-api.cellcom.co.il/api/technician/authorize/"
                "TechnicianVisitsHistory/GetTechnicianVisitsHistory")

    seen, combos = set(), []
    def _add(b, u, ut):
        key = (b, u, ut)
        if key not in seen and b:
            seen.add(key); combos.append(key)

    if ban_str.upper().startswith("PI"):
        # ה-ban שב-DB הוא בעצם userId — ה-customer_id הנומרי הוא ה-ban האמיתי
        numeric   = ban_str[2:]              # "PI9839246" → "9839246"
        real_ban  = cid_str or numeric       # customer_id נומרי מה-DB
        # ✅ הסדר הנכון שנלכד מהאפליקציה: ban=numeric, userId=PI-value, userType=JET
        _add(real_ban, ban_str,  "JET")
        _add(real_ban, ban_str,  "SIEBEL")
        _add(numeric,  ban_str,  "JET")
        _add(numeric,  ban_str,  "SIEBEL")
        # fallback — מה שניסינו לפני
        _add(ban_str,  uid_str,  "JET")
        _add(ban_str,  uid_str,  "SIEBEL")
        _add(numeric,  numeric,  "JET")
        _add(numeric,  numeric,  "SIEBEL")
    else:
        _add(ban_str, uid_str, user_type or "JET")
        if user_type != "SIEBEL":
            _add(ban_str, uid_str, "SIEBEL")
        if cid_str and cid_str not in (ban_str, uid_str):
            _add(cid_str, ban_str, "JET")

    import requests.exceptions as _rex
    _hist_headers = {**BASE_HEADERS, "Authorization": f"Bearer {token}"}
    for b, u, ut in combos:
        try:
            r = requests.post(HIST_URL, headers=_hist_headers,
                json={"DeviceModel": DEVICE_MODEL, "PhoneDeviceId": PHONE_DEVICE_ID,
                      "ban": b, "userId": u, "userType": ut},
                verify=False, timeout=8)
            data = r.json()
            if data.get("Header", {}).get("ReturnCode") == "00":
                hist = data.get("Body", {}).get("visitsHistory", [])
                if hist:
                    return hist
        except _rex.Timeout:
            break   # שרת איטי — אין טעם לנסות combo נוסף
        except Exception:
            pass
    return []


def fetch_visit_history_debug(token, ban, user_id, user_type="JET", extra_cid=None):
    """גרסת דיבאג — מחזיר גם פרטי הניסיונות כולל תיאור שגיאה.

    מבנה נכון ל-PI-prefix (לפי traffic שנלכד):
      ban=<customer_id נומרי>  userId=<PI-value>  userType=JET
    """
    ban_str  = str(ban      or "").strip()
    uid_str  = str(user_id  or "").strip()
    cid_str  = str(extra_cid or "").strip()
    # אם ban ריק — נשתמש ב-cid או uid (PI accounts שומרים ב-customer_id)
    if not ban_str:
        ban_str = cid_str or uid_str
    if not cid_str:
        cid_str = ban_str
    HIST_URL = ("https://tech-api.cellcom.co.il/api/technician/authorize/"
                "TechnicianVisitsHistory/GetTechnicianVisitsHistory")

    seen, combos = set(), []
    def _add(b, u, ut, label=""):
        key = (b, u, ut)
        if key not in seen and b:
            seen.add(key); combos.append((b, u, ut, label))

    if ban_str.upper().startswith("PI"):
        numeric  = ban_str[2:]
        real_ban = cid_str or numeric
        # ✅ הסדר הנכון שנלכד מהאפליקציה
        _add(real_ban, ban_str,  "JET",    "✅ נכון (cid+PI+JET)")
        _add(real_ban, ban_str,  "SIEBEL", "cid+PI+SIEBEL")
        _add(numeric,  ban_str,  "JET",    "numeric+PI+JET")
        _add(numeric,  ban_str,  "SIEBEL", "numeric+PI+SIEBEL")
        # מה שניסינו לפני (ה-ban/userId הפוך)
        _add(ban_str,  uid_str,  "JET",    "PI+uid+JET (ישן)")
        _add(ban_str,  uid_str,  "SIEBEL", "PI+uid+SIEBEL (ישן)")
        _add(ban_str,  ban_str,  "SIEBEL", "PI+PI+SIEBEL (ישן)")
        _add(numeric,  numeric,  "JET",    "num+num+JET (ישן)")
        _add(numeric,  numeric,  "SIEBEL", "num+num+SIEBEL (ישן)")
        if cid_str and cid_str not in (ban_str, uid_str, numeric):
            _add(real_ban, cid_str, "JET",  "cid+cid+JET")
    else:
        _add(ban_str, uid_str, user_type or "JET", "רגיל")
        if user_type != "SIEBEL":
            _add(ban_str, uid_str, "SIEBEL", "SIEBEL")
        if cid_str and cid_str not in (ban_str, uid_str):
            _add(cid_str, ban_str, "JET",    "swap cid+ban")
            _add(ban_str, cid_str, "JET",    "ban+cid")

    attempts = []
    for b, u, ut, lbl in combos:
        try:
            headers = {**BASE_HEADERS, "Authorization": f"Bearer {token}"}
            body_req = {"DeviceModel": DEVICE_MODEL, "PhoneDeviceId": PHONE_DEVICE_ID,
                        "ban": b, "userId": u, "userType": ut}
            r = requests.post(HIST_URL, headers=headers, json=body_req,
                              verify=False, timeout=15)
            try:
                data = r.json()
            except Exception:
                data = {}
            hdr  = data.get("Header", {})
            rc   = hdr.get("ReturnCode", "?")
            desc = hdr.get("ReturnCodeMessage", "") or hdr.get("ReturnDescription", "") or hdr.get("Message", "")
            hist = data.get("Body", {}).get("visitsHistory", []) if rc == "00" else []
            raw_snippet = str(data)[:300]
            attempts.append({
                "ban": b, "userId": u, "userType": ut, "label": lbl,
                "rc": rc, "desc": desc,
                "count": len(hist), "success": bool(hist),
                "raw": raw_snippet,
            })
            if hist:
                return hist, attempts
        except Exception as e:
            attempts.append({
                "ban": b, "userId": u, "userType": ut, "label": lbl,
                "rc": "exception", "count": 0,
                "error": str(e), "success": False, "desc": str(e), "raw": "",
            })
    return [], attempts

REPEAT_TASK_TYPES = {
    "שימור לקוח",
    "לא צופה ולא גולש",
    "לקוח מושבת ללא אינטרנט",
    "תקלת גלישה תשתית סיבים",
    "תקלה בקו טלפון",
}

def check_recent_visit(visits, current_task_type="", days=30, repeat_types=None):
    """מחזיר שם טכנאי מהביקור האחרון אם יש חזורת, אחרת None.
    בנק פקעות: המשימה הנוכחית V-מסומנת → חפש ביקור קודם כלשהו בתוך 30 יום.
    repeat_types: set של סוגים V-מסומנים מהמחירון; אם None — fallback ל-REPEAT_TASK_TYPES."""
    tt = (current_task_type or "").strip()
    # בדוק שסוג המשימה הנוכחית מסומן V
    if repeat_types is not None:
        if tt not in repeat_types:
            return None
    else:
        if not any(rt in tt for rt in REPEAT_TASK_TYPES):
            return None
    from datetime import timedelta
    cutoff = datetime.now() - timedelta(days=days)
    best_dt, best_tech = None, None
    for v in visits:
        raw = v.get("fullDateTimeTo", "") or v.get("dateEnd", "")
        try:
            s = str(raw).strip()
            if not s: continue
            if " " in s: dt = datetime.strptime(s[:16], "%d/%m/%Y %H:%M")
            elif "T" in s: dt = datetime.strptime(s[:19], "%Y-%m-%dT%H:%M:%S")
            elif "-" in s[:4]: dt = datetime.strptime(s[:10], "%Y-%m-%d")
            else: dt = datetime.strptime(s[:10], "%d/%m/%Y")
            if dt >= cutoff:
                if best_dt is None or dt > best_dt:
                    best_dt = dt
                    best_tech = v.get("technicianName", "") or "לא ידוע"
        except:
            continue
    return best_tech if best_tech else None

def format_visit_history(visits):
    if not visits:
        return ""
    lines = []
    cutoff = datetime.now().replace(hour=0, minute=0, second=0)
    from datetime import timedelta
    cutoff -= timedelta(days=90)
    for v in visits[:5]:
        raw = v.get("fullDateTimeTo", "") or v.get("dateEnd", "")
        try:
            s = str(raw).strip()
            if " " in s: dt = datetime.strptime(s[:16], "%d/%m/%Y %H:%M")
            elif "T" in s: dt = datetime.strptime(s[:19], "%Y-%m-%dT%H:%M:%S")
            elif "-" in s[:4]: dt = datetime.strptime(s[:10], "%Y-%m-%d")
            else: dt = datetime.strptime(s[:10], "%d/%m/%Y")
        except:
            dt = None
        date_str = v.get("dateEnd", "")
        tech = v.get("technicianName", "")
        status = v.get("status", "")
        vtype = v.get("visitType", "")
        import re as _re
        def _is_address_line(line):
            """מחזיר True אם השורה נראית ככתובת — לסינון מההיסטוריה"""
            l = line.strip()
            if not l: return True
            # מספר לקוח: PI / JET + ספרות
            if _re.match(r'^(PI|JET|pi|jet)\d+', l): return True
            # מספר טהור (מספר קריאה וכד')
            if _re.match(r'^\d{4,}$', l): return True
            # כתובת: מכיל קומה/דירה/כניסה
            if any(w in l for w in ['קומה','דירה','כניסה','ללא כניסות']): return True
            # קוד קו
            if l.startswith('קוד קו'): return True
            return False

        reason_lines = []
        for i in v.get("visitInfo", []):
            raw_r = i.get("visitCloseReason", "")
            if not raw_r: continue
            for sub in raw_r.split("\n"):
                if not _is_address_line(sub):
                    reason_lines.append(sub.strip())
        reasons = " / ".join(reason_lines) if reason_lines else ""
        parts = [p for p in [date_str, vtype, status, tech, reasons] if p]
        lines.append("  • " + " | ".join(parts))
    return "\n".join(lines)

_card_cache = {}  # call_id -> (timestamp, card)
CARD_CACHE_TTL = 300  # 5 דקות

def build_single_card(task, token, index, phone=None):
    import time as _time
    from concurrent.futures import ThreadPoolExecutor
    call_id = task.get("call_id", "")
    now = _time.time()
    if call_id and call_id in _card_cache:
        ts, cached = _card_cache[call_id]
        if now - ts < CARD_CACHE_TTL:
            cached = dict(cached)
            cached["index"] = index
            return cached
    try:
        start_dt = datetime.fromisoformat(task.get("start_date", "").replace("+03:00", ""))
        end_dt   = datetime.fromisoformat(task.get("end_date", "").replace("+03:00", ""))
        time_str = f"{start_dt.strftime('%H:%M')}–{end_dt.strftime('%H:%M')}"
        date_str = start_dt.strftime("%d/%m/%Y")
    except:
        time_str = date_str = ""
    street  = task.get("street", "")
    home_no = task.get("home_no", "")
    apt     = task.get("apartment_no", "")
    address = f"{street} {home_no}" + (f" דירה {apt}" if apt and apt != "1" else "")
    city    = task.get("city", "")
    district = task.get("district", "")
    cid     = task.get("customer_id", "")
    source  = task.get("system_source", "JET")
    # הפעלה מקבילית של שתי הקריאות לAPI
    with ThreadPoolExecutor(max_workers=2) as ex:
        f_body   = ex.submit(fetch_call_details, token, call_id, cid, source)
        f_visits = ex.submit(fetch_visit_history, token, task.get("ban", cid), task.get("user_id", cid), source)
        body   = f_body.result()
        visits = f_visits.result()
    eq      = parse_equipment(body) if body and not body.get("_error") else {}
    # repeat_types מהמחירון (אם יש phone)
    _rpt = None
    if phone:
        _prices = load_prices(phone)
        _rpt = {k.replace('__repeat','') for k,v in _prices.items() if k.endswith('__repeat') and v}
    card = {
        "index": index,
        "call_id":    call_id,
        "date":       date_str, "time": time_str,
        "name":       task.get("contact_name", ""),
        "customer_id": cid,
        "phone":      task.get("contact_phone", ""),
        "address":    f"{address}, {city}" + (f" ({district})" if district else ""),
        "city":       city,
        "task_type":  task.get("task_type", ""),
        "infra":      eq.get("infra_name", ""),
        "technology": eq.get("technology", ""),
        "planned":    eq.get("planned", ""),
        "existing":   eq.get("existing", ""),
        "tv":         eq.get("tv", ""),
        "line_code":      eq.get("line_code", ""),
        "internet_user":  eq.get("internet_user", ""),
        "router_model":   eq.get("router_model", ""),
        "router_serial":  eq.get("router_serial", ""),
        "comment":      task.get("comment_text", "") or "",
        "history":      format_visit_history(visits),
        "recent_visit": check_recent_visit(visits, task.get("task_type", ""), repeat_types=_rpt),
        "structure_type": (task.get("structureType", "") or task.get("structure_type", "") or "").strip(),
        "lat": task.get("latitude", "") or task.get("lat", ""),
        "lng": task.get("longitude", "") or task.get("lng", ""),
        "ibc_url": eq.get("ibc_url", ""),
        "infra_code": eq.get("infra_code", ""),
    }
    if call_id:
        _card_cache[call_id] = (_time.time(), card)
    return card

# ============================================================
# Message Handler
# ============================================================
def handle(text):
    text = text.strip()

    # שלב הגדרה
    if session.get("step") == "setup_phone":
        if not text.replace("-", "").isdigit():
            return "❌ מספר לא תקין. נסה שוב:"
        session["phone"] = text.replace("-", "")
        session["step"] = "setup_employee"
        return "מה מספר העובד שלך?"

    if session.get("step") == "setup_employee":
        if not text.isdigit():
            return "❌ מספר עובד לא תקין. נסה שוב:"
        session["employee_id"] = text
        session["step"] = "idle"
        # בדוק אם יש טוקן שמור תקף
        saved = get_valid_token(session["phone"])
        if saved:
            session["token"] = saved
            return "✅ נרשמת! נמצא טוקן שמור מהיום — מחובר אוטומטית 🔓\nשלח *פקעות* להתחיל."
        return "✅ נרשמת!\nשלח *טוקן* להתחברות."

    # OTP
    if session.get("step") == "awaiting_otp" and text.isdigit() and 4 <= len(text) <= 8:
        try:
            token = login_step2(session["phone"], session["employee_id"], text, session["ticket_id"])
            session["token"] = token
            session["step"] = "idle"
            save_token_for_phone(session["phone"], token)
            return "✅ מחובר! הטוקן נשמר עד 23:59 הלילה.\nשלח *פקעות* להתחיל."
        except Exception as e:
            return f"❌ {e}\nשלח *טוקן* לקוד חדש."

    # בכניסה חדשה — נסה לטעון טוקן שמור
    if session.get("phone") and not session.get("token"):
        saved = get_valid_token(session["phone"])
        if saved:
            session["token"] = saved

    if text in ("טוקן", "התחבר"):
        if not session.get("phone"):
            session["step"] = "setup_phone"
            return "👋 ברוך הבא!\nמה מספר הטלפון שלך? (למשל 0526845629)"
        # בדוק טוקן שמור תקף
        saved = get_valid_token(session["phone"])
        if saved:
            session["token"] = saved
            return "🔓 נמצא טוקן תקף מהיום — מחובר!\nשלח *פקעות* להתחיל."
        try:
            tid = login_step1(session["phone"], session["employee_id"])
            session["ticket_id"] = tid
            session["step"] = "awaiting_otp"
            return "📱 נשלח SMS — שלח את הקוד:"
        except Exception as e:
            return f"❌ {e}"

    if text in ("עזרה", "help"):
        return (
            "• *פקעות* — כל הפקעות\n"
            "• *פקעות [מחוז]* — לפי מחוז\n"
            "• *עזרה* — תפריט זה"
        )

    if text == "/reset":
        session.clear()
        return "🔄 אופס. שלח *טוקן* להתחלה."

    return "שלח *עזרה* לרשימת פקודות."

# ============================================================
# Routes
# ============================================================
@app.route("/logo")
def serve_logo():
    logo_path = os.path.join(os.path.dirname(__file__), "logo.png")
    if os.path.exists(logo_path):
        return send_file(logo_path, mimetype="image/png")
    return "", 404

@app.route("/")
def index():
    return HTML.replace('__ADMIN_PHONE__', ADMIN_PHONE)

@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json(force=True) or {}
    phone = data.get("phone", "").replace("-", "").strip()
    employee_id = data.get("employee_id", "").strip()
    # קבל טוקן מה-client אם נשלח (ניסיון התחברות שקטה)
    client_token = data.get("token", "").strip()
    if not phone or not employee_id:
        return jsonify({"status": "error", "message": "מלא את כל השדות"})
    # ── לוגין אדמין — SMS נשלח לטלפון האמיתי, נכנס כ-admin ──
    ADMIN_REAL_PHONE = "0526845629"   # הטלפון שמקבל SMS
    if phone == ADMIN_PHONE and employee_id == ADMIN_EMPLOYEE_ID:
        # רשום admin ב-DB אם לא קיים
        conn = sqlite3.connect(DB_PATH)
        conn.execute(
            "INSERT OR IGNORE INTO users (phone,name,employee_id,registered_at,trial_expires,is_admin) "
            "VALUES (?,?,?,?,?,?)",
            (ADMIN_PHONE, "מנהל", ADMIN_EMPLOYEE_ID, date.today().isoformat(), "2099-12-31", 1)
        )
        conn.commit(); conn.close()
        # שמור בסשן: phone=admin, אבל שלח SMS לטלפון האמיתי
        session["phone"] = ADMIN_PHONE
        session["employee_id"] = ADMIN_EMPLOYEE_ID
        session["admin_real_phone"] = ADMIN_REAL_PHONE
        try:
            ticket_id = login_step1(ADMIN_REAL_PHONE, ADMIN_EMPLOYEE_ID)
            session["ticket_id"] = ticket_id
            return jsonify({"status": "otp"})
        except Exception as e:
            return jsonify({"status": "error", "message": str(e)})
    # שמור בסשן רק למהלך OTP
    session["phone"] = phone
    session["employee_id"] = employee_id
    # אם הלקוח שלח טוקן — תמיד תקין (הוא שמר אותו)
    if client_token:
        return jsonify({"status": "ok", "token": client_token, "phone": phone})
    # שלח SMS
    try:
        ticket_id = login_step1(phone, employee_id)
        session["ticket_id"] = ticket_id
        return jsonify({"status": "otp"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

@app.route("/api/verify", methods=["POST"])
def api_verify():
    otp = request.json.get("otp", "").strip()
    phone = session.get("phone")
    employee_id = session.get("employee_id")
    ticket_id = session.get("ticket_id")
    admin_real_phone = session.get("admin_real_phone", "")
    if not all([phone, employee_id, ticket_id, otp]):
        return jsonify({"status": "error", "message": "חסר מידע"})
    try:
        # לאדמין — אמת מול הטלפון האמיתי אבל החזר phone=ADMIN_PHONE
        verify_phone = admin_real_phone if admin_real_phone else phone
        token = login_step2(verify_phone, employee_id, otp, ticket_id)
        save_token_for_phone(phone, token)
        return jsonify({"status": "ok", "token": token, "phone": phone})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"status": "ok"})

@app.route("/api/status")
def api_status():
    token, phone = get_auth()
    if token and phone:
        sub_status, sub_expires = get_sub_status(phone)
        user = db_get_user(phone)
        is_adm = (phone == ADMIN_PHONE) or bool(user and user.get('is_admin'))
        perms = get_effective_permissions(phone)
        return jsonify({
            "logged_in": True,
            "phone": phone,
            "sub_status": sub_status,
            "sub_expires": sub_expires,
            "is_admin": is_adm,
            "user_name": user.get('name', '') if user else '',
            "permissions": perms
        })
    return jsonify({"logged_in": False})

@app.route("/chat", methods=["POST"])
def chat():
    token, phone = get_auth()
    if not token:
        return jsonify({"type": "text", "text": "⚠️ לא מחובר. רענן את הדף."})
    msg = request.json.get("message", "")
    response = handle(msg)
    if isinstance(response, dict):
        return jsonify(response)
    return jsonify({"type": "text", "text": response})

@app.route("/api/today-schedule")
def api_today_schedule():
    """שליפת לוז לפי תאריך מ-getSchedules — מקבל date=YYYY-MM-DD מה-client"""
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    date_param = request.args.get("date", "").strip()
    try:
        target = date.fromisoformat(date_param) if date_param else date.today()
    except:
        target = date.today()
    apts = fetch_schedules_api(token, target)
    cards = []
    for apt in (apts or []):
        task = apt.get("task", {})
        cd   = apt.get("callDetails", {})
        asgn = apt.get("assignment", {})
        st   = task.get("structureType", "").strip()
        cards.append({
            "call_id":    task.get("callId", ""),
            "name":       task.get("contactName", ""),
            "phone":      task.get("contactPhoneNumber", ""),
            "address":    f"{task.get('street','')} {task.get('city','')}".strip(),
            "task_type":  task.get("taskType", ""),
            "status":     task.get("status", {}).get("displayString", "") if isinstance(task.get("status"), dict) else "",
            "time_start": task.get("formattedAppointmentStart", ""),
            "time_end":   task.get("formattedAppointmentFinish", ""),
            "comment":    task.get("comment") or asgn.get("comment") or "",
            "infra":      INFRA_MAP.get(cd.get("infrastructure",""), cd.get("infrastructure","")),
            "infra_code": cd.get("infrastructure",""),
            "structure_type": st,
            "is_vip":       task.get("isVIP","0") == "1",
            "lat":          task.get("latitude","") or task.get("lat",""),
            "lng":          task.get("longitude","") or task.get("lng",""),
            "customer_id":  task.get("customer","") or task.get("customerId",""),
            "source":       cd.get("sourceSystem","") or "JET",
            "ban":          task.get("ban","") or cd.get("ban","") or "",
            "system_source": cd.get("sourceSystem","") or "JET",
        })
    return jsonify({"schedule": cards, "count": len(cards), "date": target.strftime("%d/%m/%Y")})

def build_basic_card(task, index):
    """כרטיסיה בסיסית — ללא קריאות API, רק מהנתונים שב-getPotentialTasks"""
    try:
        start_dt = datetime.fromisoformat(task.get("start_date","").replace("+03:00",""))
        end_dt   = datetime.fromisoformat(task.get("end_date","").replace("+03:00",""))
        time_str = f"{start_dt.strftime('%H:%M')}–{end_dt.strftime('%H:%M')}"
        date_str = start_dt.strftime("%d/%m/%Y")
    except:
        time_str = date_str = ""
    street   = task.get("street","")
    home_no  = task.get("home_no","")
    apt      = task.get("apartment_no","")
    address  = f"{street} {home_no}" + (f" דירה {apt}" if apt and apt != "1" else "")
    city     = task.get("city","")
    district = task.get("district","")
    cid      = task.get("customer_id","")
    return {
        "index": index, "_num": index,
        "call_id":    task.get("call_id",""),
        "date": date_str, "time": time_str,
        "name":       task.get("contact_name",""),
        "customer_id": cid,
        "phone":      task.get("contact_phone",""),
        "address":    f"{address}, {city}" + (f" ({district})" if district else ""),
        "city": city,
        "task_type":  task.get("task_type",""),
        "comment":    task.get("comment_text","") or "",
        "structure_type": (task.get("structureType","") or task.get("structure_type","") or "").strip(),
        "lat": task.get("latitude","") or task.get("lat",""),
        "lng": task.get("longitude","") or task.get("lng",""),
        "ban":        task.get("ban",""),
        "system_source": task.get("system_source","JET"),
        "appointment_type": task.get("appointmentType","") or task.get("appointment_type",""),
        "infra":"", "technology":"", "planned":"", "existing":"",
        "tv":"", "history":"", "recent_visit":None, "ibc_url":"",
        "line_code":"", "internet_user":"", "router_model":"", "router_serial":"",
    }

@app.route("/api/tasks")
def api_tasks():
    import traceback as _tb
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    district = request.args.get("district", "")
    try:
        tasks = fetch_tasks(token)
        if not tasks:
            return jsonify({"tasks": [], "cards": [], "count": 0,
                            "info": "לא נמצאו פקעות — ייתכן שהטוקן פג תוקף"})
        if district:
            tasks = [t for t in tasks if isinstance(t, dict) and district in t.get("district", "")]
        basic_cards = []
        for i, t in enumerate(tasks):
            try:
                basic_cards.append(build_basic_card(t, i+1))
            except Exception as ce:
                basic_cards.append({"index": i+1, "_num": i+1, "call_id": str(t)[:50],
                                     "error": str(ce), "name":"", "address":"", "city":"",
                                     "task_type":"", "comment":"", "lat":"", "lng":"",
                                     "ban":"", "appointment_type":"", "structure_type":"",
                                     "system_source":"", "date":"", "time":"",
                                     "phone":"", "customer_id":"", "infra":"", "technology":"",
                                     "planned":"", "existing":"", "tv":"", "history":"",
                                     "recent_visit":None, "ibc_url":"", "line_code":"",
                                     "internet_user":"", "router_model":"", "router_serial":""})
        return jsonify({"tasks": tasks, "cards": basic_cards, "count": len(tasks)})
    except Exception as e:
        return jsonify({"error": f"שגיאה בשליפת פקעות: {e}", "trace": _tb.format_exc()}), 500

@app.route("/api/prices", methods=["GET"])
def api_get_prices():
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    prices = load_prices(phone)
    task_types = get_all_task_types()
    history = load_prices_history(phone)
    current_from = history[-1]["effective_from"] if history else None
    return jsonify({"prices": prices, "task_types": task_types, "effective_from": current_from, "history": history})

@app.route("/api/prices", methods=["POST"])
def api_save_prices():
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    prices = request.json.get("prices", {})
    effective_from = request.json.get("effective_from") or None
    save_prices(prices, phone, effective_from=effective_from)
    return jsonify({"status": "ok"})

def _overrides_file(phone):
    os.makedirs(PRICES_DIR, exist_ok=True)
    safe = "".join(c if c.isalnum() else "_" for c in (phone or "default"))
    return os.path.join(PRICES_DIR, f"overrides_{safe}.json")

def load_overrides(phone):
    path = _overrides_file(phone)
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_overrides(overrides, phone):
    path = _overrides_file(phone)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(overrides, f, ensure_ascii=False, indent=2)

@app.route("/api/reports/set-override", methods=["POST"])
def api_set_override():
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    data = request.json
    call_id   = data.get("call_id", "").strip()
    fetch_date = data.get("fetch_date", "").strip()
    price_type = data.get("price_type", "")      # "PPC" / "LSB" / "custom"
    custom_price = data.get("custom_price", None) # float or null
    if not call_id or not fetch_date:
        return jsonify({"error": "חסר מידע"}), 400
    key = f"{call_id}__{fetch_date}"
    overrides = load_overrides(phone)
    overrides[key] = {"price_type": price_type, "custom_price": custom_price}
    save_overrides(overrides, phone)
    return jsonify({"status": "ok"})

@app.route("/api/diag/proxy", methods=["POST"])
def diag_proxy():
    """פרוקסי מאובטח לגישה למכשירים ברשת מקומית (MOCA / נתב / AirTies)"""
    import re
    token, _ = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    body = request.json or {}
    url      = body.get("url", "")
    method   = body.get("method", "GET").upper()
    data     = body.get("data")        # dict → form-encoded POST
    json_body= body.get("json")        # dict → JSON POST
    auth     = body.get("auth")        # [user, pass]
    hdrs     = body.get("headers", {})

    # אבטחה: רק רשתות פרטיות
    if not re.match(r'https?://(10\.|192\.168\.|172\.(1[6-9]|2[0-9]|3[01])\.)', url):
        return jsonify({"error": "כתובת לא מאושרת — רשת פרטית בלבד"}), 403

    try:
        auth_t = tuple(auth) if auth else None
        kw = dict(headers=hdrs, auth=auth_t, timeout=12, verify=False,
                  allow_redirects=True)
        if method == "POST":
            if json_body is not None:
                r = requests.post(url, json=json_body, **kw)
            elif data is not None:
                r = requests.post(url, data=data, **kw)
            else:
                r = requests.post(url, **kw)
        else:
            r = requests.get(url, **kw)
        return jsonify({"status": r.status_code, "text": r.text[:40000],
                        "url": r.url, "ct": r.headers.get("Content-Type","")})
    except requests.Timeout:
        return jsonify({"error": "timeout — מכשיר לא מגיב"}), 504
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/reports")
def api_reports():
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    days = min(max(int(request.args.get("days", 30)), 1), 366)
    end_str = request.args.get("end", "")
    end_date   = date.fromisoformat(end_str) if end_str else date.today()
    start_date = end_date - timedelta(days=days-1)
    conn = init_db()
    fetched_new = 0
    today = date.today()
    cur_month_start = date(today.year, today.month, 1)

    def _apt_time(iso_str):
        try:
            return datetime.fromisoformat(iso_str.replace("+03:00","")).strftime("%H:%M")
        except Exception:
            return iso_str or ""

    def _task_to_apt(t):
        return {
            "task": {
                "callId":                    t.get("call_id","") or t.get("callId",""),
                "contactName":               t.get("contact_name","") or t.get("contactName",""),
                "contactPhoneNumber":        t.get("contact_phone","") or t.get("contactPhoneNumber",""),
                "taskType":                  t.get("task_type","") or t.get("taskType",""),
                "street":                    t.get("street",""),
                "city":                      t.get("city",""),
                "formattedAppointmentStart": _apt_time(t.get("start_date","")),
                "formattedAppointmentFinish":_apt_time(t.get("end_date","")),
                "status":                    t.get("status",""),
                "isVIP":                     "1" if t.get("is_vip") else "0",
                "ban":                       t.get("ban",""),
                "customer":                  t.get("customer_id","") or t.get("customer",""),
                "userId":                    t.get("user_id",""),
                "system_source":             t.get("system_source","JET"),
            },
            "callDetails": {}
        }

    # שלוף ימים במקביל (עד 8 threads — לא להעמיס על Cellcom API)
    days_range = []
    cur = start_date
    while cur <= end_date:
        days_range.append(cur)
        cur += timedelta(days=1)

    yesterday = today - timedelta(days=1)

    def _fetch_day(d):
        """מחזיר (date, appointments_or_None, needs_delete)
        כל ימי החודש הנוכחי תמיד מרועננים מה-API — אפס פספוסים."""
        if d >= cur_month_start:
            # תמיד מרענן את כל ימי החודש (כולל היום, אתמול וכל השאר)
            apts = fetch_schedules_api(token, d)
            if not apts and d == today:
                try:
                    raw = fetch_tasks(token)
                    if raw:
                        apts = [_task_to_apt(t) for t in raw]
                except Exception:
                    apts = []
            return (d, apts, True)  # needs_delete=True — תמיד מחליף
        return (d, None, False)  # חודשים קודמים — DB בלבד

    with ThreadPoolExecutor(max_workers=2) as ex:
        results = list(ex.map(_fetch_day, days_range))

    # שמור לDB בסדר כרונולוגי (לא במקביל — SQLite)
    for d, apts, needs_delete in results:
        if apts is None and not needs_delete:
            # יום ריק מה-API — סמן כנטען
            if d >= cur_month_start and not date_already_fetched(conn, d, phone):
                conn.execute("INSERT OR IGNORE INTO visits(phone,call_id,fetch_date) VALUES(?,?,?)",
                             (phone, f"EMPTY_{d.isoformat()}", d.isoformat()))
                conn.commit()
        elif apts:
            if needs_delete:
                conn.execute("DELETE FROM visits WHERE phone=? AND fetch_date=?",
                             (phone, d.isoformat()))
                conn.commit()
            save_day_to_db(conn, apts, d, phone)
            fetched_new += len(apts)
    visits = get_visits_for_range(conn, start_date, end_date, phone)
    visits = [v for v in visits if not v["call_id"].startswith("EMPTY_")]
    # 30 ימים לפני — לזיהוי חוזרות שמקורן לפני תחילת הדוח
    pre_start = start_date - timedelta(days=30)
    pre_visits = get_visits_for_range(conn, pre_start, start_date - timedelta(days=1), phone)
    pre_visits = [v for v in pre_visits if not v["call_id"].startswith("EMPTY_")]
    # 30 ימים אחרי — לזיהוי ביקורים בדוח שהובילו לחוזרת בעתיד
    post_end = end_date + timedelta(days=30)
    post_visits = get_visits_for_range(conn, end_date + timedelta(days=1), post_end, phone)
    post_visits = [v for v in post_visits if not v["call_id"].startswith("EMPTY_")]
    conn.close()
    overrides = load_overrides(phone)
    # טען את היסטוריית המחירים פעם אחת — נשתמש בה לכל ביקור לפי תאריכו
    _prices_cache = {}
    def _prices_for_visit(visit_date_str):
        if visit_date_str not in _prices_cache:
            _prices_cache[visit_date_str] = load_prices_for_date(phone, visit_date_str)
        return _prices_cache[visit_date_str]
    # הוסף מחיר לכל ביקור
    total_earnings = 0
    summary = defaultdict(lambda: {"count": 0, "earnings": 0})
    for v in visits:
        tt = v["task_type"] or ""
        is_incomplete = "לא הושלם" in (v.get("status") or "") or "לא הושלמ" in (v.get("status") or "")
        v["is_incomplete"] = is_incomplete
        ov_key = f"{v['call_id']}__{v['fetch_date']}"
        ov = overrides.get(ov_key, {})
        prices = _prices_for_visit(v.get("fetch_date", "")[:10])
        if is_incomplete:
            price = 0.0
            v["price_type"] = ""
            v["custom_price"] = None
        elif ov.get("price_type") == "custom" and ov.get("custom_price") is not None:
            price = float(ov["custom_price"])
            v["price_type"] = "custom"
            v["custom_price"] = price
        elif ("תשתית סיבים" in tt or "טריפל סיבים" in tt) and "תקלת גלישה" not in tt:
            forced = ov.get("price_type")
            if forced == "LSB" or (not forced and v.get("lsb_flag") == 1):
                price = float(prices.get(tt + "__LSB", 0))
                v["price_type"] = "LSB"
            else:
                price = float(prices.get(tt + "__PPC", 0))
                v["price_type"] = "PPC"
            v["custom_price"] = None
        else:
            price = float(prices.get(tt, 0))
            v["price_type"] = ""
            v["custom_price"] = None
        v["price"] = price
        total_earnings += price
        if tt:
            summary[tt]["count"] += 1
            summary[tt]["earnings"] += price
    summary_sorted = dict(sorted(summary.items(), key=lambda x: -x[1]["count"]))
    def _slim(v):
        return {"customer_id": v["customer_id"], "fetch_date": v["fetch_date"],
                "appt_start": v.get("appt_start",""), "task_type": v["task_type"]}
    return jsonify({
        "visits": visits,
        "pre_visits":  [_slim(v) for v in pre_visits],
        "post_visits": [_slim(v) for v in post_visits],
        "total": len(visits),
        "total_earnings": total_earnings,
        "summary": summary_sorted,
        "prices_map": load_prices(phone),
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "fetched_new": fetched_new,
    })

@app.route("/api/ibc-url")
def api_ibc_url():
    """שולף ibcBuildingUrl מ-getCallDetails לפי call_id"""
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    call_id     = request.args.get("call_id", "").strip()
    customer_id = request.args.get("customer_id", "").strip()
    source      = request.args.get("source", "JET").strip()
    if not call_id:
        return jsonify({"error": "חסר call_id"}), 400
    body = fetch_call_details(token, call_id, customer_id, source)
    if not body or body.get("_error"):
        err = (body or {}).get("_error", "לא נמצאו נתונים")
        # נסה שוב עם source ריק אם נכשל
        if source and source != "":
            body2 = fetch_call_details(token, call_id, customer_id, "")
            if body2 and not body2.get("_error"):
                body = body2
            else:
                return jsonify({"error": f"getCallDetails: {err}"}), 404
        else:
            return jsonify({"error": f"getCallDetails: {err}"}), 404
    ibc_url = _find_ibc(body)
    if not ibc_url:
        return jsonify({"error": "אין קישור IBC בקריאה זו"}), 404
    return jsonify({"ibc_url": ibc_url})

@app.route("/api/ibc-data")
def api_ibc_data():
    """שולף נתוני בניין מ-IBC API ישירות לפי addrid + token מתוך ibcBuildingUrl"""
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    ibc_url = request.args.get("url", "").strip()
    if not ibc_url or "unlimited.net.il" not in ibc_url:
        return jsonify({"error": "כתובת לא תקינה"}), 400
    # חילוץ token ו-addrid מה-URL
    from urllib.parse import urlparse, parse_qs
    parsed = urlparse(ibc_url)
    qs = parse_qs(parsed.query)
    ibc_token = qs.get("token", [""])[0]
    addrid    = qs.get("addrid", [""])[0]
    if not ibc_token or not addrid:
        return jsonify({"error": "חסר token או addrid ב-URL"}), 400
    try:
        headers = {
            "Content-Type": "application/json",
            "Accept": "application/json, text/plain, */*",
            "access-control-allow-origin": "*",
            "access-control-allow-methods": "POST",
            "token": ibc_token,
            "origin": "https://ussext-tech.unlimited.net.il",
            "referer": f"https://ussext-tech.unlimited.net.il/asset-details?addrid={addrid}",
            "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 18_7 like Mac OS X) AppleWebKit/605.1.15 Mobile/15E148 Safari/604.1",
            "username": "undefined",
        }
        body = {
            "Body": {
                "ApiCommand": "ussAPI_SearchAsset",
                "RequestBody": {"searchStr": addrid}
            },
            "Header": {
                "ProccesId": "", "ActivityId": "", "InvokerActivityId": "",
                "ApiCommand": "ussAPI_SearchAsset", "ApiCode": "",
                "From": "IBC", "To": "", "ErrorCode": "", "ErrorCodeDesc": "",
                "AddInfo": "", "UserName": ""
            }
        }
        r = requests.post(
            "https://ussext-tech.unlimited.net.il/endpoint/api/Endpoint/",
            headers=headers, json=body, timeout=10, verify=False
        )
        data = r.json()
        content = data.get("Content", [])
        if not content:
            return jsonify({"error": "לא נמצאו נתונים"}), 404
        d = content[0]
        main = d.get("mainAddrData") or d
        result = {
            "site_type":    d.get("SiteType", ""),
            "aron_num":     d.get("AronNum", ""),
            "aron_location":d.get("AronLocation", ""),
            "floor":        d.get("LocationFloor", ""),
            "floors_total": d.get("Floors", ""),
            "apartments":   d.get("Apartments", ""),
            "eqp_type":     (d.get("EqpType") or "").strip(),
            "technology":   d.get("TechnologyType", ""),
            "status":       d.get("Status", ""),
        }
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/geocode")
def api_geocode():
    addr = request.args.get("q", "").strip()
    if not addr:
        return jsonify({"error": "no address"}), 400
    conn = init_db()
    conn.execute("CREATE TABLE IF NOT EXISTS geocache (addr TEXT PRIMARY KEY, lat REAL, lng REAL)")
    row = conn.execute("SELECT lat,lng FROM geocache WHERE addr=?", (addr,)).fetchone()
    if row:
        conn.close()
        return jsonify({"lat": row[0], "lng": row[1]})
    try:
        import urllib.parse, urllib.request as ureq
        url = "https://nominatim.openstreetmap.org/search?q=" + urllib.parse.quote(addr + ", ישראל") + "&format=json&limit=1"
        req = ureq.Request(url, headers={"User-Agent": "CellcomApp/1.0"})
        with ureq.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read())
        if data:
            lat, lng = float(data[0]["lat"]), float(data[0]["lon"])
            conn.execute("INSERT OR REPLACE INTO geocache(addr,lat,lng) VALUES(?,?,?)", (addr, lat, lng))
            conn.commit()
            conn.close()
            return jsonify({"lat": lat, "lng": lng})
    except Exception as e:
        pass
    conn.close()
    return jsonify({"error": "not found"}), 404

@app.route("/api/reports/export")
def api_reports_export():
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    days    = int(request.args.get("days", 30))
    end_str = request.args.get("end", "")
    end_date   = date.fromisoformat(end_str) if end_str else date.today()
    start_date = end_date - timedelta(days=days - 1)
    conn   = init_db()
    visits = get_visits_for_range(conn, start_date, end_date, phone)
    visits = [v for v in visits if not v["call_id"].startswith("EMPTY_")]
    conn.close()
    prices = load_prices(phone)
    for v in visits:
        tt = v["task_type"] or ""
        is_incomplete = "לא הושלם" in (v.get("status") or "") or "לא הושלמ" in (v.get("status") or "")
        v["is_incomplete"] = is_incomplete
        if is_incomplete:
            v["price"] = 0.0
            v["price_type"] = ""
        elif ("תשתית סיבים" in tt or "טריפל סיבים" in tt) and "תקלת גלישה" not in tt:
            if v.get("lsb_flag") == 1:
                v["price"] = float(prices.get(tt + "__LSB", 0))
                v["price_type"] = "LSB"
            else:
                v["price"] = float(prices.get(tt + "__PPC", 0))
                v["price_type"] = "PPC"
        else:
            v["price"] = float(prices.get(tt, 0))
            v["price_type"] = ""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        import subprocess
        subprocess.run(
            [os.path.join(os.path.dirname(__file__), "venv", "bin", "pip"),
             "install", "openpyxl", "-q"], check=False)
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

    # שם מתקין
    installer_phone = phone
    installer_user  = db_get_user(installer_phone)
    installer_name  = (installer_user.get("name") or installer_phone) if installer_user else installer_phone

    BLUE       = "4472C4"
    WHITE      = "FFFFFF"
    HDR_FONT   = Font(color=WHITE, bold=True, size=10)
    HDR_FILL   = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    RIGHT      = Alignment(horizontal="right", vertical="center")
    CENTER_MID = Alignment(horizontal="center", vertical="center")
    HEBREW_MONTHS = {1:"ינואר",2:"פברואר",3:"מרץ",4:"אפריל",5:"מאי",6:"יוני",
                     7:"יולי",8:"אוגוסט",9:"ספטמבר",10:"אוקטובר",11:"נובמבר",12:"דצמבר"}

    headers = [
        "תאריך", "שם טכנאי", "מספר לקוח PSID", "מספר פקע",
        "הערות טכנאי",
        "טריפל", "טי וי בלבד", "התקנת אפליקציה",
        'התקנת שוס/מפא או מפא+שוס', "שינוי שרות",
        "פקע קלה (התקנת סופר WIFI)", "פקע קלה (שדרוג נתב)",
        "קריאת שרות מכל סוג", "הסבות שקטות", "אזור מיוחד",
        "התקנת PPC", "התקנת בניין נמוך LSB",
        "התקנת VILA טריפל", 'התקנת VILA ש"וס',
        "שם מתקין", 'סה"כ תשלום למתקין'
    ]
    col_widths = [
        12, 14, 16, 12,
        22,
        7, 8, 10,
        16, 10,
        20, 16,
        14, 12, 10,
        10, 14,
        12, 12,
        12, 12
    ]

    # מיפוי סוג משימה → דגלים בינאריים (עמודות 14–27)
    def task_type_to_flags(tt, price_type):
        flags = [None] * 14  # cols 14-27
        t = tt or ''
        is_vila = 'VILA' in t or 'וילה' in t
        if is_vila:
            if 'שוס' in t or 'מפא' in t or "ש\"וס" in t or "מפ\"א" in t or "שו\"ס" in t or "מפ\"א" in t:
                flags[13] = 1  # col 27: VILA שוס
            else:
                flags[12] = 1  # col 26: VILA טריפל
        elif 'שקט' in t or 'הסב' in t:
            flags[8] = 1   # col 22: הסבות שקטות
        elif 'מיוחד' in t:
            flags[9] = 1   # col 23: אזור מיוחד
        elif 'שינוי' in t:
            flags[4] = 1   # col 18: שינוי שרות
        elif 'אפליקציה' in t or 'app' in t.lower():
            flags[2] = 1   # col 16: התקנת אפליקציה
        elif ('WIFI' in t.upper() or 'ווייפיי' in t) and ('סופר' in t or 'קלה' in t):
            flags[5] = 1   # col 19: פקע קלה - סופר WIFI
        elif 'נתב' in t or ('שדרוג' in t and 'טריפל' not in t):
            flags[6] = 1   # col 20: פקע קלה - שדרוג נתב
        elif 'טריפל' in t:
            flags[0] = 1   # col 14: טריפל
        elif 'TV' in t or 'טלויזיה' in t or 'טלוויזיה' in t or 'tv' in t.lower():
            flags[1] = 1   # col 15: TV בלבד
        elif 'שוס' in t or 'מפא' in t or "ש\"וס" in t or "מפ\"א" in t:
            flags[3] = 1   # col 17: שוס/מפא
        else:
            flags[7] = 1   # col 21: קריאת שרות (ברירת מחדל)
        # PPC / LSB כסימון נוסף
        if price_type == 'PPC':
            flags[10] = 1  # col 24: PPC
        elif price_type == 'LSB':
            flags[11] = 1  # col 25: LSB
        return flags

    RED_FILL = PatternFill(start_color="FFD0D0", end_color="FFD0D0", fill_type="solid")

    def write_sheet(ws, rows_with_flags):
        ws.sheet_view.rightToLeft = True
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = CENTER
        ws.row_dimensions[1].height = 36
        total_price = 0.0
        for ri, (row_data, incomplete) in enumerate(rows_with_flags, 2):
            if incomplete:
                rf = RED_FILL
            else:
                fc = "EEF4FF" if ri % 2 == 0 else "FFFFFF"
                rf = PatternFill(start_color=fc, end_color=fc, fill_type="solid")
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=col, value=val)
                cell.fill = rf
                # דגלים בינאריים (cols 6-19) ממורכזים
                if 6 <= col <= 19:
                    cell.alignment = CENTER_MID
                else:
                    cell.alignment = RIGHT
            total_price += float(row_data[20] or 0)
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
        sr = len(rows_with_flags) + 3
        ws.cell(row=sr, column=1,
                value=f'סה"כ: {len(rows_with_flags)} ביקורים').font = Font(bold=True, size=12)
        if total_price > 0:
            ws.cell(row=sr, column=21, value=total_price).font = Font(bold=True, size=12)

    def make_row(v):
        flags = task_type_to_flags(v.get("task_type", ""), v.get("price_type", ""))
        row = [
            v["fetch_date"],        # 1. תאריך
            installer_name,          # 2. שם טכנאי
            v["call_id"],            # 3. מספר לקוח PSID
            v["contact_phone"],      # 4. מספר פקע
            v["contact_name"],       # 5. הערות טכנאי (שם לקוח)
        ] + flags + [               # 6-19: דגלים בינאריים
            installer_name,          # 20: שם מתקין
            v["price"] if v.get("price") else None,  # 21: סה"כ תשלום
        ]
        return row, v.get("is_incomplete", False)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # כל התקופה
    all_rows = [make_row(v) for v in visits]
    ws_all = wb.create_sheet("כל התקופה")
    write_sheet(ws_all, all_rows)

    # גיליון לכל חודש
    monthly = defaultdict(list)
    for v in visits:
        try:
            d_obj = date.fromisoformat(v["fetch_date"])
            monthly[(d_obj.year, d_obj.month)].append(make_row(v))
        except:
            pass
    for (yr, mo), m_rows in sorted(monthly.items(), reverse=True):
        ws_m = wb.create_sheet(f"{HEBREW_MONTHS[mo]} {yr}")
        write_sheet(ws_m, m_rows)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"cellcom_{start_date}_{end_date}.xlsx"
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name=filename)

@app.route("/api/reports/archive")
def api_archive_list():
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    phone_dir = os.path.join(ARCHIVE_DIR, phone.replace('+', '').replace(' ', ''))
    os.makedirs(phone_dir, exist_ok=True)
    files = []
    for fn in sorted(os.listdir(phone_dir), reverse=True):
        if fn.endswith('.xlsx'):
            fp = os.path.join(phone_dir, fn)
            files.append({
                "name": fn,
                "size": os.path.getsize(fp),
                "mtime": int(os.path.getmtime(fp))
            })
    return jsonify({"files": files})


@app.route("/api/reports/archive/save", methods=["POST"])
def api_archive_save():
    """שמור דוח Excel של חודש לארכיון"""
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    data     = request.json or {}
    days     = int(data.get("days", 30))
    end_str  = data.get("end", "")
    label    = data.get("label", "")   # e.g. "מרץ 2025"
    end_date   = date.fromisoformat(end_str) if end_str else date.today()
    start_date = end_date - timedelta(days=days - 1)

    # ייצר את ה-Excel (אותו לוגיקה כמו export אבל מחזיר bytes)
    conn   = init_db()
    visits = get_visits_for_range(conn, start_date, end_date, phone)
    visits = [v for v in visits if not v["call_id"].startswith("EMPTY_")]
    conn.close()
    prices = load_prices(phone)
    for v in visits:
        tt = v["task_type"] or ""
        is_incomplete = "לא הושלם" in (v.get("status") or "") or "לא הושלמ" in (v.get("status") or "")
        v["is_incomplete"] = is_incomplete
        if is_incomplete:
            v["price"] = 0.0; v["price_type"] = ""
        elif ("תשתית סיבים" in tt or "טריפל סיבים" in tt) and "תקלת גלישה" not in tt:
            if v.get("lsb_flag") == 1:
                v["price"] = float(prices.get(tt + "__LSB", 0)); v["price_type"] = "LSB"
            else:
                v["price"] = float(prices.get(tt + "__PPC", 0)); v["price_type"] = "PPC"
        else:
            v["price"] = float(prices.get(tt, 0)); v["price_type"] = ""

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"error": "openpyxl לא מותקן"}), 500

    from io import BytesIO
    buf = BytesIO()
    # ייצא ישר לbuf — קורא לאותה פונקציה כמו export
    # (שימוש חוזר ב-_build_report_workbook אם קיים, אחרת inline)
    wb = openpyxl.Workbook()
    ws = wb.active
    BLUE = "4472C4"; WHITE = "FFFFFF"
    HDR_FONT  = Font(color=WHITE, bold=True, size=10)
    HDR_FILL  = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    RIGHT     = Alignment(horizontal="right", vertical="center")
    headers   = ["תאריך","שם טכנאי","מספר לקוח PSID","מספר פקע","הערות טכנאי",
                 "טריפל","טי וי בלבד","התקנת אפליקציה","התקנת שוס/מפא או מפא+שוס",
                 "שינוי שרות","פקע קלה (התקנת סופר WIFI)","פקע קלה (שדרוג נתב)",
                 "קריאת שרות מכל סוג","הסבות שקטות","אזור מיוחד",
                 "התקנת PPC","התקנת בניין נמוך LSB","התקנת VILA טריפל",
                 'התקנת VILA ש"וס',"שם מתקין",'סה"כ תשלום למתקין']
    installer_name = (db_get_user(phone) or {}).get("name", phone)
    ws.title = label or f"{start_date} – {end_date}"
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER
    ws.row_dimensions[1].height = 22
    for ri, v in enumerate(visits, 2):
        tt = v["task_type"] or ""
        def flag(cond): return 1 if cond else None
        row_vals = [
            v["fetch_date"], installer_name,
            v.get("customer_id",""), v["call_id"], v.get("notes",""),
            flag("טריפל" in tt and "סיבים" not in tt and "VILA" not in tt and "שקט" not in tt),
            flag("טי וי" in tt or "TV" in tt.upper()),
            flag("אפליקציה" in tt),
            flag("שוס" in tt or "מפא" in tt),
            flag("שינוי" in tt),
            flag("WIFI" in tt.upper() and "קלה" in tt),
            flag("שדרוג" in tt and "קלה" in tt),
            flag("שרות" in tt and "VILA" not in tt and "שקט" not in tt),
            flag("שקט" in tt or "הסב" in tt),
            flag("מיוחד" in tt),
            flag("PPC" in tt.upper() or ("תשתית סיבים" in tt and v.get("price_type")=="PPC")),
            flag("LSB" in tt.upper() or v.get("price_type")=="LSB"),
            flag("VILA" in tt and "שוס" not in tt and 'ש"וס' not in tt),
            flag("VILA" in tt and ("שוס" in tt or 'ש"וס' in tt)),
            installer_name, v.get("price", 0)
        ]
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = CENTER if ci > 4 else RIGHT
    wb.save(buf); buf.seek(0)

    phone_dir = os.path.join(ARCHIVE_DIR, phone.replace('+','').replace(' ',''))
    os.makedirs(phone_dir, exist_ok=True)
    safe_label = (label or f"{start_date}_{end_date}").replace(' ','_').replace('/','').replace('\\','')
    from datetime import datetime as dt
    ts = dt.now().strftime("%Y%m%d_%H%M")
    filename = f"{safe_label}_{ts}.xlsx"
    with open(os.path.join(phone_dir, filename), 'wb') as f:
        f.write(buf.getvalue())
    return jsonify({"ok": True, "filename": filename})


@app.route("/api/reports/archive/<path:filename>")
def api_archive_download(filename):
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    phone_dir = os.path.join(ARCHIVE_DIR, phone.replace('+','').replace(' ',''))
    safe = os.path.basename(filename)
    fp = os.path.join(phone_dir, safe)
    if not os.path.isfile(fp):
        return jsonify({"error": "קובץ לא נמצא"}), 404
    return send_file(fp, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=safe)


@app.route("/api/reports/archive/<path:filename>", methods=["DELETE"])
def api_archive_delete(filename):
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    phone_dir = os.path.join(ARCHIVE_DIR, phone.replace('+','').replace(' ',''))
    safe = os.path.basename(filename)
    fp = os.path.join(phone_dir, safe)
    if os.path.isfile(fp):
        os.remove(fp)
    return jsonify({"ok": True})


@app.route("/api/reports/compare", methods=["POST"])
def api_reports_compare():
    """השוואת קובץ Excel שהועלה מול הנתונים ב-DB — מחזיר פערים"""
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    if 'file' not in request.files:
        return jsonify({"error": "לא הועלה קובץ"}), 400
    f = request.files['file']
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        return jsonify({"error": f"שגיאה בקריאת הקובץ: {e}"}), 400

    if len(rows) < 2:
        return jsonify({"error": "הקובץ ריק"}), 400

    # זהה עמודות לפי כותרת שורה ראשונה
    hdr = [str(c or '').strip() for c in rows[0]]
    def col(name):
        for i, h in enumerate(hdr):
            if name in h:
                return i
        return None

    ci_date    = col("תאריך")
    ci_callid  = col("מספר פקע")
    ci_payment = col('סה"כ')
    ci_type    = col("שם")  # שם טכנאי / שם מתקין

    if ci_callid is None:
        return jsonify({"error": 'לא נמצאה עמודת "מספר פקע" בקובץ'}), 400

    # בנה מילון: call_id → {payment, date, task_type}
    uploaded = {}
    dates_seen = []
    for r in rows[1:]:
        cid = str(r[ci_callid] or '').strip()
        if not cid or cid.startswith("EMPTY"):
            continue
        pay = float(r[ci_payment] or 0) if ci_payment is not None else 0
        d   = str(r[ci_date] or '').strip() if ci_date is not None else ''
        if d:
            try:
                from datetime import date as dclass
                parsed_d = dclass.fromisoformat(d[:10])
                dates_seen.append(parsed_d)
            except: pass
        uploaded[cid] = {"payment": pay, "date": d}

    if not uploaded:
        return jsonify({"error": "לא נמצאו שורות עם מספר פקע בקובץ"}), 400

    # שלוף נתוני DB לאותו טווח תאריכים
    conn = init_db()
    if dates_seen:
        mn = min(dates_seen); mx = max(dates_seen)
    else:
        mn = mx = date.today()
    visits = get_visits_for_range(conn, mn, mx, phone)
    conn.close()
    visits = [v for v in visits if not v["call_id"].startswith("EMPTY_")]
    prices = load_prices(phone)
    for v in visits:
        tt = v["task_type"] or ""
        is_inc = "לא הושלם" in (v.get("status") or "") or "לא הושלמ" in (v.get("status") or "")
        if is_inc:
            v["price"] = 0.0
        elif ("תשתית סיבים" in tt or "טריפל סיבים" in tt) and "תקלת גלישה" not in tt:
            lsb = v.get("lsb_flag") == 1
            v["price"] = float(prices.get(tt + ("__LSB" if lsb else "__PPC"), 0))
        else:
            v["price"] = float(prices.get(tt, 0))

    ours = {v["call_id"]: v for v in visits}

    diffs = []
    # בקובץ המועלה אבל לא אצלנו
    for cid, ud in uploaded.items():
        if cid not in ours:
            diffs.append({"type": "חסר אצלנו", "call_id": cid,
                          "date": ud["date"], "task_type": "",
                          "ours_payment": None, "their_payment": ud["payment"],
                          "diff": None})

    # אצלנו אבל לא בקובץ
    for cid, v in ours.items():
        if cid not in uploaded:
            diffs.append({"type": "חסר בקובץ", "call_id": cid,
                          "date": v["fetch_date"], "task_type": v.get("task_type",""),
                          "ours_payment": v["price"], "their_payment": None,
                          "diff": None})

    # בשניהם — בדוק פער בתשלום
    for cid in set(uploaded) & set(ours):
        op  = ours[cid]["price"]
        tp  = uploaded[cid]["payment"]
        d   = round(op - tp, 2)
        if abs(d) > 0.01:
            diffs.append({"type": "פער בתשלום", "call_id": cid,
                          "date": ours[cid]["fetch_date"],
                          "task_type": ours[cid].get("task_type",""),
                          "ours_payment": op, "their_payment": tp, "diff": d})

    diffs.sort(key=lambda x: (x["type"], x["date"] or ''))
    return jsonify({
        "diffs": diffs,
        "total_uploaded": len(uploaded),
        "total_ours": len(ours),
        "matches": len(set(uploaded) & set(ours)),
        "range": f"{mn} – {mx}"
    })


@app.route("/api/reports/compare-from-archive")
def api_reports_compare_archive():
    """השוואת דוח ארכיון קיים מול נתוני ה-DB"""
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    filename = request.args.get("file", "")
    if not filename or ".." in filename or "/" in filename or "\\" in filename:
        return jsonify({"error": "שם קובץ לא תקין"}), 400
    phone_dir = os.path.join(ARCHIVE_DIR, phone.replace('+','').replace(' ',''))
    filepath  = os.path.join(phone_dir, filename)
    if not os.path.isfile(filepath):
        return jsonify({"error": "הקובץ לא נמצא"}), 404
    # שלח ל-api_reports_compare דרך קריאה פנימית עם הקובץ
    try:
        import openpyxl
        from io import BytesIO
        with open(filepath, "rb") as f:
            raw = f.read()
        wb = openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        return jsonify({"error": f"שגיאה בקריאת הקובץ: {e}"}), 400

    if len(rows) < 2:
        return jsonify({"error": "הקובץ ריק"}), 400

    hdr = [str(c or '').strip() for c in rows[0]]
    def col(name):
        for i, h in enumerate(hdr):
            if name in h: return i
        return None

    ci_date    = col("תאריך")
    ci_callid  = col("מספר פקע")
    ci_payment = col('סה"כ')

    if ci_callid is None:
        return jsonify({"error": 'לא נמצאה עמודת "מספר פקע" בקובץ'}), 400

    uploaded = {}
    dates_seen = []
    for r in rows[1:]:
        cid = str(r[ci_callid] or '').strip()
        if not cid or cid.startswith("EMPTY"): continue
        pay = float(r[ci_payment] or 0) if ci_payment is not None else 0
        d   = str(r[ci_date] or '').strip() if ci_date is not None else ''
        if d:
            try:
                parsed_d = date.fromisoformat(d[:10])
                dates_seen.append(parsed_d)
            except: pass
        uploaded[cid] = {"payment": pay, "date": d}

    if not uploaded:
        return jsonify({"error": "לא נמצאו שורות עם מספר פקע בקובץ"}), 400

    conn = init_db()
    mn = min(dates_seen) if dates_seen else date.today()
    mx = max(dates_seen) if dates_seen else date.today()
    visits = get_visits_for_range(conn, mn, mx, phone)
    conn.close()
    visits = [v for v in visits if not v["call_id"].startswith("EMPTY_")]
    prices = load_prices(phone)
    for v in visits:
        tt = v["task_type"] or ""
        is_inc = "לא הושלם" in (v.get("status") or "") or "לא הושלמ" in (v.get("status") or "")
        if is_inc:
            v["price"] = 0.0
        elif ("תשתית סיבים" in tt or "טריפל סיבים" in tt) and "תקלת גלישה" not in tt:
            lsb = v.get("lsb_flag") == 1
            v["price"] = float(prices.get(tt + ("__LSB" if lsb else "__PPC"), 0))
        else:
            v["price"] = float(prices.get(tt, 0))

    ours = {v["call_id"]: v for v in visits}
    diffs = []
    for cid, ud in uploaded.items():
        if cid not in ours:
            diffs.append({"type": "חסר אצלנו", "call_id": cid,
                          "date": ud["date"], "task_type": "",
                          "ours_payment": None, "their_payment": ud["payment"], "diff": None})
    for cid, v in ours.items():
        if cid not in uploaded:
            diffs.append({"type": "חסר בקובץ", "call_id": cid,
                          "date": v["fetch_date"], "task_type": v.get("task_type",""),
                          "ours_payment": v["price"], "their_payment": None, "diff": None})
    for cid in set(uploaded) & set(ours):
        op = ours[cid]["price"]; tp = uploaded[cid]["payment"]
        d  = round(op - tp, 2)
        if abs(d) > 0.01:
            diffs.append({"type": "פער בתשלום", "call_id": cid,
                          "date": ours[cid]["fetch_date"],
                          "task_type": ours[cid].get("task_type",""),
                          "ours_payment": op, "their_payment": tp, "diff": d})
    diffs.sort(key=lambda x: (x["type"], x["date"] or ''))
    return jsonify({
        "diffs": diffs, "source_file": filename,
        "total_uploaded": len(uploaded), "total_ours": len(ours),
        "matches": len(set(uploaded) & set(ours)),
        "range": f"{mn} – {mx}"
    })


def _build_repeats_data(token, phone):
    """
    פונקציית עזר משותפת: בונה נתוני חוזרות לשני חודשים.
    מחזירה list של dicts עם repeats, total_visits, total_price לכל חודש.
    """
    from concurrent.futures import ThreadPoolExecutor
    HEBREW_MONTHS = {1:"ינואר",2:"פברואר",3:"מרץ",4:"אפריל",5:"מאי",6:"יוני",
                     7:"יולי",8:"אוגוסט",9:"ספטמבר",10:"אוקטובר",11:"נובמבר",12:"דצמבר"}
    today = date.today()
    prev_yr, prev_mo = (today.year, today.month - 1) if today.month > 1 else (today.year - 1, 12)
    months = [(today.year, today.month), (prev_yr, prev_mo)]

    conn = init_db()

    def _sync_month(yr, mo, is_current):
        """סנכרן ימים חסרים של חודש נתון מה-API"""
        first_d = date(yr, mo, 1)
        last_d  = date(yr, mo+1, 1) - timedelta(days=1) if mo < 12 else date(yr, 12, 31)
        end_d   = today if is_current else last_d
        cur = first_d
        while cur <= end_d:
            if is_current and cur >= today:
                # היום — תמיד רענן
                conn.execute("DELETE FROM visits WHERE phone=? AND fetch_date=?",
                             (phone, cur.isoformat()))
                conn.commit()
                apts = fetch_schedules_api(token, cur)
                if apts: save_day_to_db(conn, apts, cur, phone)
            elif not date_already_fetched(conn, cur, phone):
                apts = fetch_schedules_api(token, cur)
                if apts:
                    save_day_to_db(conn, apts, cur, phone)
                else:
                    conn.execute("INSERT OR IGNORE INTO visits(phone,call_id,fetch_date) VALUES(?,?,?)",
                                 (phone, f"EMPTY_{cur.isoformat()}", cur.isoformat()))
                    conn.commit()
            cur += timedelta(days=1)

    # סנכרן שני החודשים
    _sync_month(today.year, today.month, is_current=True)
    _sync_month(prev_yr, prev_mo, is_current=False)

    prices       = load_prices(phone)
    repeat_types = {k.replace('__repeat', '') for k, v in prices.items()
                    if k.endswith('__repeat') and v}

    def get_my_price(tt, lsb_flag):
        tt = tt or ''
        is_fiber = ('תשתית סיבים' in tt or 'טריפל סיבים' in tt) and 'תקלת גלישה' not in tt
        if is_fiber:
            return float(prices.get(tt + ('__LSB' if lsb_flag else '__PPC'), 0))
        return float(prices.get(tt, 0))

    def parse_hist_date(raw):
        if not raw: return None
        s = str(raw).strip()
        # dd/mm/yyyy או dd/mm/yyyy HH:MM
        try: return datetime.strptime(s[:10], "%d/%m/%Y").date()
        except: pass
        # ISO: yyyy-mm-dd או yyyy-mm-ddTHH:MM...
        try: return datetime.strptime(s[:10], "%Y-%m-%d").date()
        except: pass
        return None

    _hist_cache = {}

    def get_hist(ban, cid):
        key = ban or cid
        if not key: return []
        if key not in _hist_cache:
            _hist_cache[key] = fetch_visit_history(
                token, ban or cid, cid or ban, "JET",
                customer_id=cid if (ban or "").upper().startswith("PI") else None
            )
        return _hist_cache.get(key) or []

    def find_repeats(yr, mo):
        first = date(yr, mo, 1)
        last  = date(yr, mo + 1, 1) - timedelta(days=1) if mo < 12 else date(yr, 12, 31)
        all_v = [v for v in get_visits_for_range(conn, first, last, phone)
                 if not v["call_id"].startswith("EMPTY_")]
        total_visits = len(all_v)

        # שלוף היסטוריות במקביל — עבור כל הביקורים שלי
        ukeys = list({(v.get("ban",""), v.get("customer_id","")) for v in all_v
                      if v.get("ban") or v.get("customer_id")})
        with ThreadPoolExecutor(max_workers=8) as ex:
            futs = {ex.submit(get_hist, b, c): (b, c) for b, c in ukeys}
            for fut in futs:
                b, c = futs[fut]
                try: _hist_cache[b or c] = fut.result()
                except: _hist_cache[b or c] = []

        results = []
        for v in all_v:
            my_date  = date.fromisoformat(v["fetch_date"])
            my_type  = (v.get("task_type") or "").strip()
            my_price = get_my_price(my_type, v.get("lsb_flag", 0))

            # חפש את הביקור ה-V-מסומן הקרוב ביותר שהגיע אחרי שלי (תוך 30 יום)
            # ביקורים לא-V מתעלמים לחלוטין — לא מפריעים ולא ביטול
            best_follow, best_td = None, None
            for entry in get_hist(v.get("ban",""), v.get("customer_id","")):
                entry_desc = (entry.get("description") or "").strip()
                # ✅ רק ביקורים שה-description שלהם מסומן V
                if repeat_types and entry_desc not in repeat_types:
                    continue
                td = parse_hist_date(entry.get("fullDateTimeTo","") or entry.get("dateEnd",""))
                if not td: continue
                diff = (td - my_date).days   # ✅ חיובי = אחרי הביקור שלי
                if 0 < diff <= 30:
                    if best_td is None or td < best_td:   # הקרוב ביותר לאחר שלי
                        best_follow, best_td = entry, td

            if best_follow and best_td:
                results.append({
                    "my_date":    v["fetch_date"],
                    "my_type":    my_type,
                    "my_price":   my_price,
                    "cust_name":  v.get("contact_name",""),
                    "cust_phone": v.get("contact_phone",""),
                    "address":    f"{v.get('street','')} {v.get('city','')}".strip(),
                    "tech_date":  best_td.isoformat(),
                    "tech_name":  (best_follow.get("technicianName") or "לא ידוע").strip(),
                    "tech_type":  (best_follow.get("description") or best_follow.get("visitType") or "").strip(),
                    "days_diff":  (best_td - my_date).days,
                })

        results.sort(key=lambda x: x["my_date"])
        return results, total_visits

    result_months = []
    for yr, mo in months:
        repeats, total_visits = find_repeats(yr, mo)
        total_price = sum(r.get("my_price", 0) or 0 for r in repeats)
        result_months.append({
            "label":        f"{HEBREW_MONTHS[mo]} {yr}",
            "yr": yr, "mo": mo,
            "repeats":      repeats,
            "total_visits": total_visits,
            "total_price":  total_price,
        })
    conn.close()
    return result_months


def _auto_save_repeats_archive(phone, month_data):
    """שמור חודש סגור לארכיון כ-Excel (קובץ קבוע, מתעדכן)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        md = month_data
        label = md.get("label", f"{md.get('yr','')}-{md.get('mo','')}")
        safe_label = label.replace(' ', '_')
        filename = f"חוזרות_{safe_label}.xlsx"

        phone_dir = os.path.join(ARCHIVE_DIR, phone.replace('+', '').replace(' ', ''))
        os.makedirs(phone_dir, exist_ok=True)
        filepath = os.path.join(phone_dir, filename)

        BLUE   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        RED    = PatternFill(start_color="FFD0D0", end_color="FFD0D0", fill_type="solid")
        WHITE  = Font(color="FFFFFF", bold=True, size=10)
        CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
        RIGHT  = Alignment(horizontal="right", vertical="center")
        headers = ["תאריך ביקורי","סוג משימה שלי","סכום ₪","תאריך ביקור הטכנאי",
                   "שם הטכנאי השני","סוג משימה של הטכנאי","ימים אחרי ביקורי",
                   "שם לקוח","טלפון","כתובת"]
        col_widths = [14, 30, 10, 14, 20, 30, 10, 20, 14, 28]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = label
        ws.sheet_view.rightToLeft = True
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = BLUE; c.font = WHITE; c.alignment = CENTER
        ws.row_dimensions[1].height = 28
        for ri, r in enumerate(md.get("repeats", []), 2):
            for ci, val in enumerate([
                r.get("my_date",""), r.get("my_type",""), r.get("my_price", 0) or 0,
                r.get("tech_date",""), r.get("tech_name",""), r.get("tech_type",""),
                r.get("days_diff", ""), r.get("cust_name",""), r.get("cust_phone",""),
                r.get("address","")
            ], 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = RED; c.alignment = RIGHT
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        sr = len(md.get("repeats", [])) + 3
        ws.cell(row=sr,   column=1, value=f'סה"כ חוזרות: {len(md.get("repeats",[]))}').font  = Font(bold=True, size=12, color="CC0000")
        ws.cell(row=sr+1, column=1, value=f'סה"כ ביקורים: {md.get("total_visits",0)}').font  = Font(bold=True, size=12, color="1F497D")
        ws.cell(row=sr+2, column=1, value=f'סה"כ סכום: ₪{md.get("total_price",0):,.2f}').font = Font(bold=True, size=12, color="CC6600")

        buf = BytesIO()
        wb.save(buf)
        with open(filepath, 'wb') as f:
            f.write(buf.getvalue())
    except Exception:
        pass  # auto-save שקט — לא נפיל את הדוח בגלל זה


@app.route("/api/reports/repeats-stream")
def api_repeats_stream():
    """מסטרים חוזרות כ-NDJSON — מציג כל שורה מיד כשנמצאת"""
    from flask import stream_with_context
    import json as _json
    from concurrent.futures import ThreadPoolExecutor as _TPE, as_completed as _as_comp

    token, phone = get_auth()
    if not token:
        return Response('{"type":"error","error":"לא מחובר"}\n', mimetype='application/x-ndjson')

    month_offset = int(request.args.get("month", "-1"))
    prices_l     = load_prices(phone)
    today = date.today()
    if month_offset == 0:
        yr, mo = today.year, today.month
    else:
        yr  = today.year  + ((today.month + month_offset - 1) // 12)
        mo  = (today.month + month_offset - 1) % 12 + 1

    HMON = {1:"ינואר",2:"פברואר",3:"מרץ",4:"אפריל",5:"מאי",6:"יוני",
            7:"יולי",8:"אוגוסט",9:"ספטמבר",10:"אוקטובר",11:"נובמבר",12:"דצמבר"}
    label = f"{HMON[mo]} {yr}"

    def _price(tt, lsb):
        tt = tt or ''
        is_f = ('תשתית סיבים' in tt or 'טריפל סיבים' in tt) and 'תקלת גלישה' not in tt
        if is_f: return float(prices_l.get(tt+('__LSB' if lsb else '__PPC'),0))
        return float(prices_l.get(tt,0))

    def _parse_d(raw):
        if not raw: return None
        s = str(raw).strip()
        try: return datetime.strptime(s[:10],"%d/%m/%Y").date()
        except: pass
        try: return datetime.strptime(s[:10],"%Y-%m-%d").date()
        except: pass
        return None

    # חוזרת = ביקור היסטוריה עם visitType "קריאת שירות" תוך 30 יום אחרי הביקור שלי
    SERVICE_CALL_TYPE = "קריאת שירות"

    def generate():
        conn = init_db()
        try:
            first_d = date(yr, mo, 1)
            last_d  = date(yr, mo+1, 1)-timedelta(days=1) if mo<12 else date(yr,12,31)
            is_cur  = (yr==today.year and mo==today.month)
            end_d   = today if is_cur else last_d

            # מצא ימים חסרים
            missing = []
            cur = first_d
            while cur <= end_d:
                if is_cur and cur >= today:
                    missing.append((cur, True))
                elif not date_already_fetched(conn, cur, phone):
                    missing.append((cur, False))
                cur += timedelta(days=1)

            if missing:
                yield _json.dumps({"type":"sync_start","days":len(missing)},ensure_ascii=False)+"\n"
                # סנכרן במקביל
                def _fetch_day(args):
                    d, force = args
                    return d, fetch_schedules_api(token, d), force
                with _TPE(max_workers=10) as ex:
                    futs = {ex.submit(_fetch_day, a): a for a in missing}
                    done_n = 0
                    for fut in _as_comp(futs):
                        d, apts, force = fut.result()
                        if force:
                            conn.execute("DELETE FROM visits WHERE phone=? AND fetch_date=?",
                                         (phone, d.isoformat()))
                            conn.commit()
                        if apts:
                            save_day_to_db(conn, apts, d, phone)
                        else:
                            conn.execute("INSERT OR IGNORE INTO visits(phone,call_id,fetch_date) VALUES(?,?,?)",
                                         (phone, f"EMPTY_{d.isoformat()}", d.isoformat()))
                            conn.commit()
                        done_n += 1
                        yield _json.dumps({"type":"sync_progress","done":done_n,"total":len(missing)},ensure_ascii=False)+"\n"

            all_v = [v for v in get_visits_for_range(conn, first_d, last_d, phone)
                     if not v["call_id"].startswith("EMPTY_")]
            total_v = len(all_v)
            yield _json.dumps({"type":"meta","label":label,"total_visits":total_v},ensure_ascii=False)+"\n"

            if not all_v:
                yield _json.dumps({"type":"done","repeat_count":0},ensure_ascii=False)+"\n"
                return

            # שלוף היסטוריות במקביל
            ukeys = list({(v.get("ban",""), v.get("customer_id",""),
                           v.get("user_id",""), v.get("system_source","JET"))
                          for v in all_v if v.get("ban") or v.get("customer_id")})
            hist_cache = {}
            def _get_hist(args):
                ban, cid, uid, src = args
                key        = ban or cid
                effective  = ban or cid   # ban עשוי להיות ריק — נשתמש ב-cid
                u_id       = uid or cid or ban
                # העבר customer_id תמיד — fetch_visit_history ישתמש בו לקומבינציות PI
                return key, fetch_visit_history(
                    token, effective, u_id, src or "JET",
                    customer_id=cid
                )
            with _TPE(max_workers=8) as ex:
                for key, h in ex.map(_get_hist, ukeys):
                    if key: hist_cache[key] = h

            # בדוק כל ביקור שלי — חוזרת = קריאת שירות בהיסטוריה תוך 30 יום
            rep_cnt = 0
            for v in all_v:
                my_date = date.fromisoformat(v["fetch_date"])
                my_type = (v.get("task_type") or "").strip()

                # דלג על ביקורי התקנה (אם יש "התקנ" בשם המשימה)
                if "התקנ" in my_type:
                    continue

                hist = hist_cache.get(v.get("ban","") or v.get("customer_id",""), [])
                best, best_td = None, None
                for entry in hist:
                    # רק "קריאת שירות" — לא התקנות
                    if entry.get("visitType","").strip() != SERVICE_CALL_TYPE:
                        continue
                    td = _parse_d(entry.get("fullDateTimeTo","") or entry.get("dateEnd","")
                                  or entry.get("fullDateTimeFrom","") or entry.get("dateFrom",""))
                    if not td: continue
                    diff = (td - my_date).days
                    if 0 < diff <= 30:
                        if best_td is None or td < best_td:
                            best, best_td = entry, td
                if best and best_td:
                    rep_cnt += 1
                    yield _json.dumps({
                        "type":"repeat",
                        "my_date":  v["fetch_date"],
                        "my_type":  my_type,
                        "my_price": _price(my_type, v.get("lsb_flag",0)),
                        "cust_name":  v.get("contact_name",""),
                        "cust_phone": v.get("contact_phone",""),
                        "address":    f"{v.get('street','')} {v.get('city','')}".strip(),
                        "tech_date":  best_td.isoformat(),
                        "tech_name":  (best.get("technicianName") or "לא ידוע").strip(),
                        "tech_type":  (best.get("description") or best.get("visitType") or "").strip(),
                        "days_diff":  (best_td - my_date).days,
                    }, ensure_ascii=False)+"\n"

            yield _json.dumps({"type":"done","repeat_count":rep_cnt,"total_visits":total_v},
                              ensure_ascii=False)+"\n"
        except Exception as e:
            import traceback as _tb
            yield _json.dumps({"type":"error","error":str(e),"trace":_tb.format_exc()},
                              ensure_ascii=False)+"\n"
        finally:
            conn.close()

    return Response(stream_with_context(generate()),
                    mimetype='application/x-ndjson',
                    headers={'Cache-Control':'no-cache','X-Accel-Buffering':'no'})

@app.route("/api/reports/repeats-debug")
def api_repeats_debug():
    """דיבאגר חוזרות — מחזיר ניתוח מלא לכל ביקור"""
    from flask import stream_with_context
    import json as _json
    token, phone = get_auth()
    if not token:
        return Response('{"type":"error","error":"לא מחובר"}\n', mimetype='application/x-ndjson')

    month_offset = int(request.args.get("month", "-1"))
    call_id_filter = request.args.get("call_id", "").strip()  # אופציונלי — סנן ל-call_id ספציפי

    def _parse_d(raw):
        if not raw: return None
        s = str(raw).strip()
        try: return datetime.strptime(s[:10],"%d/%m/%Y").date()
        except: pass
        try: return datetime.strptime(s[:10],"%Y-%m-%d").date()
        except: pass
        return None

    def generate():
        prices = load_prices(phone)
        SERVICE_CALL_TYPE = "קריאת שירות"

        today = date.today()
        if month_offset == 0:
            yr, mo = today.year, today.month
        else:
            yr  = today.year  + ((today.month + month_offset - 1) // 12)
            mo  = (today.month + month_offset - 1) % 12 + 1

        HMON = {1:"ינואר",2:"פברואר",3:"מרץ",4:"אפריל",5:"מאי",6:"יוני",
                7:"יולי",8:"אוגוסט",9:"ספטמבר",10:"אוקטובר",11:"נובמבר",12:"דצמבר"}

        yield _json.dumps({
            "type": "config",
            "rule": "קריאת שירות תוך 30 יום (ללא V מחירון)",
            "month": f"{HMON[mo]} {yr}",
            "month_offset": month_offset
        }, ensure_ascii=False) + "\n"

        conn = init_db()
        try:
            first_d = date(yr, mo, 1)
            last_d  = date(yr, mo+1, 1)-timedelta(days=1) if mo < 12 else date(yr,12,31)

            # סנכרן ימים חסרים אוטומטית (כמו ב-repeats-stream)
            is_cur = (yr == today.year and mo == today.month)
            end_d  = today if is_cur else last_d
            missing_dbg = []
            cur = first_d
            while cur <= end_d:
                if not date_already_fetched(conn, cur, phone):
                    missing_dbg.append(cur)
                cur += timedelta(days=1)
            if missing_dbg:
                yield _json.dumps({"type":"sync_start","days":len(missing_dbg)}, ensure_ascii=False)+"\n"
                done_n = 0
                for d in missing_dbg:
                    apts = fetch_schedules_api(token, d)
                    if apts:
                        save_day_to_db(conn, apts, d, phone)
                    else:
                        conn.execute("INSERT OR IGNORE INTO visits(phone,call_id,fetch_date) VALUES(?,?,?)",
                                     (phone, f"EMPTY_{d.isoformat()}", d.isoformat()))
                        conn.commit()
                    done_n += 1
                    yield _json.dumps({"type":"sync_progress","done":done_n,"total":len(missing_dbg)}, ensure_ascii=False)+"\n"

            all_v_raw = [v for v in get_visits_for_range(conn, first_d, last_d, phone)
                         if not v["call_id"].startswith("EMPTY_")]

            # אם מחפשים call_id ספציפי — נסה חיפוש גמיש
            if call_id_filter:
                exact   = [v for v in all_v_raw if v.get("call_id","") == call_id_filter]
                partial = [v for v in all_v_raw if call_id_filter in str(v.get("call_id",""))] if not exact else exact
                all_v   = exact or partial
                sample_ids = [v.get("call_id","") for v in all_v_raw[:10]]
                if not all_v:
                    yield _json.dumps({
                        "type": "not_found",
                        "call_id": call_id_filter,
                        "total_in_db": len(all_v_raw),
                        "sample_ids": sample_ids,
                        "month": f"{first_d} – {last_d}",
                    }, ensure_ascii=False) + "\n"
                    return
            else:
                all_v = all_v_raw

            yield _json.dumps({"type":"total","count":len(all_v)}, ensure_ascii=False)+"\n"

            for v in all_v:
                my_date  = date.fromisoformat(v["fetch_date"])
                my_type  = (v.get("task_type") or "").strip()
                call_id  = v.get("call_id","")
                ban      = v.get("ban","") or v.get("customer_id","")
                cid      = v.get("customer_id","") or ban
                uid      = v.get("user_id","") or cid
                src      = v.get("system_source","JET") or "JET"

                # דלג על ביקורי התקנה
                is_install = "התקנ" in my_type

                # PI-prefix: נסה לשלוף BAN נומרי מ-getCallDetails
                effective_ban = ban
                call_det_ban  = None
                if ban.upper().startswith("PI"):
                    try:
                        det_body = fetch_call_details(token, call_id, cid or ban, src)
                        if det_body and not det_body.get("_error"):
                            found = _extract_numeric_ban(det_body)
                            if found:
                                call_det_ban  = found[1]
                                effective_ban = found[1]
                    except Exception:
                        pass

                # שלוף היסטוריה עם כל הניסיונות (debug mode)
                extra_cid = v.get("customer_id","") or ""
                hist, attempts = fetch_visit_history_debug(token, effective_ban, uid, src, extra_cid=extra_cid)

                hist_detail = []
                best_entry = None
                best_td = None
                for entry in hist:
                    visit_type = entry.get("visitType","").strip()
                    td = _parse_d(entry.get("fullDateTimeTo","") or entry.get("dateEnd","")
                                  or entry.get("fullDateTimeFrom","") or entry.get("dateFrom",""))
                    raw_date = (entry.get("fullDateTimeTo","") or entry.get("dateEnd","")
                                or entry.get("fullDateTimeFrom","") or entry.get("dateFrom",""))
                    diff = (td - my_date).days if td else None
                    is_service = (visit_type == SERVICE_CALL_TYPE)
                    in_window = (is_service and diff is not None and 0 < diff <= 30)
                    hist_detail.append({
                        "description": (entry.get("description") or "").strip(),
                        "visit_type": visit_type,
                        "tech": (entry.get("technicianName") or "").strip(),
                        "raw_date": str(raw_date),
                        "parsed_date": td.isoformat() if td else None,
                        "diff_days": diff,
                        "in_window": in_window,
                    })
                    if in_window and not is_install:
                        if best_td is None or td < best_td:
                            best_entry = entry
                            best_td = td

                if is_install:
                    status = "install_skip"
                elif best_entry:
                    status = "repeat"
                else:
                    status = "no_match"

                repeat_types_set = {k.replace('__repeat','') for k,val in prices.items()
                                     if k.endswith('__repeat') and val}
                type_match = (my_type in repeat_types_set) if repeat_types_set else True
                yield _json.dumps({
                    "type": "visit",
                    "call_id":        call_id,
                    "date":           v["fetch_date"],
                    "task_type":      my_type,
                    "type_match":     type_match,
                    "ban":            ban,
                    "effective_ban":  effective_ban,
                    "call_det_ban":   call_det_ban,
                    "user_id":        uid,
                    "source":         src,
                    "attempts":       attempts,
                    "hist_count":     len(hist),
                    "hist":           hist_detail,
                    "status":         status,
                    "best_match": {
                        "date":  best_td.isoformat() if best_td else None,
                        "tech":  (best_entry.get("technicianName") or "") if best_entry else "",
                        "desc":  (best_entry.get("description") or "") if best_entry else "",
                        "diff":  (best_td - my_date).days if best_td else None,
                    } if best_entry else None,
                }, ensure_ascii=False) + "\n"

        except Exception as e:
            import traceback as _tb
            yield _json.dumps({"type":"error","error":str(e),"trace":_tb.format_exc()},
                              ensure_ascii=False)+"\n"
        finally:
            conn.close()

    return Response(stream_with_context(generate()),
                    mimetype='application/x-ndjson',
                    headers={'Cache-Control':'no-cache','X-Accel-Buffering':'no'})


@app.route("/api/debug/tasks-test")
def api_debug_tasks_test():
    """בדיקת /api/tasks — מראה raw HTTP response כדי לאבחן שגיאות API"""
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר — הוסף ?token=...&phone=... לכתובת"}), 401
    import traceback as _tb
    url = "https://tech-api.cellcom.co.il/api/technician/authorize/callActivities/getPotentialTasks"
    body = {"DeviceModel": DEVICE_MODEL, "PhoneDeviceId": PHONE_DEVICE_ID}
    headers = {**BASE_HEADERS, "Authorization": f"Bearer {token}"}
    try:
        r = requests.post(url, headers=headers, json=body, verify=False, timeout=50)
        http_status = r.status_code
        raw_text = r.text[:1000]   # 1000 תווים ראשונים
        try:
            data = r.json()
        except Exception:
            data = {}
        rc = data.get("Header", {}).get("ReturnCode", "?") if data else "?"
        rc_msg = data.get("Header", {}).get("ReturnCodeMessage", "") if data else ""
        rd = data.get("Body", {}).get("ResponseData", {}) if data else {}
        if isinstance(rd, str):
            try: rd = json.loads(rd)
            except: pass
        tasks_obj = (rd.get("Tasks") or {}) if isinstance(rd, dict) else {}
        tasks = tasks_obj.get("Task", []) if isinstance(tasks_obj, dict) else []
        if not isinstance(tasks, list):
            tasks = [tasks] if tasks else []
        return jsonify({
            "ok": rc == "00",
            "http_status": http_status,
            "rc": rc, "rc_msg": rc_msg,
            "tasks_count": len(tasks),
            "raw_preview": raw_text,
            "header": data.get("Header", {}) if data else {},
            "sample": tasks[:1] if tasks else [],
        })
    except requests.exceptions.Timeout:
        return jsonify({"ok": False, "error": "Timeout — שרת סלקום לא הגיב תוך 15 שניות"})
    except requests.exceptions.ConnectionError as e:
        return jsonify({"ok": False, "error": f"שגיאת חיבור: {e}"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "trace": _tb.format_exc()})

@app.route("/api/debug/server-log")
def api_debug_server_log():
    """מציג שורות אחרונות מלוג השרת"""
    token, _ = get_auth()
    if not token:
        return "לא מחובר", 401
    log_path = os.path.join(os.path.dirname(__file__), "server.log")
    if not os.path.exists(log_path):
        return "<pre>לא נמצא לוג</pre>"
    with open(log_path, "r", encoding="utf-8", errors="replace") as f:
        lines = f.readlines()
    last = "".join(lines[-80:])
    return f"<pre style='direction:ltr;font-size:12px'>{last}</pre>"

@app.route("/api/debug/reports-error")
def api_debug_reports_error():
    """מנסה להריץ את לוגיקת /api/reports ומחזיר traceback מלא."""
    import traceback
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    try:
        conn = init_db()
        today = date.today()
        start_date = date(today.year, today.month, 1)
        visits = get_visits_for_range(conn, start_date, today, phone)
        conn.close()
        prices = load_prices_for_date(phone, today.isoformat())
        return jsonify({"ok": True, "visits_count": len(visits), "prices_keys": list(prices.keys())[:5]})
    except Exception as e:
        return jsonify({"error": str(e), "traceback": traceback.format_exc()})

@app.route("/api/debug/schedule-count")
def api_debug_schedule_count():
    """מחזיר ספירת ביקורים ו-callIds בלבד — ללא מידע רגיש."""
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    date_str = request.args.get("date", date.today().isoformat())
    try:
        target = date.fromisoformat(date_str)
    except Exception:
        return jsonify({"error": "תאריך לא תקין"}), 400
    apts = fetch_schedules_api(token, target)
    call_ids = [a.get("task", {}).get("callId", "") for a in apts]
    # גם בדוק DB
    conn = init_db()
    db_rows = conn.execute(
        "SELECT call_id, task_type, appt_start, status FROM visits WHERE fetch_date=? AND phone=? ORDER BY appt_start",
        (date_str, phone)
    ).fetchall()
    conn.close()
    return jsonify({
        "api_count": len(apts),
        "api_call_ids": call_ids,
        "db_count": len(db_rows),
        "db_visits": [{"call_id": r[0], "task_type": r[1], "appt_start": r[2], "status": r[3]} for r in db_rows]
    })

@app.route("/api/debug/schedule-raw")
def api_debug_schedule_raw():
    """מציג JSON גולמי של תגובת getSchedules לתאריך נתון.
    ?date=2026-04-12  (token ו-phone דרך headers/query)
    מטרה: לגלות אילו שדות מכיל task עבור ביקורים מסוג PI
    """
    token, phone = get_auth()
    if not token:
        return Response("<p style='color:red'>לא מחובר — הוסף ?token=...&phone=...</p>",
                        mimetype="text/html")

    date_str = request.args.get("date","").strip()
    call_id  = request.args.get("call_id","").strip()   # אופציונלי — סנן לפקע ספציפי

    if not date_str:
        return Response("<p style='color:orange'>חסר ?date=2026-04-12</p>",
                        mimetype="text/html")

    try:
        target = date.fromisoformat(date_str)
    except Exception:
        return Response(f"<p style='color:red'>תאריך לא תקין: {date_str}</p>",
                        mimetype="text/html")

    raw = fetch_schedules_api(token, target)

    # סנן לפקע ספציפי אם ביקשו
    if call_id:
        raw = [a for a in raw
               if str(a.get("task",{}).get("callId","")) == call_id
               or call_id in str(a.get("task",{}).get("callId",""))]

    # סיכום שדות task (כולל כל השדות הפוטנציאליים ל-BAN)
    ban_fields = ["ban","BAN","billingAccountNumber","accountNumber","jetBan",
                  "jetAccountNumber","accountId","billingAccount","customerId",
                  "customer","customerNumber","userId","siebelId","siebel_id",
                  "sourceSystem","systemSource","userType","serviceOrderId"]
    summary_rows = ""
    for apt in raw:
        task = apt.get("task",{})
        cd   = apt.get("callDetails",{})
        cid_val = task.get("callId","?")
        row_cells = "".join(
            f"<td style='padding:3px 8px;color:#c084fc'>{task.get(f,'') or cd.get(f,'') or '—'}</td>"
            for f in ban_fields
        )
        summary_rows += f"<tr><td style='padding:3px 8px;color:#53bdeb;white-space:nowrap'>{cid_val}</td>{row_cells}</tr>"

    headers_html = "".join(f"<th style='padding:3px 8px;color:#6b94b8;white-space:nowrap'>{f}</th>" for f in ban_fields)

    body = json.dumps(raw, ensure_ascii=False, indent=2)

    html = f"""<!DOCTYPE html><html dir="ltr"><head>
    <meta charset="utf-8"><title>Schedule Raw — {date_str}</title>
    <style>body{{background:#0a0f18;color:#c9d8e8;font-family:monospace;font-size:12px;padding:16px}}
    pre{{background:#111;border:1px solid #334;border-radius:8px;padding:14px;white-space:pre-wrap;word-break:break-all;overflow-x:auto}}
    h2{{color:#53bdeb}} input,button{{background:#1a2d42;color:#eaf4ff;border:1px solid #3d6fa8;padding:6px 10px;border-radius:6px;font-size:13px}}
    table{{border-collapse:collapse;background:#0d1b2a;border-radius:8px;overflow:hidden;width:100%}}
    tr:nth-child(even){{background:#0a1520}}
    </style></head><body>
    <h2>📋 Schedule Raw — {date_str} ({len(raw)} ביקורים)</h2>
    <form method="get" style="margin-bottom:12px">
      <input type="hidden" name="token" value="{request.args.get('token','')}">
      <input type="hidden" name="phone" value="{request.args.get('phone','')}">
      <input name="date" value="{date_str}" placeholder="2026-04-12" style="width:130px">
      <input name="call_id" value="{call_id}" placeholder="call_id (אופציונלי)" style="width:160px">
      <button type="submit">🔍 טען</button>
    </form>
    <h3 style="color:#f9c846">🔍 שדות רלוונטיים ל-BAN לכל ביקור:</h3>
    <div style="overflow-x:auto;margin-bottom:20px">
    <table><tr><th style="padding:3px 8px;color:#3dba6f">callId</th>{headers_html}</tr>
    {summary_rows}
    </table></div>
    <h3 style="color:#6b94b8">📄 JSON מלא:</h3>
    <pre>{body}</pre>
    </body></html>"""

    return Response(html, mimetype="text/html",
                    headers={"Cache-Control":"no-cache"})


@app.route("/api/reports/history-probe")
def api_history_probe():
    """דף HTML ישיר לבדיקת API היסטוריה — מאפשר לראות את התגובה הגולמית
    פרמטרים: ?call_id=4447662&month=-1  (token ו-phone דרך headers או query)"""
    token, phone = get_auth()

    call_id_filter = request.args.get("call_id","").strip()
    month_offset   = int(request.args.get("month","-1"))
    custom_ban     = request.args.get("ban","").strip()
    custom_uid     = request.args.get("uid","").strip()
    custom_utype   = request.args.get("utype","").strip()
    HIST_URL = "https://tech-api.cellcom.co.il/api/technician/authorize/TechnicianVisitsHistory/GetTechnicianVisitsHistory"

    rows_html = ""

    if not token:
        rows_html = "<p style='color:red'>❌ לא מחובר — יש לשלוח token ו-phone בquery-string: ?token=...&phone=...</p>"
    else:
        # חפש בDB
        today = date.today()
        yr  = today.year  + ((today.month + month_offset - 1) // 12)
        mo  = (today.month + month_offset - 1) % 12 + 1
        first_d = date(yr, mo, 1)
        last_d  = date(yr, mo+1, 1)-timedelta(days=1) if mo < 12 else date(yr,12,31)
        conn = init_db()
        all_v_raw = [v for v in get_visits_for_range(conn, first_d, last_d, phone)
                     if not v["call_id"].startswith("EMPTY_")]
        conn.close()

        if call_id_filter:
            found = [v for v in all_v_raw if call_id_filter in str(v.get("call_id",""))]
        else:
            found = all_v_raw[:5]   # הצג 5 ראשונים אם לא נבחר call_id

        if not found:
            sample = [v.get("call_id","") for v in all_v_raw[:15]]
            rows_html = f"<p style='color:orange'>⚠️ פקע {call_id_filter} לא נמצא ב-DB ({len(all_v_raw)} פקעות בחודש זה)</p>"
            rows_html += "<p style='color:#999'>דוגמאות: " + ", ".join(sample) + "</p>"
        else:
            for v in found:
                ban  = v.get("ban","") or v.get("customer_id","")
                cid  = v.get("customer_id","") or ban
                uid  = custom_uid or v.get("user_id","") or cid
                src  = custom_utype or v.get("system_source","JET") or "JET"
                b_use= custom_ban or ban

                rows_html += f"""<div style='border:1px solid #444;border-radius:8px;padding:12px;margin-bottom:16px;background:#111'>
                  <h3 style='color:#c084fc;margin:0 0 8px'>📋 פקע {v.get('call_id','')} — {v.get('fetch_date','')} | {v.get('task_type','')}</h3>
                  <p style='font-family:monospace;color:#9bbdd8;font-size:12px;margin:4px 0'>
                    ban={ban} | customer_id={cid} | user_id={v.get('user_id','')} | system_source={v.get('system_source','')}
                  </p>"""

                # כל הניסיונות
                _, attempts = fetch_visit_history_debug(token, b_use, uid, src, extra_cid=cid)
                for i, a in enumerate(attempts):
                    ok = a.get("success")
                    rc = a.get("rc","?")
                    desc = (a.get("desc","") or a.get("error","") or "")[:200]
                    raw  = (a.get("raw","") or "")[:400]
                    bg   = "#1a3d22" if ok else "#2a0a0a"
                    icon = "✅" if ok else "❌"
                    rows_html += f"""<div style='background:{bg};border-radius:5px;padding:8px;margin-top:6px;font-family:monospace;font-size:12px'>
                      <b style='color:{"#3dba6f" if ok else "#f15c6e"}'>{icon} #{i+1} ban={a['ban']} userId={a['userId']} userType={a['userType']} → RC={rc}</b>
                      {"<br><span style='color:#f9c846'>הודעה: "+desc+"</span>" if desc else ""}
                      {"<br><span style='color:#556677;font-size:10px;word-break:break-all'>Raw: "+raw+"</span>" if raw else ""}
                      {"<br><span style='color:#3dba6f;font-size:13px'>✅ "+str(a.get('count',0))+" ביקורים היסטוריה</span>" if ok else ""}
                    </div>"""
                rows_html += "</div>"

    form_html = f"""<!DOCTYPE html><html dir="rtl"><head>
    <meta charset="utf-8"><title>History Probe</title>
    <style>body{{background:#0a0f18;color:#eaf4ff;font-family:Arial,sans-serif;padding:20px;direction:rtl}}
    input,select{{background:#1a2d42;color:#eaf4ff;border:1px solid #3d6fa8;padding:6px 10px;border-radius:6px;font-size:14px}}
    button{{background:#1a3a5c;border:1px solid #3d6fa8;color:#7db4e0;padding:8px 18px;border-radius:6px;cursor:pointer;font-size:14px}}
    </style></head><body>
    <h2>🔬 History Probe — בדיקת API היסטוריה</h2>
    <form method="get">
      <input type="hidden" name="token" value="{request.args.get('token','')}">
      <input type="hidden" name="phone" value="{request.args.get('phone','')}">
      <label>מספר פקע: <input name="call_id" value="{call_id_filter}" placeholder="4447662"></label> &nbsp;
      <label>חודש: <select name="month">
        <option value="-1" {"selected" if month_offset==-1 else ""}>קודם</option>
        <option value="0"  {"selected" if month_offset==0  else ""}>נוכחי</option>
      </select></label> &nbsp;
      <label>BAN ידני: <input name="ban" value="{custom_ban}" placeholder="(מה-DB)"></label> &nbsp;
      <label>userId ידני: <input name="uid" value="{custom_uid}" placeholder="(מה-DB)"></label> &nbsp;
      <label>userType: <input name="utype" value="{custom_utype}" placeholder="JET/SIEBEL"></label> &nbsp;
      <button type="submit">🔍 בדוק</button>
    </form>
    <hr style='border-color:#334'>
    {rows_html}
    </body></html>"""

    return Response(form_html, mimetype="text/html",
                    headers={"Cache-Control":"no-cache"})


@app.route("/api/reports/visit-history-direct")
def api_visit_history_direct():
    """שולף היסטוריית ביקורים ללקוח של ביקור נתון.
    מחפש call_id בכל החודשים ב-DB. אם לא נמצא — מקבל ban/cid ידני.
    מחזיר JSON: {found_in_db, visit, attempts, history}
    """
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401

    call_id   = request.args.get("call_id","").strip()
    manual_ban= request.args.get("ban","").strip()
    manual_cid= request.args.get("cid","").strip()

    conn = init_db()
    visit_row = None

    if call_id:
        # חפש ב-DB בכל החודשים — EXACT ואז PARTIAL
        cur = conn.execute(
            "SELECT phone,call_id,fetch_date,customer_id,task_type,"
            "contact_name,contact_phone,street,city,"
            "COALESCE(ban,'') AS ban,"
            "COALESCE(system_source,'JET') AS system_source,"
            "COALESCE(user_id,'') AS user_id "
            "FROM visits WHERE phone=? AND call_id=? LIMIT 1",
            (phone, call_id)
        )
        row = cur.fetchone()
        if not row:
            cur2 = conn.execute(
                "SELECT phone,call_id,fetch_date,customer_id,task_type,"
                "contact_name,contact_phone,street,city,"
                "COALESCE(ban,'') AS ban,"
                "COALESCE(system_source,'JET') AS system_source,"
                "COALESCE(user_id,'') AS user_id "
                "FROM visits WHERE phone=? AND call_id LIKE ? LIMIT 1",
                (phone, f"%{call_id}%")
            )
            row = cur2.fetchone()
            if row:
                cols = [d[0] for d in cur2.description]
            else:
                cols = []
        else:
            cols = [d[0] for d in cur.description]

        if row and cols:
            visit_row = dict(zip(cols, row))

    conn.close()

    # הגדר ban/cid/uid/src
    ban = manual_ban or (visit_row.get("ban","") if visit_row else "") or ""
    cid = manual_cid or (visit_row.get("customer_id","") if visit_row else "") or ""
    uid = (visit_row.get("user_id","") if visit_row else "") or ""
    src = (visit_row.get("system_source","JET") if visit_row else "JET") or "JET"

    # אם ban ריק — השתמש ב-customer_id כ-ban (נפוץ ל-PI accounts)
    if not ban:
        ban = cid or uid
    if not cid:
        cid = ban

    if not ban:
        return jsonify({
            "error": "לא נמצא הביקור ב-DB ולא סופק BAN ידני",
            "found_in_db": False,
            "call_id": call_id,
        })

    uid = uid or cid or ban

    # ── שלב חדש: אם BAN הוא PI-prefix, נסה לשלוף BAN נומרי מ-getCallDetails ──
    call_details_ban = None
    call_details_raw = {}
    call_details_ban_field = None
    if ban.upper().startswith("PI") and call_id:
        # שיטה 1: חפש BAN ב-getCallDetails
        try:
            details_body = fetch_call_details(token, call_id, cid or ban, src)
            if details_body and not details_body.get("_error"):
                call_details_raw = details_body
                found = _extract_numeric_ban(details_body)
                if found:
                    call_details_ban = found[1]
                    call_details_ban_field = found[0]
        except Exception as e:
            call_details_raw = {"_exception": str(e)}

        # שיטה 2: שלוף raw schedule ביום הביקור וחפש BAN ב-actions[]
        method2_dbg = {"tried": False, "fetch_date": None, "apt_count": 0,
                       "matched_callid": False, "actions_count": 0, "error": None}
        if not call_details_ban and visit_row and visit_row.get("fetch_date"):
            method2_dbg["tried"] = True
            method2_dbg["fetch_date"] = visit_row["fetch_date"]
            try:
                import re as _re2
                from datetime import date as _date_cls
                fetch_d = _date_cls.fromisoformat(visit_row["fetch_date"])
                raw_apts = fetch_schedules_api(token, fetch_d)
                method2_dbg["apt_count"] = len(raw_apts)
                for apt_r in raw_apts:
                    t_r = apt_r.get("task", {})
                    apt_cid = str(t_r.get("callId","")).strip()
                    if apt_cid != str(call_id).strip():
                        continue
                    method2_dbg["matched_callid"] = True
                    actions = apt_r.get("actions") or []
                    method2_dbg["actions_count"] = len(actions)
                    for action in actions:
                        a_ban = (action.get("ban","") or action.get("Ban","") or "").strip()
                        if a_ban and a_ban.isdigit() and len(a_ban) >= 6:
                            call_details_ban = a_ban
                            call_details_ban_field = "actions[].ban"
                            break
                        jp = action.get("jsonParams","") or ""
                        if jp and isinstance(jp, str):
                            m = _re2.search(r'"Ban"\s*:\s*"(\d{6,})"', jp)
                            if m:
                                call_details_ban = m.group(1)
                                call_details_ban_field = "actions[].jsonParams.Ban"
                                break
                    if call_details_ban:
                        break
            except Exception as _e2:
                method2_dbg["error"] = str(_e2)
        else:
            method2_dbg = {"tried": False, "reason": "no fetch_date or ban not PI"}

    # אם מצאנו BAN נומרי — השתמש בו; אחרת — PI value
    effective_ban = call_details_ban or ban

    hist, attempts = fetch_visit_history_debug(token, effective_ban, uid, src, extra_cid=cid)

    # אם עדיין ריק ו-BAN שונה מה-original — נסה גם עם ה-PI ban
    if not hist and call_details_ban and call_details_ban != ban:
        hist2, attempts2 = fetch_visit_history_debug(token, ban, uid, src, extra_cid=cid)
        attempts = attempts + attempts2
        if hist2:
            hist = hist2

    # מיצוי שדות BAN מכל המבנה לתצוגת דיבאג
    ban_fields_found = {}
    if call_details_raw and not call_details_raw.get("_exception") and not call_details_raw.get("_error"):
        for field in _BAN_FIELDS:
            def _get_field_recursive(obj, fld, depth=0):
                if depth > 8: return None
                if isinstance(obj, dict):
                    if fld in obj and obj[fld]: return str(obj[fld])
                    for v in obj.values():
                        r = _get_field_recursive(v, fld, depth+1)
                        if r: return r
                elif isinstance(obj, list):
                    for item in obj:
                        r = _get_field_recursive(item, fld, depth+1)
                        if r: return r
                return None
            val = _get_field_recursive(call_details_raw, field)
            if val:
                ban_fields_found[field] = val

    return jsonify({
        "call_id":               call_id,
        "found_in_db":           visit_row is not None,
        "visit":                 visit_row,
        "ban":                   ban,
        "effective_ban":         effective_ban,
        "call_details_ban":      call_details_ban,
        "call_details_ban_field":call_details_ban_field,
        "call_details_ban_fields": ban_fields_found,
        "method2_dbg":           method2_dbg,
        "customer_id":           cid,
        "user_id":               uid,
        "source":                src,
        "attempts":              attempts,
        "history":               hist,
        "hist_count":            len(hist),
    })


@app.route("/api/reports/update-visit-ban", methods=["POST"])
def api_update_visit_ban():
    """עדכון BAN ידני לביקור ב-DB לפי call_id.
    Body: {call_id, ban, customer_id (optional)}"""
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    data   = request.get_json(force=True) or {}
    call_id = str(data.get("call_id","")).strip()
    new_ban = str(data.get("ban","")).strip()
    new_cid = str(data.get("customer_id","")).strip()
    if not call_id or not new_ban:
        return jsonify({"error": "חסר call_id או ban"}), 400
    conn = sqlite3.connect(DB_PATH)
    try:
        # עדכן את כל השורות של פקע זה
        if new_cid:
            conn.execute(
                "UPDATE visits SET ban=?, customer_id=? WHERE phone=? AND call_id=?",
                (new_ban, new_cid, phone, call_id)
            )
        else:
            conn.execute(
                "UPDATE visits SET ban=? WHERE phone=? AND call_id=?",
                (new_ban, phone, call_id)
            )
        rows = conn.total_changes
        conn.commit()
    finally:
        conn.close()
    return jsonify({"ok": True, "call_id": call_id, "ban": new_ban, "rows_updated": rows})


@app.route("/api/reports/repeats-data")
def api_repeats_data():
    """מחזיר נתוני חוזרות כ-JSON לתצוגה במסך.
    ?month=0  → החודש הנוכחי בלבד
    ?month=-1 → החודש הקודם בלבד
    (ללא פרמטר) → שני החודשים"""
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    month_offset = request.args.get("month", None)   # "0" / "-1" / None
    # בדוק שיש סימוני V במחירון
    _prices_chk = load_prices(phone)
    _rpt_chk = {k.replace('__repeat','') for k,v in _prices_chk.items() if k.endswith('__repeat') and v}
    if not _rpt_chk:
        return jsonify({"no_v_marks": True, "months": []})
    try:
        months = _build_repeats_data(token, phone)
        # פילטר לפי חודש ספציפי אם ביקשו
        if month_offset is not None:
            try:
                mo_off = int(month_offset)
                today = date.today()
                t_yr = today.year + ((today.month + mo_off - 1) // 12)
                t_mo = (today.month + mo_off - 1) % 12 + 1
                months = [m for m in months if m.get("yr") == t_yr and m.get("mo") == t_mo]
            except:
                pass
        # שמור אוטומטית את החודש הקודם (הסגור) לארכיון
        today = date.today()
        for m in months:
            yr, mo = m.get("yr"), m.get("mo")
            if not yr or not mo: continue
            if yr == today.year and mo == today.month: continue
            last_day = date(yr, mo + 1, 1) - timedelta(days=1) if mo < 12 else date(yr, 12, 31)
            if (today - last_day).days >= 30:
                _auto_save_repeats_archive(phone, m)
        return jsonify({"months": months})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/reports/export-repeats")
def api_export_repeats():
    """ייצוא דוח חוזרות ל-Excel"""
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    try:
        months_data = _build_repeats_data(token, phone)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    RED    = PatternFill(start_color="FFD0D0", end_color="FFD0D0", fill_type="solid")
    BLUE   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    WHITE  = Font(color="FFFFFF", bold=True, size=10)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    RIGHT  = Alignment(horizontal="right", vertical="center")
    headers    = ["תאריך ביקורי","סוג משימה שלי","סכום ₪","תאריך ביקור הטכנאי",
                  "שם הטכנאי השני","סוג משימה של הטכנאי","ימים אחרי ביקורי",
                  "שם לקוח","טלפון","כתובת"]
    col_widths = [14, 30, 10, 14, 20, 30, 10, 20, 14, 28]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for md in months_data:
        ws = wb.create_sheet(md["label"])
        ws.sheet_view.rightToLeft = True
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = BLUE; c.font = WHITE; c.alignment = CENTER
        ws.row_dimensions[1].height = 28
        for ri, r in enumerate(md["repeats"], 2):
            for ci, val in enumerate([r["my_date"],r["my_type"],r.get("my_price",0) or 0,
                                      r["tech_date"],r["tech_name"],r["tech_type"],r["days_diff"],
                                      r["cust_name"],r["cust_phone"],r["address"]], 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = RED; c.alignment = RIGHT
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        sr = len(md["repeats"]) + 3
        ws.cell(row=sr,   column=1, value=f'סה"כ חוזרות: {len(md["repeats"])}').font   = Font(bold=True, size=12, color="CC0000")
        ws.cell(row=sr+1, column=1, value=f'סה"כ ביקורים: {md["total_visits"]}').font  = Font(bold=True, size=12, color="1F497D")
        ws.cell(row=sr+2, column=1, value=f'סה"כ סכום: ₪{md["total_price"]:,.2f}').font = Font(bold=True, size=12, color="CC6600")

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    today = date.today()
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"חוזרות_{today.strftime('%Y-%m')}.xlsx")

@app.route("/api/inventory")
def api_inventory():
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    categories = fetch_inventory(token)
    if not categories:
        return jsonify({"items": [], "last_updated": "", "total": 0})
    items = process_inventory_snapshot(categories)
    hist  = load_inventory_history()
    return jsonify({
        "items": items,
        "last_updated": hist.get("last_updated", ""),
        "total": len(items),
        "received": hist.get("received", {}),
    })

@app.route("/api/inventory/received", methods=["POST"])
def api_inventory_received():
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    data   = request.get_json() or {}
    serial = data.get("serial", "")
    status = data.get("status", "received")
    desc   = data.get("desc", "")
    if not serial:
        return jsonify({"error": "חסר serial"}), 400
    hist = load_inventory_history()
    if "received" not in hist:
        hist["received"] = {}
    hist["received"][serial] = {"status": status, "date": date.today().isoformat(), "desc": desc}
    save_inventory_history(hist)
    return jsonify({"ok": True})

@app.route("/api/save-return", methods=["POST"])
def api_save_return():
    import base64 as b64mod
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    data = request.get_json() or {}
    serial      = data.get("serial", "")
    description = data.get("description", "")
    photo_b64   = data.get("photo", "")
    quantity    = int(data.get("quantity", 1) or 1)
    now = datetime.now()
    photo_filename = ""
    if photo_b64:
        try:
            os.makedirs(RETURNS_PHOTOS_DIR, exist_ok=True)
            safe_serial = "".join(c if c.isalnum() else "_" for c in serial)
            photo_filename = f"{safe_serial}_{now.strftime('%Y%m%d_%H%M%S')}.jpg"
            raw = b64mod.b64decode(photo_b64.split(",")[-1])
            with open(os.path.join(RETURNS_PHOTOS_DIR, photo_filename), "wb") as f:
                f.write(raw)
        except Exception:
            photo_filename = ""
    conn = init_db()
    conn.execute(
        "INSERT INTO returns (phone, serial, description, return_date, photo_filename, created_at, quantity) VALUES (?,?,?,?,?,?,?)",
        (phone, serial, description, now.date().isoformat(), photo_filename, now.isoformat(), quantity)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/returns")
def api_get_returns():
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    days = int(request.args.get("days", 30))
    since = (date.today() - timedelta(days=days)).isoformat()
    conn = init_db()
    rows = conn.execute(
        "SELECT id, serial, description, return_date, photo_filename, created_at, quantity FROM returns WHERE return_date >= ? AND phone=? ORDER BY created_at DESC",
        (since, phone)
    ).fetchall()
    conn.close()
    return jsonify({"returns": [
        {"id": r[0], "serial": r[1], "description": r[2],
         "return_date": r[3], "has_photo": bool(r[4]), "photo_filename": r[4],
         "quantity": r[6] or 1}
        for r in rows
    ]})

@app.route("/api/equipment-report")
def api_equipment_report():
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    # קבלות
    hist = load_inventory_history()
    received_by_date = {}
    for serial, entry in hist.get("received", {}).items():
        if isinstance(entry, dict) and entry.get("status") == "received":
            d = entry.get("date", "")
            received_by_date.setdefault(d, []).append({
                "serial": serial, "desc": entry.get("desc", "")
            })
    # החזרות
    conn = init_db()
    rows = conn.execute(
        "SELECT serial, description, return_date, quantity FROM returns WHERE phone=? ORDER BY return_date DESC, rowid DESC",
        (phone,)
    ).fetchall()
    conn.close()
    returned_by_date = {}
    for r in rows:
        returned_by_date.setdefault(r[2], []).append({
            "serial": r[0] or "", "desc": r[1] or "", "quantity": r[3] or 1
        })
    return jsonify({"received_by_date": received_by_date, "returned_by_date": returned_by_date})

@app.route("/api/equipment-report/export")
def api_equipment_report_export():
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    report_date = request.args.get("date", "")
    report_type = request.args.get("type", "received")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook(); ws = wb.active
    ws.sheet_view.rightToLeft = True
    BLUE = "4472C4"; WHITE = "FFFFFF"
    hdr_font = Font(color=WHITE, bold=True, size=11)
    hdr_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    ctr  = Alignment(horizontal="center", vertical="center")
    rght = Alignment(horizontal="right",  vertical="center")
    def hdr(ws, cols):
        for i,(h,w) in enumerate(cols,1):
            c=ws.cell(row=1,column=i,value=h); c.font=hdr_font; c.fill=hdr_fill; c.alignment=ctr
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
        ws.row_dimensions[1].height=22
    if report_type == "received":
        ws.title = f"קבלות {report_date}"
        hdr(ws,[("תאריך",12),("מספר סריאל",20),("תיאור",30)])
        hist = load_inventory_history()
        ri=2
        for serial,entry in hist.get("received",{}).items():
            if isinstance(entry,dict) and entry.get("status")=="received" and entry.get("date")==report_date:
                for i,v in enumerate([report_date,serial,entry.get("desc","")],1):
                    c=ws.cell(row=ri,column=i,value=v); c.alignment=rght
                ri+=1
        ws.cell(row=ri+1,column=1,value=f'סה"כ: {ri-2} פריטים').font=Font(bold=True)
    else:
        ws.title = f"החזרות {report_date}"
        hdr(ws,[("תאריך",12),("מספר סריאל",20),("תיאור",30),("כמות",8)])
        conn=init_db()
        rows=conn.execute(
            "SELECT serial,description,quantity FROM returns WHERE phone=? AND return_date=? ORDER BY rowid DESC",
            (phone,report_date)
        ).fetchall(); conn.close()
        for ri,r in enumerate(rows,2):
            for i,v in enumerate([report_date,r[0],r[1],r[2] or 1],1):
                c=ws.cell(row=ri,column=i,value=v); c.alignment=rght
        ws.cell(row=len(rows)+3,column=1,value=f'סה"כ: {len(rows)} פריטים').font=Font(bold=True)
    buf=BytesIO(); wb.save(buf); buf.seek(0)
    fn=f"equipment_{report_type}_{report_date}.xlsx"
    return send_file(buf,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,download_name=fn)

@app.route("/api/equipment-report/delete-date", methods=["DELETE"])
def api_equipment_report_delete_date():
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    data        = request.get_json() or {}
    report_date = data.get("date", "")
    report_type = data.get("type", "")   # "received" | "returned"
    if not report_date or not report_type:
        return jsonify({"error": "חסרים פרמטרים"}), 400
    if report_type == "received":
        hist = load_inventory_history()
        hist["received"] = {
            s: e for s, e in hist.get("received", {}).items()
            if not (isinstance(e, dict) and e.get("date") == report_date)
        }
        save_inventory_history(hist)
    elif report_type == "returned":
        conn = init_db()
        conn.execute("DELETE FROM returns WHERE phone=? AND return_date=?", (phone, report_date))
        conn.commit(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/delete-return/<int:rid>", methods=["DELETE"])
def api_delete_return(rid):
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    conn = init_db()
    row = conn.execute("SELECT photo_filename FROM returns WHERE id=? AND phone=?", (rid, phone)).fetchone()
    if row:
        photo_filename = row[0]
        if photo_filename:
            try:
                os.remove(os.path.join(RETURNS_PHOTOS_DIR, photo_filename))
            except Exception:
                pass
        conn.execute("DELETE FROM returns WHERE id=? AND phone=?", (rid, phone))
        conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/return-photo/<filename>")
def api_return_photo(filename):
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    # אבטחה — רק שם קובץ, בלי path traversal
    safe = os.path.basename(filename)
    path = os.path.join(RETURNS_PHOTOS_DIR, safe)
    if not os.path.exists(path):
        return jsonify({"error": "לא נמצא"}), 404
    return send_file(path, mimetype="image/jpeg")

@app.route("/qr-payment")
def serve_qr_payment():
    if os.path.exists(QR_PAYMENT_FILE):
        return send_file(QR_PAYMENT_FILE, mimetype="image/png")
    return "", 404

@app.route("/api/register", methods=["POST"])
def api_register():
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    data = request.json or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"status": "error", "message": "הכנס שם"})
    employee_id = data.get("employee_id", "") or session.get("employee_id", "")
    db_register_user(phone, name, employee_id)
    return jsonify({"status": "ok"})

@app.route("/api/payment-confirm", methods=["POST"])
def api_payment_confirm():
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    user = db_get_user(phone)
    if not user:
        return jsonify({"status": "error", "message": "לא רשום"})
    new_exp = db_extend_subscription(phone, SUB_DAYS)
    return jsonify({"status": "ok", "expires": new_exp})

@app.route("/api/admin/users")
def api_admin_users():
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    user = db_get_user(phone)
    if phone != ADMIN_PHONE and not (user and user.get('is_admin')):
        return jsonify({"error": "אין הרשאה"}), 403
    conn = sqlite3.connect(DB_PATH)
    rows = conn.execute(
        "SELECT phone,name,employee_id,registered_at,trial_expires,subscription_expires,is_admin,permissions,premium_expires "
        "FROM users ORDER BY registered_at DESC"
    ).fetchall()
    conn.close()
    users_list = []
    for row in rows:
        u = dict(zip(['phone','name','employee_id','registered_at',
                      'trial_expires','subscription_expires','is_admin','permissions','premium_expires'], row))
        try:
            u['permissions'] = json.loads(u['permissions'] or '{}')
        except:
            u['permissions'] = {}
        st, exp = get_sub_status(u['phone'])
        u['sub_status'] = st; u['sub_expires'] = exp
        # merge with defaults — שמור ערכים מקוריים (false/true/date-string)
        merged = dict(DEFAULT_PERMISSIONS)
        merged.update({k: v for k, v in u['permissions'].items() if k in DEFAULT_PERMISSIONS})
        u['permissions'] = merged
        u['premium_expires'] = u.get('premium_expires') or ''
        users_list.append(u)
    return jsonify({"users": users_list, "permission_keys": list(DEFAULT_PERMISSIONS.keys())})

@app.route("/api/admin/permissions/<target_phone>", methods=["POST"])
def api_admin_set_permissions(target_phone):
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    user = db_get_user(phone)
    if phone != ADMIN_PHONE and not (user and user.get('is_admin')):
        return jsonify({"error": "אין הרשאה"}), 403
    data = request.json or {}
    # v יכול להיות: false, true, או מחרוזת ISO date
    new_perms = {}
    for k, v in data.items():
        if k not in DEFAULT_PERMISSIONS: continue
        if v is False or v is None:
            new_perms[k] = False
        elif v is True:
            new_perms[k] = True
        elif isinstance(v, str) and v:
            new_perms[k] = v   # ISO date
        else:
            new_perms[k] = bool(v)
    conn = sqlite3.connect(DB_PATH)
    conn.execute("UPDATE users SET permissions=? WHERE phone=?",
                 (json.dumps(new_perms, ensure_ascii=False), target_phone))
    conn.commit(); conn.close()
    return jsonify({"status": "ok", "permissions": new_perms})

@app.route("/api/admin/copy-prices", methods=["POST"])
def api_admin_copy_prices():
    """מעתיק מחירון admin ליוזר ספציפי"""
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    if not _is_admin_phone(phone):
        return jsonify({"error": "אין הרשאה"}), 403
    target = (request.json or {}).get("phone", "").strip()
    if not target:
        return jsonify({"error": "חסר phone"}), 400
    admin_prices = load_prices(ADMIN_PHONE)
    if not admin_prices:
        return jsonify({"error": "אין מחירון לאדמין"}), 404
    save_prices(admin_prices, target)
    return jsonify({"status": "ok", "copied_keys": len(admin_prices), "to": target})

@app.route("/api/admin/extend", methods=["POST"])
def api_admin_extend():
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    user = db_get_user(phone)
    if phone != ADMIN_PHONE and not (user and user.get('is_admin')):
        return jsonify({"error": "אין הרשאה"}), 403
    data = request.json or {}
    target = data.get("phone", "").strip()
    days = int(data.get("days", 30))
    if not target:
        return jsonify({"status": "error", "message": "חסר מספר טלפון"})
    new_exp = db_extend_subscription(target, days)
    return jsonify({"status": "ok", "expires": new_exp})

@app.route("/api/check-repeats-batch", methods=["POST"])
def api_check_repeats_batch():
    """בדיקת חוזרות לרשימת משימות — מחזיר {call_id: tech_name or null}"""
    from concurrent.futures import ThreadPoolExecutor
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    tasks = (request.json or {}).get("tasks", [])
    if not tasks:
        return jsonify({})
    prices = load_prices(phone)
    repeat_types = {k.replace('__repeat','') for k,v in prices.items()
                    if k.endswith('__repeat') and v}
    # סנן — רק משימות שה-task_type שלהן V-מסומן
    to_check = [t for t in tasks
                if not repeat_types or (t.get("task_type","") or "").strip() in repeat_types]
    if not to_check:
        return jsonify({})
    def _check(t):
        cid     = t.get("customer_id","") or ""
        ban     = t.get("ban","") or cid
        tt      = (t.get("task_type","") or "").strip()
        call_id = t.get("call_id","") or ""
        src     = t.get("system_source","") or t.get("source","") or "JET"
        visits  = fetch_visit_history(token, ban, cid, src,
                    customer_id=cid if (ban or "").upper().startswith("PI") else None)
        recent  = check_recent_visit(visits, tt, repeat_types=repeat_types)
        return call_id, recent
    results = {}
    with ThreadPoolExecutor(max_workers=min(8, len(to_check))) as ex:
        futs = {ex.submit(_check, t): t for t in to_check}
        for fut in futs:
            try:
                cid_key, recent = fut.result()
                if cid_key:
                    results[cid_key] = recent
            except:
                pass
    return jsonify(results)

@app.route("/api/task-detail", methods=["POST"])
def api_task_detail():
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    data = request.json or {}
    task = data.get("task", {})
    idx  = data.get("index", 1)
    # DEBUG — log raw call-details response
    call_id = task.get("call_id","")
    cid     = task.get("customer_id","")
    source  = task.get("system_source","JET")
    raw = fetch_call_details(token, call_id, cid, source)
    import sys; sys.stderr.write(f"[TASK-DETAIL-DEBUG] call_id={call_id} cid={cid} source={source} raw_keys={list(raw.keys()) if isinstance(raw,dict) else raw}\n"); sys.stderr.flush()
    card = build_single_card(task, token, idx, phone=phone)
    return jsonify(card)

@app.route("/api/tasks/all-details", methods=["POST"])
def api_tasks_all_details():
    """מושך פרטים לכל הפקעות במקביל — קריאה אחת במקום 44"""
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    tasks = (request.json or {}).get("tasks", [])
    if not tasks:
        return jsonify({"cards": []})
    from concurrent.futures import ThreadPoolExecutor
    def get_card(args):
        task, idx = args
        try:
            return build_single_card(task, token, idx + 1, phone=phone)
        except Exception:
            return None
    with ThreadPoolExecutor(max_workers=min(20, len(tasks))) as executor:
        results = list(executor.map(get_card, [(t, i) for i, t in enumerate(tasks)]))
    return jsonify({"cards": results})

# ============================================================
# צ'אט פנימי
# ============================================================
def _is_admin_phone(phone):
    u = db_get_user(phone)
    return phone == ADMIN_PHONE or bool(u and u.get('is_admin'))

@app.route("/api/chat/unread")
def api_chat_unread():
    token, phone = get_auth()
    if not token:
        return jsonify({"count": 0})
    conn = init_db()
    row = conn.execute(
        "SELECT COUNT(*) FROM messages WHERE to_phone=? AND is_read=0", (phone,)
    ).fetchone()
    conn.close()
    return jsonify({"count": row[0] if row else 0})

@app.route("/api/chat/conversations")
def api_chat_conversations():
    """רשימת שיחות למנהל — משתמשים שיש איתם הודעות"""
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    if not _is_admin_phone(phone):
        return jsonify({"error": "אין הרשאה"}), 403
    conn = init_db()
    rows = conn.execute(
        "SELECT DISTINCT CASE WHEN from_phone=? THEN to_phone ELSE from_phone END AS other "
        "FROM messages WHERE from_phone=? OR to_phone=? ORDER BY other",
        (phone, phone, phone)
    ).fetchall()
    result = []
    for (other,) in rows:
        if other == phone: continue
        unread = conn.execute(
            "SELECT COUNT(*) FROM messages WHERE from_phone=? AND to_phone=? AND is_read=0",
            (other, phone)
        ).fetchone()[0]
        last = conn.execute(
            "SELECT body,created_at FROM messages WHERE (from_phone=? AND to_phone=?) OR (from_phone=? AND to_phone=?) ORDER BY id DESC LIMIT 1",
            (other, phone, phone, other)
        ).fetchone()
        u = db_get_user(other)
        name = (u.get('name') or other) if u else other
        perms = get_effective_permissions(other)
        prem_exp = (u.get('premium_expires') or '') if u else ''
        result.append({
            "phone": other, "name": name,
            "unread": unread,
            "last_msg": last[0] if last else "",
            "last_at":  last[1][:16] if last else "",
            "premium": perms.get('premium', False),
            "premium_expires": prem_exp
        })
    conn.close()
    result.sort(key=lambda x: x['last_at'], reverse=True)
    return jsonify({"conversations": result})

@app.route("/api/chat/messages")
def api_chat_messages():
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    is_adm = _is_admin_phone(phone)
    with_phone = request.args.get("with", ADMIN_PHONE if not is_adm else "")
    if not with_phone:
        return jsonify({"messages": []})
    conn = init_db()
    rows = conn.execute(
        "SELECT id,from_phone,to_phone,body,created_at,is_read FROM messages "
        "WHERE (from_phone=? AND to_phone=?) OR (from_phone=? AND to_phone=?) ORDER BY id ASC",
        (phone, with_phone, with_phone, phone)
    ).fetchall()
    # סמן כנקרא
    conn.execute(
        "UPDATE messages SET is_read=1 WHERE to_phone=? AND from_phone=? AND is_read=0",
        (phone, with_phone)
    )
    conn.commit(); conn.close()
    msgs = [{"id":r[0],"from":r[1],"to":r[2],"body":r[3],"at":r[4][:16],"read":bool(r[5])} for r in rows]
    return jsonify({"messages": msgs})

@app.route("/api/chat/send", methods=["POST"])
def api_chat_send():
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    data = request.json or {}
    body = data.get("body", "").strip()
    if not body:
        return jsonify({"error": "הודעה ריקה"}), 400
    is_adm = _is_admin_phone(phone)
    to_phone = data.get("to", ADMIN_PHONE if not is_adm else "")
    if not to_phone:
        return jsonify({"error": "חסר נמען"}), 400
    now = datetime.now().isoformat()
    conn = init_db()
    conn.execute(
        "INSERT INTO messages (from_phone,to_phone,body,created_at) VALUES (?,?,?,?)",
        (phone, to_phone, body, now)
    )
    conn.commit(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/admin/grant-premium", methods=["POST"])
def api_admin_grant_premium():
    token, phone = get_auth()
    if not token:
        return jsonify({"error": "לא מחובר"}), 401
    if not _is_admin_phone(phone):
        return jsonify({"error": "אין הרשאה"}), 403
    data = request.json or {}
    target = data.get("phone", "").strip()
    months = max(1, int(data.get("months", 1)))
    if not target:
        return jsonify({"error": "חסר טלפון"}), 400
    conn = sqlite3.connect(DB_PATH)
    row = conn.execute("SELECT premium_expires FROM users WHERE phone=?", (target,)).fetchone()
    today = date.today()
    base = today
    if row and row[0]:
        try: base = max(date.fromisoformat(row[0]), today)
        except: pass
    new_exp = (base + timedelta(days=months * 30)).isoformat()
    conn.execute("UPDATE users SET premium_expires=? WHERE phone=?", (new_exp, target))
    conn.commit(); conn.close()
    # הודעת מערכת לאותו משתמש
    now = datetime.now().isoformat()
    conn2 = init_db()
    conn2.execute(
        "INSERT INTO messages (from_phone,to_phone,body,created_at) VALUES (?,?,?,?)",
        (phone, target, f"✅ הוענקה גישת פרמיום ל-{months} חודשים (עד {new_exp})", now)
    )
    conn2.commit(); conn2.close()
    return jsonify({"status": "ok", "premium_expires": new_exp})

# ============================================================
# HTML — Mobile-first
# ============================================================
HTML = r"""<!DOCTYPE html>
<html dir="rtl" lang="he">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1, maximum-scale=1">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="apple-mobile-web-app-title" content="WorkManager">
<link rel="apple-touch-icon" href="/logo">
<link rel="icon" type="image/png" href="/logo">
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<title>WorkManager</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:#0d1b2a;color:#ddeeff;height:100dvh;display:flex;flex-direction:column;overflow:hidden}

/* ─── Login / OTP screens ─── */
.screen{display:none;flex:1;flex-direction:column;align-items:center;justify-content:center;padding:24px 20px;gap:0}
.screen.active{display:flex}
.logo{margin-bottom:16px}
.logo img{width:90px;height:90px;border-radius:18px}
.screen h1{font-size:21px;font-weight:700;margin-bottom:6px;text-align:center;color:#eaf4ff}
.screen p{font-size:14px;color:#6b94b8;margin-bottom:28px;text-align:center}
.card{background:#1a2d42;border-radius:16px;padding:26px 22px;width:100%;max-width:380px;display:flex;flex-direction:column;gap:14px;box-shadow:0 8px 32px rgba(0,0,0,.5)}
.field{display:flex;flex-direction:column;gap:7px}
.field label{font-size:14px;color:#7baac8;font-weight:500}
.field input{background:#223d58;border:1px solid #284461;border-radius:10px;padding:13px 14px;color:#ddeeff;font-size:16px;outline:none;direction:rtl;width:100%}
.field input:focus{border-color:#3dba6f;box-shadow:0 0 0 2px rgba(61,186,111,.2)}
.field input::placeholder{color:#4a6a88}
.btn-main{background:#3dba6f;border:none;border-radius:10px;padding:14px;color:#fff;font-size:16px;font-weight:700;cursor:pointer;width:100%;margin-top:4px;letter-spacing:.3px}
.btn-main:active{background:#2e9458}
.btn-main:disabled{background:#223d58;color:#4a6a88;cursor:not-allowed}
.err{color:#f07080;font-size:13px;text-align:center;min-height:18px}
.back-link{font-size:13px;color:#6b94b8;text-align:center;cursor:pointer;text-decoration:underline;margin-top:10px}

/* ─── Chat screen ─── */
#chat-screen{display:none;flex:1;flex-direction:column;overflow:hidden}
#chat-screen.active{display:flex}
#header{background:#0d1b2a;padding:12px 16px;display:flex;align-items:center;gap:10px;flex-shrink:0;border-bottom:1px solid #1c3450}
.av{width:42px;height:42px;border-radius:10px;overflow:hidden;flex-shrink:0;display:flex;align-items:center;justify-content:center}
.ht{font-size:15px;font-weight:700;flex:1;color:#eaf4ff}
.hs{font-size:12px;color:#6b94b8}
#logout-btn{background:transparent;border:1px solid #284461;color:#6b94b8;border-radius:8px;padding:5px 12px;font-size:13px;cursor:pointer}
#logout-btn:active{background:#223d58}
#msgs{flex:1;overflow-y:auto;padding:0 10px 12px;display:flex;flex-direction:column;gap:4px;background:#0d1b2a}
#sched-tabs{position:sticky;top:0;z-index:10;margin:0 -10px;width:calc(100% + 20px)}
#msgs::-webkit-scrollbar{width:4px}
#msgs::-webkit-scrollbar-thumb{background:#284461;border-radius:4px}
.msg{max-width:85%;padding:8px 12px 5px;border-radius:8px;font-size:14px;line-height:1.6;white-space:pre-wrap;word-break:break-word}
.user{background:#0d4a32;align-self:flex-start;border-radius:8px 8px 0 8px}
.bot{background:#d0e4f7;align-self:flex-end;border-radius:8px 8px 8px 0}
.msg b{font-weight:700}
.mtime{font-size:11px;color:#6b94b8;text-align:left;margin-top:2px}
.qr{display:flex;gap:6px;flex-wrap:wrap;justify-content:flex-end;margin:8px 0 6px;direction:rtl}
.qb{background:#1e3a5c;border:1px solid #3d6a9a;color:#ddeeff;border-radius:20px;padding:9px 18px;font-size:14px;font-weight:600;cursor:pointer;white-space:nowrap;box-shadow:0 2px 10px rgba(0,0,0,.5)}
.qb:active{background:#122a44;border-color:#2a5080}
#typing{display:none;align-self:flex-end;background:#d0e4f7;padding:10px 14px;border-radius:8px 8px 8px 0;margin-bottom:2px}
#typing span{display:inline-block;width:7px;height:7px;background:#6b94b8;border-radius:50%;animation:b 1.2s infinite;margin:0 2px}
#typing span:nth-child(2){animation-delay:.2s}
#typing span:nth-child(3){animation-delay:.4s}
@keyframes b{0%,60%,100%{transform:translateY(0)}30%{transform:translateY(-6px)}}
#bar{background:#0d1b2a;padding:8px 10px;display:flex;gap:8px;align-items:center;flex-shrink:0;border-top:1px solid #1c3450}
#inp{flex:1;background:#223d58;border:none;border-radius:20px;padding:10px 16px;color:#ddeeff;font-size:15px;outline:none;direction:rtl}
#inp::placeholder{color:#6b94b8}
#send-btn{width:42px;height:42px;background:#3dba6f;border:none;border-radius:50%;cursor:pointer;display:flex;align-items:center;justify-content:center;flex-shrink:0}
#send-btn svg{fill:white;width:20px;height:20px}
/* ─── Tabs ─── */
#tabs{background:#0d1b2a;border-bottom:1px solid #1c3450}
.tab-btn{flex:1;min-width:0;padding:12px 6px;background:transparent;border:none;color:#6b94b8;font-size:14px;font-weight:600;cursor:pointer;border-bottom:2px solid transparent;white-space:nowrap;text-align:center}
.tab-btn.active{color:#3dba6f;border-bottom:2px solid #3dba6f;font-weight:700}
#tabs::-webkit-scrollbar{display:none}
/* ─── Schedule sub-tabs ─── */
.sched-tab{flex:1;padding:8px 4px;background:transparent;border:none;border-bottom:2px solid transparent;color:#6b94b8;font-size:13px;font-weight:600;cursor:pointer;text-align:center;white-space:nowrap}
.sched-tab.active{color:#53bdeb;border-bottom:2px solid #53bdeb;font-weight:700}
.sched-tab:hover:not(.active){color:#9bbdd8;background:#0f2033}
/* ─── Inventory cards ─── */
.inv-card{background:#d0e4f7;color:#0d1b2a;border-radius:12px;padding:14px 16px;margin:6px 0;direction:rtl;font-size:15.5px;font-weight:500;line-height:2;box-shadow:0 2px 12px rgba(0,0,0,.5)}
.inv-card.ok{border-right:4px solid #3dba6f}
.inv-card.blocked{border-right:4px solid #f07080}
.inv-badge{display:inline-block;padding:2px 9px;border-radius:10px;font-size:12px;font-weight:700}
.badge-ok{background:#0a3d22;color:#3dba6f}
.badge-blocked{background:#3d0d16;color:#f07080}
.badge-new{background:#0a2a4a;color:#5ab8f5;font-size:11px;margin-right:6px}
/* ─── Accordion ─── */
.acc{border-top:1px solid #1c3450;margin-top:8px;cursor:pointer}
.acc-hdr{color:#6b94b8;font-size:13px;display:flex;justify-content:space-between;padding:6px 0;user-select:none}
.acc-arr{transition:transform .2s;display:inline-block}
.acc-body{display:none;color:#9bbdd8;font-size:13px;padding:8px 10px 8px;line-height:1.8;background:#1a2d42;border-radius:0 0 8px 8px;margin-top:2px}
.acc.open .acc-body{display:block}
.acc.open .acc-arr{transform:rotate(180deg)}
/* ─── Task cards ─── */
.task-card{background:#d0e4f7 !important;border-right-color:#3dba6f !important}
/* ─── Diagnostics tab ─── */
.diag-section{background:#162d44;border-radius:12px;padding:14px;margin-bottom:12px;direction:rtl}
.diag-title{font-size:15px;font-weight:700;color:#eaf4ff;margin-bottom:4px}
.diag-subtitle{font-size:12px;color:#8696a0;margin:4px 0 10px;line-height:1.7}
.diag-btn-primary{display:block;background:#1f5c9e;border-radius:10px;padding:12px;text-align:center;color:#eaf4ff;text-decoration:none;font-weight:700;font-size:14px;border:none;cursor:pointer}
.diag-btn-primary:hover{background:#2a72bf}
.diag-btn-add{background:#3dba6f;color:#fff;border:none;border-radius:8px;padding:8px 14px;font-size:13px;font-weight:700;cursor:pointer;white-space:nowrap}
.diag-btn-check{width:100%;background:#1f5c9e;color:#fff;border:none;border-radius:10px;padding:11px;font-size:14px;font-weight:700;cursor:pointer}
.diag-btn-check:hover{background:#2a72bf}
.diag-btn-warn{width:100%;background:#7a4a00;color:#f9c846;border:none;border-radius:10px;padding:11px;font-size:14px;font-weight:700;cursor:pointer}
.setting-row{display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin-bottom:4px}
.setting-label{color:#9bbdd8;font-size:13px;min-width:130px}
.setting-input{flex:1;min-width:80px;background:#0d1b2a;border:1px solid #284461;border-radius:8px;padding:8px;color:#eaf4ff;font-size:14px;text-align:center}
.fw-row{display:flex;align-items:center;gap:10px;padding:10px 12px;border-radius:10px;direction:rtl}
.fw-row.fw-good{background:#1a3d22}
.fw-row.fw-bad{background:#3d1a1a}
.fw-badge{display:inline-block;padding:3px 10px;border-radius:12px;font-size:12px;font-weight:700}
.fw-rev{background:#6d1a1a;color:#f07080}
.fw-vsco{background:#1a3d22;color:#3dba6f}
.fw-mmxl{background:#1a3d5c;color:#53bdeb}
.fw-update-btn{margin-right:auto;background:#f9c846;color:#1a1200;border:none;border-radius:8px;padding:5px 12px;font-size:12px;font-weight:700;cursor:pointer}
.step-item{background:#0d1b2a;border-radius:8px;padding:8px 12px;color:#eaf4ff;margin-bottom:4px}
.logo-btn{background:#223d58;border:1px solid #284461;color:#8696a0;border-radius:8px;padding:6px 14px;font-size:13px;cursor:pointer}
.logo-btn-active{background:#3dba6f;border:1px solid #3dba6f;color:#fff;border-radius:8px;padding:6px 14px;font-size:13px;font-weight:700;cursor:pointer}
.checklist-item{display:flex;align-items:flex-start;gap:10px;padding:8px 10px;border-bottom:1px solid #1c3450;font-size:13px;cursor:pointer;user-select:none}
.checklist-item:last-child{border-bottom:none}
.checklist-item input[type=checkbox]{width:18px;height:18px;flex-shrink:0;cursor:pointer;accent-color:#3dba6f}
.checklist-item.done{opacity:0.55}
.checklist-item.done span{text-decoration:line-through;color:#6b94b8}
.dsl-cable-badge{display:inline-block;padding:3px 10px;border-radius:12px;font-size:13px;font-weight:700;margin-right:6px}
</style>
</head>
<body>

<!-- ══ מסך כניסה ══ -->
<div class="screen" id="login-screen">
  <div class="logo"><img src="/logo" onerror="this.outerHTML='<span style=font-size:52px>🔧</span>'"></div>
  <h1>WorkManager</h1>
  <p>הכנס את הפרטים שלך כדי להמשיך</p>
  <div class="card">
    <div class="field">
      <label>מספר טלפון</label>
      <input id="l-phone" type="tel" placeholder="05XXXXXXXX" inputmode="numeric" />
    </div>
    <div class="field">
      <label>מספר עובד</label>
      <input id="l-emp" type="text" placeholder="מספר עובד" inputmode="numeric" />
    </div>
    <div class="err" id="l-err"></div>
    <button class="btn-main" id="l-btn" onclick="doLogin()">כניסה</button>
  </div>
</div>

<!-- ══ מסך OTP ══ -->
<div class="screen" id="otp-screen">
  <div class="logo">📱</div>
  <h1>אימות SMS</h1>
  <p>נשלח קוד למספר שהזנת.<br>הכנס אותו כאן:</p>
  <div class="card">
    <div class="field">
      <label>קוד אימות</label>
      <input id="otp-inp" type="text" placeholder="123456" inputmode="numeric" maxlength="8"
             onkeypress="if(event.key==='Enter')doVerify()" />
    </div>
    <div class="err" id="otp-err"></div>
    <button class="btn-main" id="otp-btn" onclick="doVerify()">אמת קוד</button>
    <div class="back-link" onclick="showScreen('login')">← חזור לכניסה</div>
  </div>
</div>

<!-- ══ מסך הרשמה ══ -->
<div class="screen" id="reg-screen">
  <div class="logo"><img src="/logo" onerror="this.outerHTML='<span style=font-size:52px>🔧</span>'"></div>
  <h1>ברוך הבא! 👋</h1>
  <p>הרשמה ראשונה — <b style="color:#3dba6f">7 ימי ניסיון חינם</b></p>
  <div class="card">
    <div class="field">
      <label>השם שלך</label>
      <input id="reg-name" type="text" placeholder="שם מלא" autocomplete="name" />
    </div>
    <div class="err" id="reg-err"></div>
    <button class="btn-main" id="reg-btn" onclick="doRegister()">🚀 התחל ניסיון חינם</button>
    <div style="font-size:12px;color:#8fa5b5;text-align:center;margin-top:4px">לאחר 7 ימים — 100 ₪/חודש</div>
  </div>
</div>


<!-- ══ מסך צ'אט ══ -->
<div id="chat-screen">
  <div id="header">
    <div class="av"><img src="/logo" style="width:42px;height:42px;object-fit:cover;border-radius:10px" onerror="this.parentElement.innerHTML='🔧'"></div>
    <div style="flex:1"><div class="ht">סלקום טכנאים</div><div class="hs" id="st">מחובר</div></div>
    <button id="logout-btn" onclick="doLogout()">יציאה</button>
  </div>
  <div id="tabs" style="display:flex;background:#0d1b2a;border-bottom:1px solid #284461;flex-shrink:0;overflow-x:auto;scrollbar-width:none;-ms-overflow-style:none">
    <button class="tab-btn active" id="tab-tasks" onclick="switchTab('tasks')">📋 פקעות</button>
    <button class="tab-btn" id="tab-diag" onclick="switchTab('diag')">🔧 בדיקות</button>
    <button class="tab-btn" id="tab-inv" onclick="switchTab('inv')">🏭 מחסן</button>
    <button class="tab-btn" id="tab-rep" onclick="switchTab('rep')">📊 דוחות</button>
    <button class="tab-btn" id="tab-price" onclick="switchTab('price')">💰 מחירון</button>
    <button class="tab-btn" id="tab-admin" onclick="switchTab('admin')" style="display:none">🛡️ מנהל</button>
  </div>
  <div id="msgs">
    <div id="sched-tabs" style="display:flex;border-bottom:1px solid #284461;background:#0d1b2a;flex-shrink:0;direction:rtl">
      <button class="sched-tab" id="sched-tab-prev" onclick="switchSchedTab(-1)">אתמול</button>
      <button class="sched-tab active" id="sched-tab-today" onclick="switchSchedTab(0)">היום</button>
      <button class="sched-tab" id="sched-tab-next" onclick="switchSchedTab(1)">מחר</button>
      <button class="sched-tab" id="sched-tab-bank" onclick="switchSchedTab('bank')">📋 בנק פקעות</button>
    </div>
    <div class="qr" id="tasks-qr" style="display:none">
      <button class="qb" id="sort-btn" onclick="toggleSortByCity()">📍 מיין לפי מיקום</button>
      <button class="qb" onclick="loadTasks()">📋 רענן</button>
    </div>
  </div>
  <div id="inv-panel" style="display:none;flex:1;overflow-y:auto;direction:rtl">
    <!-- תת-לשוניות מחסן -->
    <div style="display:flex;border-bottom:1px solid #284461;background:#0d1b2a;position:sticky;top:0;z-index:5">
      <button id="sub-tab-inv" onclick="switchSubTab('inv')"
        style="flex:1;padding:10px;background:transparent;border:none;border-bottom:2px solid #3dba6f;color:#3dba6f;font-size:13px;font-weight:700;cursor:pointer">📦 מלאי</button>
      <button id="sub-tab-eqrep" onclick="switchSubTab('eqrep')"
        style="flex:1;padding:10px;background:transparent;border:none;border-bottom:2px solid transparent;color:#6b94b8;font-size:13px;cursor:pointer">🗂️ ציוד</button>
    </div>
    <!-- תוכן מלאי -->
    <div id="inv-content" style="padding:12px 10px"></div>
    <!-- תוכן ציוד -->
    <div id="eqrep-content" style="display:none;padding:12px 10px">
      <div style="display:flex;gap:8px;margin-bottom:12px">
        <button id="eqrep-tab-rcv" onclick="switchEqTab('rcv')"
          style="flex:1;padding:9px;border-radius:8px;border:none;background:#1f7a4a;color:#fff;font-size:14px;font-weight:700;cursor:pointer">
          📥 קיבלתי</button>
        <button id="eqrep-tab-ret" onclick="switchEqTab('ret')"
          style="flex:1;padding:9px;border-radius:8px;border:none;background:#223d58;color:#6b94b8;font-size:14px;cursor:pointer">
          📤 החזרות</button>
      </div>
      <div id="eqrep-rcv-list"></div>
      <div id="eqrep-ret-list" style="display:none"></div>
    </div>
  </div>
  <div id="rep-panel" style="display:none;flex:1;overflow-y:auto;padding:12px 10px;direction:rtl"></div>
  <div id="price-panel" style="display:none;flex:1;overflow-y:auto;padding:12px 10px;direction:rtl"></div>
  <div id="admin-panel" style="display:none;flex:1;overflow-y:auto;padding:12px 10px;direction:rtl"></div>
  <div id="diag-panel" style="display:none;flex:1;overflow-y:auto;padding:12px 10px;direction:rtl"></div>

  <!-- מודל סריקת MAC/סריאל -->
  <div id="scan-modal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,0.85);z-index:9000;flex-direction:column;align-items:center;justify-content:center;padding:20px">
    <div style="background:#d0e4f7;border-radius:16px;padding:18px;width:100%;max-width:420px;direction:rtl">
      <div style="font-size:17px;font-weight:700;margin-bottom:12px;color:#e9edef">📷 זיהוי ציוד</div>
      <img id="scan-preview" src="" style="width:100%;border-radius:10px;max-height:220px;object-fit:contain;background:#0d1b2a;display:none">
      <div id="scan-status" style="margin-top:10px;font-size:14px;min-height:40px;text-align:center"></div>
      <div id="scan-result" style="margin-top:4px"></div>
    </div>
  </div>
</div>

<script>
/* ── helpers ── */
function showScreen(name){
  ['login-screen','otp-screen','chat-screen','reg-screen','payment-screen'].forEach(id=>{
    const el=document.getElementById(id);
    if(el) el.classList.remove('active');
  });
  document.getElementById('chat-screen').classList.remove('active');
  if(name==='login') document.getElementById('login-screen').classList.add('active');
  else if(name==='otp') document.getElementById('otp-screen').classList.add('active');
  else if(name==='chat') document.getElementById('chat-screen').classList.add('active');
  else if(name==='reg') document.getElementById('reg-screen').classList.add('active');
  else if(name==='payment') document.getElementById('payment-screen').classList.add('active');
}

let _isAdmin=false;
let _userName='';
let _userPerms={premium:false,full_card:false,inventory:false,equipment_report:false,reports:false,prices:false};

/* ── apiFetch — מוסיף X-Token/X-Phone לכל בקשת API ── */
function apiFetch(url, opts={}){
  const token=localStorage.getItem('cc_token')||'';
  const phone=localStorage.getItem('cc_phone')||'';
  opts.headers=Object.assign({'X-Token':token,'X-Phone':phone},opts.headers||{});
  return fetch(url,opts).then(res=>{
    if(res.status===401){
      localStorage.removeItem('cc_token');
      showScreen('login');
      document.getElementById('l-err').textContent='פג תוקף החיבור — התחבר מחדש';
    }
    return res;
  });
}

function saveAuthLocally(token, phone){
  if(token) localStorage.setItem('cc_token', token);
  if(phone) localStorage.setItem('cc_phone', phone);
}

function showAdminTab(){
  const tb=document.getElementById('tab-admin');
  if(tb) tb.style.display='';
}

function applyPermissions(){
  // הלשוניות תמיד גלויות — רק sub-tab ציוד מוסתר לפי הרשאה
  const subEqrep=document.getElementById('sub-tab-eqrep');
  if(subEqrep) subEqrep.style.display=(_isAdmin||_userPerms.equipment_report||_userPerms.premium)?'':'none';
}

const NO_PERM_HTML=`<div style="display:flex;flex-direction:column;align-items:center;justify-content:center;height:100%;padding:40px 20px;text-align:center;gap:16px">
  <div style="font-size:52px">🔒</div>
  <div style="font-size:17px;font-weight:700;color:#edf2f5">אין הרשאה</div>
  <div style="font-size:14px;color:#8fa5b5;max-width:260px">פנה למנהל לקבלת גישה ללשונית זו</div>
</div>`;

async function handleSubStatus(sub_status){
  if(sub_status==='not_registered'){showScreen('reg');return true;}
  return false;
}

async function attemptLogin(phone, emp, silent){
  const errEl=document.getElementById('l-err');
  const btn=document.getElementById('l-btn');
  errEl.textContent='';
  if(!silent){btn.disabled=true;btn.textContent='מתחבר...';}
  try{
    // שלח טוקן שמור (אם יש) — השרת יאמת אותו
    const savedToken=localStorage.getItem('cc_token')||'';
    const r=await apiFetch('/api/login',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({phone,employee_id:emp,token:savedToken})});
    const d=await r.json();
    if(d.status==='ok'){
      // שמור טוקן שהשרת החזיר
      if(d.token) saveAuthLocally(d.token, d.phone||phone);
      const sd=await apiFetch('/api/status').then(r=>r.json());
      _isAdmin=sd.is_admin||false; _userName=sd.user_name||'';
      if(sd.permissions) _userPerms=Object.assign({},_userPerms,sd.permissions);
      if(_isAdmin) showAdminTab();
      applyPermissions();
      const handled=await handleSubStatus(sd.sub_status);
      if(!handled){showScreen('chat');initChat();}
    } else if(d.status==='otp'){
      showScreen('otp');
      setTimeout(()=>document.getElementById('otp-inp').focus(),100);
    } else {
      errEl.textContent=d.message||'שגיאה';
      showScreen('login');
    }
  }catch(e){
    errEl.textContent='שגיאת רשת';
    showScreen('login');
  }
  if(!silent){btn.disabled=false;btn.textContent='כניסה';}
}

/* ── on load ── */
(async function(){
  // נסה טוקן שמור ב-localStorage
  const phone=localStorage.getItem('cc_phone');
  const token=localStorage.getItem('cc_token');
  const emp=localStorage.getItem('cc_emp');
  if(phone && token){
    let d={logged_in:false};
    try{d=await apiFetch('/api/status').then(r=>r.json());}catch(e){}
    if(d.logged_in){
      _isAdmin=d.is_admin||false; _userName=d.user_name||'';
      if(d.permissions) _userPerms=Object.assign({},_userPerms,d.permissions);
      if(_isAdmin) showAdminTab();
      applyPermissions();
      const handled=await handleSubStatus(d.sub_status);
      if(!handled) showScreen('chat');
      return;
    }
  }
  // אין טוקן — נסה כניסה שקטה עם phone+emp
  if(phone && emp){
    const ok=await attemptLoginSilent(phone, emp);
    if(ok) return;
    const masked='05'+'*'.repeat(Math.max(0,phone.length-2));
    document.getElementById('l-phone').value='';
    document.getElementById('l-phone').placeholder=masked;
    document.getElementById('l-emp').value='';
    document.getElementById('l-err').textContent='';
    const hint=document.createElement('div');
    hint.id='quick-login-hint';
    hint.style.cssText='text-align:center;font-size:13px;color:#8fa5b5;margin-top:-6px';
    hint.innerHTML=`מחובר בד"כ כ <b style="color:#dde8ee">${masked}</b> — <span style="color:#3dba6f;cursor:pointer;text-decoration:underline" onclick="quickLogin()">כניסה מהירה</span>`;
    const card=document.querySelector('#login-screen .card');
    card.insertBefore(hint, card.firstChild);
  }
  showScreen('login');
})();

async function attemptLoginSilent(phone, emp){
  try{
    const savedToken=localStorage.getItem('cc_token')||'';
    const r=await apiFetch('/api/login',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({phone,employee_id:emp,token:savedToken})});
    const d=await r.json();
    if(d.status==='ok'){
      if(d.token) saveAuthLocally(d.token, d.phone||phone);
      const sd=await apiFetch('/api/status').then(r=>r.json());
      _isAdmin=sd.is_admin||false; _userName=sd.user_name||'';
      if(sd.permissions) _userPerms=Object.assign({},_userPerms,sd.permissions);
      if(_isAdmin) showAdminTab();
      applyPermissions();
      const handled=await handleSubStatus(sd.sub_status);
      if(!handled){showScreen('chat');initChat();}
      return true;
    }
    return false;
  }catch(e){return false;}
}

async function quickLogin(){
  const phone=localStorage.getItem('cc_phone');
  const emp=localStorage.getItem('cc_emp');
  if(phone&&emp) await attemptLogin(phone,emp,false);
}

/* ── login ── */
async function doLogin(){
  const phone=document.getElementById('l-phone').value.replace(/-/g,'').trim();
  const emp=document.getElementById('l-emp').value.trim();
  const errEl=document.getElementById('l-err');
  errEl.textContent='';
  if(!phone||!emp){errEl.textContent='יש למלא את כל השדות';return;}
  // save for future auto-login
  localStorage.setItem('cc_phone', phone);
  localStorage.setItem('cc_emp', emp);
  await attemptLogin(phone, emp, false);
}
document.getElementById('l-phone').addEventListener('keypress',e=>{if(e.key==='Enter')document.getElementById('l-emp').focus()});
document.getElementById('l-emp').addEventListener('keypress',e=>{if(e.key==='Enter')doLogin()});

/* ── OTP ── */
async function doVerify(){
  const otp=document.getElementById('otp-inp').value.trim();
  const errEl=document.getElementById('otp-err');
  const btn=document.getElementById('otp-btn');
  errEl.textContent='';
  if(!otp){errEl.textContent='הכנס קוד';return;}
  btn.disabled=true;btn.textContent='מאמת...';
  try{
    const r=await apiFetch('/api/verify',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({otp})});
    const d=await r.json();
    if(d.status==='ok'){
      // שמור טוקן ב-localStorage
      if(d.token) saveAuthLocally(d.token, d.phone||localStorage.getItem('cc_phone'));
      document.getElementById('otp-inp').value='';
      const sd=await apiFetch('/api/status').then(r=>r.json());
      _isAdmin=sd.is_admin||false; _userName=sd.user_name||'';
      if(sd.permissions) _userPerms=Object.assign({},_userPerms,sd.permissions);
      if(_isAdmin) showAdminTab();
      applyPermissions();
      const handled=await handleSubStatus(sd.sub_status);
      if(!handled){showScreen('chat');initChat();}
    } else {errEl.textContent=d.message||'קוד שגוי';}
  }catch(e){errEl.textContent='שגיאת רשת';}
  btn.disabled=false;btn.textContent='אמת קוד';
}

/* ── logout ── */
async function doLogout(){
  await apiFetch('/api/logout',{method:'POST'});
  localStorage.removeItem('cc_phone');
  localStorage.removeItem('cc_emp');
  localStorage.removeItem('cc_token');
  _isAdmin=false; _userName='';
  _userPerms={premium:false,full_card:false,inventory:false,equipment_report:false,reports:false,prices:false};
  const adminTab=document.getElementById('tab-admin');
  if(adminTab) adminTab.style.display='none';
  const msgsEl=document.getElementById('msgs');
  msgsEl.dataset.schedLoaded='';
  _resetMsgsContainer();
  showScreen('login');
}

/* ── chat ── */
function ts(){return new Date().toLocaleTimeString('he-IL',{hour:'2-digit',minute:'2-digit'})}
function md(t){return t.replace(/\*(.*?)\*/g,'<b>$1</b>')}
function add(t,type){
  const b=document.getElementById('msgs');
  const d=document.createElement('div');
  d.className='msg '+type;
  const safe=t.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
  d.innerHTML=md(safe)+'<div class="mtime">'+ts()+'</div>';
  b.appendChild(d);b.scrollTop=9999;
}
function s(t){document.getElementById('inp').value=t;go()}
function toggleAcc(el){
  el.classList.toggle('open');
  if(el.classList.contains('open')){
    const body=el.querySelector('.acc-body');
    if(body){
      const ibcDiv=body.querySelector('[id^="ibc_"]');
      if(ibcDiv&&!ibcDiv.dataset.ibcLoaded){
        ibcDiv.dataset.ibcLoaded='1';
        const elemId=ibcDiv.id.replace('ibc_','');
        if(ibcDiv.dataset.ibcUrl){
          // בנק פקעות — URL ישיר
          loadIbcData(elemId, ibcDiv.dataset.ibcUrl);
        } else if(ibcDiv.dataset.callId){
          // לוז — שולפים URL תחילה מ-getCallDetails
          loadIbcViaCallId(elemId, ibcDiv.dataset.callId, ibcDiv.dataset.customerId||'', ibcDiv.dataset.source||'JET');
        }
      }
    }
  }
}
async function loadIbcViaCallId(elemId, callId, customerId, source){
  const el=document.getElementById('ibc_'+elemId);
  if(!el) return;
  try{
    const r=await apiFetch(`/api/ibc-url?call_id=${encodeURIComponent(callId)}&customer_id=${encodeURIComponent(customerId)}&source=${encodeURIComponent(source)}`);
    const d=await r.json();
    if(d.error){
      // נסה שוב עם customer_id ריק אם לא הצליח
      const r2=await apiFetch(`/api/ibc-url?call_id=${encodeURIComponent(callId)}&customer_id=&source=`);
      const d2=await r2.json();
      if(d2.error){el.innerHTML=`<span style="color:#f15c6e;font-size:12px">IBC: ${d2.error}</span>`;return;}
      loadIbcData(elemId, d2.ibc_url);
      return;
    }
    loadIbcData(elemId, d.ibc_url);
  }catch(e){
    el.innerHTML=`<span style="color:#6b94b8;font-size:12px">IBC לא זמין</span>`;
  }
}
function _buildCardHTML(c){
  const repeatBadge=c.recent_visit
    ?`<span style="color:#f15c6e;font-size:12px;font-weight:700">⚠️ חשד לחזורת — <span style="text-decoration:underline">${c.recent_visit}</span></span>`
    :'';
  let structBadge='';
  if(_isAdmin && c.structure_type){
    const st=c.structure_type.toUpperCase();
    if(st.includes('LSB')) structBadge=`<span style="background:#1a4a2e;color:#3dba6f;border:1px solid #3dba6f;border-radius:6px;padding:1px 7px;font-size:11px;font-weight:700;margin-right:6px">LSB</span>`;
    else if(st==='PPC') structBadge=`<span style="background:#1a2d4a;color:#53bdeb;border:1px solid #53bdeb;border-radius:6px;padding:1px 7px;font-size:11px;font-weight:700;margin-right:6px">PPC</span>`;
  }
  let html='';
  if(_isAdmin||_userPerms.full_card){
    html=`<div style="font-weight:700;font-size:15px;color:#25d366;margin-bottom:6px;display:flex;justify-content:space-between;align-items:center">
      <span>${c._num?`<span style="color:#8696a0;font-size:13px;margin-left:6px">${c._num}.</span>`:''}📋 קריאה ${c.call_id}${structBadge}</span>${repeatBadge}</div>`;
    if(c.date)    html+=`<div>📅 ${c.date}&nbsp;&nbsp;⏰ ${c.time}</div>`;
    html+=`<div>👤 ${c.name}&nbsp;&nbsp;<span style="color:#8696a0;font-size:12px">${c.customer_id}</span></div>`;
    if(c.phone)   html+=`<div>📞 <a href="tel:${c.phone}" style="color:#53bdeb">${c.phone}</a></div>`;
    html+=`<div>📍 ${c.address}${c._dist?` <span style="color:#53bdeb;font-size:12px">(${c._dist} ק"מ)</span>`:''}</div>`;
    html+=`<div>🔧 ${c.task_type}</div>`;
    if(c.infra)   html+=`<div>🌐 ${c.infra}${c.technology?' · '+c.technology:''}</div>`;
    const _isIbcCard=(c.ibc_url&&_isAdmin&&(c.infra_code==='FB'||c.infra_code==='IB'||!c.infra_code));
    const _isBezeq=_isAdmin&&(c.infra_code==='BN'||c.infra_code==='BF'||c.infra_code==='NV');
    if(_isIbcCard){
      const ibcId='ibc_'+c.call_id;
      html+=`<div class="acc" onclick="toggleAcc(this)">
        <div class="acc-hdr"><span>🏢 נתוני IBC</span><span class="acc-arr">▾</span></div>
        <div class="acc-body"><div id="${ibcId}" style="font-size:13px"><span style="color:#6b94b8">⏳ טוען...</span></div></div>
      </div>`;
    }
    if(_isBezeq&&(c.line_code||c.internet_user||c.router_model||c.technology)){
      html+=`<div style="margin:6px 0;padding:10px 12px;background:#ddeeff;border-radius:8px;border-right:3px solid #1a6aaa;font-size:14px">`;
      html+=`<div style="font-weight:700;color:#0d3a6a;margin-bottom:4px">🏗️ נתוני תשתית בזק</div>`;
      if(c.technology)    html+=`<div>⚡ טכנולוגיה: <b>${c.technology}</b></div>`;
      if(c.line_code)     html+=`<div>🔑 קוד קו: <span style="font-family:monospace;font-size:14px;background:#c8e0f8;padding:1px 6px;border-radius:4px">${c.line_code}</span></div>`;
      if(c.internet_user) html+=`<div>👤 משתמש: <span style="font-family:monospace;font-size:13px">${c.internet_user}</span></div>`;
      if(c.router_model)  html+=`<div>📡 נתב: ${c.router_model}${c.router_serial?` <span style="color:#5a7a9a;font-size:12px">(${c.router_serial})</span>`:''}</div>`;
      html+=`</div>`;
    }
    if(c.comment) html+=`<div>💬 ${c.comment}</div>`;
    if(_isAdmin&&(c.planned||c.existing||c.tv)){
      html+=`<div class="acc" onclick="toggleAcc(this)"><div class="acc-hdr"><span>⚙️ ציוד</span><span class="acc-arr">▾</span></div><div class="acc-body">`;
      if(c.planned)  html+=`<div>📦 להתקנה: ${c.planned}</div>`;
      if(c.existing) html+=`<div>🖥 קיים: ${c.existing}</div>`;
      if(c.tv)       html+=`<div>📺 TV: ${c.tv}</div>`;
      html+=`</div></div>`;
    }
    if(_isAdmin&&!_isBezeq&&(c.line_code||c.internet_user||c.router_model)){
      html+=`<div class="acc" onclick="toggleAcc(this)"><div class="acc-hdr"><span>🌐 נתוני רשת</span><span class="acc-arr">▾</span></div><div class="acc-body">`;
      if(c.line_code)     html+=`<div>🔑 קוד קו: <span style="font-family:monospace;font-size:14px">${c.line_code}</span></div>`;
      if(c.internet_user) html+=`<div>👤 משתמש אינטרנט: <span style="font-family:monospace;font-size:13px">${c.internet_user}</span></div>`;
      if(c.router_model){
        const serial=c.router_serial?` <span style="color:#8696a0;font-size:12px">(${c.router_serial})</span>`:'';
        html+=`<div>📡 נתב: ${c.router_model}${serial}</div>`;
      }
      html+=`</div></div>`;
    }
    if(c.history) html+=`<div class="acc" onclick="toggleAcc(this)"><div class="acc-hdr"><span>📜 היסטוריה</span><span class="acc-arr">▾</span></div><div class="acc-body">${c.history.replace(/\n/g,'<br>')}</div></div>`;
  } else {
    html=`<div style="font-weight:700;font-size:15px;color:#25d366;margin-bottom:6px;display:flex;justify-content:space-between;align-items:center">
      <span>${c._num?`<span style="color:#8696a0;font-size:13px;margin-left:6px">${c._num}.</span>`:''}📋 קריאה ${c.call_id}</span>${repeatBadge}</div>`;
    if(c.date)  html+=`<div>📅 ${c.date}&nbsp;&nbsp;⏰ ${c.time}</div>`;
    html+=`<div>👤 ${c.name}</div>`;
    if(c.phone) html+=`<div>📞 <a href="tel:${c.phone}" style="color:#53bdeb">${c.phone}</a></div>`;
    html+=`<div>📍 ${c.address}${c._dist?` <span style="color:#53bdeb;font-size:12px">(${c._dist} ק"מ)</span>`:''}</div>`;
    html+=`<div>🔧 ${c.task_type}</div>`;
    if(c.history) html+=`<div class="acc" onclick="toggleAcc(this)"><div class="acc-hdr"><span>📜 היסטוריה</span><span class="acc-arr">▾</span></div><div class="acc-body">${c.history.replace(/\n/g,'<br>')}</div></div>`;
  }
  return html;
}

async function openBankCard(c){
  let ov=document.getElementById('bank-card-overlay');
  if(!ov){
    ov=document.createElement('div');
    ov.id='bank-card-overlay';
    ov.style.cssText='position:fixed;inset:0;background:#0a1826;z-index:8000;overflow-y:auto;direction:rtl;padding:12px';
    document.body.appendChild(ov);
  }
  // ── שלב 1: הצג מיד עם נתונים בסיסיים (אפס המתנה) ──
  ov.innerHTML=`
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;position:sticky;top:0;background:#0a1826;padding:6px 0;z-index:1">
      <button onclick="closeBankCard()" style="background:none;border:1px solid #3d5166;color:#8696a0;border-radius:8px;padding:6px 14px;cursor:pointer;font-size:14px">← חזרה</button>
    </div>
    <div id="bank-card-inner" style="background:#d0e4f7;color:#0d1b2a;border-radius:10px;padding:14px 16px;direction:rtl;border-right:4px solid #25d366;font-size:15.5px;font-weight:500;line-height:2">
      ${_buildCardHTML(c)}
      <div id="card-detail-loader" style="text-align:center;color:#4a7aaa;font-size:12px;padding:8px 0;margin-top:4px">⏳ טוען תשתית והיסטוריה...</div>
    </div>`;
  ov.style.display='block';
  ov.ontouchstart=null; ov.ontouchend=null;
  // ── שלב 2: טען תשתית+היסטוריה ברקע — מחליף את ה-loader בלבד, לא את כל הכרטיס ──
  try{
    const r=await apiFetch('/api/task-detail',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({task:c,index:c._num||1})
    });
    const detail=await r.json();
    // עדכן c רק בשדות שנטענו חדש (תשתית, היסטוריה וכד')
    const enrichFields=['infra','technology','planned','existing','tv','history',
      'recent_visit','ibc_url','infra_code','line_code','internet_user',
      'router_model','router_serial','structure_type'];
    for(const k of enrichFields){
      if(detail[k]!==null && detail[k]!==undefined && detail[k]!=='' && detail[k]!==false)
        c[k]=detail[k];
    }
    // בנה רק את החלק שנטען (תשתית + היסטוריה) ושים במקום ה-loader
    const loader=document.getElementById('card-detail-loader');
    if(!loader) return;
    let extraHtml='';
    const iib=(c.ibc_url&&_isAdmin&&(c.infra_code==='FB'||c.infra_code==='IB'||!c.infra_code));
    const isBezeq=_isAdmin&&(c.infra_code==='BN'||c.infra_code==='BF'||c.infra_code==='NV');
    if(c.infra) extraHtml+=`<div>🌐 ${c.infra}${c.technology?' · '+c.technology:''}</div>`;
    if(iib){
      const ibcId='ibc_'+c.call_id;
      extraHtml+=`<div class="acc" onclick="toggleAcc(this)">
        <div class="acc-hdr"><span>🏢 נתוני IBC</span><span class="acc-arr">▾</span></div>
        <div class="acc-body"><div id="${ibcId}" style="font-size:13px"><span style="color:#6b94b8">⏳ טוען...</span></div></div>
      </div>`;
    }
    if(isBezeq&&(c.line_code||c.internet_user||c.router_model||c.technology)){
      extraHtml+=`<div style="margin:6px 0;padding:10px 12px;background:#ddeeff;border-radius:8px;border-right:3px solid #1a6aaa;font-size:14px">`;
      extraHtml+=`<div style="font-weight:700;color:#0d3a6a;margin-bottom:4px">🏗️ נתוני תשתית בזק</div>`;
      if(c.technology)    extraHtml+=`<div>⚡ טכנולוגיה: <b>${c.technology}</b></div>`;
      if(c.line_code)     extraHtml+=`<div>🔑 קוד קו: <span style="font-family:monospace;font-size:14px;background:#c8e0f8;padding:1px 6px;border-radius:4px">${c.line_code}</span></div>`;
      if(c.internet_user) extraHtml+=`<div>👤 משתמש: <span style="font-family:monospace;font-size:13px">${c.internet_user}</span></div>`;
      if(c.router_model)  extraHtml+=`<div>📡 נתב: ${c.router_model}${c.router_serial?` <span style="color:#5a7a9a;font-size:12px">(${c.router_serial})</span>`:''}</div>`;
      extraHtml+=`</div>`;
    }
    if(_isAdmin&&!isBezeq&&(c.line_code||c.internet_user||c.router_model)){
      extraHtml+=`<div class="acc" onclick="toggleAcc(this)"><div class="acc-hdr"><span>🌐 נתוני רשת</span><span class="acc-arr">▾</span></div><div class="acc-body">`;
      if(c.line_code)     extraHtml+=`<div>🔑 קוד קו: <span style="font-family:monospace;font-size:14px">${c.line_code}</span></div>`;
      if(c.internet_user) extraHtml+=`<div>👤 משתמש אינטרנט: <span style="font-family:monospace;font-size:13px">${c.internet_user}</span></div>`;
      if(c.router_model)  extraHtml+=`<div>📡 נתב: ${c.router_model}${c.router_serial?` <span style="color:#8696a0;font-size:12px">(${c.router_serial})</span>`:''}</div>`;
      extraHtml+=`</div></div>`;
    }
    if(_isAdmin&&(c.planned||c.existing||c.tv)){
      extraHtml+=`<div class="acc" onclick="toggleAcc(this)"><div class="acc-hdr"><span>⚙️ ציוד</span><span class="acc-arr">▾</span></div><div class="acc-body">`;
      if(c.planned)  extraHtml+=`<div>📦 להתקנה: ${c.planned}</div>`;
      if(c.existing) extraHtml+=`<div>🖥 קיים: ${c.existing}</div>`;
      if(c.tv)       extraHtml+=`<div>📺 TV: ${c.tv}</div>`;
      extraHtml+=`</div></div>`;
    }
    if(c.history) extraHtml+=`<div class="acc" onclick="toggleAcc(this)"><div class="acc-hdr"><span>📜 היסטוריה</span><span class="acc-arr">▾</span></div><div class="acc-body">${c.history.replace(/\n/g,'<br>')}</div></div>`;
    // החלף את ה-loader בתוכן החדש
    loader.outerHTML=extraHtml||'';
    // IBC lazy-load
    if(iib){
      const ibcEl=document.getElementById('ibc_'+c.call_id);
      if(ibcEl) ibcEl.dataset.ibcUrl=c.ibc_url;
    }
    // עדכן חשד לחוזרת אם נוסף
    if(detail.recent_visit){
      const hdr=document.querySelector('#bank-card-inner [data-repeat-badge]');
      if(!hdr){
        const firstDiv=document.querySelector('#bank-card-inner div');
        if(firstDiv) firstDiv.insertAdjacentHTML('afterend',
          `<div data-repeat-badge style="color:#c0392b;font-size:13px;font-weight:700">⚠️ חשד לחוזרת — <span style="text-decoration:underline">${detail.recent_visit}</span></div>`);
      }
    }
  }catch(e){
    const loader=document.getElementById('card-detail-loader');
    if(loader) loader.textContent='⚠️ לא ניתן לטעון פרטי תשתית';
  }
}

function closeBankCard(){
  const ov=document.getElementById('bank-card-overlay');
  if(ov) ov.style.display='none';
}

function addCard(c,box){
  if(!box) box=document.getElementById('msgs');
  const card=document.createElement('div');
  card.className='task-card';
  card.dataset.callid=c.call_id||'';

  // בנק פקעות — כרטיסייה קומפקטית + לחיצה פותחת overlay
  if(_schedTabOffset==='bank'){
    card.style.cssText='background:#d0e4f7;color:#0d1b2a;border-radius:10px;padding:11px 14px;direction:rtl;border-right:4px solid #25d366;font-size:14px;font-weight:500;line-height:1.8;margin:4px 0;cursor:pointer;user-select:none';
    const dist=c._dist?` <span style="color:#1a6a9a;font-size:12px">${c._dist} ק"מ</span>`:'';
    const repBadge=c.recent_visit
      ?`<div style="color:#c0392b;font-size:12px;font-weight:700;margin-top:2px">⚠️ חשד לחוזרת — <span style="text-decoration:underline">${c.recent_visit}</span></div>`
      :'';
    card.innerHTML=`
      <div style="display:flex;justify-content:space-between;align-items:center">
        <span style="font-weight:700;color:#1a3a5c;font-size:13px">${c._num?`<span style="color:#8696a0">${c._num}. </span>`:''}${c.call_id}</span>
        <span style="font-size:12px;color:#1a6a9a">${c.task_type||''}</span>
      </div>
      <div style="font-size:13px;color:#2a4a6a">📍 ${c.address||''}${dist}</div>
      ${c.date?`<div style="font-size:12px;color:#4a6a8a">📅 ${c.date} ⏰ ${c.time||''}</div>`:''}
      ${repBadge}`;
    card.onclick=()=>openBankCard(c);
    box.appendChild(card);
    return;
  }

  card.style.cssText='background:#d0e4f7;color:#0d1b2a;border-radius:10px;padding:14px 16px;direction:rtl;border-right:4px solid #25d366;font-size:15.5px;font-weight:500;line-height:2;margin:4px 0';
  card.innerHTML=_buildCardHTML(c);
  box.appendChild(card);
  box.scrollTop=9999;
  if(c.ibc_url&&_isAdmin){
    const ibcEl=document.getElementById('ibc_'+c.call_id);
    if(ibcEl) ibcEl.dataset.ibcUrl=c.ibc_url;
  }
}

async function loadIbcData(callId, ibcUrl){
  const el=document.getElementById('ibc_'+callId);
  if(!el) return;
  try{
    const r=await apiFetch('/api/ibc-data?url='+encodeURIComponent(ibcUrl));
    const d=await r.json();
    if(d.error){el.innerHTML=`<span style="color:#f15c6e;font-size:12px">IBC: ${d.error}</span>`;return;}
    const st=d.site_type||'';
    const stColor=st.includes('LSB')?'#3dba6f':st==='PPC'?'#53bdeb':'#8fa5b5';
    const stLabel=st.includes('LSB')?'LSB':st==='PPC'?'PPC':st;
    let html=`<div style="color:#25d366;font-weight:700;font-size:12px;margin-bottom:5px">🏢 נתוני IBC</div>`;
    html+=`<div style="display:grid;grid-template-columns:1fr 1fr;gap:3px 10px;font-size:12px;color:#c5d5e0">`;
    if(stLabel) html+=`<div>סוג: <b style="color:${stColor}">${stLabel}</b></div>`;
    if(d.aron_num) html+=`<div>ארון: <b>${d.aron_num}</b></div>`;
    if(d.aron_location) html+=`<div>מיקום: <b>${d.aron_location}</b></div>`;
    if(d.floor) html+=`<div>קומת ארון: <b>${d.floor}</b></div>`;
    if(d.floors_total) html+=`<div>קומות: <b>${d.floors_total}</b></div>`;
    if(d.apartments) html+=`<div>דירות: <b>${d.apartments}</b></div>`;
    if(d.technology) html+=`<div>טכנולוגיה: <b>${d.technology}</b></div>`;
    if(d.eqp_type) html+=`<div>ציוד: <b>${d.eqp_type}</b></div>`;
    html+=`</div>`;
    el.innerHTML=html;
  }catch(e){
    el.innerHTML=`<span style="color:#6b94b8;font-size:12px">IBC לא זמין</span>`;
  }
}

function _resetMsgsContainer(){
  /* rebuilds sched-tabs + qr bar, leaves the rest empty */
  const box=document.getElementById('msgs');
  box.innerHTML=`
    <div id="sched-tabs" style="display:flex;border-bottom:1px solid #284461;background:#0d1b2a;flex-shrink:0;direction:rtl">
      <button class="sched-tab" id="sched-tab-prev" onclick="switchSchedTab(-1)">אתמול</button>
      <button class="sched-tab" id="sched-tab-today" onclick="switchSchedTab(0)">היום</button>
      <button class="sched-tab" id="sched-tab-next" onclick="switchSchedTab(1)">מחר</button>
      <button class="sched-tab" id="sched-tab-bank" onclick="switchSchedTab('bank')">📋 בנק פקעות</button>
    </div>
    <div class="qr" id="tasks-qr" style="display:none">
      <button class="qb" id="sort-btn" onclick="toggleSortByCity()">📍 מיין לפי מיקום</button>
      <button class="qb" onclick="loadTasks()">📋 רענן</button>
    </div>`;
  return box;
}

function switchSchedTab(v){
  _schedTabOffset=v;
  // update active tab styling
  const tabs=['prev','today','next','bank'];
  const map={'-1':'prev','0':'today','1':'next','bank':'bank'};
  const key=String(v);
  tabs.forEach(t=>{
    const el=document.getElementById('sched-tab-'+t);
    if(el) el.classList.toggle('active', map[key]===t);
  });
  if(v==='bank'){
    // hide sort bar initially (loadTasks will show it)
    const qr=document.getElementById('tasks-qr');
    if(qr) qr.style.display='none';
    loadTasks();
  } else {
    const qr=document.getElementById('tasks-qr');
    if(qr) qr.style.display='none';
    loadTodaySchedule(Number(v));
  }
}

async function loadTodaySchedule(offset=0){
  // ביטול טעינה קודמת
  if(_loadSchedAbort){_loadSchedAbort.abort();}
  if(_loadTasksAbort){_loadTasksAbort.abort();}
  _loadSchedAbort=new AbortController();
  const sig=_loadSchedAbort.signal;

  const box=document.getElementById('msgs');
  const existing=document.getElementById('sched-tabs');
  if(!existing){_resetMsgsContainer();}
  Array.from(box.children).forEach(c=>{
    if(c.id!=='sched-tabs'&&c.id!=='tasks-qr') c.remove();
  });
  const qr=document.getElementById('tasks-qr');
  if(qr) qr.style.display='none';
  _allCards=[];_sortedByCity=false;

  const d0=new Date(); d0.setDate(d0.getDate()+offset);
  const dateStr=d0.getFullYear()+'-'+String(d0.getMonth()+1).padStart(2,'0')+'-'+String(d0.getDate()).padStart(2,'0');
  const hdr=document.createElement('div');
  hdr.className='tasks-hdr';
  hdr.style.cssText='color:#8696a0;font-size:13px;text-align:right;margin:6px 0';
  hdr.textContent='טוען לוז מסלקום... (עשוי לקחת עד דקה)';
  box.appendChild(hdr);
  try{
    const ctrl2=new AbortController();
    const to2=setTimeout(()=>ctrl2.abort(),50000);
    let r;
    try{ r=await apiFetch('/api/today-schedule?date='+dateStr,{signal:ctrl2.signal}); }
    finally{ clearTimeout(to2); }
    if(sig.aborted) return;
    if(r.status===401){hdr.textContent='❌ טוקן פג תוקף — התנתק והתחבר מחדש';return;}
    if(!r.ok){hdr.textContent=`❌ שגיאת שרת ${r.status}`;console.error('sched err',r.status,await r.text());return;}
    const d=await r.json();
    if(sig.aborted) return;
    if(d.error){hdr.textContent=`❌ ${d.error}`;return;}
    if(!d.schedule||d.schedule.length===0){hdr.textContent=`אין ביקורים מתוזמנים (${d.date||''})`;return;}
    hdr.textContent=`📅 לוז ${d.date} — ${d.count} ביקורים`;
    d.schedule.forEach((v,i)=>{
      if(sig.aborted) return;
      const card=document.createElement('div');
      card.className='task-card';
      card.dataset.callid=v.call_id||'';
      card.style.cssText='background:#d0e4f7;color:#0d1b2a;border-radius:10px;padding:14px 16px;direction:rtl;border-right:4px solid #f9c846;font-size:15.5px;font-weight:500;line-height:2;margin:4px 0';
      const st=v.structure_type||'';
      let structBadge='';
      if(_isAdmin&&st){
        if(st.includes('LSB')) structBadge=`<span style="background:#1a4a2e;color:#3dba6f;border:1px solid #3dba6f;border-radius:6px;padding:1px 7px;font-size:11px;font-weight:700;margin-right:6px">LSB</span>`;
        else if(st==='PPC') structBadge=`<span style="background:#1a2d4a;color:#53bdeb;border:1px solid #53bdeb;border-radius:6px;padding:1px 7px;font-size:11px;font-weight:700;margin-right:6px">PPC</span>`;
      }
      const statusColor=v.status==='הושלם'?'#3dba6f':v.status==='נשלח לטכנאי'?'#f9c846':'#8fa5b5';
      let html=`<div style="font-weight:700;font-size:15px;color:#0d4a8a;margin-bottom:6px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:4px">
        <span><span style="color:#5a7a9a;font-size:13px;margin-left:6px">${i+1}.</span>📅 קריאה ${v.call_id}${structBadge}</span>
        <span style="font-size:12px;color:${statusColor}">${v.status||''}</span>
        <span id="rep_sched_${v.call_id}" style="display:none;color:#f15c6e;font-size:12px;font-weight:700;width:100%"></span>
      </div>`;
      html+=`<div>⏰ ${v.time_start||'—'}–${v.time_end||'—'}</div>`;
      html+=`<div>👤 ${v.name||''}${v.is_vip?' ⭐':''}</div>`;
      if(v.phone) html+=`<div>📞 <a href="tel:${v.phone}" style="color:#53bdeb">${v.phone}</a></div>`;
      html+=`<div>📍 ${v.address||''}</div>`;
      html+=`<div>🔧 ${v.task_type||''}</div>`;
      if(v.infra) html+=`<div>🌐 ${v.infra}</div>`;
      if(v.comment) html+=`<div>💬 ${v.comment}</div>`;
      const isIbc=(v.infra_code==='FB'||v.infra_code==='IB');
      if(isIbc){
        const ibcId='ibc_sched_'+v.call_id;
        html+=`<div class="acc" onclick="toggleAcc(this)">
          <div class="acc-hdr"><span>🏢 נתוני IBC</span><span class="acc-arr">▾</span></div>
          <div class="acc-body"><div id="${ibcId}" style="font-size:13px"><span style="color:#6b94b8">⏳ טוען...</span></div></div>
        </div>`;
      }
      card.innerHTML=html;
      if(isIbc){
        const ibcEl=card.querySelector('[id^="ibc_sched_"]');
        if(ibcEl){
          ibcEl.dataset.callId=v.call_id;
          ibcEl.dataset.customerId=v.customer_id||'';
          ibcEl.dataset.source=v.source||'JET';
        }
      }
      // כפתור בדיקות — רק בלוז (היום/מחר/אתמול), לא בבנק פקעות
      _cardStore[v.call_id||''] = v;
      const diagBtn=document.createElement('button');
      diagBtn.style.cssText='margin-top:8px;width:100%;background:linear-gradient(135deg,#0d2a1a,#0d1b2a);border:1px solid #3dba6f;color:#3dba6f;border-radius:8px;padding:9px 12px;font-size:13px;font-weight:700;cursor:pointer;direction:rtl;display:flex;align-items:center;justify-content:center;gap:6px';
      diagBtn.innerHTML='🔧 בדיקות לפקעה זו';
      diagBtn.onclick=()=>openDiagForCard(v.call_id||'');
      card.appendChild(diagBtn);

      _allCards.push(v);
      box.appendChild(card);
    });
    if(!sig.aborted){
      box.scrollTop=0;
      // בדוק חוזרות בbackground — מעדכן badge אחרי
      _checkScheduleRepeats(d.schedule, sig);
    }
  }catch(e){
    if(e.name==='AbortError') return;
    hdr.textContent='❌ שגיאה בטעינת לוז';
  }
}

async function _checkScheduleRepeats(tasks, sig){
  if(!tasks||!tasks.length) return;
  try{
    const r=await apiFetch('/api/check-repeats-batch',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({tasks}),
      signal:sig
    });
    if(!r.ok) return;
    const results=await r.json();
    for(const [callId, techName] of Object.entries(results)){
      if(!techName) continue;
      const el=document.getElementById('rep_sched_'+callId);
      if(el){
        el.textContent=`⚠️ חשד לחוזרת — ${techName}`;
        el.style.display='block';
      }
    }
  }catch(e){}
}

let _loadTasksAbort=null;
let _loadSchedAbort=null;
let _schedTabOffset=0;   // לשונית פעילה: -1 אתמול | 0 היום | 1 מחר | 'bank' בנק
let _allCards=[];
let _sortedByCity=false;
let _districtFilter='';  // לא בשימוש (נשאר לתאימות)
let _cityFilter='';      // עיר פעילה לסינון
let _searchFilter='';    // לא בשימוש (נשאר לתאימות)
let _visitTypeFilter=''; // '' | 'קריאת שירות' | 'התקנה'

function renderCards(cards,box,hdr,count){
  const existingCards=box.querySelectorAll('.task-card');
  existingCards.forEach(el=>el.remove());
  cards.forEach(c=>addCard(c,box));
  hdr.textContent=`📋 בנק פקעות — ${new Date().toLocaleDateString('he-IL')} | סה"כ: ${count} פקעות`;
}

function _filteredCards(){
  let cards=_allCards;
  if(_cityFilter){const cq=_cityFilter.toLowerCase();cards=cards.filter(c=>(c.city||'').toLowerCase().includes(cq));}
  if(_visitTypeFilter){cards=cards.filter(c=>(c.appointment_type||c.task_type||'').includes(_visitTypeFilter));}
  return cards;
}

function _renderWithFilters(){
  const box=document.getElementById('msgs');
  const hdr=box.querySelector('.tasks-hdr');
  const cards=_filteredCards();
  box.querySelectorAll('.task-card').forEach(el=>el.remove());
  cards.forEach(c=>addCard(c,box));
  const parts=[];
  if(_cityFilter) parts.push(_cityFilter);
  if(_visitTypeFilter) parts.push(_visitTypeFilter);
  const label=parts.length?parts.join(' › ')+' — ':'';
  if(hdr) hdr.textContent=`📋 בנק פקעות — ${new Date().toLocaleDateString('he-IL')} | ${label}סה"כ: ${cards.length} פקעות`;
}

function _buildSearchBar(){
  // הוסר — נשאר רק סינון לפי עיר
  document.getElementById('bank-search-bar')?.remove();
  _buildCityBar();
}

function applySearchFilter(q){ /* לא בשימוש */ }

function _buildDistrictBar(allCards){
  // הוסר — נשאר רק סינון לפי עיר
  document.getElementById('district-filter-bar')?.remove();
  _buildCityBar();
}

function applyDistrictFilter(v){ /* לא בשימוש */ }

function _buildCityBar(){
  document.getElementById('city-filter-bar')?.remove();
  const qr=document.getElementById('tasks-qr');
  if(!qr) return;
  const cBar=document.createElement('div');
  cBar.id='city-filter-bar';
  cBar.style.cssText='display:flex;gap:6px;align-items:center;padding:6px 0;direction:rtl;flex-wrap:wrap';
  cBar.innerHTML=`
    <input id="city-search-input" type="text" placeholder="🏙️ סינון לפי עיר"
      value="${_cityFilter}"
      style="flex:1;min-width:120px;background:#0d1f33;border:1px solid #1e3a5f;color:#c9d8e8;border-radius:8px;
             padding:7px 12px;font-size:13px;direction:rtl;outline:none"
      oninput="applyCityFilter(this.value)"
      onkeydown="if(event.key==='Escape'){applyCityFilter('');this.value='';}">
    <button id="vt-btn-service" onclick="toggleVisitType('קריאת שירות')"
      style="padding:7px 13px;border-radius:8px;border:1px solid #3d6fa8;font-size:13px;cursor:pointer;flex-shrink:0;
             background:${_visitTypeFilter==='קריאת שירות'?'#1a3d5c':'#0d1f33'};
             color:${_visitTypeFilter==='קריאת שירות'?'#53bdeb':'#6b94b8'}">📞 קריאת שירות</button>
    <button id="vt-btn-install" onclick="toggleVisitType('התקנה')"
      style="padding:7px 13px;border-radius:8px;border:1px solid #3d6fa8;font-size:13px;cursor:pointer;flex-shrink:0;
             background:${_visitTypeFilter==='התקנה'?'#1a3d5c':'#0d1f33'};
             color:${_visitTypeFilter==='התקנה'?'#53bdeb':'#6b94b8'}">🔧 התקנה</button>`;
  qr.after(cBar);
}

function applyCityFilter(v){
  _cityFilter=v.trim();
  _renderWithFilters();
}

function toggleVisitType(type){
  _visitTypeFilter=(_visitTypeFilter===type)?'':type;
  // עדכן סגנון הכפתורים
  const sBtn=document.getElementById('vt-btn-service');
  const iBtn=document.getElementById('vt-btn-install');
  if(sBtn){
    const on=_visitTypeFilter==='קריאת שירות';
    sBtn.style.background=on?'#1a3d5c':'#0d1f33';
    sBtn.style.color=on?'#53bdeb':'#6b94b8';
    sBtn.style.borderColor=on?'#53bdeb':'#3d6fa8';
  }
  if(iBtn){
    const on=_visitTypeFilter==='התקנה';
    iBtn.style.background=on?'#1a3d5c':'#0d1f33';
    iBtn.style.color=on?'#53bdeb':'#6b94b8';
    iBtn.style.borderColor=on?'#53bdeb':'#3d6fa8';
  }
  _renderWithFilters();
}
function haversine(lat1,lon1,lat2,lon2){
  const R=6371, dLat=(lat2-lat1)*Math.PI/180, dLon=(lon2-lon1)*Math.PI/180;
  const a=Math.sin(dLat/2)**2+Math.cos(lat1*Math.PI/180)*Math.cos(lat2*Math.PI/180)*Math.sin(dLon/2)**2;
  return R*2*Math.atan2(Math.sqrt(a),Math.sqrt(1-a));
}
async function toggleSortByCity(){
  const box=document.getElementById('msgs');
  const btn=document.getElementById('sort-btn');
  if(_sortedByCity){
    _sortedByCity=false;
    btn.textContent='📍 מיין לפי מיקום';
    const hdr=box.querySelector('.tasks-hdr');
    renderCards(_allCards,box,hdr,_allCards.length);
    return;
  }
  btn.textContent='⏳ מאתר מיקום...';
  btn.disabled=true;
  try{
    const pos=await new Promise((res,rej)=>navigator.geolocation.getCurrentPosition(res,rej,{timeout:8000}));
    const uLat=pos.coords.latitude, uLng=pos.coords.longitude;
    btn.textContent='⏳ מחשב מרחקים...';
    const cards=await Promise.all(_allCards.map(async c=>{
      if(c._lat!=null) return c;
      const q=encodeURIComponent((c.city||c.address||''));
      try{const r=await apiFetch('/api/geocode?q='+q);const d=await r.json();if(d.lat){c._lat=d.lat;c._lng=d.lng;}}catch(e){}
      return c;
    }));
    const sorted=[...cards].sort((a,b)=>{
      const da=a._lat!=null?haversine(uLat,uLng,a._lat,a._lng):9999;
      const db=b._lat!=null?haversine(uLat,uLng,b._lat,b._lng):9999;
      return da-db;
    });
    sorted.forEach(c=>{if(c._lat!=null)c._dist=haversine(uLat,uLng,c._lat,c._lng).toFixed(1);});
    _sortedByCity=true;
    const hdr=box.querySelector('.tasks-hdr');
    renderCards(sorted,box,hdr,_allCards.length);
    btn.textContent='🔢 מיון מקורי';
  }catch(e){
    btn.textContent='📍 מיין לפי מיקום';
    alert('לא ניתן לאתר מיקום. אפשר גישה ל-GPS בהגדרות.');
  }
  btn.disabled=false;
}
async function loadTasks(district=''){
  if(_loadSchedAbort){_loadSchedAbort.abort();}
  if(_loadTasksAbort){_loadTasksAbort.abort();}
  _loadTasksAbort=new AbortController();
  _allCards=[];_sortedByCity=false;
  const sig=_loadTasksAbort.signal;
  const box=document.getElementById('msgs');
  /* preserve sched-tabs, clear content */
  const existing2=document.getElementById('sched-tabs');
  if(!existing2){_resetMsgsContainer();}
  Array.from(box.children).forEach(c=>{if(c.id!=='sched-tabs'&&c.id!=='tasks-qr')c.remove();});
  const qr2=document.getElementById('tasks-qr');
  if(qr2){qr2.style.display='flex';const sb=document.getElementById('sort-btn');if(sb)sb.textContent='📍 מיין לפי מיקום';}
  document.getElementById('st').textContent='טוען...';
  const hdr=document.createElement('div');
  hdr.className='tasks-hdr';
  hdr.style.cssText='color:#8696a0;font-size:13px;text-align:right;margin:6px 0';
  hdr.textContent='מושך פקעות מסלקום... (עשוי לקחת עד דקה)';
  box.appendChild(hdr);
  box.scrollTop=9999;
  try{
    const url='/api/tasks'+(district?'?district='+encodeURIComponent(district):'');
    const ctrl=new AbortController();
    const timeout=setTimeout(()=>ctrl.abort(),65000);  // 65s — server timeout is 50s
    let r;
    try{ r=await apiFetch(url,{signal:ctrl.signal}); }
    finally{ clearTimeout(timeout); }
    if(r.status===401){hdr.textContent='❌ טוקן פג תוקף — התנתק והתחבר מחדש';document.getElementById('st').textContent='מחובר';return;}
    if(!r.ok){
      let errMsg=`שגיאת שרת ${r.status}`;
      try{const errBody=JSON.parse(await r.text());if(errBody.error)errMsg=errBody.error;if(errBody.trace)console.error('TRACE:\n'+errBody.trace);}catch(ex){}
      hdr.textContent=`❌ ${errMsg}`;document.getElementById('st').textContent='שגיאה';return;
    }
    const d=await r.json();
    if(d.error){hdr.textContent=`❌ ${d.error}`;document.getElementById('st').textContent='מחובר';return;}
    if(!d.tasks||d.tasks.length===0){
      hdr.textContent='לא נמצאו פקעות.';
      document.getElementById('st').textContent='מחובר';
      return;
    }
    // שלב 1: הצג כרטיסיות בסיסיות מיידית (ללא API calls נוספים)
    _districtFilter=''; _cityFilter=''; _searchFilter=''; _visitTypeFilter=''; // אפס סינון בכל טעינה חדשה
    const basicCards=d.cards||[];
    basicCards.forEach((card,i)=>{card._num=i+1;_allCards.push(card);addCard(card,box);});
    hdr.textContent=`📋 בנק פקעות — ${new Date().toLocaleDateString('he-IL')} | סה"כ: ${d.count} פקעות`;
    document.getElementById('st').textContent='מחובר';
    const qr=document.getElementById('tasks-qr');
    if(qr&&!document.getElementById('tasks-map-btn')){
      const mb=document.createElement('button');mb.className='qb';mb.id='tasks-map-btn';
      mb.textContent='🗺️ הצג במפה';mb.onclick=showTasksMap;qr.appendChild(mb);
    }
    _buildCityBar();
    hdr.textContent=`📋 בנק פקעות — ${new Date().toLocaleDateString('he-IL')} | סה"כ: ${d.count} פקעות`;
  }catch(e){
    if(e.name==='AbortError'){hdr.textContent='❌ זמן קצוב — שרת לא מגיב';return;}
    hdr.textContent=`❌ שגיאה: ${e.message}`;
    document.getElementById('st').textContent='מחובר';
  }
}
/* ── tabs ── */
let _activeSubTab='inv';
function switchSubTab(sub){
  _activeSubTab=sub;
  const invContent=document.getElementById('inv-content');
  const eqContent=document.getElementById('eqrep-content');
  const btnInv=document.getElementById('sub-tab-inv');
  const btnEq=document.getElementById('sub-tab-eqrep');
  if(sub==='inv'){
    invContent.style.display='block'; eqContent.style.display='none';
    btnInv.style.color='#3dba6f'; btnInv.style.borderBottomColor='#3dba6f'; btnInv.style.fontWeight='700';
    btnEq.style.color='#8696a0';  btnEq.style.borderBottomColor='transparent'; btnEq.style.fontWeight='400';
  } else {
    invContent.style.display='none'; eqContent.style.display='block';
    btnEq.style.color='#3dba6f'; btnEq.style.borderBottomColor='#3dba6f'; btnEq.style.fontWeight='700';
    btnInv.style.color='#8696a0'; btnInv.style.borderBottomColor='transparent'; btnInv.style.fontWeight='400';
    if(!document.getElementById('eqrep-content').dataset.loaded){
      document.getElementById('eqrep-content').dataset.loaded='1';
      loadEqReport();
    }
  }
}

function canAccess(tab){
  if(_isAdmin) return true;
  if(tab==='tasks') return true;
  // premium = גישה לכל הלשוניות; הרשאות פרטניות עובדות גם בלעדיה
  if(_userPerms.premium) return true;
  if(tab==='inv')    return !!_userPerms.inventory;
  if(tab==='rep')    return !!_userPerms.reports;
  if(tab==='price')  return !!_userPerms.prices;
  return false;
}

function switchTab(tab){
  ['tasks','inv','rep','price','admin','diag'].forEach(t=>{
    const el=document.getElementById('tab-'+t);
    if(el) el.classList.toggle('active',tab===t);
  });
  const msgs=document.getElementById('msgs');
  const inv=document.getElementById('inv-panel');
  const rep=document.getElementById('rep-panel');
  const price=document.getElementById('price-panel');
  const adm=document.getElementById('admin-panel');
  const diag=document.getElementById('diag-panel');
  msgs.style.display='none';
  inv.style.display='none';
  rep.style.display='none'; price.style.display='none'; adm.style.display='none'; diag.style.display='none';
  if(tab==='tasks'){
    msgs.style.display='flex';
    switchSchedTab(_schedTabOffset);
  } else if(tab==='inv'){
    inv.style.display='flex'; inv.style.flexDirection='column';
    if(!canAccess('inv')){inv.innerHTML=NO_PERM_HTML;return;}
    if(!inv.dataset.loaded) loadInventory();
  } else if(tab==='rep'){
    rep.style.display='block';
    if(!canAccess('rep')){rep.innerHTML=NO_PERM_HTML;return;}
    if(!rep.dataset.loaded) initReports();
    else loadReports();
  } else if(tab==='price'){
    price.style.display='block';
    if(!canAccess('price')){price.innerHTML=NO_PERM_HTML;return;}
    if(!price.dataset.loaded) loadPriceList();
  } else if(tab==='admin'){
    adm.style.display='block';
    if(!adm.dataset.loaded) loadAdminPanel();
  } else if(tab==='diag'){
    diag.style.display='block';
    if(!diag.dataset.loaded) loadDiagPanel();
    else if(_diagCard) syncDiagFromCard(_diagCard);
    else showAllDiagSections();
  }
}

/* ── inventory ── */
async function loadInventory(){
  const panel=document.getElementById('inv-content');
  panel.innerHTML='<div style="color:#8696a0;text-align:center;padding:30px">טוען מלאי...</div>';
  try{
    const r=await apiFetch('/api/inventory');
    if(r.status===401){document.getElementById('inv-content').innerHTML='<div style="color:#f15c6e;text-align:center;padding:30px">❌ לא מחובר</div>';return;}
    const d=await r.json();
    if(!d.items||d.items.length===0){panel.innerHTML='<div style="color:#8696a0;text-align:center;padding:30px">לא נמצא מלאי</div>';return;}
    panel.dataset.loaded='1';
    const today=new Date().toISOString().slice(0,10);

    // שמור סריאלים לזיהוי OCR
    _invSerials=[];
    d.items.forEach(item=>{ if(item.serial) _invSerials.push({serial:item.serial,desc:item.description||''}); });

    // קיבוץ לפי קטגוריה → תיאור
    const cats={};
    d.items.forEach(item=>{
      if(!cats[item.category]) cats[item.category]={};
      const key=item.description||'ללא שם';
      if(!cats[item.category][key]) cats[item.category][key]=[];
      cats[item.category][key].push(item);
    });

    let html=`<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;flex-wrap:wrap;gap:6px">
      <div style="color:#8696a0;font-size:13px">עודכן: ${d.last_updated} | סה"כ: ${d.total} פריטים</div>
      <div style="display:flex;gap:6px">
        <button id="scan-mac-btn" class="qb" style="font-size:13px;padding:6px 14px;background:#1a3a5c;border-color:#3d5166;display:none" onclick="scanEquipment()">📷 סרוק ציוד</button>
        <button id="return-mode-btn" class="qb" style="font-size:13px;padding:6px 14px" onclick="toggleReturnMode()">📦 החזרת ציוד</button>
      </div>
    </div>`;
    let uid=0;

    // ══ תבליט ציוד חדש ══
    const newItems=d.items.filter(i=>i.first_seen===today);
    if(newItems.length>0){
      const receivedMap=d.received||{};
      html+=`<div style="background:#071f07;border:2px solid #3dba6f;border-radius:14px;padding:14px 16px;margin-bottom:16px">
        <div style="font-weight:700;color:#3dba6f;font-size:15px;margin-bottom:10px">🆕 ציוד חדש התקבל — ${today} &nbsp;(${newItems.length} יח')</div>`;
      newItems.forEach(item=>{
        const escSn=(item.serial||'').replace(/'/g,"\\'");
        const escDesc=(item.description||'').replace(/'/g,"\\'");
        const nbId='nb_'+(item.serial||('no'+(uid++))).replace(/[^a-zA-Z0-9]/g,'_');
        const rcvEntry=receivedMap[item.serial];
        const rcvStatus=rcvEntry&&rcvEntry.date===today?rcvEntry.status:(rcvEntry===today?'received':null);
        html+=`<div style="display:flex;justify-content:space-between;align-items:center;padding:8px 0;border-bottom:1px solid #0f3a0f;gap:8px">
          <div style="flex:1;min-width:0">
            <div style="color:#ddeeff;font-size:14px;font-weight:600">${item.description||'—'}</div>
            ${item.serial
              ?`<div style="color:#6b94b8;font-size:12px;margin-top:2px">🔢 ${item.serial}</div>`
              :`<div style="color:#6b94b8;font-size:12px;margin-top:2px">כמות: ${item.amount||1}</div>`}
          </div>
          <div style="display:flex;gap:6px;flex-shrink:0">
            ${item.serial?(
              rcvStatus==='received'
                ?`<button disabled style="background:#1f7a4a;border:1px solid #1f7a4a;color:#fff;border-radius:8px;padding:6px 14px;font-size:13px;font-weight:700">✅ קיבלתי</button>`
                :rcvStatus==='not_received'
                  ?`<button disabled style="background:#7a1f1f;border:1px solid #7a1f1f;color:#fff;border-radius:8px;padding:6px 14px;font-size:13px;font-weight:700">❌ לא קיבלתי</button>`
                  :`<button id="${nbId}_yes" onclick="event.stopPropagation();markRcv('${escSn}','${escDesc}','${nbId}','received')"
                      style="background:#0d3320;border:2px solid #3dba6f;color:#3dba6f;border-radius:8px;padding:6px 14px;cursor:pointer;font-size:13px;font-weight:700">✓ קיבלתי</button>
                    <button id="${nbId}_no" onclick="event.stopPropagation();markRcv('${escSn}','${escDesc}','${nbId}','not_received')"
                      style="background:#330d0d;border:2px solid #f15c6e;color:#f15c6e;border-radius:8px;padding:6px 14px;cursor:pointer;font-size:13px;font-weight:700">✗ לא קיבלתי</button>`
            ):''}
          </div>
        </div>`;
      });
      html+=`</div>`;
    }

    for(const [cat, descMap] of Object.entries(cats)){
      const isBlocked=cat.includes('חסום')||cat.includes('לא תקין');
      const catColor=isBlocked?'#f15c6e':'#25d366';
      const totalInCat=Object.values(descMap).reduce((s,a)=>s+a.length,0);
      html+=`<div style="font-weight:700;color:${catColor};margin:14px 0 6px;font-size:15px">${cat} — ${totalInCat} יח'</div>`;

      for(const [desc, items] of Object.entries(descMap)){
        const count=items.reduce((s,i)=>s+(i.serial?1:(i.amount||1)),0);
        const catalog=items[0].catalog||'';
        const hasNew=items.some(i=>i.first_seen===today);
        const cls=isBlocked?'inv-card blocked':'inv-card ok';
        const badge=isBlocked?'<span class="inv-badge badge-blocked">חסום</span>':'<span class="inv-badge badge-ok">תקין</span>';
        const newBadge=hasNew?'<span class="inv-badge badge-new">חדש היום</span>':'';
        const id='inv'+uid++;

        html+=`<div class="${cls}" style="cursor:pointer" onclick="toggleInvAcc('${id}')">
          <div style="display:flex;justify-content:space-between;align-items:center">
            <span style="font-weight:700">${desc}</span>
            <span style="display:flex;gap:6px;align-items:center">${newBadge}${badge}<span style="background:#223d58;color:#ddeeff;border-radius:12px;padding:2px 10px;font-size:13px;font-weight:700">${count}</span></span>
          </div>`;
        if(catalog) html+=`<div style="color:#8696a0;font-size:12px">מק"ט: ${catalog}</div>`;
        html+=`<div id="${id}" style="display:none;margin-top:8px;border-top:1px solid #284461;padding-top:8px">`;
        items.forEach(item=>{
          const isNew=item.first_seen===today;
          const receivedMap=d.received||{};
          html+=`<div style="padding:5px 0;border-bottom:1px solid #1a252e;font-size:13px">`;
          if(item.serial){
            const safeId='cam_'+item.serial.replace(/[^a-zA-Z0-9]/g,'_');
            const rcvId='rcv_'+item.serial.replace(/[^a-zA-Z0-9]/g,'_');
            const escSerial=item.serial.replace(/'/g,"\\'");
            const escDesc=(desc||'').replace(/'/g,"\\'");
            const rcvEntry=receivedMap[item.serial];
            const rcvStatus=rcvEntry&&rcvEntry.date===today?rcvEntry.status:(rcvEntry===today?'received':null);
            html+=`<div style="display:flex;justify-content:space-between;align-items:center;gap:6px">
              <span>🔢 ${item.serial}</span>
              <div style="display:flex;gap:5px;align-items:center">
                ${isNew?(rcvStatus==='received'?
                  `<button disabled style="background:#1f7a4a;border:1px solid #1f7a4a;color:#e9edef;border-radius:8px;padding:3px 10px;font-size:12px;white-space:nowrap">✅ קיבלתי</button>`:
                  rcvStatus==='not_received'?
                  `<button disabled style="background:#7a1f1f;border:1px solid #7a1f1f;color:#e9edef;border-radius:8px;padding:3px 10px;font-size:12px;white-space:nowrap">❌ לא קיבלתי</button>`:
                  `<button id="${rcvId}_yes" onclick="event.stopPropagation();markRcv('${escSerial}','${escDesc}','${rcvId}','received')"
                    style="background:#1a3a5c;border:1px solid #3d5166;color:#e9edef;border-radius:8px;padding:3px 10px;cursor:pointer;font-size:12px;white-space:nowrap">קיבלתי</button>
                   <button id="${rcvId}_no" onclick="event.stopPropagation();markRcv('${escSerial}','${escDesc}','${rcvId}','not_received')"
                    style="background:#3a1a1a;border:1px solid #7a3a3a;color:#e9edef;border-radius:8px;padding:3px 10px;cursor:pointer;font-size:12px;white-space:nowrap">לא קיבלתי</button>`
                ):''}
                <button class="return-cam-btn" id="${safeId}"
                  onclick="event.stopPropagation();takeReturnPhoto('${escSerial}','${escDesc}')"
                  style="display:none;background:#223d58;border:1px solid #284461;color:#ddeeff;border-radius:8px;padding:3px 10px;cursor:pointer;font-size:15px">📷</button>
              </div>
            </div>`;
          } else {
            const maxQty=item.amount||1;
            const nsId='ns_'+id+'_'+items.indexOf(item);
            const safeNsId=nsId.replace(/[^a-zA-Z0-9]/g,'_');
            const escDescNS=(desc||'').replace(/'/g,"\\'");
            html+=`<div style="display:flex;justify-content:space-between;align-items:center">
              <span style="color:#8696a0">כמות: ${maxQty}</span>
              <div class="return-qty-row" id="qrow_${safeNsId}" style="display:none;gap:6px;align-items:center">
                <span style="color:#8696a0;font-size:11px">להחזיר:</span>
                <input type="number" id="qty_${safeNsId}" min="0" max="${maxQty}" value="0"
                  onclick="event.stopPropagation()"
                  oninput="event.stopPropagation();markNsReturn('${safeNsId}','${escDescNS}',${maxQty})"
                  style="width:50px;background:#223d58;border:1px solid #284461;border-radius:6px;color:#ddeeff;padding:3px 5px;font-size:14px;text-align:center">
                <button id="cam_${safeNsId}"
                  onclick="event.stopPropagation();takeReturnPhoto('${safeNsId}','${escDescNS}')"
                  style="background:#223d58;border:1px solid #284461;color:#ddeeff;border-radius:8px;padding:3px 9px;cursor:pointer;font-size:15px">📷</button>
              </div>
            </div>`;
          }
          html+=`<div style="color:${isNew?'#53bdeb':'#8696a0'}">📅 ${item.first_seen} (${item.days_held} ימים)${isNew?' ← חדש היום':''}</div>`;
          html+=`</div>`;
        });
        html+=`</div></div>`;
      }
    }

    html+=`<div style="margin:14px 0"><button class="qb" onclick="refreshInventory()">🔄 רענן מלאי</button></div>`;
    html+=`<div id="returns-history-box"></div>`;
    document.getElementById('inv-content').innerHTML=html;
    const invPanel=document.getElementById('inv-panel');
    invPanel.dataset.loaded='1';
    loadReturnHistory(d.items||[]);
  }catch(e){
    document.getElementById('inv-content').innerHTML='<div style="color:#f15c6e;text-align:center;padding:30px">❌ שגיאה בטעינה</div>';
  }
}

function toggleInvAcc(id){
  const el=document.getElementById(id);
  el.style.display=el.style.display==='none'?'block':'none';
}

async function markRcv(serial, desc, rcvId, status){
  const btnYes=document.getElementById(rcvId+'_yes');
  const btnNo=document.getElementById(rcvId+'_no');
  if(btnYes) btnYes.disabled=true;
  if(btnNo)  btnNo.disabled=true;
  try{
    await apiFetch('/api/inventory/received',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({serial, desc, status})
    });
    if(status==='received'){
      if(btnYes){ btnYes.textContent='✅ קיבלתי'; btnYes.style.background='#1f7a4a'; btnYes.style.borderColor='#1f7a4a'; }
      if(btnNo)  btnNo.style.display='none';
    } else {
      if(btnNo){ btnNo.textContent='❌ לא קיבלתי'; btnNo.style.background='#7a1f1f'; btnNo.style.borderColor='#7a1f1f'; }
      if(btnYes) btnYes.style.display='none';
    }
  }catch(e){
    if(btnYes){ btnYes.disabled=false; }
    if(btnNo)  { btnNo.disabled=false; }
  }
}

/* ── החזרת ציוד ── */
let _returnMode=false;
let _pendingReturns={};
let _invSerials=[];  // [{serial, desc}] — מאוכלס בטעינת מלאי

function toggleReturnMode(){
  const btn=document.getElementById('return-mode-btn');
  if(!btn) return;
  if(!_returnMode){
    _returnMode=true;
    _pendingReturns={};
    btn.textContent='✅ סיימתי';
    btn.style.background='#005c4b';
    btn.style.borderColor='#25d366';
    // הוסף כפתור ביטול
    let cancelBtn=document.getElementById('return-cancel-btn');
    if(!cancelBtn){
      cancelBtn=document.createElement('button');
      cancelBtn.id='return-cancel-btn';
      cancelBtn.className='qb';
      cancelBtn.textContent='❌ ביטול';
      cancelBtn.style.cssText='font-size:13px;padding:6px 14px;background:#223d58;border-color:#f15c6e;color:#f15c6e';
      cancelBtn.onclick=cancelReturnMode;
      btn.parentNode.insertBefore(cancelBtn,btn.nextSibling);
    }
    cancelBtn.style.display='inline-block';
    const scanBtn=document.getElementById('scan-mac-btn');
    if(scanBtn) scanBtn.style.display='inline-block';
    document.querySelectorAll('.return-cam-btn').forEach(b=>b.style.display='inline-block');
    document.querySelectorAll('.return-qty-row').forEach(r=>r.style.display='flex');
  } else {
    submitReturns();
  }
}

function cancelReturnMode(){
  _returnMode=false;
  _pendingReturns={};
  const btn=document.getElementById('return-mode-btn');
  if(btn){btn.textContent='📦 החזרת ציוד';btn.style.background='';btn.style.borderColor='';}
  const cancelBtn=document.getElementById('return-cancel-btn');
  if(cancelBtn) cancelBtn.style.display='none';
  document.querySelectorAll('.return-cam-btn').forEach(b=>{
    b.style.display='none';b.textContent='📷';b.style.background='#223d58';b.style.borderColor='#284461';
  });
  document.querySelectorAll('.return-qty-row').forEach(r=>{
    r.style.display='none';
    const inp=r.querySelector('input[type="number"]');
    if(inp) inp.value='0';
    const camBtn=r.querySelector('button');
    if(camBtn){camBtn.textContent='📷';camBtn.style.background='#223d58';camBtn.style.borderColor='#284461';}
  });
}

async function submitReturns(){
  const btn=document.getElementById('return-mode-btn');
  if(btn){btn.disabled=true;btn.textContent='שולח...';}
  // אסוף גם פריטי כמות (NOSN) שסומנו
  document.querySelectorAll('.return-qty-row').forEach(row=>{
    const inp=row.querySelector('input[type="number"]');
    if(!inp) return;
    const qty=parseInt(inp.value)||0;
    if(qty<=0) return;
    const nsId=row.id.replace('qrow_','');
    if(!_pendingReturns[nsId]){
      // מצא שם מהתיאור בכפתור המצלמה
      const camBtn=row.querySelector('button');
      _pendingReturns[nsId]={description:'',photo:'',quantity:qty};
    } else {
      _pendingReturns[nsId].quantity=qty;
    }
  });
  const entries=Object.entries(_pendingReturns);
  let ok=0;
  for(const [serial,data] of entries){
    try{
      await apiFetch('/api/save-return',{method:'POST',headers:{'Content-Type':'application/json'},
        body:JSON.stringify({serial,description:data.description,photo:data.photo||'',quantity:data.quantity||1})});
      ok++;
    }catch(e){}
  }
  _returnMode=false;
  _pendingReturns={};
  if(btn){btn.disabled=false;btn.textContent='📦 החזרת ציוד';btn.style.background='';btn.style.borderColor='';}
  const cancelBtn=document.getElementById('return-cancel-btn');
  if(cancelBtn) cancelBtn.style.display='none';
  const scanBtnS=document.getElementById('scan-mac-btn');
  if(scanBtnS) scanBtnS.style.display='none';
  document.querySelectorAll('.return-cam-btn').forEach(b=>{
    b.style.display='none';b.textContent='📷';b.style.background='#223d58';b.style.borderColor='#284461';
  });
  document.querySelectorAll('.return-qty-row').forEach(r=>{
    r.style.display='none';
    const inp=r.querySelector('input[type="number"]');
    if(inp) inp.value='0';
    const camBtn=r.querySelector('button');
    if(camBtn){camBtn.textContent='📷';camBtn.style.background='#223d58';camBtn.style.borderColor='#284461';}
  });
  if(ok>0){
    loadReturnHistory([]);
    const box=document.getElementById('returns-history-box');
    if(box) box.scrollIntoView({behavior:'smooth'});
  }
}

/* ── סריקת MAC/סריאל עם OCR ── */
function normalizeSerial(s){ return (s||'').replace(/[:\-\.\s]/g,'').toUpperCase(); }

function findSerialInOcr(ocrText){
  const cleanFull = normalizeSerial(ocrText);
  // חפש רצפים hex של 6+ תווים בטקסט
  const hexSeqs = cleanFull.match(/[0-9A-F]{6,}/g) || [];
  // נסה גם עם נקודותיים (MAC בפורמט XX:XX:XX:...)
  const macPatterns = ocrText.toUpperCase().match(/([0-9A-F]{2}[:\-]){5}[0-9A-F]{2}/g) || [];
  const normMacs = macPatterns.map(normalizeSerial);
  const allCandidates = [...hexSeqs, ...normMacs];
  for(const item of _invSerials){
    const ns = normalizeSerial(item.serial);
    if(!ns) continue;
    // התאמה מלאה
    if(allCandidates.some(c => c===ns || c.includes(ns) || ns.includes(c))) return item;
    // התאמה חלקית (8+ תווים מתוך הסריאל)
    if(ns.length>=8 && allCandidates.some(c => c.length>=8 && ns.includes(c.substring(0,8)))) return item;
  }
  return null;
}

async function scanEquipment(){
  // טען Tesseract.js בעת הצורך
  if(!window.Tesseract){
    const scr=document.createElement('script');
    scr.src='https://cdn.jsdelivr.net/npm/tesseract.js@4/dist/tesseract.min.js';
    document.head.appendChild(scr);
    await new Promise((res,rej)=>{ scr.onload=res; scr.onerror=rej; });
  }
  const input=document.createElement('input');
  input.type='file'; input.accept='image/*'; input.capture='environment';
  input.onchange=async(e)=>{
    const file=e.target.files[0]; if(!file) return;
    // הצג מודל טעינה
    const modal=document.getElementById('scan-modal');
    const preview=document.getElementById('scan-preview');
    const status=document.getElementById('scan-status');
    const result=document.getElementById('scan-result');
    modal.style.display='flex';
    preview.style.display='none';
    result.innerHTML='';
    status.innerHTML='<div style="color:#53bdeb;font-size:15px">🔍 מנתח תמונה...</div>';
    try{
      // הצג preview
      const objUrl=URL.createObjectURL(file);
      preview.src=objUrl; preview.style.display='block';
      // הרץ OCR
      const ocr=await Tesseract.recognize(objUrl,'eng',{logger:()=>{}});
      const ocrText=ocr.data.text;
      URL.revokeObjectURL(objUrl);
      // כווץ תמונה לאחסון
      const reader=new FileReader();
      reader.onload=(ev)=>{
        const img=new Image();
        img.onload=()=>{
          const MAX=800; let w=img.width,h=img.height;
          if(w>MAX||h>MAX){ if(w>h){h=Math.round(h*MAX/w);w=MAX;}else{w=Math.round(w*MAX/h);h=MAX;} }
          const c=document.createElement('canvas'); c.width=w; c.height=h;
          c.getContext('2d').drawImage(img,0,0,w,h);
          const photoData=c.toDataURL('image/jpeg',0.75);
          const matched=findSerialInOcr(ocrText);
          if(matched){
            status.innerHTML=`<div style="color:#25d366;font-size:15px">✅ זוהה!</div>
              <div style="font-weight:700;font-size:14px;margin-top:4px">${matched.serial}</div>
              <div style="color:#8696a0;font-size:12px">${matched.desc}</div>`;
            result.innerHTML=`
              <button onclick="confirmScan('${matched.serial.replace(/'/g,"\\'")}','${(matched.desc||'').replace(/'/g,"\\'")}',\`${photoData.replace(/`/g,'\\`')}\`)"
                style="width:100%;padding:12px;background:#25d366;border:none;color:#111b21;border-radius:10px;font-size:15px;font-weight:700;cursor:pointer;margin-top:10px">
                ✅ אשר החזרה</button>
              <button onclick="closeScanModal()"
                style="width:100%;padding:10px;background:#223d58;border:1px solid #284461;color:#ddeeff;border-radius:10px;font-size:14px;cursor:pointer;margin-top:6px">
                ביטול</button>`;
          } else {
            status.innerHTML=`
              <div style="font-size:40px;margin-bottom:8px">⚠️</div>
              <div style="color:#f15c6e;font-size:16px;font-weight:700">ציוד לא מזוהה</div>
              <div style="color:#8696a0;font-size:13px;margin-top:6px">הציוד שצולם אינו שייך למחסן שלך</div>`;
            result.innerHTML=`<button onclick="closeScanModal()"
              style="width:100%;padding:12px;background:#3a1a1a;border:1px solid #7a3a3a;color:#f15c6e;border-radius:10px;font-size:15px;font-weight:700;cursor:pointer;margin-top:10px">
              סגור</button>`;
          }
        };
        img.src=ev.target.result;
      };
      reader.readAsDataURL(file);
    }catch(err){
      status.innerHTML='<div style="color:#f15c6e">❌ שגיאה בזיהוי</div>';
      document.getElementById('scan-result').innerHTML=`<button onclick="closeScanModal()"
        style="width:100%;padding:10px;background:#223d58;border:1px solid #284461;color:#ddeeff;border-radius:10px;font-size:14px;cursor:pointer;margin-top:8px">סגור</button>`;
    }
  };
  input.click();
}

function confirmScan(serial, desc, photoData){
  _pendingReturns[serial]={description:desc, photo:photoData};
  const safeId='cam_'+serial.replace(/[^a-zA-Z0-9]/g,'_');
  const camBtn=document.getElementById(safeId);
  if(camBtn){camBtn.textContent='✅';camBtn.style.background='#005c4b';camBtn.style.borderColor='#25d366';}
  closeScanModal();
  const toast=document.createElement('div');
  toast.style.cssText='position:fixed;bottom:80px;left:50%;transform:translateX(-50%);background:#25d366;color:#111b21;padding:10px 22px;border-radius:20px;font-weight:700;z-index:9999;font-size:14px';
  toast.textContent=`✅ ${serial} נוסף להחזרה`;
  document.body.appendChild(toast);
  setTimeout(()=>toast.remove(),3000);
}

function closeScanModal(){
  document.getElementById('scan-modal').style.display='none';
}

function takeReturnPhoto(serial,desc){
  const input=document.createElement('input');
  input.type='file';input.accept='image/*';input.capture='environment';
  input.onchange=(e)=>{
    const file=e.target.files[0];
    if(!file) return;
    const reader=new FileReader();
    reader.onload=(ev)=>{
      const img=new Image();
      img.onload=()=>{
        const MAX=800;
        let w=img.width,h=img.height;
        if(w>MAX||h>MAX){
          if(w>h){h=Math.round(h*MAX/w);w=MAX;}
          else{w=Math.round(w*MAX/h);h=MAX;}
        }
        const c=document.createElement('canvas');
        c.width=w;c.height=h;
        c.getContext('2d').drawImage(img,0,0,w,h);
        const photoData=c.toDataURL('image/jpeg',0.75);
        _pendingReturns[serial]={description:desc,photo:photoData};
        const safeId='cam_'+serial.replace(/[^a-zA-Z0-9]/g,'_');
        const camBtn=document.getElementById(safeId);
        if(camBtn){camBtn.textContent='✅';camBtn.style.background='#005c4b';camBtn.style.borderColor='#25d366';}
      };
      img.src=ev.target.result;
    };
    reader.readAsDataURL(file);
  };
  input.click();
}

/* ── דוח ציוד ── */
let _eqActiveTab='rcv';

function switchEqTab(t){
  _eqActiveTab=t;
  const btnRcv=document.getElementById('eqrep-tab-rcv');
  const btnRet=document.getElementById('eqrep-tab-ret');
  const listRcv=document.getElementById('eqrep-rcv-list');
  const listRet=document.getElementById('eqrep-ret-list');
  if(t==='rcv'){
    btnRcv.style.background='#1f7a4a'; btnRcv.style.color='#fff'; btnRcv.style.fontWeight='700';
    btnRet.style.background='#223d58'; btnRet.style.color='#6b94b8'; btnRet.style.fontWeight='400';
    listRcv.style.display='block'; listRet.style.display='none';
  } else {
    btnRet.style.background='#c0392b'; btnRet.style.color='#fff'; btnRet.style.fontWeight='700';
    btnRcv.style.background='#223d58'; btnRcv.style.color='#6b94b8'; btnRcv.style.fontWeight='400';
    listRcv.style.display='none'; listRet.style.display='block';
  }
}

async function loadEqReport(){
  const listRcv=document.getElementById('eqrep-rcv-list');
  const listRet=document.getElementById('eqrep-ret-list');
  listRcv.innerHTML=listRet.innerHTML='<div style="text-align:center;padding:20px;color:#8696a0">טוען...</div>';
  try{
    const r=await apiFetch('/api/equipment-report');
    const d=await r.json();
    renderEqList(listRcv, d.received_by_date||{}, 'received');
    renderEqList(listRet, d.returned_by_date||{}, 'returned');
  }catch(e){
    listRcv.innerHTML=listRet.innerHTML='<div style="color:#f15c6e;text-align:center;padding:20px">❌ שגיאה</div>';
  }
}

function renderEqList(container, byDate, type){
  const dates=Object.keys(byDate).sort().reverse();
  if(!dates.length){
    container.innerHTML='<div style="text-align:center;padding:30px;color:#8696a0">אין פריטים</div>';
    return;
  }
  const icon=type==='received'?'📥':'📤';
  let html='';
  for(const d of dates){
    const items=byDate[d];
    const encDate=encodeURIComponent(d);
    const dayId=`eq_${type}_${d.replace(/-/g,'')}`;
    html+=`<div id="card_${dayId}" style="background:#d0e4f7;border-radius:10px;margin-bottom:10px;overflow:hidden;color:#0d1b2a">
      <div style="display:flex;justify-content:space-between;align-items:center;padding:10px 12px;border-bottom:1px solid #b0c8e0;cursor:pointer"
           onclick="toggleEqDay('${dayId}')">
        <span style="font-weight:700;font-size:15px;color:#0d1b2a">${icon} ${d} <span style="color:#4a6a88;font-size:12px;font-weight:400">(${items.length} פריטים)</span></span>
        <div style="display:flex;gap:6px;align-items:center" onclick="event.stopPropagation()">
          <a href="/api/equipment-report/export?date=${encDate}&type=${type}&token=${encodeURIComponent(localStorage.getItem('cc_token')||'')}&phone=${encodeURIComponent(localStorage.getItem('cc_phone')||'')}"
             style="background:#1a3a5c;border:1px solid #3d5166;color:#e9edef;border-radius:8px;padding:4px 10px;font-size:12px;text-decoration:none;white-space:nowrap">
            📥 Excel</a>
          <button onclick="deleteEqDate('${d}','${type}','card_${dayId}')"
            style="background:#3a1a1a;border:1px solid #7a3a3a;color:#f15c6e;border-radius:8px;padding:4px 9px;font-size:13px;cursor:pointer"
            title="מחק דוח">🗑️</button>
        </div>
      </div>
      <div id="eq_${type}_${d.replace(/-/g,'')}" style="display:none;padding:8px 12px">`;
    for(const item of items){
      html+=`<div style="display:flex;justify-content:space-between;padding:5px 0;border-bottom:1px solid #b0c8e0;font-size:13px">
        <span style="color:#0d1b2a;font-weight:600">${item.serial||'—'}</span>
        <span style="color:#4a6a88;font-size:12px">${item.desc||item.description||''}</span>
        ${item.quantity&&item.quantity>1?`<span style="color:#53bdeb">×${item.quantity}</span>`:''}
      </div>`;
    }
    html+=`</div></div>`;
  }
  container.innerHTML=html;
}

function toggleEqDay(id){
  const el=document.getElementById(id);
  if(el) el.style.display=el.style.display==='none'?'block':'none';
}

async function deleteEqDate(date, type, cardId){
  if(!confirm(`למחוק את הדוח של ${date}?`)) return;
  try{
    const r=await apiFetch('/api/equipment-report/delete-date',{
      method:'DELETE',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({date, type})
    });
    const d=await r.json();
    if(d.ok){
      const card=document.getElementById(cardId);
      if(card) card.remove();
    }
  }catch(e){ alert('שגיאה במחיקה'); }
}

async function loadReturnHistory(currentItems){
  const box=document.getElementById('returns-history-box');
  if(!box) return;
  try{
    const r=await apiFetch('/api/returns?days=30');
    if(!r.ok) return;
    const d=await r.json();
    if(!d.returns||d.returns.length===0){box.innerHTML='';return;}
    const filtered=d.returns;
    if(filtered.length===0){box.innerHTML='';return;}
    let html=`<div style="margin-top:18px;border-top:2px solid #284461;padding-top:14px">
      <div style="font-weight:700;color:#f15c6e;margin-bottom:10px;font-size:15px">📤 הוחזר (30 ימים אחרונים)</div>`;
    filtered.forEach(ret=>{
      const isNS=ret.serial.startsWith('ns_')||ret.serial.startsWith('NOSN');
      const serialLine=isNS
        ?`<span style="font-weight:700">📦 ${ret.description||'ציוד'} × ${ret.quantity||1}</span>`
        :`<span style="font-weight:700">🔢 ${ret.serial}</span>`;
      const descLine=isNS?'':`<div style="color:#4a6a88;margin-top:2px">${ret.description||''}</div>`;
      const qtyLine=(!isNS&&(ret.quantity||1)>1)?`<div style="color:#8696a0;font-size:12px">כמות: ${ret.quantity}</div>`:'';
      html+=`<div id="ret-row-${ret.id}" style="background:#d0e4f7;color:#0d1b2a;border-radius:8px;padding:10px 14px;margin-bottom:8px;font-size:13px;direction:rtl;border-right:3px solid #f15c6e">
        <div style="display:flex;justify-content:space-between;align-items:center">
          ${serialLine}
          <div style="display:flex;gap:8px;align-items:center">
            <span style="color:#4a6a88;font-size:12px">📅 ${ret.return_date}</span>
            <button onclick="deleteReturn(${ret.id})" style="background:none;border:none;color:#f15c6e;font-size:18px;cursor:pointer;padding:0 4px;line-height:1" title="מחק">🗑</button>
          </div>
        </div>
        ${descLine}${qtyLine}
        ${ret.has_photo?`<div style="margin-top:4px"><a href="/api/return-photo/${ret.photo_filename}" target="_blank" style="color:#53bdeb;font-size:12px">📷 צפה בתמונה</a></div>`:''}
      </div>`;
    });
    html+=`</div>`;
    box.innerHTML=html;
  }catch(e){box.innerHTML='';}
}

function markNsReturn(safeId,desc,maxQty){
  const inp=document.getElementById('qty_'+safeId);
  if(!inp) return;
  let qty=parseInt(inp.value)||0;
  if(qty<0){qty=0;inp.value=0;}
  if(qty>maxQty){qty=maxQty;inp.value=maxQty;}
  if(qty>0){
    if(!_pendingReturns[safeId]) _pendingReturns[safeId]={description:desc,photo:''};
    _pendingReturns[safeId].quantity=qty;
    _pendingReturns[safeId].description=desc;
  } else {
    delete _pendingReturns[safeId];
  }
}

async function deleteReturn(id){
  if(!confirm('למחוק את ההחזרה הזו?')) return;
  try{
    const r=await apiFetch('/api/delete-return/'+id,{method:'DELETE'});
    if(r.ok){
      const row=document.getElementById('ret-row-'+id);
      if(row) row.remove();
    }
  }catch(e){}
}

/* ── reports ── */
function isoDay(offset){
  const d=new Date(); d.setDate(d.getDate()+offset);
  // תאריך מקומי (לא UTC) — חשוב לאזור ישראל UTC+3
  return d.getFullYear()+'-'+String(d.getMonth()+1).padStart(2,'0')+'-'+String(d.getDate()).padStart(2,'0');
}
function _repBtn(id,label,onclick){
  return `<button id="${id}" onclick="${onclick}" style="flex:1;background:#223d58;border:1px solid #284461;border-radius:8px;padding:9px 4px;color:#ddeeff;font-size:14px;cursor:pointer;font-weight:600">${label}</button>`;
}
function _repSetActive(id){
  ['rep-btn-prev','rep-btn-today','rep-btn-yesterday','rep-btn-tomorrow','rep-btn-curmonth','rep-btn-prevmonth'].forEach(b=>{
    const el=document.getElementById(b);
    if(!el) return;
    if(b===id){el.style.background='#25d366';el.style.color='#fff';el.style.border='none';}
    else{el.style.background='#223d58';el.style.color='#ddeeff';el.style.border='1px solid #284461';}
  });
}
function isoOf(d){return d.getFullYear()+'-'+String(d.getMonth()+1).padStart(2,'0')+'-'+String(d.getDate()).padStart(2,'0');}
function setRepDay(offset){
  const d=new Date(); d.setDate(d.getDate()+offset);
  const s=isoOf(d);
  document.getElementById('rep-from').value=s;
  document.getElementById('rep-to').value=s;
  _repSetActive(offset===-1?'rep-btn-yesterday':offset===1?'rep-btn-tomorrow':'rep-btn-today');
  loadReports();
}
function setRepMonth(offset){
  // offset=0 → החודש הנוכחי, offset=-1 → חודש קודם
  const now=new Date();
  const y=now.getFullYear(), m=now.getMonth()+offset;
  const first=new Date(y,m,1);
  const last=offset===0?now:new Date(y,m+1,0);
  document.getElementById('rep-from').value=isoOf(first);
  document.getElementById('rep-to').value=isoOf(last);
  _repSetActive(offset===0?'rep-btn-curmonth':'rep-btn-prevmonth');
  loadReports();
}

function initReports(){
  const panel=document.getElementById('rep-panel');
  const today=isoOf(new Date());
  const now=new Date();
  const firstCur=isoOf(new Date(now.getFullYear(),now.getMonth(),1));
  panel.dataset.loaded='1';
  panel.innerHTML=`
    <div style="display:flex;border-bottom:1px solid #284461;margin-bottom:12px;direction:rtl;width:100%">
      <button id="rep-tab-visits" onclick="switchRepTab('visits')"
        style="flex:1;padding:10px 4px;background:transparent;border:none;border-bottom:2px solid #25d366;color:#25d366;font-size:14px;font-weight:700;cursor:pointer;text-align:center">📊 ביקורים</button>
      <button id="rep-tab-repeats" onclick="switchRepTab('repeats')"
        style="flex:1;padding:10px 4px;background:transparent;border:none;border-bottom:2px solid transparent;color:#6b94b8;font-size:14px;font-weight:600;cursor:pointer;text-align:center">🔁 חוזרות</button>
      <button id="rep-tab-archive" onclick="switchRepTab('archive')"
        style="flex:1;padding:10px 4px;background:transparent;border:none;border-bottom:2px solid transparent;color:#6b94b8;font-size:14px;font-weight:600;cursor:pointer;text-align:center">📂 דוחות עבר</button>
    </div>

    <div id="rep-pane-visits">
      <div style="background:#1a2d42;border-radius:10px;padding:14px;margin-bottom:12px">
        <div style="display:flex;gap:6px;margin-bottom:10px;direction:rtl;flex-wrap:wrap">
          ${_repBtn('rep-btn-yesterday','אתמול','setRepDay(-1)')}
          ${_repBtn('rep-btn-today','היום','setRepDay(0)')}
          ${_repBtn('rep-btn-tomorrow','מחר','setRepDay(1)')}
          ${_repBtn('rep-btn-curmonth','החודש','setRepMonth(0)')}
          ${_repBtn('rep-btn-prevmonth','חודש קודם','setRepMonth(-1)')}
        </div>
        <input id="rep-from" type="hidden" value="${firstCur}">
        <input id="rep-to"   type="hidden" value="${today}">
        <div style="display:flex;gap:8px;flex-wrap:wrap;align-items:center;justify-content:flex-end">
          <button class="btn-main" id="rep-export-btn" style="padding:9px 14px;font-size:14px;width:auto;background:#1a3a5c;border:1px solid #3d5166;display:none" onclick="exportReports()">📥 Excel</button>
          <button class="btn-main" id="rep-archive-save-btn" style="padding:9px 14px;font-size:14px;width:auto;background:#1a3d22;border:1px solid #3dba6f;color:#3dba6f;display:none" onclick="saveToArchive()">💾 שמור לארכיון</button>
        </div>
      </div>
      <div id="rep-content"></div>
    </div>

    <div id="rep-pane-repeats" style="display:none">
      <div style="display:flex;gap:8px;margin-bottom:14px;direction:rtl">
        <button id="rep-rpt-cur" onclick="loadRepeatsPane(0)"
          style="flex:1;padding:10px;background:#223d58;border:1px solid #284461;border-radius:8px;color:#ddeeff;font-size:14px;font-weight:600;cursor:pointer">📅 החודש הנוכחי</button>
        <button id="rep-rpt-prev" onclick="loadRepeatsPane(-1)"
          style="flex:1;padding:10px;background:#25d366;border:none;border-radius:8px;color:#fff;font-size:14px;font-weight:700;cursor:pointer">📅 החודש הקודם</button>
        <button onclick="toggleRepDbg()"
          style="padding:10px 14px;background:#2a1a3d;border:1px solid #6b3fa0;border-radius:8px;color:#c084fc;font-size:13px;font-weight:600;cursor:pointer">🔬 דיבאגר</button>
      </div>

      <!-- דיבאגר -->
      <div id="rep-dbg-panel" style="display:none;background:#0f1c2e;border:1px solid #6b3fa0;border-radius:10px;padding:14px;margin-bottom:14px;direction:rtl">
        <div style="font-weight:700;color:#c084fc;margin-bottom:10px;font-size:14px">🔬 היסטוריית ביקורים לפי מספר פקע</div>
        <div style="display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin-bottom:6px">
          <input id="dbg-callid" type="text" placeholder="מספר פקע (לדוגמה: 4447662)"
            style="flex:1;min-width:160px;background:#0d1b2a;border:1px solid #6b3fa0;border-radius:8px;padding:9px 12px;color:#eaf4ff;font-size:14px;direction:ltr">
          <button onclick="runRepDbg()"
            style="padding:9px 18px;background:#4a1a7a;border:1px solid #c084fc;border-radius:8px;color:#e9d5ff;font-size:14px;font-weight:700;cursor:pointer">🔍 שלוף היסטוריה</button>
        </div>
        <!-- BAN ידני — מוצג גם אם לא נמצא ב-DB וגם אם BAN ריק -->
        <div id="dbg-manual-ban-row" style="display:none;gap:8px;flex-wrap:wrap;align-items:center;margin-bottom:6px">
          <span style="color:#f9c846;font-size:12px">🔧 BAN ידני:</span>
          <input id="dbg-ban" type="text" placeholder="BAN נומרי (לדוגמה: 335497519)"
            style="width:150px;background:#0d1b2a;border:1px solid #f9a846;border-radius:6px;padding:7px 10px;color:#eaf4ff;font-size:13px;direction:ltr">
          <button onclick="runRepDbgManual()"
            style="padding:7px 14px;background:#4a3000;border:1px solid #f9a846;border-radius:6px;color:#f9c846;font-size:13px;cursor:pointer">🔍 שלוף היסטוריה</button>
          <button id="save-ban-btn" onclick="saveVisitBan()"
            style="padding:7px 14px;background:#1a2d42;border:1px solid #3d6fa8;border-radius:6px;color:#7db4e0;font-size:13px;cursor:pointer"
            title="שמור את ה-BAN הנומרי ב-DB כדי שדוח החוזרות יעבוד">💾 שמור BAN ל-DB</button>
        </div>
        <div id="rep-dbg-result"></div>
      </div>

      <div id="rep-rpt-content" style="color:#c9d8e8;font-size:14px;text-align:center;padding:20px;color:#6b94b8">
        בחר חודש לטעינת הדוח
      </div>
    </div>

    <div id="rep-pane-archive" style="display:none">
      <!-- רשימת ארכיון -->
      <div style="background:#1a2d42;border-radius:10px;padding:14px;margin-bottom:12px">
        <div style="font-weight:700;color:#f9c846;margin-bottom:10px;font-size:15px">📂 ארכיון דוחות חודשיים</div>
        <div id="archive-list-box">
          <div style="color:#6b94b8;font-size:13px">טוען...</div>
        </div>
      </div>

      <!-- השוואת קובץ -->
      <div style="background:#1a2d42;border-radius:10px;padding:14px;margin-bottom:12px">
        <div style="font-weight:700;color:#53bdeb;margin-bottom:10px;font-size:15px">🔍 השוואת קובץ Excel</div>
        <div style="font-size:12px;color:#9bbdd8;margin-bottom:10px;direction:rtl;line-height:1.6">
          השווה מול הנתונים שלנו ותקבל פירוט פערים
        </div>
        <!-- Toggle source -->
        <div style="display:flex;gap:6px;margin-bottom:12px;direction:rtl">
          <button id="cmp-src-file" onclick="setCmpSource('file')"
            style="flex:1;padding:8px 4px;border-radius:8px;font-size:13px;font-weight:700;cursor:pointer;
                   background:#25d366;border:none;color:#fff">
            📂 העלה קובץ
          </button>
          <button id="cmp-src-arch" onclick="setCmpSource('archive')"
            style="flex:1;padding:8px 4px;border-radius:8px;font-size:13px;font-weight:700;cursor:pointer;
                   background:#223d58;border:1px solid #284461;color:#ddeeff">
            🗂️ מדוחות שמורים
          </button>
        </div>
        <!-- Upload file panel -->
        <div id="cmp-panel-file" style="direction:rtl">
          <input type="file" id="compare-file-input" accept=".xlsx,.xls"
            style="background:#0d1b2a;border:1px solid #284461;border-radius:8px;padding:8px;color:#eaf4ff;font-size:13px;width:100%;box-sizing:border-box;margin-bottom:10px">
        </div>
        <!-- Archive panel -->
        <div id="cmp-panel-archive" style="display:none;direction:rtl">
          <select id="cmp-archive-select"
            style="width:100%;background:#0d1b2a;border:1px solid #284461;border-radius:8px;padding:9px;color:#eaf4ff;font-size:13px;margin-bottom:10px">
            <option value="">טוען דוחות...</option>
          </select>
        </div>
        <button onclick="runCompare()"
          style="width:100%;padding:11px;background:linear-gradient(135deg,#1a3d5c,#1a2d4a);border:1px solid #53bdeb;color:#eaf4ff;border-radius:10px;font-size:14px;font-weight:700;cursor:pointer">
          🔍 השווה
        </button>
        <div id="compare-result" style="margin-top:12px"></div>
      </div>
    </div>`;
  setRepMonth(0);
}

function switchRepTab(tab){
  const panes={visits:'rep-pane-visits',repeats:'rep-pane-repeats',archive:'rep-pane-archive'};
  const btns ={visits:'rep-tab-visits', repeats:'rep-tab-repeats', archive:'rep-tab-archive'};
  const colors={visits:'#25d366',repeats:'#f9c846',archive:'#53bdeb'};
  for(const [k,id] of Object.entries(panes)){
    const el=document.getElementById(id);
    if(el) el.style.display=(k===tab)?'':'none';
  }
  for(const [k,id] of Object.entries(btns)){
    const el=document.getElementById(id);
    if(!el) continue;
    if(k===tab){el.style.borderBottomColor=colors[k];el.style.color=colors[k];el.style.fontWeight='700';}
    else{el.style.borderBottomColor='transparent';el.style.color='#6b94b8';el.style.fontWeight='600';}
  }
  if(tab==='archive') loadArchiveList();
  if(tab==='repeats') loadRepeatsPane(-1);
}

let _repRptLoading=false;
let _repRptAbort=null;
async function loadRepeatsPane(offset){
  function _setRptBtn(el, active){
    if(!el) return;
    el.style.background=active?'#25d366':'#223d58';
    el.style.border=active?'none':'1px solid #284461';
    el.style.color=active?'#fff':'#ddeeff';
    el.style.fontWeight=active?'700':'600';
  }
  _setRptBtn(document.getElementById('rep-rpt-cur'),  offset===0);
  _setRptBtn(document.getElementById('rep-rpt-prev'), offset===-1);
  if(_repRptAbort){_repRptAbort.abort();_repRptAbort=null;}
  _repRptLoading=false;
  const box=document.getElementById('rep-rpt-content');
  if(!box) return;
  box.innerHTML='<div style="text-align:center;padding:30px;color:#6b94b8"><div style="font-size:26px;margin-bottom:8px">⏳</div><div id="rpt-status-msg">מתחיל טעינה...</div></div>';
  _repRptAbort=new AbortController();
  _repRptLoading=true;

  /* ── עזרים לעדכון stats בזמן אמת ── */
  let _repCount=0, _repPrice=0, _totalVisits=0, _monthLabel='';
  let _tbodyEl=null, _statsCountEl=null, _statsPriceEl=null;

  function _updateStats(){
    if(_statsCountEl) _statsCountEl.textContent=_repCount;
    if(_statsPriceEl && _repPrice>0) _statsPriceEl.parentElement.style.display='';
    if(_statsPriceEl) _statsPriceEl.textContent='₪'+_repPrice.toFixed(2);
  }

  function _buildTableStructure(label, totalVis){
    _monthLabel=label; _totalVisits=totalVis;
    const statsId='rpt-stat-'+Date.now();
    box.innerHTML=`
      <div style="display:flex;gap:16px;flex-wrap:wrap;margin-bottom:16px;padding:12px;background:#0a1826;border-radius:10px;border:1px solid #1e3a5f;justify-content:center;direction:rtl">
        <div style="text-align:center"><div style="font-size:22px;font-weight:700;color:#f9c846">${label}</div><div style="color:#6b94b8;font-size:12px">חודש</div></div>
        <div style="text-align:center"><div id="rpt-cnt" style="font-size:22px;font-weight:700;color:#f15c6e">0</div><div style="color:#6b94b8;font-size:12px">חוזרות</div></div>
        <div style="text-align:center"><div style="font-size:22px;font-weight:700;color:#7db4e0">${totalVis}</div><div style="color:#6b94b8;font-size:12px">ביקורים בחודש</div></div>
        <div id="rpt-price-wrap" style="text-align:center;display:none"><div id="rpt-price" style="font-size:22px;font-weight:700;color:#3dba6f">₪0.00</div><div style="color:#6b94b8;font-size:12px">סה"כ</div></div>
      </div>
      <div style="overflow-x:auto">
      <table style="width:100%;border-collapse:collapse;font-size:13px;direction:rtl">
        <thead>
          <tr style="background:#0a1826;color:#6b94b8;text-align:right">
            <th style="padding:8px 10px;border-bottom:1px solid #1e3a5f">תאריך שלי</th>
            <th style="padding:8px 10px;border-bottom:1px solid #1e3a5f">סוג משימה שלי</th>
            <th style="padding:8px 10px;border-bottom:1px solid #1e3a5f;color:#3dba6f">₪</th>
            <th style="padding:8px 10px;border-bottom:1px solid #1e3a5f">תאריך חזרה</th>
            <th style="padding:8px 10px;border-bottom:1px solid #1e3a5f">טכנאי</th>
            <th style="padding:8px 10px;border-bottom:1px solid #1e3a5f">סוג חזרה</th>
            <th style="padding:8px 10px;border-bottom:1px solid #1e3a5f;text-align:center">ימים</th>
            <th style="padding:8px 10px;border-bottom:1px solid #1e3a5f">לקוח</th>
            <th style="padding:8px 10px;border-bottom:1px solid #1e3a5f">כתובת</th>
          </tr>
        </thead>
        <tbody id="rpt-tbody"></tbody>
      </table></div>
      <div id="rpt-footer" style="text-align:center;margin-top:14px"></div>`;
    _tbodyEl      = document.getElementById('rpt-tbody');
    _statsCountEl = document.getElementById('rpt-cnt');
    _statsPriceEl = document.getElementById('rpt-price');
  }

  function _appendRow(r){
    if(!_tbodyEl) return;
    const price=r.my_price?`<span style="color:#3dba6f">₪${Number(r.my_price).toFixed(2)}</span>`:'—';
    const tr=document.createElement('tr');
    tr.style.borderBottom='1px solid #0d1f33';
    tr.onmouseover=()=>tr.style.background='#132538';
    tr.onmouseout=()=>tr.style.background='';
    tr.innerHTML=`
      <td style="padding:7px 10px;white-space:nowrap">${r.my_date||''}</td>
      <td style="padding:7px 10px">${r.my_type||''}</td>
      <td style="padding:7px 10px">${price}</td>
      <td style="padding:7px 10px;white-space:nowrap">${r.tech_date||''}</td>
      <td style="padding:7px 10px">${r.tech_name||''}</td>
      <td style="padding:7px 10px">${r.tech_type||''}</td>
      <td style="padding:7px 10px;text-align:center;color:#f9c846">${r.days_diff||''}</td>
      <td style="padding:7px 10px">${r.cust_name||''}</td>
      <td style="padding:7px 10px;font-size:12px;color:#6b94b8">${r.address||''}</td>`;
    _tbodyEl.appendChild(tr);
  }

  function _setStatus(msg){
    const el=document.getElementById('rpt-status-msg');
    if(el) el.textContent=msg;
  }

  try{
    const resp=await apiFetch('/api/reports/repeats-stream?month='+offset,
                              {signal:_repRptAbort.signal});
    if(!resp.ok) throw new Error('שגיאת שרת '+resp.status);

    const reader=resp.body.getReader();
    const dec=new TextDecoder();
    let buf='';

    while(true){
      const {done,value}=await reader.read();
      if(done) break;
      buf+=dec.decode(value,{stream:true});
      let nl;
      while((nl=buf.indexOf('\n'))!==-1){
        const line=buf.slice(0,nl).trim();
        buf=buf.slice(nl+1);
        if(!line) continue;
        let msg;
        try{msg=JSON.parse(line);}catch{continue;}

        if(msg.type==='no_v_marks'){
          // לא בשימוש יותר — ממשיכים ללא V
          continue;
        }
        if(msg.type==='error'){
          box.innerHTML=`<div style="color:#f15c6e;padding:20px">❌ ${msg.error||'שגיאה'}</div>`;
          return;
        }
        if(msg.type==='sync_start'){
          _setStatus(`מסנכרן ${msg.days} ימים...`);
        }
        if(msg.type==='sync_progress'){
          _setStatus(`מסנכרן... ${msg.done}/${msg.total} ימים`);
        }
        if(msg.type==='meta'){
          _buildTableStructure(msg.label, msg.total_visits);
        }
        if(msg.type==='repeat'){
          _repCount++;
          _repPrice+=(msg.my_price||0);
          _appendRow(msg);
          _updateStats();
        }
        if(msg.type==='done'){
          const footer=document.getElementById('rpt-footer');
          if(footer){
            if(msg.repeat_count===0){
              footer.innerHTML=`<div style="color:#6b94b8;padding:20px;text-align:center">לא נמצאו חוזרות בחודש זה<br><span style="font-size:12px">(${_totalVisits} ביקורים נבדקו)</span></div>`;
            } else {
              footer.innerHTML=`<button onclick="doExportRepeatsXL()" style="background:#1a3a5c;border:1px solid #3d6fa8;color:#7db4e0;padding:9px 22px;border-radius:8px;cursor:pointer;font-size:14px">📥 ייצוא Excel</button>`;
            }
          }
        }
      }
    }

    /* אם meta לא הגיע (חודש ריק לפני meta) — הצג הודעה */
    if(!_tbodyEl && _repCount===0){
      const statusEl=document.getElementById('rpt-status-msg');
      if(statusEl) statusEl.textContent='לא נמצאו חוזרות בחודש זה';
    }

  }catch(e){
    if(e.name==='AbortError') return;
    box.innerHTML=`<div style="color:#f15c6e;padding:20px">❌ ${e.message}</div>`;
  }finally{
    _repRptLoading=false;
  }
}

function toggleRepDbg(){
  const p=document.getElementById('rep-dbg-panel');
  if(p) p.style.display=p.style.display==='none'?'':'none';
}

function openProbe(){
  const callId=(document.getElementById('dbg-callid').value||'').trim();
  const token=localStorage.getItem('cc_token')||'';
  const phone=localStorage.getItem('cc_phone')||'';
  const url='/api/reports/history-probe?token='+encodeURIComponent(token)
    +'&phone='+encodeURIComponent(phone)
    +(callId?'&call_id='+encodeURIComponent(callId):'');
  window.open(url,'_blank');
}

/* פותח את ה-Raw Schedule לתאריך של הביקור שנמצא ב-DB */
function openRawSchedule(dateStr){
  const callId=(document.getElementById('dbg-callid').value||'').trim();
  const token=localStorage.getItem('cc_token')||'';
  const phone=localStorage.getItem('cc_phone')||'';
  const d=dateStr||prompt('תאריך (YYYY-MM-DD):','');
  if(!d) return;
  const url='/api/debug/schedule-raw?token='+encodeURIComponent(token)
    +'&phone='+encodeURIComponent(phone)
    +'&date='+encodeURIComponent(d)
    +(callId?'&call_id='+encodeURIComponent(callId):'');
  window.open(url,'_blank');
}

function _renderVisitHistory(d, box){
  let h='';

  // כרטיס ביקור — אם נמצא ב-DB
  if(d.visit){
    const vv=d.visit;
    h+=`<div style="background:#0f1c2e;border:1px solid #3d6fa8;border-radius:8px;padding:10px;margin-bottom:10px;font-size:13px">
      <div style="display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:6px;margin-bottom:6px">
        <div style="font-weight:700;color:#53bdeb">📋 פקע ${d.call_id} — ${vv.fetch_date||''} | ${vv.task_type||'(ריק)'}</div>
        <button onclick="openRawSchedule('${vv.fetch_date||''}')"
          style="padding:4px 10px;background:#1a2d42;border:1px solid #3d6fa8;border-radius:6px;color:#7db4e0;font-size:11px;cursor:pointer"
          title="הצג את ה-JSON הגולמי מה-API לתאריך זה — כדי לראות את כל שדות הפקע">📋 Raw JSON</button>
      </div>
      <div style="font-family:monospace;color:#6b94b8;font-size:11px">
        ban=<span style="color:#c084fc">${d.ban||'—'}</span> &nbsp;|&nbsp;
        customer_id=<span style="color:#c084fc">${d.customer_id||'—'}</span> &nbsp;|&nbsp;
        user_id=<span style="color:#c084fc">${d.user_id||'—'}</span> &nbsp;|&nbsp;
        source=<span style="color:#53bdeb">${d.source||'JET'}</span>
      </div>
      ${d.call_details_ban?`<div style="font-family:monospace;color:#3dba6f;font-size:11px;margin-top:3px">
        ✅ BAN נומרי נמצא: <span style="color:#3dba6f;font-weight:700">${d.call_details_ban}</span>
        <span style="color:#6b94b8">(מ: ${d.call_details_ban_field||'?'})</span>
      </div>`:
      (d.ban&&d.ban.toUpperCase().startsWith('PI'))?`<div style="font-family:monospace;color:#f9c846;font-size:11px;margin-top:3px">
        ⚠️ לא נמצא BAN נומרי
        ${(()=>{const m=d.method2_dbg||{};
          if(!m.tried) return '— שיטה 2 לא נוסתה';
          if(m.error) return `— שגיאה: ${m.error}`;
          if(!m.matched_callid) return `— schedule חזר ${m.apt_count} ביקורים, callId לא נמצא`;
          return `— schedule: ${m.apt_count} ביקורים, ${m.actions_count} actions, לא נמצא BAN נומרי בשום action`;
        })()}
      </div>`:''
      }

    </div>`;
  } else {
    h+=`<div style="background:#2a1a00;border:1px solid #f9c846;border-radius:8px;padding:8px;margin-bottom:10px;font-size:12px;color:#f9c846">
      ⚠️ פקע ${d.call_id||''} לא נמצא ב-DB — שולף לפי BAN ידני
      &nbsp;(ban=${d.ban||'—'} cid=${d.customer_id||'—'})
    </div>`;
  }

  // ניסיונות API
  if(d.attempts&&d.attempts.length){
    const successAttempt=d.attempts.find(a=>a.success);
    h+=`<details style="margin-bottom:10px">
      <summary style="cursor:pointer;color:${successAttempt?'#3dba6f':'#f15c6e'};font-size:13px;font-weight:700;user-select:none">
        ${successAttempt?'✅ API הצליח':'❌ כל ניסיונות ה-API נכשלו'} — לחץ לפרטים
      </summary>
      <div style="overflow-x:auto;margin-top:6px"><table style="border-collapse:collapse;font-size:11px;font-family:monospace;width:100%">
        <tr style="background:#0a1826;color:#6b94b8">
          <th style="padding:3px 6px">#</th>
          <th style="padding:3px 6px;text-align:left">תיאור</th>
          <th style="padding:3px 6px;text-align:left">ban</th>
          <th style="padding:3px 6px;text-align:left">userId</th>
          <th style="padding:3px 6px;text-align:left">userType</th>
          <th style="padding:3px 6px;text-align:center">RC</th>
          <th style="padding:3px 6px;text-align:center">תוצאה</th>
        </tr>`;
    d.attempts.forEach((a,i)=>{
      const ok=a.success, rc=a.rc||'?', lbl=a.label||'', isOk=lbl.startsWith('✅');
      h+=`<tr style="border-bottom:1px solid #0d1f33;background:${ok?'#1a3d22':isOk?'#1a2a1a':''}">
        <td style="padding:3px 6px;color:#6b94b8">${i+1}</td>
        <td style="padding:3px 6px;color:${isOk?'#3dba6f':'#8696a0'};font-size:10px">${lbl}</td>
        <td style="padding:3px 6px;color:#c084fc">${a.ban}</td>
        <td style="padding:3px 6px;color:#c084fc">${a.userId}</td>
        <td style="padding:3px 6px;color:#53bdeb">${a.userType}</td>
        <td style="padding:3px 6px;text-align:center;color:${rc==='00'?'#3dba6f':'#f15c6e'}">${rc}</td>
        <td style="padding:3px 6px;text-align:center;color:${ok?'#3dba6f':'#f15c6e'}">${ok?'✅ '+a.count:'❌ '+(a.desc||'ריק').substring(0,50)}</td>
      </tr>`;
    });
    h+=`</table></div></details>`;
  }

  // היסטוריית ביקורים
  if(!d.history||!d.history.length){
    h+=`<div style="color:#f15c6e;padding:12px;background:#1a0a0a;border-radius:8px;font-size:13px">
      ❌ לא נמצאה היסטוריה ללקוח זה
    </div>`;
  } else {
    h+=`<div style="font-weight:700;color:#eaf4ff;font-size:14px;margin-bottom:8px">
      📜 היסטוריית ביקורים — <span style="color:#3dba6f">${d.history.length} רשומות</span>
    </div>
    <div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:13px">
      <tr style="background:#0a1826;position:sticky;top:0">
        <th style="padding:6px 10px;color:#6b94b8;text-align:right">תיאור</th>
        <th style="padding:6px 10px;color:#6b94b8;text-align:center">סוג ביקור</th>
        <th style="padding:6px 10px;color:#6b94b8;text-align:center">תאריך</th>
        <th style="padding:6px 10px;color:#6b94b8;text-align:center">שעה</th>
        <th style="padding:6px 10px;color:#6b94b8;text-align:right">טכנאי</th>
        <th style="padding:6px 10px;color:#6b94b8;text-align:center">סטטוס</th>
      </tr>`;
    d.history.forEach(e=>{
      const isSvc = (e.visitType||'').trim()==='קריאת שירות';
      const typeColor = isSvc?'#3dba6f':'#6b94b8';
      const statusOk  = (e.status||'').includes('הושלם');
      h+=`<tr style="border-bottom:1px solid #0d1f33;background:${isSvc?'#0a1a10':''}">
        <td style="padding:5px 10px;color:#eaf4ff">${e.description||'—'}</td>
        <td style="padding:5px 10px;text-align:center;color:${typeColor};font-weight:${isSvc?'700':'400'}">${e.visitType||'—'}</td>
        <td style="padding:5px 10px;text-align:center;color:#9bbdd8;font-family:monospace">${e.dateEnd||e.fullDateTimeTo?.substring(0,10)||'—'}</td>
        <td style="padding:5px 10px;text-align:center;color:#6b94b8;font-family:monospace">${e.timeEnd||'—'}</td>
        <td style="padding:5px 10px;color:#c084fc">${e.technicianName||'—'}</td>
        <td style="padding:5px 10px;text-align:center;color:${statusOk?'#3dba6f':'#f9c846'}">${e.status||'—'}</td>
      </tr>`;
    });
    h+=`</table></div>`;
  }

  box.innerHTML=h;
}

async function runRepDbg(){
  const callId=(document.getElementById('dbg-callid').value||'').trim();
  const box=document.getElementById('rep-dbg-result');
  if(!box) return;
  if(!callId){box.innerHTML='<div style="color:#f9c846;padding:10px">הכנס מספר פקע</div>';return;}
  box.innerHTML='<div style="color:#c084fc;padding:10px">⏳ מחפש פקע '+callId+' ושולף היסטוריה...</div>';
  const manualRow=document.getElementById('dbg-manual-ban-row');
  if(manualRow) manualRow.style.display='none';
  try{
    const r=await apiFetch('/api/reports/visit-history-direct?call_id='+encodeURIComponent(callId));
    if(!r.ok){box.innerHTML=`<div style="color:#f15c6e">❌ שגיאת שרת ${r.status}</div>`;return;}
    const d=await r.json();
    if(d.error){
      box.innerHTML=`<div style="color:#f15c6e;padding:10px">❌ ${d.error}</div>`;
      if(manualRow) manualRow.style.display='flex';
      return;
    }
    // הצג BAN ידני אם: לא ב-DB, או BAN ריק/PI
    const banVal=d.ban||'';
    if(manualRow && (!d.found_in_db || !banVal || banVal.toUpperCase().startsWith('PI')))
      manualRow.style.display='flex';
    _renderVisitHistory(d, box);
  }catch(e){
    box.innerHTML=`<div style="color:#f15c6e">❌ שגיאה: ${e.message}</div>`;
  }
}

async function runRepDbgManual(){
  const callId=(document.getElementById('dbg-callid').value||'').trim();
  const ban=(document.getElementById('dbg-ban').value||'').trim();
  const cid=(document.getElementById('dbg-cid').value||'').trim();
  const box=document.getElementById('rep-dbg-result');
  if(!box) return;
  if(!ban&&!cid){box.innerHTML='<div style="color:#f9c846;padding:10px">הכנס BAN</div>';return;}
  box.innerHTML='<div style="color:#c084fc;padding:10px">⏳ שולף היסטוריה עם BAN ידני...</div>';
  try{
    const url='/api/reports/visit-history-direct?call_id='+encodeURIComponent(callId)
      +(ban?'&ban='+encodeURIComponent(ban):'')
      +(cid?'&cid='+encodeURIComponent(cid):'');
    const r=await apiFetch(url);
    if(!r.ok){box.innerHTML=`<div style="color:#f15c6e">❌ שגיאת שרת ${r.status}</div>`;return;}
    const d=await r.json();
    if(d.error){box.innerHTML=`<div style="color:#f15c6e;padding:10px">❌ ${d.error}</div>`;return;}
    _renderVisitHistory(d, box);
  }catch(e){
    box.innerHTML=`<div style="color:#f15c6e">❌ שגיאה: ${e.message}</div>`;
  }
}

async function saveVisitBan(){
  const callId=(document.getElementById('dbg-callid').value||'').trim();
  const ban=(document.getElementById('dbg-ban').value||'').trim();
  const cid=(document.getElementById('dbg-cid').value||'').trim();
  if(!callId||!ban){alert('חסר מספר פקע או BAN');return;}
  try{
    const r=await apiFetch('/api/reports/update-visit-ban',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({call_id:callId,ban:ban,customer_id:cid||undefined})
    });
    const d=await r.json();
    if(d.ok){
      const btn=document.getElementById('save-ban-btn');
      if(btn){btn.textContent='✅ נשמר!';btn.style.background='#1a3d22';setTimeout(()=>{btn.textContent='💾 שמור BAN ל-DB';btn.style.background='';},2000);}
    } else {
      alert('שגיאה: '+(d.error||'?'));
    }
  }catch(e){alert('שגיאה: '+e.message);}
}

async function loadArchiveList(){
  const box=document.getElementById('archive-list-box');
  if(!box) return;
  box.innerHTML='<div style="color:#6b94b8;font-size:13px">טוען...</div>';
  try{
    const r=await apiFetch('/api/reports/archive');
    const d=await r.json();
    if(!d.files||d.files.length===0){
      box.innerHTML='<div style="color:#6b94b8;font-size:13px;text-align:center;padding:10px">אין דוחות בארכיון</div>';
      return;
    }
    let h='<div style="display:flex;flex-direction:column;gap:8px">';
    d.files.forEach(f=>{
      const kb=Math.round(f.size/1024);
      const dt=new Date(f.mtime*1000).toLocaleDateString('he-IL');
      const label=f.name.replace(/_\d{8}_\d{4}\.xlsx$/,'').replace(/_/g,' ');
      h+=`<div style="background:#0d1b2a;border:1px solid #284461;border-radius:10px;padding:10px 12px;direction:rtl;display:flex;align-items:center;justify-content:space-between;gap:8px">
        <div>
          <div style="font-size:13px;font-weight:700;color:#eaf4ff">📄 ${label}</div>
          <div style="font-size:11px;color:#6b94b8">📅 ${dt} · ${kb} KB</div>
        </div>
        <div style="display:flex;gap:6px">
          <a href="/api/reports/archive/${encodeURIComponent(f.name)}?token=${encodeURIComponent(localStorage.getItem('cc_token')||'')}&phone=${encodeURIComponent(localStorage.getItem('cc_phone')||'')}"
            style="background:#1a3a5c;border:1px solid #53bdeb;color:#53bdeb;border-radius:8px;padding:6px 10px;font-size:12px;font-weight:700;text-decoration:none">📥</a>
          <button onclick="deleteArchive('${f.name}',this)"
            style="background:#3d1a1a;border:1px solid #f15c6e;color:#f15c6e;border-radius:8px;padding:6px 10px;font-size:12px;font-weight:700;cursor:pointer">🗑</button>
        </div>
      </div>`;
    });
    h+='</div>';
    box.innerHTML=h;
  }catch(e){
    box.innerHTML='<div style="color:#f15c6e;font-size:13px">שגיאה בטעינת ארכיון</div>';
  }
}

async function deleteArchive(filename, btn){
  if(!confirm('למחוק את הדוח?')) return;
  btn.disabled=true;
  try{
    await apiFetch('/api/reports/archive/'+encodeURIComponent(filename),{method:'DELETE'});
    loadArchiveList();
  }catch(e){ btn.disabled=false; alert('שגיאה במחיקה'); }
}

async function saveToArchive(){
  const from=document.getElementById('rep-from').value;
  const to=document.getElementById('rep-to').value;
  if(!from||!to){alert('בחר טווח תאריכים קודם');return;}
  const fromD=new Date(from), toD=new Date(to);
  const days=Math.round((toD-fromD)/86400000)+1;
  const HEBREW_MONTHS=['','ינואר','פברואר','מרץ','אפריל','מאי','יוני','יולי','אוגוסט','ספטמבר','אוקטובר','נובמבר','דצמבר'];
  const label=`${HEBREW_MONTHS[toD.getMonth()+1]} ${toD.getFullYear()}`;
  const btn=document.getElementById('rep-archive-save-btn');
  btn.disabled=true; btn.textContent='שומר...';
  try{
    const r=await apiFetch('/api/reports/archive/save',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({days,end:to,label})
    });
    const d=await r.json();
    if(d.ok){
      btn.textContent='✅ נשמר!';
      setTimeout(()=>{btn.textContent='💾 שמור לארכיון';btn.disabled=false;},2000);
    } else {
      alert(d.error||'שגיאה בשמירה'); btn.textContent='💾 שמור לארכיון'; btn.disabled=false;
    }
  }catch(e){ alert('שגיאה בשמירה'); btn.textContent='💾 שמור לארכיון'; btn.disabled=false; }
}

let _cmpSource='file';
function setCmpSource(src){
  _cmpSource=src;
  const btnFile=document.getElementById('cmp-src-file');
  const btnArch=document.getElementById('cmp-src-arch');
  const panelFile=document.getElementById('cmp-panel-file');
  const panelArch=document.getElementById('cmp-panel-archive');
  if(src==='file'){
    if(btnFile){btnFile.style.background='#25d366';btnFile.style.border='none';btnFile.style.color='#fff';}
    if(btnArch){btnArch.style.background='#223d58';btnArch.style.border='1px solid #284461';btnArch.style.color='#ddeeff';}
    if(panelFile) panelFile.style.display='';
    if(panelArch) panelArch.style.display='none';
  } else {
    if(btnArch){btnArch.style.background='#25d366';btnArch.style.border='none';btnArch.style.color='#fff';}
    if(btnFile){btnFile.style.background='#223d58';btnFile.style.border='1px solid #284461';btnFile.style.color='#ddeeff';}
    if(panelArch) panelArch.style.display='';
    if(panelFile) panelFile.style.display='none';
    // טעון רשימת ארכיון
    _loadCmpArchiveList();
  }
}

async function _loadCmpArchiveList(){
  const sel=document.getElementById('cmp-archive-select');
  if(!sel) return;
  try{
    const r=await apiFetch('/api/reports/archive');
    const d=await r.json();
    if(!d.files||!d.files.length){
      sel.innerHTML='<option value="">אין דוחות שמורים</option>';
      return;
    }
    sel.innerHTML='<option value="">בחר דוח להשוואה...</option>';
    d.files.forEach(f=>{
      const label=f.name.replace(/_\d{8}_\d{4}\.xlsx$/,'').replace(/_/g,' ');
      const opt=document.createElement('option');
      opt.value=f.name; opt.textContent=label;
      sel.appendChild(opt);
    });
  }catch(e){sel.innerHTML='<option value="">שגיאה בטעינת דוחות</option>';}
}

function _renderCompareResult(d, box){
  if(!d.diffs||d.diffs.length===0){
    box.innerHTML=`<div style="background:#1a3d22;border:1px solid #3dba6f;border-radius:10px;padding:14px;text-align:center;direction:rtl">
      <div style="font-size:18px;margin-bottom:6px">✅</div>
      <div style="font-weight:700;color:#3dba6f;font-size:14px">אין פערים!</div>
      <div style="color:#9bbdd8;font-size:12px;margin-top:4px">${d.matches} קריאות תואמות · ${d.range}</div>
    </div>`;
    return;
  }
  const missing_us=d.diffs.filter(x=>x.type==='חסר אצלנו');
  const missing_them=d.diffs.filter(x=>x.type==='חסר בקובץ');
  const pay_diff=d.diffs.filter(x=>x.type==='פער בתשלום');
  const srcLabel=d.source_file?` · ${d.source_file.replace(/_\d{8}_\d{4}\.xlsx$/,'').replace(/_/g,' ')}`:'';
  let h=`<div style="background:#1a2d42;border-radius:10px;padding:12px;direction:rtl">
    <div style="font-weight:700;color:#eaf4ff;margin-bottom:10px;font-size:14px">📊 תוצאות השוואה — ${d.range}${srcLabel}</div>
    <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px">
      <div style="background:#3d1a1a;border:1px solid #f15c6e;border-radius:8px;padding:6px 12px;font-size:12px;color:#f15c6e;font-weight:700">${missing_us.length} חסרים אצלנו</div>
      <div style="background:#3d2a00;border:1px solid #f9c846;border-radius:8px;padding:6px 12px;font-size:12px;color:#f9c846;font-weight:700">${missing_them.length} חסרים בקובץ</div>
      <div style="background:#1a2d4a;border:1px solid #53bdeb;border-radius:8px;padding:6px 12px;font-size:12px;color:#53bdeb;font-weight:700">${pay_diff.length} פערי תשלום</div>
      <div style="background:#1a3d22;border:1px solid #3dba6f;border-radius:8px;padding:6px 12px;font-size:12px;color:#3dba6f;font-weight:700">${d.matches} תואמות</div>
    </div>`;

  if(pay_diff.length>0){
    h+=`<div style="font-weight:700;color:#53bdeb;margin-bottom:6px;font-size:13px">💰 פערי תשלום (${pay_diff.length})</div>
    <div style="overflow-x:auto;margin-bottom:12px"><table style="width:100%;border-collapse:collapse;font-size:12px">
      <tr style="background:#0d1b2a">
        <th style="padding:6px 8px;color:#6b94b8;text-align:right">מספר פקע</th>
        <th style="padding:6px 8px;color:#6b94b8;text-align:right">סוג משימה</th>
        <th style="padding:6px 8px;color:#6b94b8;text-align:center">תאריך</th>
        <th style="padding:6px 8px;color:#6b94b8;text-align:center">אצלנו ₪</th>
        <th style="padding:6px 8px;color:#6b94b8;text-align:center">בקובץ ₪</th>
        <th style="padding:6px 8px;color:#6b94b8;text-align:center">פער ₪</th>
      </tr>`;
    pay_diff.forEach(x=>{
      const diffColor=x.diff>0?'#3dba6f':'#f15c6e';
      h+=`<tr style="border-bottom:1px solid #162d44">
        <td style="padding:5px 8px;color:#eaf4ff;font-weight:700">${x.call_id}</td>
        <td style="padding:5px 8px;color:#9bbdd8;font-size:11px">${x.task_type||'—'}</td>
        <td style="padding:5px 8px;color:#9bbdd8;text-align:center">${x.date||''}</td>
        <td style="padding:5px 8px;color:#eaf4ff;text-align:center">${x.ours_payment!=null?'₪'+x.ours_payment:'—'}</td>
        <td style="padding:5px 8px;color:#eaf4ff;text-align:center">${x.their_payment!=null?'₪'+x.their_payment:'—'}</td>
        <td style="padding:5px 8px;color:${diffColor};text-align:center;font-weight:700">${x.diff>0?'+':''}${x.diff}</td>
      </tr>`;
    });
    h+='</table></div>';
  }

  if(missing_us.length>0){
    h+=`<div style="font-weight:700;color:#f15c6e;margin-bottom:6px;font-size:13px">❌ חסרים אצלנו (${missing_us.length})</div>
    <div style="overflow-x:auto;margin-bottom:12px"><table style="width:100%;border-collapse:collapse;font-size:12px">
      <tr style="background:#0d1b2a">
        <th style="padding:6px 8px;color:#6b94b8;text-align:right">מספר פקע</th>
        <th style="padding:6px 8px;color:#6b94b8;text-align:center">תאריך</th>
        <th style="padding:6px 8px;color:#6b94b8;text-align:center">בקובץ ₪</th>
        <th style="padding:6px 8px;color:#6b94b8;text-align:center">פער ₪</th>
      </tr>`;
    missing_us.forEach(x=>{
      const pay=x.their_payment||0;
      h+=`<tr style="border-bottom:1px solid #162d44">
        <td style="padding:5px 8px;color:#f15c6e;font-weight:700">${x.call_id}</td>
        <td style="padding:5px 8px;color:#9bbdd8;text-align:center">${x.date||''}</td>
        <td style="padding:5px 8px;color:#eaf4ff;text-align:center">${pay?'₪'+pay:'—'}</td>
        <td style="padding:5px 8px;color:#f15c6e;text-align:center;font-weight:700">${pay?'-₪'+pay:'—'}</td>
      </tr>`;
    });
    h+='</table></div>';
  }

  if(missing_them.length>0){
    h+=`<div style="font-weight:700;color:#f9c846;margin-bottom:6px;font-size:13px">⚠️ חסרים בקובץ (${missing_them.length})</div>
    <div style="overflow-x:auto;margin-bottom:12px"><table style="width:100%;border-collapse:collapse;font-size:12px">
      <tr style="background:#0d1b2a">
        <th style="padding:6px 8px;color:#6b94b8;text-align:right">מספר פקע</th>
        <th style="padding:6px 8px;color:#6b94b8;text-align:right">סוג משימה</th>
        <th style="padding:6px 8px;color:#6b94b8;text-align:center">תאריך</th>
        <th style="padding:6px 8px;color:#6b94b8;text-align:center">אצלנו ₪</th>
        <th style="padding:6px 8px;color:#6b94b8;text-align:center">פער ₪</th>
      </tr>`;
    missing_them.forEach(x=>{
      const pay=x.ours_payment||0;
      h+=`<tr style="border-bottom:1px solid #162d44">
        <td style="padding:5px 8px;color:#f9c846;font-weight:700">${x.call_id}</td>
        <td style="padding:5px 8px;color:#9bbdd8;font-size:11px">${x.task_type||'—'}</td>
        <td style="padding:5px 8px;color:#9bbdd8;text-align:center">${x.date||''}</td>
        <td style="padding:5px 8px;color:#eaf4ff;text-align:center">${pay?'₪'+pay:'—'}</td>
        <td style="padding:5px 8px;color:#3dba6f;text-align:center;font-weight:700">${pay?'+₪'+pay:'—'}</td>
      </tr>`;
    });
    h+='</table></div>';
  }
  h+='</div>';
  box.innerHTML=h;
}

async function runCompare(){
  const box=document.getElementById('compare-result');
  box.innerHTML='<div style="color:#6b94b8;font-size:13px;text-align:center;padding:10px">⏳ מעבד...</div>';
  const tk=localStorage.getItem('cc_token')||'';
  const ph=localStorage.getItem('cc_phone')||'';
  try{
    let d;
    if(_cmpSource==='archive'){
      const sel=document.getElementById('cmp-archive-select');
      const fn=sel?sel.value:'';
      if(!fn){box.innerHTML='<div style="color:#f15c6e;font-size:13px">בחר דוח קודם</div>';return;}
      const r=await apiFetch('/api/reports/compare-from-archive?file='+encodeURIComponent(fn));
      d=await r.json();
    } else {
      const input=document.getElementById('compare-file-input');
      if(!input.files||!input.files[0]){box.innerHTML='<div style="color:#f15c6e;font-size:13px">בחר קובץ קודם</div>';return;}
      const fd=new FormData();
      fd.append('file', input.files[0]);
      const r=await fetch('/api/reports/compare',{
        method:'POST', headers:{'X-Token':tk,'X-Phone':ph}, body:fd
      });
      d=await r.json();
    }
    if(d.error){box.innerHTML=`<div style="color:#f15c6e;font-size:13px">❌ ${d.error}</div>`;return;}
    _renderCompareResult(d, box);
  }catch(e){ box.innerHTML=`<div style="color:#f15c6e;font-size:13px">❌ שגיאה: ${e.message}</div>`; }
}

async function loadReports(){
  const from=document.getElementById('rep-from').value;
  const to=document.getElementById('rep-to').value;
  const box=document.getElementById('rep-content');
  if(!from||!to){box.innerHTML='<div style="color:#f15c6e">בחר תאריכים</div>';return;}
  const fromD=new Date(from), toD=new Date(to);
  const days=Math.round((toD-fromD)/86400000)+1;
  box.innerHTML=`<div style="color:#8696a0;text-align:center;padding:20px">טוען... (${days} ימים)</div>`;
  try{
    const r=await apiFetch('/api/reports?days='+days+'&end='+to);
    if(r.status===401){box.innerHTML='<div style="color:#f15c6e">❌ לא מחובר</div>';return;}
    const d=await r.json();
    if(!d.visits||d.visits.length===0){box.innerHTML='<div style="color:#8696a0;text-align:center;padding:20px">לא נמצאו ביקורים בטווח זה</div>';return;}

    // שמור ביקורים גלובלית למפה
    _lastVisits=d.visits;

    // הצג כפתורי ייצוא ומפה
    const expBtn=document.getElementById('rep-export-btn');
    if(expBtn) expBtn.style.display='inline-block';
    const archBtn=document.getElementById('rep-archive-save-btn');
    if(archBtn) archBtn.style.display='inline-block';

    const hasEarnings=d.total_earnings>0;
    let html=`<div style="background:#1a2d42;border-radius:10px;padding:12px;margin-bottom:10px">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
        <div style="font-weight:700;color:#25d366">סיכום — ${d.total} ביקורים</div>
        ${hasEarnings?`<div style="font-weight:700;font-size:16px;color:#25d366">₪${d.total_earnings.toLocaleString('he-IL',{minimumFractionDigits:0,maximumFractionDigits:2})}</div>`:''}
      </div>`;
    for(const [type,info] of Object.entries(d.summary)){
      const pct=Math.round(info.count/d.total*100);
      html+=`<div style="display:flex;justify-content:space-between;font-size:13px;padding:3px 0;border-bottom:1px solid #284461">
        <span>${type} <span style="color:#6b94b8">${info.count} (${pct}%)</span></span>
        ${info.earnings>0?`<span style="color:#25d366">₪${info.earnings.toLocaleString('he-IL',{minimumFractionDigits:0,maximumFractionDigits:2})}</span>`:'<span style="color:#7a9ab0">ללא מחיר</span>'}
      </div>`;
    }
    if(!hasEarnings) html+=`<div style="color:#6b94b8;font-size:12px;margin-top:8px">💡 הגדר מחירים בלשונית 💰 מחירון</div>`;
    html+=`</div>`;

    // זיהוי חוזרות — מסמן את הביקור המקורי באדום (is_repeat_source)
    // כולל בדיקה 30 יום לפני ואחרי טווח הדוח
    const REPEAT_FAULTS=[
      'דחוף שימור לקוח',
      'לא צופה ולא גולש מלא',
      'לקוח מושבת',
      'תקלה בקו טלפון',
      'תקלת גלישה תשתית סיבים',
    ];
    const isRepeatFault=tt=>REPEAT_FAULTS.some(f=>(tt||'').includes(f));
    // בנה map מ-call_id לאובייקט ביקור (רק ביקורי הדוח)
    const visitById={};
    d.visits.forEach(v=>{ if(v.call_id) visitById[v.call_id]=v; });
    d.visits.forEach(v=>{ v.is_repeat_source=false; });
    // קיבוץ כל הביקורים (pre + דוח + post) לפי לקוח
    const byCustomer={};
    const addToCust=(v,tag)=>{
      const cid=v.customer_id||''; if(!cid) return;
      if(!byCustomer[cid]) byCustomer[cid]=[];
      byCustomer[cid].push({...v,_tag:tag});
    };
    (d.pre_visits||[]).forEach(v=>addToCust(v,'pre'));
    d.visits.forEach(v=>addToCust(v,'main'));
    (d.post_visits||[]).forEach(v=>addToCust(v,'post'));
    Object.values(byCustomer).forEach(cvs=>{
      cvs.sort((a,b)=>a.fetch_date.localeCompare(b.fetch_date)||(a.appt_start||'').localeCompare(b.appt_start||''));
      for(let i=1;i<cvs.length;i++){
        const prev=cvs[i-1], cur=cvs[i];
        const diff=(new Date(cur.fetch_date)-new Date(prev.fetch_date))/86400000;
        if(diff<=30 && isRepeatFault(cur.task_type)){
          // הביקור המקורי (prev) מסומן — אם הוא בדוח
          if(prev._tag==='main' && visitById[prev.call_id]){
            visitById[prev.call_id].is_repeat_source=true;
          }
        }
      }
    });

    // קיבוץ לפי תאריך
    const byDate={};
    d.visits.forEach(v=>{if(!byDate[v.fetch_date])byDate[v.fetch_date]=[];byDate[v.fetch_date].push(v);});
    for(const [dt, rows] of Object.entries(byDate)){
      const dateObj=new Date(dt+'T12:00:00');
      const dayName=['ראשון','שני','שלישי','רביעי','חמישי','שישי','שבת'][dateObj.getDay()];
      const dayTotal=rows.reduce((s,v)=>s+(v.price||0),0);
      html+=`<div style="display:flex;justify-content:space-between;align-items:center;margin:12px 0 6px">
        <div style="font-weight:700;color:#53bdeb;font-size:13px">יום ${dayName} ${dt} (${rows.length})</div>
        <div id="day_total_${dt}" style="font-weight:700;color:#25d366;font-size:14px">${dayTotal>0?`₪${dayTotal.toLocaleString('he-IL',{minimumFractionDigits:0,maximumFractionDigits:2})}`:''}
        </div>
      </div>`;
      rows.forEach((v,vi)=>{
        const isVip=v.is_vip==='כן';
        const ppc=d.prices_map?d.prices_map[(v.task_type||'')+'__PPC']||0:0;
        const lsb=d.prices_map?d.prices_map[(v.task_type||'')+'__LSB']||0:0;
        const cardId=`vc_${dt}_${vi}`;
        const callId=v.call_id||'';
        const isIncomplete=v.is_incomplete===true;
        const isRepeat=v.is_repeat_source===true;
        const isCustom=v.price_type==='custom';
        const isFiber=((v.task_type||'').includes('תשתית סיבים')||(v.task_type||'').includes('טריפל סיבים'))&&!(v.task_type||'').includes('תקלת גלישה');
        const borderColor=isRepeat?'#f15c6e':isIncomplete?'#f15c6e':isVip?'#f9c846':'#25d366';
        const bgColor=isRepeat?'#fce8e8':'#d0e4f7';
        html+=`<div class="inv-card ok" data-dt="${dt}" data-price="${v.price||0}" data-repeat="${isRepeat?1:0}" style="margin:4px 0;border-right-color:${borderColor};background:${bgColor}" id="${cardId}">
          <div style="display:flex;justify-content:space-between;align-items:center">
            <span style="font-weight:700;font-size:13px">${isRepeat?'🔁 ':''}${v.task_type||'—'}</span>
            <span style="display:flex;gap:6px;align-items:center">
              ${isVip?'<span class="inv-badge badge-blocked">VIP</span>':''}
              ${isIncomplete
                ?'<span style="color:#f15c6e;font-weight:700;font-size:13px">לא הושלם ₪0</span>'
                :isCustom
                  ?`<span id="${cardId}_price" style="color:#f9c846;font-weight:700;font-size:13px">✏️ ₪${v.price}</span>`
                  :(isFiber&&(ppc||lsb))?`
                <button id="${cardId}_ppc" onclick="setPriceType('${cardId}','PPC',${ppc},'${dt}','${callId}')"
                  style="background:${v.price_type==='PPC'?'#005c4b':'#1a3a5c'};border:1px solid ${v.price_type==='PPC'?'#25d366':'#3d5166'};color:${v.price_type==='PPC'?'#25d366':'#53bdeb'};border-radius:6px;padding:3px 8px;font-size:12px;cursor:pointer">
                  ${v.price_type==='PPC'?'✅ ':''}PPC ₪${ppc}</button>
                <button id="${cardId}_lsb" onclick="setPriceType('${cardId}','LSB',${lsb},'${dt}','${callId}')"
                  style="background:${v.price_type==='LSB'?'#005c4b':'#1a3a5c'};border:1px solid ${v.price_type==='LSB'?'#25d366':'#3d5166'};color:${v.price_type==='LSB'?'#25d366':'#53bdeb'};border-radius:6px;padding:3px 8px;font-size:12px;cursor:pointer">
                  ${v.price_type==='LSB'?'✅ ':''}LSB ₪${lsb}</button>`
                  :(v.price>0?`<span id="${cardId}_price" style="color:#25d366;font-size:13px">₪${v.price}</span>`:'')
              }
              ${!isIncomplete?`<button onclick="showOverride('${cardId}','${callId}','${dt}',${v.price||0})"
                style="background:transparent;border:1px solid #6b94b8;color:#6b94b8;border-radius:6px;padding:2px 7px;font-size:11px;cursor:pointer">✏️ חריגה</button>`:''}
            </span>
          </div>
          <div style="color:#0d1b2a;font-size:13px">${v.contact_name||''}</div>
          <div style="color:#4a6a88;font-size:12px">📍 ${v.street||''} ${v.city||''}</div>
          ${v.appt_start?`<div style="color:#4a6a88;font-size:12px">⏰ ${v.appt_start}–${v.appt_finish}</div>`:''}
          ${v.status&&!isIncomplete?`<div style="color:#4a6a88;font-size:12px">סטטוס: ${v.status}</div>`:''}
          ${v.infrastructure?`<div style="color:#4a6a88;font-size:12px">🌐 ${v.infrastructure}</div>`:''}
          <div id="${cardId}_ovr" style="font-size:12px;display:none;margin-top:6px;gap:6px;align-items:center">
            <input type="number" id="${cardId}_ovr_inp" placeholder="סכום חריגה ₪" min="0" step="0.01"
              style="background:#223d58;border:1px solid #284461;border-radius:8px;padding:5px 8px;color:#ddeeff;font-size:13px;width:130px">
            <button onclick="saveOverride('${cardId}','${callId}','${dt}')"
              style="background:#3dba6f;border:none;border-radius:8px;padding:5px 10px;color:#fff;font-size:12px;cursor:pointer">שמור</button>
            <button onclick="clearOverride('${cardId}','${callId}','${dt}')"
              style="background:transparent;border:1px solid #f15c6e;border-radius:8px;padding:5px 10px;color:#f15c6e;font-size:12px;cursor:pointer">בטל</button>
          </div>
          <div id="${cardId}_sel" style="font-size:12px;color:#25d366;margin-top:4px"></div>
        </div>`;
      });
    }
    // סיכום חוזרות
    const repeatVisits=d.visits.filter(v=>v.is_repeat_source===true);
    if(repeatVisits.length>0){
      const repEarnings=repeatVisits.reduce((s,v)=>s+(v.price||0),0);
      html+=`<div style="background:#3d1a1a;border:1px solid #f15c6e;border-radius:10px;padding:12px;margin-top:14px;direction:rtl">
        <div style="font-weight:700;color:#f15c6e;font-size:14px;margin-bottom:6px">🔁 סיכום חוזרות — ${repeatVisits.length} ביקורים</div>`;
      const repByType={};
      repeatVisits.forEach(v=>{const tt=v.task_type||'—';if(!repByType[tt])repByType[tt]={count:0,earn:0};repByType[tt].count++;repByType[tt].earn+=(v.price||0);});
      for(const [tt,info] of Object.entries(repByType)){
        html+=`<div style="display:flex;justify-content:space-between;font-size:13px;padding:3px 0;border-bottom:1px solid #5a2020;color:#f0d0d0">
          <span>${tt} <span style="color:#f15c6e">(${info.count})</span></span>
          ${info.earn>0?`<span style="color:#f9c846">₪${info.earn.toLocaleString('he-IL',{minimumFractionDigits:0,maximumFractionDigits:2})}</span>`:'<span style="color:#7a5050">ללא מחיר</span>'}
        </div>`;
      }
      if(repEarnings>0) html+=`<div style="font-weight:700;color:#f9c846;font-size:15px;margin-top:8px;text-align:left">סה"כ חוזרות: ₪${repEarnings.toLocaleString('he-IL',{minimumFractionDigits:0,maximumFractionDigits:2})}</div>`;
      html+=`</div>`;
    }
    box.innerHTML=html;
  }catch(e){box.innerHTML='<div style="color:#f15c6e;text-align:center;padding:20px">❌ שגיאה בטעינה<br><small style="font-size:11px;opacity:0.7">'+String(e)+'</small></div>';console.error('loadReports error:',e);}
}

async function exportRepeats(){
  // Show the repeats modal overlay
  let overlay=document.getElementById('repeats-overlay');
  if(!overlay){
    overlay=document.createElement('div');
    overlay.id='repeats-overlay';
    overlay.style.cssText='position:fixed;inset:0;background:rgba(0,0,0,0.85);z-index:9999;overflow-y:auto;padding:20px 10px;direction:rtl';
    document.body.appendChild(overlay);
  }
  overlay.innerHTML=`
    <div style="max-width:900px;margin:0 auto;background:#0d1f33;border-radius:14px;padding:20px;border:1px solid #1e3a5f">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px">
        <div style="font-size:18px;font-weight:700;color:#f9c846">🔁 דוח חוזרות</div>
        <button onclick="document.getElementById('repeats-overlay').remove()"
          style="background:none;border:1px solid #f15c6e;color:#f15c6e;border-radius:6px;padding:4px 12px;cursor:pointer;font-size:14px">✕ סגור</button>
      </div>
      <div id="repeats-body" style="color:#c9d8e8;font-size:14px;text-align:center;padding:30px">
        <div style="font-size:28px;margin-bottom:10px">⏳</div>
        <div>טוען נתונים, זה יכול לקחת כמה שניות...</div>
      </div>
    </div>`;
  overlay.style.display='block';

  try{
    const token=localStorage.getItem('cc_token')||'';
    const phone=localStorage.getItem('cc_phone')||'';
    const res=await fetch('/api/reports/repeats-data?token='+encodeURIComponent(token)+'&phone='+encodeURIComponent(phone));
    const data=await res.json();
    if(data.error){
      document.getElementById('repeats-body').innerHTML=`<div style="color:#f15c6e;padding:20px">❌ שגיאה: ${data.error}</div>`;
      return;
    }
    const months=data.months||[];
    if(!months.length||months.every(m=>!(m.repeats||[]).length)){
      document.getElementById('repeats-body').innerHTML='<div style="color:#6b94b8;padding:30px;text-align:center">לא נמצאו חוזרות</div>';
      return;
    }

    let html='';
    let grandTotal=0, grandVisits=0, grandPrice=0;

    for(const m of months){
      const reps=m.repeats||[];
      if(!reps.length) continue;
      const mPrice=reps.reduce((s,r)=>s+(r.my_price||0),0);
      grandTotal+=reps.length;
      grandVisits+=(m.total_visits||0);
      grandPrice+=mPrice;

      html+=`<div style="margin-bottom:20px">
        <div style="font-weight:700;color:#f9c846;font-size:15px;margin-bottom:8px;padding-bottom:6px;border-bottom:1px solid #1e3a5f">
          📅 ${m.label||m.month} — ${reps.length} חוזרות מתוך ${m.total_visits||0} ביקורים
          ${mPrice>0?`<span style="color:#3dba6f;margin-right:10px">₪${mPrice.toFixed(2)}</span>`:''}
        </div>
        <div style="overflow-x:auto">
        <table style="width:100%;border-collapse:collapse;font-size:13px">
          <thead>
            <tr style="background:#0a1826;color:#6b94b8;text-align:right">
              <th style="padding:7px 10px;border-bottom:1px solid #1e3a5f">תאריך ביקורי</th>
              <th style="padding:7px 10px;border-bottom:1px solid #1e3a5f">סוג משימה שלי</th>
              <th style="padding:7px 10px;border-bottom:1px solid #1e3a5f;color:#3dba6f">סכום ₪</th>
              <th style="padding:7px 10px;border-bottom:1px solid #1e3a5f">תאריך טכנאי</th>
              <th style="padding:7px 10px;border-bottom:1px solid #1e3a5f">שם הטכנאי</th>
              <th style="padding:7px 10px;border-bottom:1px solid #1e3a5f">סוג משימה טכנאי</th>
              <th style="padding:7px 10px;border-bottom:1px solid #1e3a5f">ימים אחרי</th>
              <th style="padding:7px 10px;border-bottom:1px solid #1e3a5f">לקוח</th>
              <th style="padding:7px 10px;border-bottom:1px solid #1e3a5f">כתובת</th>
            </tr>
          </thead>
          <tbody>`;
      for(const r of reps){
        const price=r.my_price?`<span style="color:#3dba6f">₪${r.my_price.toFixed(2)}</span>`:'—';
        html+=`<tr style="border-bottom:1px solid #0d1f33;transition:background 0.15s" onmouseover="this.style.background='#132538'" onmouseout="this.style.background=''">
          <td style="padding:7px 10px;white-space:nowrap">${r.my_date||''}</td>
          <td style="padding:7px 10px">${r.my_type||''}</td>
          <td style="padding:7px 10px">${price}</td>
          <td style="padding:7px 10px;white-space:nowrap">${r.tech_date||''}</td>
          <td style="padding:7px 10px">${r.tech_name||''}</td>
          <td style="padding:7px 10px">${r.tech_type||''}</td>
          <td style="padding:7px 10px;text-align:center;color:#f9c846">${r.days_diff||''}</td>
          <td style="padding:7px 10px">${r.cust_name||''}</td>
          <td style="padding:7px 10px;font-size:12px;color:#6b94b8">${r.address||''}</td>
        </tr>`;
      }
      html+=`</tbody></table></div></div>`;
    }

    // Grand summary
    html+=`<div style="margin-top:16px;padding:14px;background:#0a1826;border-radius:10px;border:1px solid #1e3a5f;display:flex;gap:24px;flex-wrap:wrap;justify-content:center">
      <div style="text-align:center"><div style="font-size:22px;font-weight:700;color:#f15c6e">${grandTotal}</div><div style="color:#6b94b8;font-size:12px">סה"כ חוזרות</div></div>
      <div style="text-align:center"><div style="font-size:22px;font-weight:700;color:#7db4e0">${grandVisits}</div><div style="color:#6b94b8;font-size:12px">סה"כ ביקורים</div></div>
      <div style="text-align:center"><div style="font-size:22px;font-weight:700;color:#3dba6f">₪${grandPrice.toFixed(2)}</div><div style="color:#6b94b8;font-size:12px">סה"כ סכום</div></div>
    </div>`;

    // Export button
    html+=`<div style="text-align:center;margin-top:16px">
      <button onclick="doExportRepeatsXL()" style="background:#1a3a5c;border:1px solid #3d6fa8;color:#7db4e0;padding:10px 24px;border-radius:8px;cursor:pointer;font-size:14px">📥 ייצוא Excel</button>
    </div>`;

    document.getElementById('repeats-body').innerHTML=html;
  }catch(e){
    document.getElementById('repeats-body').innerHTML=`<div style="color:#f15c6e;padding:20px">❌ שגיאה: ${e.message}</div>`;
  }
}

function doExportRepeatsXL(){
  const token=localStorage.getItem('cc_token')||'';
  const phone=localStorage.getItem('cc_phone')||'';
  window.location.href=`/api/reports/export-repeats?token=${encodeURIComponent(token)}&phone=${encodeURIComponent(phone)}`;
}
function exportReports(){
  const from=document.getElementById('rep-from').value;
  const to=document.getElementById('rep-to').value;
  if(!from||!to) return;
  const fromD=new Date(from), toD=new Date(to);
  const days=Math.round((toD-fromD)/86400000)+1;
  const tk=localStorage.getItem('cc_token')||'';
  const ph=localStorage.getItem('cc_phone')||'';
  window.location.href='/api/reports/export?days='+days+'&end='+to+'&token='+encodeURIComponent(tk)+'&phone='+encodeURIComponent(ph);
}

/* ── price list ── */
async function loadPriceList(){
  const panel=document.getElementById('price-panel');
  panel.dataset.loaded='1';
  panel.innerHTML='<div style="color:#8696a0;text-align:center;padding:20px">טוען...</div>';
  try{
    const r=await apiFetch('/api/prices');
    const d=await r.json();
    const prices=d.prices||{};
    const types=d.task_types||[];
    // מיזוג — סוגים שיש מחיר אבל לא בDB
    const allTypes=[...new Set([...types,...Object.keys(prices).filter(k=>!k.includes('__PPC')&&!k.includes('__LSB'))])].sort();
    let html=`<div style="background:#1a2d42;border-radius:10px;padding:14px;margin-bottom:12px">
      <div style="font-weight:700;color:#25d366;margin-bottom:12px;font-size:15px">💰 מחירון לפי סוג משימה</div>
      <div style="color:#6b94b8;font-size:12px;margin-bottom:12px">הזן מחיר לכל סוג — ישמשו לחישוב הדוחות</div>`;
    if(allTypes.length===0){
      html+=`<div style="color:#6b94b8;font-size:13px">טען דוחות קודם כדי לראות את סוגי המשימות</div>`;
    } else {
      allTypes.forEach(t=>{
        const isFiber=(t.includes('תשתית סיבים')||t.includes('טריפל סיבים'))&&!t.includes('תקלת גלישה');
        if(isFiber){
          const valPPC=prices[t+'__PPC']||'';
          const valLSB=prices[t+'__LSB']||'';
          const isRepeatF=prices[t+'__repeat']?true:false;
          html+=`<div style="padding:8px 0;border-bottom:1px solid #284461">
            <div style="display:flex;align-items:center;gap:10px;margin-bottom:6px">
              <span style="font-size:14px;flex:1">${t}</span>
              <label style="display:flex;align-items:center;gap:4px;cursor:pointer;white-space:nowrap;font-size:12px;color:${isRepeatF?'#f9c846':'#6b94b8'}">
                <input type="checkbox" id="repeat_${encodeURIComponent(t)}" ${isRepeatF?'checked':''}
                  onchange="this.closest('label').style.color=this.checked?'#f9c846':'#6b94b8'"
                  style="width:15px;height:15px;cursor:pointer;accent-color:#f9c846">
                חוזרת
              </label>
            </div>
            <div style="display:flex;gap:8px">
              <div style="flex:1">
                <div style="font-size:11px;color:#6b94b8;margin-bottom:3px">PPC</div>
                <div style="display:flex;align-items:center;gap:4px">
                  <input type="number" id="price_${encodeURIComponent(t+'__PPC')}" value="${valPPC}" placeholder="0"
                    style="width:100%;background:#223d58;border:1px solid #284461;border-radius:8px;padding:7px 10px;color:#ddeeff;font-size:14px;text-align:left"
                    min="0" step="0.01">
                  <span style="color:#6b94b8;font-size:13px">₪</span>
                </div>
              </div>
              <div style="flex:1">
                <div style="font-size:11px;color:#6b94b8;margin-bottom:3px">LSB</div>
                <div style="display:flex;align-items:center;gap:4px">
                  <input type="number" id="price_${encodeURIComponent(t+'__LSB')}" value="${valLSB}" placeholder="0"
                    style="width:100%;background:#223d58;border:1px solid #284461;border-radius:8px;padding:7px 10px;color:#ddeeff;font-size:14px;text-align:left"
                    min="0" step="0.01">
                  <span style="color:#6b94b8;font-size:13px">₪</span>
                </div>
              </div>
            </div>
          </div>`;
        } else {
          const val=prices[t]||'';
          const isRepeat=prices[t+'__repeat']?true:false;
          html+=`<div style="display:flex;justify-content:space-between;align-items:center;padding:8px 0;border-bottom:1px solid #284461;gap:8px">
            <span style="font-size:14px;flex:1">${t}</span>
            <div style="display:flex;align-items:center;gap:10px">
              <label style="display:flex;align-items:center;gap:4px;cursor:pointer;white-space:nowrap;font-size:12px;color:${isRepeat?'#f9c846':'#6b94b8'}">
                <input type="checkbox" id="repeat_${encodeURIComponent(t)}" ${isRepeat?'checked':''}
                  onchange="this.closest('label').style.color=this.checked?'#f9c846':'#6b94b8'"
                  style="width:15px;height:15px;cursor:pointer;accent-color:#f9c846">
                חוזרת
              </label>
              <input type="number" id="price_${encodeURIComponent(t)}" value="${val}" placeholder="0"
                style="width:90px;background:#223d58;border:1px solid #284461;border-radius:8px;padding:7px 10px;color:#ddeeff;font-size:14px;text-align:left"
                min="0" step="0.01">
              <span style="color:#6b94b8;font-size:13px">₪</span>
            </div>
          </div>`;
        }
      });
    }
    html+=`<button class="btn-main" style="margin-top:14px" onclick="savePrices()">💾 שמור מחירון</button>
    <div id="price-msg" style="color:#25d366;font-size:13px;text-align:center;min-height:18px;margin-top:8px"></div>
    </div>`;
    panel.innerHTML=html;
  }catch(e){
    panel.innerHTML='<div style="color:#f15c6e;text-align:center;padding:20px">❌ שגיאה</div>';
  }
}

async function savePrices(){
  const msg=document.getElementById('price-msg');
  const inputs=document.querySelectorAll('#price-panel input[type=number]');
  const prices={};
  inputs.forEach(inp=>{
    const key=decodeURIComponent(inp.id.replace('price_',''));
    const val=parseFloat(inp.value);
    if(!isNaN(val)&&val>0) prices[key]=val;
  });
  // שמור סימוני חוזרת
  document.querySelectorAll('#price-panel input[type=checkbox]').forEach(cb=>{
    const key=decodeURIComponent(cb.id.replace('repeat_',''))+'__repeat';
    if(cb.checked) prices[key]=1;
  });
  try{
    const r=await apiFetch('/api/prices',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({prices})});
    const d=await r.json();
    msg.textContent=d.status==='ok'?'✅ נשמר!':'❌ שגיאה';
    setTimeout(()=>msg.textContent='',3000);
  }catch(e){msg.textContent='❌ שגיאת רשת';}
}

async function setPriceType(cardId, type, price, dt, callId){
  const ppcBtn=document.getElementById(cardId+'_ppc');
  const lsbBtn=document.getElementById(cardId+'_lsb');
  if(ppcBtn&&lsbBtn){
    const active='background:#005c4b;border:1px solid #25d366;color:#25d366;border-radius:6px;padding:3px 8px;font-size:12px;cursor:pointer';
    const inactive='background:#1a3a5c;border:1px solid #3d5166;color:#53bdeb;border-radius:6px;padding:3px 8px;font-size:12px;cursor:pointer';
    if(type==='PPC'){
      ppcBtn.style.cssText=active; ppcBtn.textContent=`✅ PPC ₪${price}`;
      lsbBtn.style.cssText=inactive; lsbBtn.textContent=lsbBtn.textContent.replace('✅ ','');
    } else {
      lsbBtn.style.cssText=active; lsbBtn.textContent=`✅ LSB ₪${price}`;
      ppcBtn.style.cssText=inactive; ppcBtn.textContent=ppcBtn.textContent.replace('✅ ','');
    }
  }
  // שמור בשרת
  if(callId){
    await apiFetch('/api/reports/set-override',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({call_id:callId,fetch_date:dt,price_type:type,custom_price:null})});
  }
  recalcDayTotal(dt, cardId, price);
}

function recalcDayTotal(dt, changedCard, newPrice){
  // עדכן מחיר הכרטיסייה המשתנה ב-dataset
  if(changedCard) document.getElementById(changedCard).dataset.price=newPrice;
  // חשב מחדש סכום יומי
  const dayEl=document.getElementById('day_total_'+dt);
  if(!dayEl) return;
  let total=0;
  document.querySelectorAll('.inv-card[data-dt="'+dt+'"]').forEach(c=>{
    total+=parseFloat(c.dataset.price||0);
  });
  dayEl.textContent=total>0?`₪${total.toLocaleString('he-IL',{minimumFractionDigits:0,maximumFractionDigits:2})}`:'';
}

function showOverride(cardId, callId, dt, currentPrice){
  const ovr=document.getElementById(cardId+'_ovr');
  if(!ovr) return;
  const isVisible=ovr.style.display==='flex';
  ovr.style.display=isVisible?'none':'flex';
  if(!isVisible){
    const inp=document.getElementById(cardId+'_ovr_inp');
    if(inp){inp.value=currentPrice||'';inp.focus();}
  }
}

async function saveOverride(cardId, callId, dt){
  const inp=document.getElementById(cardId+'_ovr_inp');
  if(!inp) return;
  const val=parseFloat(inp.value);
  if(isNaN(val)||val<0){inp.style.borderColor='#f15c6e';return;}
  await apiFetch('/api/reports/set-override',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({call_id:callId,fetch_date:dt,price_type:'custom',custom_price:val})});
  // עדכן תצוגה
  const priceEl=document.getElementById(cardId+'_price');
  if(priceEl){priceEl.textContent=`✏️ ₪${val}`;priceEl.style.color='#f9c846';}
  document.getElementById(cardId+'_ovr').style.display='none';
  document.getElementById(cardId).dataset.price=val;
  recalcDayTotal(dt);
}

async function clearOverride(cardId, callId, dt){
  await apiFetch('/api/reports/set-override',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({call_id:callId,fetch_date:dt,price_type:'',custom_price:null})});
  document.getElementById(cardId+'_ovr').style.display='none';
  loadReports(); // טען מחדש לחישוב מחיר מקורי
}

async function refreshInventory(){
  document.getElementById('inv-panel').dataset.loaded='';
  loadInventory();
}

/* ══════════════════════════════
   MAP — מפת ביקורים
══════════════════════════════ */
let _lastVisits=[];
let _mapInstance=null;
let _mapLoaded=false;

function _geoCache(){try{return JSON.parse(localStorage.getItem('_geo_cache')||'{}');}catch{return {};}}
function _saveGeoCache(c){try{localStorage.setItem('_geo_cache',JSON.stringify(c));}catch{}}

async function _loadLeaflet(){
  if(_mapLoaded) return;
  await new Promise((res,rej)=>{
    const s=document.createElement('script');
    s.src='https://unpkg.com/leaflet@1.9.4/dist/leaflet.js';
    s.onload=res; s.onerror=rej;
    document.head.appendChild(s);
  });
  _mapLoaded=true;
}

async function showReportsMap(){
  const modal=document.getElementById('map-modal');
  modal.style.display='flex';
  const status=document.getElementById('map-status');
  status.textContent='טוען מפה...';

  try{ await _loadLeaflet(); }catch(e){status.textContent='שגיאה בטעינת מפה';return;}

  // אתחול מפה (פעם ראשונה בלבד)
  if(!_mapInstance){
    _mapInstance=L.map('map-container',{zoomControl:true}).setView([31.8,35.0],8);
    L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png',{
      attribution:'© <a href="https://openstreetmap.org">OpenStreetMap</a>',
      maxZoom:19
    }).addTo(_mapInstance);
  } else {
    // נקה סמנים קודמים
    _mapInstance.eachLayer(l=>{if(l.options&&(l.options.radius!==undefined||l._icon)) _mapInstance.removeLayer(l);});
  }
  // גרום למפה לחשב גודל נכון אחרי שהמודל נפתח
  setTimeout(()=>_mapInstance.invalidateSize(),100);

  const bounds=[];

  // מיקום המשתמש
  status.textContent='מאתר מיקומך...';
  if(navigator.geolocation){
    try{
      const pos=await new Promise((res,rej)=>navigator.geolocation.getCurrentPosition(res,rej,{timeout:8000}));
      const ul=[pos.coords.latitude,pos.coords.longitude];
      bounds.push(ul);
      L.circleMarker(ul,{radius:13,fillColor:'#3dba6f',color:'#fff',weight:3,fillOpacity:0.95,zIndexOffset:1000})
        .addTo(_mapInstance)
        .bindPopup('<div dir="rtl"><b>📍 המיקום שלי</b></div>');
      _mapInstance.setView(ul,11);
    }catch(e){status.textContent='לא ניתן לאתר מיקום';}
  }

  if(!_lastVisits||!_lastVisits.length){
    status.textContent='אין ביקורים לתצוגה — טען דוח תחילה';
    return;
  }

  // גאוקוד כתובות (עם cache)
  const cache=_geoCache();
  let done=0, placed=0;
  const total=_lastVisits.length;

  for(const v of _lastVisits){
    const addr=((v.street||'')+' '+(v.city||'')+' ישראל').trim();
    if(!addr||addr==='ישראל'){done++;continue;}

    let lat,lng;
    if(cache[addr]){
      lat=cache[addr][0]; lng=cache[addr][1];
    } else {
      try{
        const r=await fetch('https://nominatim.openstreetmap.org/search?q='+encodeURIComponent(addr)+'&format=json&limit=1&countrycodes=il');
        const data=await r.json();
        if(data.length){
          lat=parseFloat(data[0].lat); lng=parseFloat(data[0].lon);
          cache[addr]=[lat,lng];
          _saveGeoCache(cache);
        }
      }catch(e){}
      // המתן כדי לכבד מגבלת Nominatim (1 בקשה/שנייה)
      await new Promise(r=>setTimeout(r,1100));
    }

    if(lat&&lng){
      bounds.push([lat,lng]);
      placed++;
      const isInc=v.is_incomplete===true;
      const isVip=v.is_vip==='כן';
      const col=isInc?'#f15c6e':isVip?'#f9c846':'#53bdeb';
      const priceStr=v.price>0?`<br>💰 ₪${v.price}`:'';
      L.circleMarker([lat,lng],{radius:9,fillColor:col,color:'#0d1b2a',weight:2,fillOpacity:0.9})
        .addTo(_mapInstance)
        .bindPopup(`<div dir="rtl" style="min-width:170px;font-family:Arial,sans-serif">
          <b style="font-size:13px">${v.task_type||''}</b><br>
          👤 ${v.contact_name||''}<br>
          📍 ${v.street||''} ${v.city||''}<br>
          📅 ${v.fetch_date||''} ${v.appt_start?'⏰ '+v.appt_start:''}
          ${priceStr}
          ${isInc?'<br><span style="color:#f15c6e;font-weight:700">⚠️ לא הושלם</span>':''}
          ${isVip?'<br><span style="color:#f9c846;font-weight:700">⭐ VIP</span>':''}
        </div>`);
    }

    done++;
    status.textContent=`ממפה... ${done}/${total}`;
  }

  if(bounds.length>1) _mapInstance.fitBounds(bounds,{padding:[40,40],maxZoom:14});
  status.textContent=`✅ ${placed} ביקורים על המפה`;
}

function closeMap(){
  document.getElementById('map-modal').style.display='none';
}

async function showTasksMap(){
  // מציג פקעות מבנק הפקעות על המפה — משתמש בקואורדינטות ישירות מה-API (ללא geocoding)
  const modal=document.getElementById('map-modal');
  modal.style.display='flex';
  const status=document.getElementById('map-status');
  status.textContent='טוען מפה...';

  try{ await _loadLeaflet(); }catch(e){status.textContent='שגיאה בטעינת מפה';return;}

  if(!_mapInstance){
    _mapInstance=L.map('map-container',{zoomControl:true}).setView([31.8,35.0],8);
    L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png',{
      attribution:'© <a href="https://openstreetmap.org">OpenStreetMap</a>',maxZoom:19
    }).addTo(_mapInstance);
  } else {
    _mapInstance.eachLayer(l=>{if(l.options&&(l.options.radius!==undefined||l._icon)) _mapInstance.removeLayer(l);});
  }
  setTimeout(()=>_mapInstance.invalidateSize(),100);

  const bounds=[];

  // מיקום המשתמש
  if(navigator.geolocation){
    try{
      const pos=await new Promise((res,rej)=>navigator.geolocation.getCurrentPosition(res,rej,{timeout:8000}));
      const ul=[pos.coords.latitude,pos.coords.longitude];
      bounds.push(ul);
      L.circleMarker(ul,{radius:13,fillColor:'#3dba6f',color:'#fff',weight:3,fillOpacity:0.95,zIndexOffset:1000})
        .addTo(_mapInstance).bindPopup('<div dir="rtl"><b>📍 המיקום שלי</b></div>');
      _mapInstance.setView(ul,12);
    }catch(e){}
  }

  if(!_allCards||!_allCards.length){
    status.textContent='טען פקעות תחילה (בנק פקעות)';return;
  }

  let placed=0;
  const total=_allCards.length;

  for(const c of _allCards){
    // קואורדינטות ישירות מה-API — ללא geocoding
    const lat=parseFloat(c.lat);
    const lng=parseFloat(c.lng);
    if(!lat||!lng||isNaN(lat)||isNaN(lng)) continue;

    bounds.push([lat,lng]);
    placed++;

    const isVip=(c.name||'').includes('VIP')||(c.task_type||'').includes('VIP');
    const color=isVip?'#f9c846':'#53bdeb';

    L.circleMarker([lat,lng],{radius:10,fillColor:color,color:'#0d1b2a',weight:2,fillOpacity:0.9})
      .addTo(_mapInstance)
      .bindPopup(`<div dir="rtl" style="min-width:180px;font-family:Arial,sans-serif;line-height:1.6">
        <b style="font-size:13px;color:#005c8a">📋 קריאה ${c.call_id||''}</b><br>
        👤 ${c.name||''}<br>
        📍 ${c.address||''}<br>
        🔧 ${c.task_type||''}<br>
        📅 ${c.date||''}${c.time?' ⏰ '+c.time:''}<br>
        ${c.phone?`📞 <a href="tel:${c.phone}" style="color:#0077b6">${c.phone}</a>`:''}
      </div>`);
  }

  if(bounds.length>1) _mapInstance.fitBounds(bounds,{padding:[40,40],maxZoom:14});
  status.textContent=`✅ ${placed} פקעות על המפה`;
}

/* ── הרשמה ── */
async function doRegister(){
  const name=document.getElementById('reg-name').value.trim();
  const errEl=document.getElementById('reg-err');
  const btn=document.getElementById('reg-btn');
  errEl.textContent='';
  if(!name){errEl.textContent='הכנס שם מלא';return;}
  btn.disabled=true;btn.textContent='שומר...';
  try{
    const r=await apiFetch('/api/register',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({name,employee_id:localStorage.getItem('cc_emp')||''})});
    const d=await r.json();
    if(d.status==='ok'){
      const st=await apiFetch('/api/status');
      const sd=await st.json();
      _isAdmin=sd.is_admin||false; _userName=sd.user_name||'';
      if(sd.permissions) _userPerms=Object.assign({},_userPerms,sd.permissions);
      if(_isAdmin) showAdminTab();
      applyPermissions();
      showScreen('chat');
    } else {
      errEl.textContent=d.message||'שגיאה';
    }
  }catch(e){errEl.textContent='שגיאת רשת';}
  btn.disabled=false;btn.textContent='🚀 התחל ניסיון חינם';
}
document.getElementById('reg-name').addEventListener('keypress',e=>{if(e.key==='Enter')doRegister()});


/* ── אדמין ── */
const PERM_LABELS={
  premium:'⭐ פרמיום (גישה לכל הלשוניות)',
  full_card:'כרטיס פקעה מלא',
  inventory:'לשונית מחסן',
  equipment_report:'דוח ציוד (תת-לשונית)',
  reports:'לשונית דוחות',
  prices:'לשונית מחירון'
};

/* ── permission months modal ── */
let _permModalResolve=null;
function showPermMonthsModal(label){
  return new Promise(resolve=>{
    _permModalResolve=resolve;
    document.getElementById('perm-modal-label').textContent='הרשאה: '+label;
    document.getElementById('perm-modal').style.display='flex';
  });
}
function closePermModal(confirmed){
  document.getElementById('perm-modal').style.display='none';
  if(_permModalResolve){
    const months=confirmed?parseFloat(document.getElementById('perm-months-sel').value):null;
    _permModalResolve(months===null?null:months);
    _permModalResolve=null;
  }
}

function toggleUserCard(hdr){
  const details=hdr.nextElementSibling;
  const arrow=hdr.querySelector('.usr-arr');
  const isOpen=details.style.display!=='none';
  details.style.display=isOpen?'none':'block';
  if(arrow) arrow.style.transform=isOpen?'':'rotate(180deg)';
}

function permExpiry(months){
  if(months===0||months==='0') return true;  // ללא הגבלה
  const days=Math.round(parseFloat(months)*30);
  const d=new Date(); d.setDate(d.getDate()+days);
  return d.toISOString().slice(0,10);
}

function daysLeft(isoDate){
  try{
    const exp=new Date(isoDate+'T00:00:00');
    const now=new Date(); now.setHours(0,0,0,0);
    return Math.round((exp-now)/(1000*60*60*24));
  }catch(e){return 0;}
}

function permValueLabel(v){
  if(v===false||v===null||v===undefined) return '';
  if(v===true) return '<span style="font-size:10px;color:#3dba6f;font-weight:600">✓ ללא הגבלה</span>';
  try{
    const days=daysLeft(v);
    if(days<0) return `<span style="font-size:10px;color:#f15c6e">פג תוקף (${v})</span>`;
    const color=days<=7?'#f9c846':days<=30?'#53bdeb':'#3dba6f';
    return `<span style="font-size:10px;color:${color};font-weight:600">נותרו ${days} ימים</span>`;
  }catch(e){return '';}
}

async function revokePerm(phone, permKey, cbId, spanId){
  if(!confirm('לבטל את ההרשאה "' + (PERM_LABELS[permKey]||permKey) + '"?')) return;
  const cb=document.getElementById(cbId);
  if(cb) cb.disabled=true;
  await saveSinglePerm(phone, permKey, false);
  if(cb){ cb.checked=false; cb.disabled=false; }
  document.getElementById(spanId).innerHTML='';
  document.getElementById('rb-'+cbId).innerHTML='';
  document.getElementById('eb-'+cbId).innerHTML='';
}

async function editPermTimer(phone, permKey, cbId, spanId){
  const label=PERM_LABELS[permKey]||permKey;
  const months=await showPermMonthsModal(label);
  if(months===null) return;
  const value=permExpiry(months);
  await saveSinglePerm(phone, permKey, value);
  document.getElementById(spanId).innerHTML=permValueLabel(value);
  const cb=document.getElementById(cbId);
  if(cb) cb._permValue=value;
}

async function togglePerm(phone, permKey, checkbox, labelSpanId){
  if(!checkbox.checked){
    // כיבוי — שמור מיד ללא חלון
    checkbox.disabled=true;
    await saveSinglePerm(phone, permKey, false);
    document.getElementById(labelSpanId).innerHTML='';
    checkbox.disabled=false;
    return;
  }
  // הפעלה — שאל כמה חודשים
  checkbox.checked=false; // החזר זמנית עד אישור
  const label=PERM_LABELS[permKey]||permKey;
  const months=await showPermMonthsModal(label);
  if(months===null){
    checkbox.checked=false; // ביטול
    return;
  }
  const value=permExpiry(months);
  checkbox.disabled=true;
  await saveSinglePerm(phone, permKey, value);
  checkbox.checked=true;
  document.getElementById(labelSpanId).innerHTML=permValueLabel(value);
  checkbox.disabled=false;
}

async function saveSinglePerm(phone, permKey, value){
  // שלוף הרשאות נוכחיות מה-DOM
  const card=document.querySelector(`[data-usrphone="${phone}"]`);
  const perms={};
  if(card){
    card.querySelectorAll('input[data-perm]').forEach(cb=>{
      perms[cb.dataset.perm]=cb._permValue!==undefined?cb._permValue:cb.checked;
    });
  }
  perms[permKey]=value;
  // שמור _permValue על ה-checkbox
  if(card){
    const cb=card.querySelector(`input[data-perm="${permKey}"]`);
    if(cb) cb._permValue=value;
  }
  try{
    await apiFetch('/api/admin/permissions/'+encodeURIComponent(phone),{
      method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify(perms)
    });
  }catch(e){}
}

async function loadAdminPanel(){
  const panel=document.getElementById('admin-panel');
  panel.dataset.loaded='1';
  panel.innerHTML='<div style="color:#8fa5b5;text-align:center;padding:24px">טוען...</div>';
  try{
    const r=await apiFetch('/api/admin/users');
    if(r.status===403){panel.innerHTML='<div style="color:#f07080;text-align:center;padding:24px">❌ אין הרשאת מנהל</div>';return;}
    const d=await r.json();
    const users=d.users||[];
    const permKeys=d.permission_keys||Object.keys(PERM_LABELS);
    let html=`<div style="font-weight:700;color:#3dba6f;font-size:16px;margin-bottom:14px">🛡️ לוח מנהל — ${users.length} משתמשים</div>`;
    if(!users.length){
      html+='<div style="color:#8fa5b5;text-align:center;padding:20px">אין משתמשים רשומים עדיין</div>';
    }
    users.forEach((u,i)=>{
      const maskedPhone=u.phone?u.phone.slice(0,3)+'****'+u.phone.slice(-2):'-';
      const isAdminUser=!!u.is_admin;
      const perms=u.permissions||{};
      let permsHtml='';
      if(!isAdminUser){
        permsHtml+=`<div style="margin-top:10px;border-top:1px solid #284461;padding-top:10px">
          <div style="font-size:12px;color:#8fa5b5;margin-bottom:8px;font-weight:600">🔐 הרשאות:</div>
          <div style="display:flex;flex-direction:column;gap:9px">`;
        permKeys.forEach(k=>{
          const v=perms.hasOwnProperty(k)?perms[k]:false;
          const isOn=v===true||(typeof v==='string'&&v>=(new Date().toISOString().slice(0,10)));
          const expLbl=permValueLabel(v);
          const spanId=`pv-${u.phone.replace(/\D/g,'')}-${k}`;
          const cbId=`cb-${u.phone.replace(/\D/g,'')}-${k}`;
          const label=PERM_LABELS[k]||k;
          const editBtn=isOn
            ?`<button onclick="editPermTimer('${u.phone}','${k}','${cbId}','${spanId}')"
                title="שנה טיימר"
                style="background:none;border:1px solid #53bdeb;color:#53bdeb;border-radius:6px;padding:2px 7px;font-size:11px;cursor:pointer;line-height:1.4">✏️</button>`
            :'';
          const revokeBtn=isOn
            ?`<button onclick="revokePerm('${u.phone}','${k}','${cbId}','${spanId}')"
                title="בטל הרשאה"
                style="background:none;border:1px solid #f15c6e;color:#f15c6e;border-radius:6px;padding:2px 7px;font-size:12px;cursor:pointer;line-height:1.4">✕ בטל</button>`
            :'';
          permsHtml+=`<div style="display:flex;align-items:center;gap:8px">
            <input id="${cbId}" type="checkbox" data-perm="${k}" ${isOn?'checked':''}
              onchange="togglePerm('${u.phone}','${k}',this,'${spanId}')"
              style="width:17px;height:17px;accent-color:#3dba6f;cursor:pointer">
            <span style="font-size:13px;color:#dde8ee;flex:1">${label}</span>
            <span id="${spanId}" style="font-size:11px">${expLbl}</span>
            <span id="eb-${cbId}">${editBtn}</span>
            <span id="rb-${cbId}">${revokeBtn}</span>
          </div>`;
        });
        permsHtml+=`</div></div>`;
      } else {
        permsHtml+=`<div style="margin-top:8px;font-size:12px;color:#3dba6f">✅ מנהל — גישה מלאה</div>`;
      }
      // תוקף מנוי: ההרשאה הפעילה הקרובה לפקוע
      let subSummary='';
      if(!isAdminUser){
        const activeDates=Object.values(perms).filter(v=>typeof v==='string'&&v);
        if(activeDates.length){
          const minDays=Math.min(...activeDates.map(daysLeft));
          const col=minDays<0?'#f15c6e':minDays<=7?'#f9c846':minDays<=30?'#53bdeb':'#3dba6f';
          subSummary=`<div style="margin-top:4px;font-size:12px;color:${col}">⏱ תוקף מנוי: ${minDays<0?'פג תוקף':`נותרו ${minDays} ימים`}</div>`;
        } else if(Object.values(perms).some(v=>v===true)){
          subSummary=`<div style="margin-top:4px;font-size:12px;color:#3dba6f">⏱ מנוי ללא הגבלה</div>`;
        } else {
          subSummary=`<div style="margin-top:4px;font-size:12px;color:#8fa5b5">⏱ אין מנוי פעיל</div>`;
        }
      }
      html+=`<div data-usrphone="${u.phone}" style="background:#1a2d42;border-radius:10px;margin-bottom:10px;overflow:hidden">
        <div onclick="toggleUserCard(this)" style="padding:12px 14px;cursor:pointer;display:flex;justify-content:space-between;align-items:center;gap:8px">
          <div style="flex:1;min-width:0">
            <div style="font-weight:700;font-size:15px;color:#edf2f5">${u.name||'ללא שם'}${isAdminUser?' <span style="font-size:11px;color:#3dba6f">🛡</span>':''}</div>
            <div style="color:#3dba6f;font-size:12px;margin-top:2px">נרשם: ${u.registered_at||'—'}</div>
            ${subSummary}
          </div>
          <span class="usr-arr" style="color:#6b94b8;font-size:20px;display:inline-block;transition:transform .25s;flex-shrink:0">▾</span>
        </div>
        <div style="display:none;padding:0 14px 12px;border-top:1px solid #284461">
          <div style="color:#9eabb5;font-size:13px;margin-top:8px">📱 ${maskedPhone}</div>
          ${permsHtml}
        </div>
      </div>`;
    });
    panel.innerHTML=html;
  }catch(e){
    panel.innerHTML='<div style="color:#f07080;text-align:center;padding:24px">❌ שגיאה בטעינה</div>';
  }
}

/* ════════════════════════════════════════
   צ'אט פנימי
   ════════════════════════════════════════ */
let _chatOpen=false;
let _chatWith='';       // phone of conversation partner (for admin)
let _chatPollTimer=null;

function initChat(){
  // הצג כפתור צ'אט לכל מי שמחובר
  const btn=document.getElementById('chat-float-btn');
  if(btn) btn.style.display='flex';
  startChatPoll();
}

function startChatPoll(){
  if(_chatPollTimer) clearInterval(_chatPollTimer);
  _chatPollTimer=setInterval(async()=>{
    if(_chatOpen && _chatWith){
      await refreshChatMessages();
    } else {
      await refreshUnreadBadge();
    }
  }, 8000);
}

async function refreshUnreadBadge(){
  try{
    const d=await apiFetch('/api/chat/unread').then(r=>r.json());
    const badge=document.getElementById('chat-unread-badge');
    if(!badge) return;
    const n=d.count||0;
    badge.textContent=n;
    badge.style.display=n>0?'flex':'none';
  }catch(e){}
}

function openChatModal(){
  _chatOpen=true;
  document.getElementById('chat-modal').style.display='flex';
  if(_isAdmin){
    showChatConvList();
  } else {
    document.getElementById('chat-input-bar').style.display='flex';
    updateChatHeader('💬 צ\'אט עם מנהל', false);
    openChatConversation('');
  }
}

function closeChatModal(){
  _chatOpen=false;
  _chatWith='';
  document.getElementById('chat-modal').style.display='none';
  refreshUnreadBadge();
}

async function showChatConvList(){
  _chatWith='';
  document.getElementById('chat-input-bar').style.display='none';
  const body=document.getElementById('chat-modal-body');
  body.innerHTML='<div style="color:#8fa5b5;text-align:center;padding:24px">טוען...</div>';
  updateChatHeader('💬 שיחות', false);
  try{
    const d=await apiFetch('/api/chat/conversations').then(r=>r.json());
    const convs=d.conversations||[];
    if(!convs.length){
      body.innerHTML='<div style="color:#8fa5b5;text-align:center;padding:40px;font-size:14px">אין שיחות עדיין</div>';
      return;
    }
    let html='';
    convs.forEach(c=>{
      const unreadBadge=c.unread>0?`<span style="background:#f15c6e;color:#fff;border-radius:50%;min-width:20px;height:20px;display:inline-flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;margin-right:4px">${c.unread}</span>`:'';
      const premBadge=c.premium?`<span style="font-size:11px;color:#f9c846;margin-right:4px">⭐ פרמיום${c.premium_expires?' עד '+c.premium_expires:''}</span>`:`<span style="font-size:11px;color:#8fa5b5">ללא פרמיום</span>`;
      html+=`<div onclick="openChatConversation('${c.phone}')" style="background:#d0e4f7;border-radius:10px;padding:12px 14px;margin-bottom:8px;cursor:pointer;display:flex;justify-content:space-between;align-items:center">
        <div>
          <div style="font-weight:700;color:#edf2f5;font-size:14px">${c.name} ${unreadBadge}</div>
          <div style="font-size:12px;color:#8fa5b5;margin-top:2px">${c.last_msg?c.last_msg.slice(0,40)+(c.last_msg.length>40?'...':''):''}</div>
          <div style="margin-top:4px">${premBadge}</div>
        </div>
        <span style="color:#8fa5b5;font-size:18px">›</span>
      </div>`;
    });
    body.innerHTML=html;
  }catch(e){
    body.innerHTML='<div style="color:#f07080;text-align:center;padding:24px">שגיאה בטעינה</div>';
  }
}

async function openChatConversation(withPhone){
  _chatWith=withPhone||'';
  document.getElementById('chat-input-bar').style.display='flex';
  const body=document.getElementById('chat-modal-body');
  body.innerHTML='<div style="color:#8fa5b5;text-align:center;padding:24px">טוען...</div>';
  const title=_isAdmin?(await getUserName(withPhone)):'צ\'אט עם מנהל';
  updateChatHeader(title, true, withPhone);
  await refreshChatMessages();
}

async function getUserName(phone){
  // try from conv list cache or return masked
  return phone?phone.slice(0,3)+'****'+phone.slice(-2):'—';
}

function updateChatHeader(title, showBack, withPhone=''){
  const hdr=document.getElementById('chat-modal-header');
  let grantBtn='';
  if(_isAdmin && withPhone){
    grantBtn=`<button onclick="showGrantModal('${withPhone}')" style="background:#3dba6f;color:#fff;border:none;border-radius:8px;padding:6px 12px;font-size:12px;font-weight:700;cursor:pointer;margin-right:8px">⭐ הענק פרמיום</button>`;
  }
  hdr.innerHTML=`
    <div style="display:flex;align-items:center;gap:10px">
      ${showBack&&_isAdmin?`<button onclick="showChatConvList()" style="background:none;border:none;color:#3dba6f;font-size:22px;cursor:pointer;padding:0">‹</button>`:''}
      <span style="font-weight:700;font-size:16px;color:#edf2f5">${title}</span>
    </div>
    <div style="display:flex;align-items:center">
      ${grantBtn}
      <button onclick="closeChatModal()" style="background:none;border:none;color:#8fa5b5;font-size:22px;cursor:pointer;padding:0 4px">✕</button>
    </div>`;
}

async function refreshChatMessages(){
  const body=document.getElementById('chat-modal-body');
  if(!body) return;
  const toPhone=_chatWith||(_isAdmin?'':ADMIN_PHONE_JS);
  if(!toPhone && _isAdmin) return;
  try{
    const d=await apiFetch('/api/chat/messages?with='+encodeURIComponent(toPhone)).then(r=>r.json());
    const msgs=d.messages||[];
    const wasAtBottom=body.scrollHeight-body.clientHeight<=body.scrollTop+40;
    const html=msgs.length?msgs.map(m=>{
      const mine=m.from===(_isAdmin?ADMIN_PHONE_JS:localStorage.getItem('cc_phone'));
      return `<div style="display:flex;justify-content:${mine?'flex-start':'flex-end'};margin-bottom:8px">
        <div style="max-width:80%;background:${mine?'#0d4a32':'#d0e4f7'};border-radius:${mine?'4px 14px 14px 14px':'14px 4px 14px 14px'};padding:8px 12px;font-size:14px;color:#ddeeff;line-height:1.5">
          ${m.body.replace(/\n/g,'<br>')}
          <div style="font-size:10px;color:#8fa5b5;margin-top:3px;text-align:${mine?'left':'right'}">${m.at}</div>
        </div>
      </div>`;
    }).join('')
    :'<div style="color:#8fa5b5;text-align:center;padding:30px;font-size:14px">אין הודעות עדיין — שלח הודעה למנהל</div>';
    body.innerHTML=html;
    if(wasAtBottom||!msgs.length) body.scrollTop=body.scrollHeight;
    await refreshUnreadBadge();
  }catch(e){}
}

async function sendChatMessage(){
  const inp=document.getElementById('chat-inp');
  const body=inp.value.trim();
  if(!body) return;
  inp.value='';
  const toPhone=_chatWith||(_isAdmin?'':ADMIN_PHONE_JS);
  if(!toPhone && _isAdmin) return;
  try{
    await apiFetch('/api/chat/send',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({body,to:toPhone})});
    await refreshChatMessages();
  }catch(e){inp.value=body;}
}

let _grantTarget='';
function showGrantModal(phone){
  _grantTarget=phone;
  document.getElementById('grant-modal').style.display='flex';
}
function closeGrantModal(){
  document.getElementById('grant-modal').style.display='none';
}
async function doGrantPremium(){
  const months=parseInt(document.getElementById('grant-months').value)||1;
  if(!_grantTarget) return;
  const btn=document.getElementById('grant-btn');
  btn.disabled=true; btn.textContent='מעניק...';
  try{
    const d=await apiFetch('/api/admin/grant-premium',{method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({phone:_grantTarget,months})}).then(r=>r.json());
    if(d.status==='ok'){
      closeGrantModal();
      await refreshChatMessages();
      // רענן שיחות
      if(!_chatWith) showChatConvList();
    }
  }catch(e){}
  btn.disabled=false; btn.textContent='הענק';
}

async function go(){
  const i=document.getElementById('inp');
  const t=i.value.trim();if(!t)return;
  i.value='';
  if(t.startsWith('פקעות')){
    add(t,'user');
    const parts=t.split(' ');
    loadTasks(parts[1]||'');
    return;
  }
  add(t,'user');
  const ty=document.getElementById('typing');
  document.getElementById('msgs').appendChild(ty);
  ty.style.display='block';
  document.getElementById('st').textContent='מעבד...';
  document.getElementById('msgs').scrollTop=9999;
  try{
    const r=await apiFetch('/chat',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({message:t})});
    const d=await r.json();
    ty.style.display='none';
    add(d.text||d.response||'','bot');
  }catch(e){ty.style.display='none';add('❌ שגיאה','bot');}
  document.getElementById('st').textContent='מחובר';
}
/* ══════════════════════════════════════
   🤖 AUTOMATION ENGINE
   ══════════════════════════════════════ */

/* --- helpers --- */
function autoLog(msg, type='info'){
  const el=document.getElementById('auto-log');
  if(!el) return;
  const colors={info:'#9bbdd8',ok:'#3dba6f',err:'#f15c6e',warn:'#f9c846',action:'#53bdeb'};
  const icons={info:'ℹ',ok:'✅',err:'❌',warn:'⚠️',action:'🔧'};
  const ts=new Date().toLocaleTimeString('he-IL',{hour:'2-digit',minute:'2-digit',second:'2-digit'});
  el.innerHTML+=`<div style="color:${colors[type]||colors.info};margin:2px 0"><span style="color:#4a6080">${ts}</span> ${icons[type]||''} ${msg}</div>`;
  el.scrollTop=el.scrollHeight;
}
function autoLogClear(){
  const el=document.getElementById('auto-log');
  if(el) el.innerHTML='';
  const s=document.getElementById('auto-summary');
  if(s) s.innerHTML='';
}
function autoSummary(items){
  const s=document.getElementById('auto-summary');
  if(!s||!items.length) return;
  const ok=items.filter(i=>i.ok).length;
  const bad=items.length-ok;
  s.innerHTML=`<div style="background:#162d44;border-radius:10px;padding:10px;direction:rtl;font-size:13px">
    <div style="font-weight:700;margin-bottom:6px;color:#eaf4ff">📋 סיכום: ${ok} תקין | ${bad} בעיה</div>
    ${items.map(i=>`<div style="color:${i.ok?'#3dba6f':i.warn?'#f9c846':'#f15c6e'};margin:2px 0">${i.ok?'✅':i.warn?'⚠️':'❌'} ${i.label}: ${i.value}</div>`).join('')}
  </div>`;
}

async function proxyFetch(url, method='GET', data=null, auth=null, json_body=null){
  try{
    const body={url, method};
    if(data) body.data=data;
    if(json_body) body.json=json_body;
    if(auth) body.auth=auth;
    const r=await apiFetch('/api/diag/proxy',{method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify(body)});
    if(!r.ok){
      const e=await r.json().catch(()=>({}));
      return {error: e.error||r.status};
    }
    return await r.json();
  }catch(e){ return {error:String(e)}; }
}

function extractField(html, patterns){
  for(const p of patterns){
    const m=html.match(p);
    if(m) return m[1]||m[0];
  }
  return null;
}

/* --- MOCA automation --- */
const MOCA_CREDS=[['admin','maxlinear'],['admin','entropic'],['admin','admin']];
const MOCA_TARGETS={lof:1450, txpow:8, beacon:10};

async function runMocaAuto(){
  autoLogClear();
  const ip=(document.getElementById('auto-moca-ip').value||'').trim();
  if(!ip){ autoLog('הזן IP של מוקה','warn'); return; }
  await _mocaAutoCore(ip);
}

async function _mocaAutoCore(ip){
  autoLog(`📡 מתחבר למוקה ${ip}...`,'action');
  let cred=null;
  for(const [u,p] of MOCA_CREDS){
    const r=await proxyFetch(`http://${ip}/`,  'GET', null, [u,p]);
    if(r && !r.error && r.status===200){ cred=[u,p]; break; }
  }
  if(!cred){ autoLog(`לא ניתן להתחבר ל-${ip} — בדוק IP וסיסמה`,'err'); return null; }
  autoLog(`חיבור הצליח (${cred[0]})`, 'ok');

  // === קרא דף ראשי ===
  const home=await proxyFetch(`http://${ip}/`,  'GET', null, cred);
  const html=home?.text||'';

  // === זיהוי Firmware ===
  let fw='unknown';
  if(/VSCO|vsco|94[^0-9]/i.test(html)) fw='vsco';
  else if(/MMXL|mmxl|2\.12/i.test(html)) fw='mmxl';
  else if(/REV\b|rev\b|\b89\b/i.test(html)) fw='rev';
  autoLog(`Firmware: ${fw.toUpperCase()}`, fw==='rev'?'err':fw==='unknown'?'warn':'ok');

  // === בדוק PHY Rates ===
  const phyMatches=[...html.matchAll(/(\d{3,4})\s*Mbps/gi)];
  if(phyMatches.length){
    const rates=phyMatches.map(m=>parseInt(m[1])).filter(v=>v>0);
    const thr=fw==='mmxl'?1300:600;
    const bad=rates.filter(r=>r<thr);
    if(bad.length) autoLog(`PHY Rates נמוכים: ${bad.join(', ')} Mbps (סף: ${thr})`,'warn');
    else autoLog(`PHY Rates תקינים (${rates.join(', ')} Mbps)`,'ok');
  }

  // === נסה לקרוא הגדרות מתקדמות ===
  const advUrls=['/moca_advanced.htm','/advanced.htm','/mocaadvanced.html',
                 '/settings/advanced','/cgi-bin/moca_adv'];
  let advHtml='';
  for(const u of advUrls){
    const r=await proxyFetch(`http://${ip}${u}`, 'GET', null, cred);
    if(r&&!r.error&&r.status===200&&r.text&&r.text.length>200){
      advHtml=r.text; break;
    }
  }

  const results=[];

  if(advHtml){
    // LOF
    const lofVal=extractField(advHtml,[/lof[^0-9]*(\d{4})/i,/name="lof"[^>]*value="(\d+)"/i,/LOF[^0-9]*(\d{4})/]);
    const txVal =extractField(advHtml,[/tx.power[^0-9]*(\d+)/i,/name="txpower"[^>]*value="(\d+)"/i,/Tx Power[^0-9]*(\d+)/i]);
    const beaconVal=extractField(advHtml,[/beacon.power[^0-9]*(\d+)/i,/name="beacon"[^>]*value="(\d+)"/i]);

    if(lofVal) autoLog(`LOF: ${lofVal} (יעד: 1450)`, parseInt(lofVal)===1450?'ok':'warn');
    if(txVal)  autoLog(`Tx Power: ${txVal} dBm (יעד: 8)`, parseInt(txVal)===8?'ok':'warn');
    if(beaconVal) autoLog(`Beacon: ${beaconVal} (יעד: 10)`, parseInt(beaconVal)===10?'ok':'warn');

    // תיקון הגדרות אם צריך
    const needFix=(lofVal&&parseInt(lofVal)!==1450)||(txVal&&parseInt(txVal)!==8)||(beaconVal&&parseInt(beaconVal)!==10);
    if(needFix){
      autoLog('מנסה לתקן הגדרות...','action');
      const fixUrls=['/goform/mocaSettings','/cgi-bin/apply','/apply.cgi','/settings/save'];
      for(const fu of fixUrls){
        const fixData={lof:1450, txpower:8, beacon_power:10};
        const fr=await proxyFetch(`http://${ip}${fu}`,'POST',fixData,cred);
        if(fr&&!fr.error&&fr.status<400){ autoLog('✅ הגדרות תוקנו','ok'); break; }
      }
    } else if(lofVal&&txVal&&beaconVal){
      autoLog('כל ההגדרות תקינות!','ok');
    }
  } else {
    autoLog('לא נמצאו הגדרות מתקדמות — בדוק ידנית','warn');
  }

  results.push({label:'Firmware',value:fw.toUpperCase(),ok:fw!=='rev',warn:fw==='unknown'});
  return {fw, ip, cred};
}

async function runMocaFwCheck(){
  autoLogClear();
  const ip=(document.getElementById('auto-moca-ip').value||'').trim();
  if(!ip){ autoLog('הזן IP של מוקה','warn'); return; }
  autoLog(`📡 מתחבר למוקה ${ip}...`,'action');
  let cred=null;
  for(const [u,p] of MOCA_CREDS){
    const r=await proxyFetch(`http://${ip}/`, 'GET', null, [u,p]);
    if(r&&!r.error&&r.status===200){ cred=[u,p]; break; }
  }
  if(!cred){ autoLog('לא ניתן להתחבר','err'); return; }

  const home=await proxyFetch(`http://${ip}/`, 'GET', null, cred);
  const html=home?.text||'';
  const isRev=/REV\b|\b89\b/i.test(html);

  if(!isRev){ autoLog('Firmware אינו REV — אין צורך בעדכון','ok'); return; }
  autoLog('זוהה Firmware REV — מתחיל עדכון ל-VSCO...','warn');
  autoLog('עיין בהוראות: Advanced → Upgrade → העלה קובץ עדכון מוקה.bin','action');
  autoLog('לא ניתן לעדכן Firmware אוטומטית — נדרש העלאת קובץ ידנית','warn');
  autoLog('פתח http://'+ip+' ← Advanced ← Upgrade ← בחר עדכון מוקה.bin','action');
}

/* --- Router scan --- */
async function runRouterScan(){
  autoLogClear();
  const ip=(document.getElementById('auto-router-ip').value||'10.100.102.1').trim();
  autoLog(`🔍 סורק נתב ${ip}...`,'action');
  const creds=[['net013','lz5jbzto'],['Admin','lz5jbzto'],['admin','admin'],['user','user']];
  let cred=null, html='';
  for(const [u,p] of creds){
    const r=await proxyFetch(`http://${ip}/`, 'GET', null, [u,p]);
    if(r&&!r.error&&r.status===200){ cred=[u,p]; html=r.text||''; break; }
  }
  if(!cred){ autoLog(`לא ניתן להתחבר לנתב ${ip}`,'err'); return null; }
  autoLog(`חיבור הצליח (${cred[0]})`,'ok');

  // זהה דגם
  let model='unknown';
  if(/VV5822|VV5820/i.test(html)) model='ADB VV5822/5820';
  else if(/SagemCom|F@ST 5657|5657/i.test(html)) model='SagemCom 5657';
  else if(/SagemCom|F@ST 5670|5670/i.test(html)) model='SagemCom 5670 ULTRA';
  else if(/266V2|MESH.*ULTRA/i.test(html)) model='SagemCom 266V2 MESH';
  else if(/SwissCom|Swisscom/i.test(html)) model='SwissCom';
  else if(/B35/i.test(html)) model='נתב B35';
  autoLog(`דגם נתב: ${model}`,'ok');

  // קרא SSID
  const ssid=extractField(html,[/SSID[^"]*"([^"]{2,32})"/,/ssid[^>]*value="([^"]{2,32})"/i,
    /Network Name[^>]*>([^<]{2,32})</i,/Primary SSID[^>]*value="([^"]+)"/i]);
  if(ssid){ autoLog(`SSID: ${ssid}`,'ok'); window._routerSSID=ssid; }

  const pass=extractField(html,[/WPA.*Passphrase[^>]*value="([^"]+)"/i,
    /password[^>]*value="([^"]{6,})"(?!.*confirm)/i]);
  if(pass){ autoLog(`סיסמה WiFi: ${pass}`,'ok'); window._routerPass=pass; }

  return {model, ssid, pass, cred, ip};
}

/* --- AirTies WiFi sync --- */
async function runAirtiesSync(){
  autoLogClear();
  const aip=(document.getElementById('auto-airties-ip').value||'').trim();
  const rip=(document.getElementById('auto-router-ip').value||'10.100.102.1').trim();

  autoLog('📖 קורא SSID וסיסמה מהנתב...','action');
  const router=await runRouterScan();
  const ssid=router?.ssid||window._routerSSID||null;
  const pass=router?.pass||window._routerPass||null;

  if(!ssid||!pass){
    autoLog('לא נמצאו SSID/סיסמה בנתב — הזן ידנית','warn');
    const ssidM=prompt('הזן שם רשת WiFi (SSID):');
    const passM=prompt('הזן סיסמה:');
    if(!ssidM||!passM){ autoLog('בוטל','warn'); return; }
    window._routerSSID=ssidM; window._routerPass=passM;
  }

  if(!aip){ autoLog('הזן IP של AirTies','warn'); return; }
  autoLog(`📡 מתחבר לאירטייז ${aip}...`,'action');
  const aHome=await proxyFetch(`http://${aip}/main.html`, 'GET');
  if(!aHome||aHome.error||aHome.status!==200){
    autoLog(`לא ניתן להתחבר לאירטייז ${aip}`,'err'); return;
  }
  autoLog('מחובר לאירטייז','ok');

  const finalSSID=window._routerSSID||ssid;
  const finalPass=window._routerPass||pass;

  // נסה API נפוצים של AirTies
  const airtiesApis=[
    {url:`http://${aip}/api/wireless/settings`, json:{ssid:finalSSID, password:finalPass, mode:'ap'}},
    {url:`http://${aip}/goform/wifi_settings`,  data:{ssid:finalSSID, key:finalPass, mode:'ap'}},
    {url:`http://${aip}/cgi-bin/setup`,          data:{ssid:finalSSID, psk:finalPass, wifi_mode:'ap'}},
  ];
  let synced=false;
  for(const api of airtiesApis){
    autoLog(`מנסה ${api.url}...`,'info');
    const r=api.json
      ? await proxyFetch(api.url,'POST',null,null,api.json)
      : await proxyFetch(api.url,'POST',api.data);
    if(r&&!r.error&&r.status<400){ synced=true; autoLog(`✅ SSID/סיסמה עודכנו ב-AirTies`,'ok'); break; }
  }
  if(!synced){
    autoLog('לא הצלחתי לדחוף אוטומטית — פתח AirTies ידנית','warn');
    autoLog(`SSID: ${finalSSID} | סיסמה: ${finalPass}`,'action');
    autoLog(`Quick Setup → Access Point → הזן ← Save`,'action');
  }
}

/* --- Full automation --- */
async function runFullAuto(){
  autoLogClear();
  autoLog('🚀 מתחיל בדיקה מלאה...','action');
  const summary=[];

  // 1. סריקת נתב
  autoLog('─── שלב 1: נתב ───','info');
  const router=await runRouterScan();
  if(router) summary.push({label:'נתב',value:router.model||'?',ok:true});
  else summary.push({label:'נתב',value:'לא מגיב',ok:false});

  // 2. מוקה
  const mip=(document.getElementById('auto-moca-ip').value||'').trim();
  if(mip){
    autoLog('─── שלב 2: מוקה ───','info');
    const moca=await _mocaAutoCore(mip);
    if(moca) summary.push({label:'מוקה Firmware',value:moca.fw.toUpperCase(),ok:moca.fw!=='rev',warn:moca.fw==='unknown'});
    else summary.push({label:'מוקה',value:'לא מגיב',ok:false});
  }

  // 3. AirTies
  const aip=(document.getElementById('auto-airties-ip').value||'').trim();
  if(aip&&router?.ssid){
    autoLog('─── שלב 3: AirTies WiFi sync ───','info');
    await runAirtiesSync();
    summary.push({label:'AirTies WiFi',value:'סונכרן',ok:true});
  }

  autoLog('✅ בדיקה מלאה הסתיימה','ok');
  autoSummary(summary);
}

/* ══ DSL / נחושת functions ══ */

function detectDslCable(){
  const ds=parseFloat(document.getElementById('dsl-ds').value)||0;
  const us=parseFloat(document.getElementById('dsl-us').value)||0;
  const div=document.getElementById('dsl-cable-result');
  if(!ds&&!us){div.innerHTML='';return;}
  let cable='',color='',notes='',ok=true;
  if(ds>=150000){cable='T200';color='#53bdeb';notes='150K-200K Kbps הורדה | 5K-10K Kbps העלאה';}
  else if(ds>=55000){cable='T100';color='#3dba6f';notes='55K-102K Kbps הורדה | 3K-5K Kbps העלאה';}
  else if(ds>=10000){cable='T15';color='#f9c846';notes='~18,938 Kbps הורדה | ~925 Kbps העלאה';}
  else{cable='נמוך מדי';color='#f15c6e';ok=false;notes='סנכרון מתחת ל-10 Mbps — אסור להשלים התקנה';}
  const usOk=us>0;
  div.innerHTML=`
    <div style="background:${ok?'#162d44':'#3d1a1a'};border-radius:10px;padding:12px;direction:rtl">
      <div style="display:flex;align-items:center;gap:10px;margin-bottom:6px">
        <span class="dsl-cable-badge" style="background:${ok?'#0d3a5c':'#5c0d0d'};color:${color}">${cable}</span>
        <span style="font-weight:700;color:${color};font-size:15px">${ok?'✅ תקין':'❌ בעיה'}</span>
      </div>
      <div style="font-size:12px;color:#9bbdd8">${notes}</div>
      ${ds>0?'<div style="font-size:13px;margin-top:6px;color:#eaf4ff">הורדה: <b>'+Math.round(ds/1000)+' Mbps</b> | העלאה: <b>'+Math.round(us/1000)+' Mbps</b></div>':''}
      ${!ok?'<div style="margin-top:8px;font-size:12px;color:#f15c6e">⚠️ לא לבצע התקנה — מהירות מינימום 15 Mbps</div>':''}
    </div>`;
}

function checkDslQuality(){
  const snrDs=parseFloat(document.getElementById('snr-ds').value);
  const snrUs=parseFloat(document.getElementById('snr-us').value);
  const attDs=parseFloat(document.getElementById('att-ds').value);
  const checks=[];
  if(!isNaN(snrDs)){
    const ok=snrDs>=7;
    checks.push({label:'Noise Margin Downstream',val:snrDs+' dB',ok,hint:ok?'':'צריך ≥ 7 dB'});
  }
  if(!isNaN(snrUs)){
    const ok=snrUs>=7;
    checks.push({label:'Noise Margin Upstream',val:snrUs+' dB',ok,hint:ok?'':'צריך ≥ 7 dB'});
  }
  if(!isNaN(attDs)){
    const ok=attDs<=40;
    checks.push({label:'Line Attenuation DS',val:attDs+' dB',ok,hint:ok?'':'ערך גבוה — בעיה בקו'});
  }
  if(!checks.length){
    document.getElementById('dsl-quality-result').innerHTML='<div style="color:#f9c846;font-size:13px;text-align:center">הזן ערכים לבדיקה</div>';
    return;
  }
  let rows='',allOk=true;
  checks.forEach(c=>{
    if(!c.ok) allOk=false;
    rows+=`<div style="display:flex;justify-content:space-between;padding:8px 10px;border-bottom:1px solid #1c3450;font-size:13px;direction:rtl">
      <span style="color:#9bbdd8">${c.label}</span>
      <span style="color:${c.ok?'#3dba6f':'#f15c6e'}">${c.ok?'✅':'❌'} ${c.val} ${c.hint?'('+c.hint+')':''}</span>
    </div>`;
  });
  document.getElementById('dsl-quality-result').innerHTML=
    '<div style="background:#0d1b2a;border-radius:10px;overflow:hidden">'+rows+'</div>'
    +(allOk?'<div style="color:#3dba6f;text-align:center;margin-top:8px;font-size:13px">✅ איכות קו תקינה!</div>':'');
}

const DSL_ERR_THRESHOLDS=[
  {id:'err-fec',idus:'err-fec-us',label:'FEC',thr:400,desc:'שגיאות תיקון אוטומטי'},
  {id:'err-crc',idus:'err-crc-us',label:'CRC',thr:5,desc:'שגיאות שלא תוקנו — חמורות'},
  {id:'err-es',idus:'err-es-us',label:'ES',thr:5,desc:'שניות עם לפחות CRC אחד'},
  {id:'err-ses',idus:'err-ses-us',label:'SES',thr:5,desc:'שניות עם 8 שגיאות CRC'},
  {id:'err-uas',idus:'err-uas-us',label:'UAS',thr:1,desc:'שניות ללא סנכרון'},
  {id:'err-los',idus:'err-los-us',label:'LOS',thr:1,desc:'כמות אבדן סנכרון'},
  {id:'err-lom',idus:'err-lom-us',label:'LOM',thr:1,desc:'שינוי פתאומי ב-SNR'},
];

function checkDslErrors(){
  let rows='',allOk=true,anyEntered=false;
  DSL_ERR_THRESHOLDS.forEach(e=>{
    const dsV=parseFloat(document.getElementById(e.id).value);
    const usV=parseFloat(document.getElementById(e.idus).value);
    [['DS',dsV],['US',usV]].forEach(([dir,val])=>{
      if(isNaN(val)) return;
      anyEntered=true;
      const ok=val<=e.thr;
      if(!ok) allOk=false;
      rows+=`<div style="display:flex;justify-content:space-between;align-items:center;padding:7px 10px;border-bottom:1px solid #1c3450;font-size:12px;direction:rtl">
        <span style="color:#9bbdd8"><b>${e.label}</b> ${dir} <span style="font-size:11px;opacity:0.7">${e.desc}</span></span>
        <span style="color:${ok?'#3dba6f':'#f15c6e'};font-weight:${ok?'400':'700'}">${ok?'✅':'❌'} ${val} <span style="font-size:11px;opacity:0.7">(סף: ${e.thr})</span></span>
      </div>`;
    });
  });
  const div=document.getElementById('dsl-errors-result');
  if(!anyEntered){div.innerHTML='<div style="color:#f9c846;font-size:13px;text-align:center">הזן ערכים מ-Advanced Statistics</div>';return;}
  div.innerHTML='<div style="background:#0d1b2a;border-radius:10px;overflow:hidden">'+rows+'</div>'
    +(allOk?'<div style="color:#3dba6f;text-align:center;margin-top:8px;font-size:13px">✅ כל הערכים תקינים!</div>':'<div style="color:#f15c6e;text-align:center;margin-top:8px;font-size:12px">⚠️ ערכים גבוהים — בדוק תיקון חיבורים, פילטרים, גידים</div>');
}

const NO_SYNC_STEPS=[
  'בדיקת סנכרון בגידים מהקיר לכיוון נקודת חיבור בלבד',
  'ודא תג בזק ע"פ נהלי עבודה',
  'בצע Port Reset באפליקציה: "נתוני קו" / "חבילה תפעולית"',
  'בצע מדידת מהירות T15 באפליקציית טכנאים',
  'ודא אין שגיאות — הכנס ב"נתוני רשת" באפליקציה',
  'בדיקת סנכרון פתיחת תקלת חוסר סנכרון באפליקציה — ודא גישה לתיבה',
  'פנה למוקד בזק: 037278074',
];

function renderNoSyncChecklist(){
  const div=document.getElementById('no-sync-checklist');
  if(!div) return;
  div.innerHTML=NO_SYNC_STEPS.map((s,i)=>`
    <div class="checklist-item" id="nsc-${i}" onclick="toggleNoSync(${i})">
      <input type="checkbox" id="nsc-cb-${i}" onclick="event.stopPropagation();toggleNoSync(${i})">
      <span>${i+1}. ${s}</span>
    </div>`).join('');
}

function toggleNoSync(i){
  const item=document.getElementById('nsc-'+i);
  const cb=document.getElementById('nsc-cb-'+i);
  if(!item||!cb) return;
  cb.checked=!cb.checked;
  item.classList.toggle('done',cb.checked);
}

function resetNoSyncList(){
  NO_SYNC_STEPS.forEach((_,i)=>{
    const cb=document.getElementById('nsc-cb-'+i);
    const item=document.getElementById('nsc-'+i);
    if(cb) cb.checked=false;
    if(item) item.classList.remove('done');
  });
}

/* ══════════════════════════════════════
   💾 ROUTER FIRMWARE DATABASE
   ══════════════════════════════════════
   עדכן גרסאות כאן כשיוצאות גרסאות חדשות
   ══════════════════════════════════════ */
const ROUTER_FW_DB = {
  /* ══ B35 — SagemCom F@ST 5359 (נחושת VDSL) ══ */
  'b35': {
    name: 'נתב B35 (SagemCom 5359)',
    model_patterns: [/5359/i, /B35/i, /F@ST.*5359/i],
    versions: [
      { version: 'v1.28.13', label: 'v1.28.13 ✅ עדכני',     good: true  },
      { version: 'v0.29.00', label: 'v0.29.00 ❌ ישן',        good: false },
      { version: 'v0.24.0',  label: 'v0.24.0  ❌ ישן מאוד',   good: false },
    ],
    bad_versions: ['0.24.0', '0.29.00'],
    fw_url_patterns: ['/info.html', '/status.html', '/'],
    fw_extract: [/5359CC-(v[\d.]+)/i, /[Ff]irmware[^>]*>([\d.v]+)/i, /SW Version[^>]*>([^<]+)</i],
    update_url: '/upgrade.html',
    login: {user: 'net013', pass: 'lz5jbzto'},
    notes: 'SagemCom F@ST 5359 — נחושת VDSL'
  },

  /* ══ AIO — SagemCom F@ST 5657 (FTTH סופרפייבר) ══ */
  'aio': {
    name: 'SagemCom AIO (5657)',
    model_patterns: [/5657/i, /AIO\b(?!.*ULTRA)/i, /F@ST.*5657/i],
    versions: [
      { version: 'sw07.84.64',  label: 'sw07.84.64  ✅ עדכני',    good: true  },
      { version: 'sw07.84.63',  label: 'sw07.84.63  ❌ ישן',       good: false },
      { version: 'sw07.84.57.2',label: 'sw07.84.57.2 ❌ ישן מאוד', good: false },
    ],
    bad_versions: ['07.84.57', '07.84.63'],
    fw_url_patterns: ['/home.html', '/info.html', '/'],
    fw_extract: [/sw(07\.[\d.]+)/i, /[Ff]irmware Version[^>]*>([^<]+)</i, /SW[^>]*>([^<]+)</i],
    update_url: '/upgrade.html',
    login: {user: 'net013', pass: 'lz5jbzto'},
    notes: 'SagemCom F@ST 5657 — FTTH סופרפייבר'
  },

  /* ══ AIO ULTRA — SagemCom F@ST 5670 (WiFi 6) ══ */
  'aio_ultra': {
    name: 'SagemCom AIO ULTRA (5670)',
    model_patterns: [/5670/i, /AIO.*ULTRA/i, /F@ST.*5670/i],
    versions: [
      { version: 'v0.62.0', label: 'v0.62.0 ✅ עדכני',  good: true  },
      { version: 'v0.52.0', label: 'v0.52.0 ❌ ישן',     good: false },
      { version: 'v0.37.0', label: 'v0.37.0 ❌ ישן',     good: false },
      { version: 'v0.34.0', label: 'v0.34.0 ❌ ישן מאוד',good: false },
    ],
    bad_versions: ['0.16.0', '0.19.0', '0.34.0', '0.37.0', '0.52.0'],
    fw_url_patterns: ['/home.html', '/info.html', '/'],
    fw_extract: [/5670CC-(v[\d.]+)/i, /[Ff]irmware Version[^>]*>([^<]+)</i],
    update_url: '/upgrade.html',
    login: {user: 'net013', pass: 'lz5jbzto'},
    notes: 'SagemCom 5670 AIO ULTRA — WiFi 6, עד 5.2Gbps'
  },

  /* ══ STAR PLUS — SagemCom 5674 ══ */
  'star_plus': {
    name: 'STAR PLUS (SagemCom 5674)',
    model_patterns: [/5674/i, /STAR.?PLUS(?!.*HEIT)/i, /StarPlus(?!.*Heit)/i],
    versions: [
      { version: 'V1037.0.27', label: 'V1037.0.27 ✅ עדכני', good: true  },
      { version: 'V1037.0.14', label: 'V1037.0.14 ❌ ישן',    good: false },
    ],
    bad_versions: ['1037.0.14'],
    fw_url_patterns: ['/home.html', '/info.html', '/'],
    fw_extract: [/(V\d{4}\.\d+\.\d+)/i, /[Ff]irmware[^>]*>([^<]+)</i],
    update_url: '/upgrade.html',
    login: {user: 'net013', pass: 'lz5jbzto'},
    notes: 'SagemCom 5674 STAR PLUS'
  },

  /* ══ STAR PLUS HEITS — SagemCom 5698 (StarPro) ══ */
  'star_plus_heits': {
    name: 'STAR PLUS HEITS (SagemCom 5698)',
    model_patterns: [/5698/i, /STAR.?PLUS.*HEIT/i, /HEITS/i, /StarPro/i],
    versions: [
      { version: 'V1033.0', label: 'V1033.0 ✅ עדכני', good: true },
    ],
    bad_versions: [],
    fw_url_patterns: ['/home.html', '/info.html', '/'],
    fw_extract: [/(V\d{4}\.\d+)/i, /[Ff]irmware[^>]*>([^<]+)</i],
    update_url: '/upgrade.html',
    login: {user: 'net013', pass: 'lz5jbzto'},
    notes: 'SagemCom 5698 STAR PLUS HEITS / StarPro'
  },

  /* ══ ADB VVS822/5823 ══ */
  'adb_vvs822': {
    name: 'ADB VVS822/5823',
    model_patterns: [/VV5822|VVS822|VV5823|VVS823/i],
    versions: [
      { version: '7.5.0.0039', label: '7.5.0.0039 ❌ Hard Reset נדרש', good: false },
    ],
    bad_versions: ['0039'],
    fw_url_patterns: ['/home.html', '/'],
    fw_extract: [/[Ff]irmware Version[^:]*:\s*([^\s<]+)/i, /VV58\d+_CLC_([^\s<"]+)/i],
    update_url: '/upgrade.html',
    login: {user: 'Admin', pass: 'lz5jbzto'},
    notes: 'ADB VVS822/5823 — גרסה 0039 דורשת Hard Reset'
  },

  /* ══ SwissCom ST6840 ══ */
  'swisscom': {
    name: 'SwissCom (ST6840)',
    model_patterns: [/ST6840/i, /SwissCom/i, /Swiss/i],
    versions: [
      { version: '7.8.1.0009', label: '7.8.1.0009 ✅ עדכני', good: true },
    ],
    bad_versions: [],
    fw_url_patterns: ['/home.html', '/info.html', '/'],
    fw_extract: [/ST6840_CLC_([\d.]+)/i, /[Ff]irmware[^>]*>([^<]+)</i],
    update_url: '/upgrade.html',
    login: {user: 'net013', pass: 'lz5jbzto'},
    notes: 'SwissCom ST6840 — נחושת'
  },

  /* ══ AirTies Air4920IL — יחידת MESH ══ */
  'air4920': {
    name: 'AirTies Air4920IL',
    model_patterns: [/Air4920/i, /4920/i],
    versions: [
      { version: '3.112.14.4.1461', label: '3.112.14.4.1461 ✅ עדכני', good: true },
    ],
    bad_versions: [],
    fw_url_patterns: ['/api/v1/system/info', '/cgi-bin/status.cgi', '/status.html', '/'],
    fw_extract: [/"firmware[_\-]?version"\s*:\s*"([^"]+)"/i, /[Ff]irmware[^>]*>([\d.]+)/i, /([\d]+\.[\d]+\.[\d]+\.[\d]+\.[\d]+)/],
    update_url: '/api/v1/system/upgrade',
    login: {user: '', pass: ''},
    notes: 'AirTies Air4920IL — יחידת Super Wi-Fi MESH'
  },

  /* ══ AirTies Air4930IL — יחידת MESH ══ */
  'air4930': {
    name: 'AirTies Air4930IL',
    model_patterns: [/Air4930/i, /4930/i],
    versions: [
      { version: '2.112.14.4.528', label: '2.112.14.4.528 ✅ עדכני', good: true },
    ],
    bad_versions: [],
    fw_url_patterns: ['/api/v1/system/info', '/cgi-bin/status.cgi', '/status.html', '/'],
    fw_extract: [/"firmware[_\-]?version"\s*:\s*"([^"]+)"/i, /[Ff]irmware[^>]*>([\d.]+)/i, /([\d]+\.[\d]+\.[\d]+\.[\d]+\.[\d]+)/],
    update_url: '/api/v1/system/upgrade',
    login: {user: '', pass: ''},
    notes: 'AirTies Air4930IL — יחידת Super Wi-Fi MESH'
  }
};

/* בדיקת firmware לנתב */
async function checkRouterFirmware(ip, modelKey) {
  const db = ROUTER_FW_DB[modelKey];
  if (!db) { autoLog(`דגם לא מוכר: ${modelKey}`, 'err'); return null; }

  autoLog(`🔍 בודק firmware נתב ${db.name} ב-${ip}...`, 'action');
  const cred = [db.login.user, db.login.pass];

  let html = '';
  for (const urlPath of db.fw_url_patterns) {
    const r = await proxyFetch(`http://${ip}${urlPath}`, 'GET', null, cred);
    if (r && !r.error && r.status === 200 && r.text) { html = r.text; break; }
  }
  if (!html) {
    // נסה ללא auth
    const r = await proxyFetch(`http://${ip}/`, 'GET');
    if (r && !r.error && r.status === 200) html = r.text || '';
  }
  if (!html) { autoLog(`לא ניתן לקרוא נתב ${ip}`, 'err'); return null; }

  // חלץ גרסה
  let fwVer = null;
  for (const pat of db.fw_extract) {
    const m = html.match(pat);
    if (m) { fwVer = (m[1] || m[0]).trim(); break; }
  }

  // זיהוי אוטומטי מה-HTML
  if (!fwVer) {
    const autoMatch = html.match(/[Ff]irmware[^<]{0,50}([\d.]+[-\w]*)/);
    if (autoMatch) fwVer = autoMatch[1];
  }

  if (!fwVer) { autoLog('לא זוהתה גרסת firmware בדף', 'warn'); return null; }
  autoLog(`Firmware נוכחי: ${fwVer}`, 'info');

  // בדוק אם גרסה ישנה
  const isBad = db.bad_versions.some(bv => fwVer.includes(bv));
  const goodVersions = db.versions.filter(v => v.good);
  const latestGood = goodVersions[goodVersions.length - 1];

  if (isBad) {
    autoLog(`❌ גרסה ישנה! ${db.notes}`, 'err');
    if (db.name.includes('ADB') || db.name.includes('VVS')) {
      autoLog('פתרון: Hard Reset → המתן 5 דקות לגרסה רולינגית', 'action');
    } else {
      autoLog(`פתרון: נווט ל http://${ip}${db.update_url} לעדכון firmware`, 'action');
    }
    return { fwVer, status: 'bad', model: db.name };
  } else {
    autoLog(`✅ Firmware תקין (${fwVer})`, 'ok');
    return { fwVer, status: 'ok', model: db.name };
  }
}

/* ══ Router Firmware — UI helpers ══ */
function renderFwVersionsTable(){
  const div=document.getElementById('fw-versions-table');
  if(!div) return;
  let h='<div style="background:#0d1b2a;border-radius:10px;overflow:hidden">';
  Object.entries(ROUTER_FW_DB).forEach(([k,db])=>{
    const goodVers=db.versions.filter(v=>v.good).map(v=>v.version||'—');
    const goodStr=goodVers.length?goodVers.join(', '):'לא הוגדר';
    const badStr=db.bad_versions.length?db.bad_versions.join(', '):'—';
    const hasBad=db.bad_versions.length>0;
    h+=`<div style="padding:10px 12px;border-bottom:1px solid #1c3450;direction:rtl">
      <div style="display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:4px">
        <div>
          <div style="color:#9b7fff;font-weight:700;font-size:13px">${db.name}</div>
          <div style="color:#6b94b8;font-size:11px;margin-top:2px">${db.notes}</div>
        </div>
        <div style="text-align:left;min-width:120px">
          <div style="color:#3dba6f;font-size:12px">✅ תקין: ${goodStr}</div>
          ${hasBad?`<div style="color:#f15c6e;font-size:11px;margin-top:2px">❌ דרוש עדכון אם: ${badStr}</div>`:''}
        </div>
      </div>
    </div>`;
  });
  h+='</div>';
  div.innerHTML=h;
}

async function runRouterFwCheckSingle(){
  const ip=(document.getElementById('fw-router-ip').value||'10.100.102.1').trim();
  const modelKey=document.getElementById('fw-router-model').value;
  const resultDiv=document.getElementById('fw-single-result');
  resultDiv.innerHTML='<div style="color:#9bbdd8;font-size:13px;direction:rtl">⏳ בודק...</div>';
  autoLogClear();
  const result=await checkRouterFirmware(ip,modelKey);
  if(!result){
    resultDiv.innerHTML='<div style="background:#3d0d0d;border-radius:10px;padding:10px 12px;color:#f15c6e;font-size:13px;direction:rtl">❌ לא ניתן להתחבר לנתב — בדוק IP</div>';
    return;
  }
  const ok=result.status==='ok';
  resultDiv.innerHTML=`<div style="background:${ok?'#0d3a22':'#3d0d0d'};border-radius:10px;padding:12px;direction:rtl">
    <div style="font-size:16px;font-weight:700;color:${ok?'#3dba6f':'#f15c6e'}">${ok?'✅':'❌'} ${result.model}</div>
    <div style="font-size:13px;color:#eaf4ff;margin-top:6px">גרסה נוכחית: <b style="color:#f9c846">${result.fwVer}</b></div>
    <div style="font-size:12px;color:${ok?'#9bbdd8':'#f9c846'};margin-top:4px">${ok?'Firmware תקין ✅':'⚠️ גרסה ישנה — דרוש עדכון!'}</div>
  </div>`;
}

async function runRouterFwScan(){
  const ip=(document.getElementById('fw-scan-ip').value||'10.100.102.1').trim();
  const resultDiv=document.getElementById('fw-scan-result');
  resultDiv.innerHTML='<div style="color:#9bbdd8;font-size:13px;direction:rtl">⏳ סורק...</div>';
  autoLogClear();
  autoLog(`🔍 סורק נתב ${ip} — בודק כל הדגמים...`,'action');

  // נסה לזהות דגם אוטומטית
  let html='';
  const creds=[['net013','lz5jbzto'],['Admin','lz5jbzto'],['admin','admin'],['user','user']];
  const paths=['/','/info.html','/home.html','/status.html'];
  for(const p of paths){
    for(const [u,pw] of creds){
      const r=await proxyFetch(`http://${ip}${p}`,'GET',null,[u,pw]);
      if(r&&!r.error&&r.status===200&&r.text){html=r.text;break;}
    }
    if(html) break;
  }

  if(!html){
    resultDiv.innerHTML='<div style="background:#3d0d0d;border-radius:10px;padding:10px 12px;color:#f15c6e;font-size:13px;direction:rtl">❌ לא ניתן להתחבר — בדוק IP</div>';
    return;
  }

  // זהה דגם
  let detectedKey=null;
  for(const [key,db] of Object.entries(ROUTER_FW_DB)){
    if(db.model_patterns.some(p=>p.test(html))){detectedKey=key;break;}
  }

  if(!detectedKey){
    // נסה לחלץ גרסה גנרית
    const m=html.match(/[Ff]irmware[^<"]{0,50}([\d]+\.[\d.]+[-\w]*)/);
    const ver=m?m[1]:'לא זוהה';
    resultDiv.innerHTML=`<div style="background:#162d44;border-radius:10px;padding:12px;direction:rtl">
      <div style="color:#f9c846;font-weight:700;font-size:14px">⚠️ דגם לא זוהה אוטומטית</div>
      <div style="color:#eaf4ff;font-size:13px;margin-top:6px">גרסה שנמצאה: <b>${ver}</b></div>
      <div style="color:#9bbdd8;font-size:12px;margin-top:4px">בחר דגם ידנית ב"בדוק Firmware — אוטומטי"</div>
    </div>`;
    return;
  }

  const result=await checkRouterFirmware(ip,detectedKey);
  if(!result){
    resultDiv.innerHTML='<div style="background:#3d0d0d;border-radius:10px;padding:10px;color:#f15c6e;font-size:13px;direction:rtl">❌ לא הצלחתי לקרוא גרסה</div>';
    return;
  }
  const ok=result.status==='ok';
  resultDiv.innerHTML=`<div style="background:${ok?'#0d3a22':'#3d0d0d'};border-radius:10px;padding:12px;direction:rtl">
    <div style="font-size:16px;font-weight:700;color:${ok?'#3dba6f':'#f15c6e'}">${ok?'✅':'❌'} ${result.model}</div>
    <div style="font-size:13px;color:#eaf4ff;margin-top:6px">גרסה נוכחית: <b style="color:#f9c846">${result.fwVer}</b></div>
    <div style="font-size:12px;color:${ok?'#9bbdd8':'#f9c846'};margin-top:4px">${ok?'Firmware תקין ✅':'⚠️ גרסה ישנה — דרוש עדכון!'}</div>
  </div>`;
}

/* ══ AirTies Firmware Check ══ */
async function checkAirtiesFw(){
  const ip=(document.getElementById('airties-fw-ip').value||'').trim();
  const modelKey=document.getElementById('airties-fw-model').value;
  const resultDiv=document.getElementById('airties-fw-result');
  if(!ip){
    resultDiv.innerHTML='<div style="background:#3d2a0a;border-radius:8px;padding:8px 12px;color:#f9c846;font-size:13px;direction:rtl">⚠️ הכנס IP של יחידת AirTies</div>';
    return;
  }
  resultDiv.innerHTML='<div style="color:#9bbdd8;font-size:13px;direction:rtl">⏳ מתחבר ל-AirTies...</div>';
  autoLogClear();
  autoLog(`📡 בודק AirTies ${ip}...`, 'action');

  const db = ROUTER_FW_DB[modelKey];
  if(!db){ resultDiv.innerHTML='<div style="color:#f15c6e;font-size:13px">דגם לא מוכר</div>'; return; }

  let html='', foundVer=null;
  // נסה endpoints שונים
  for(const path of db.fw_url_patterns){
    const r=await proxyFetch(`http://${ip}${path}`,'GET');
    if(r&&!r.error&&r.status===200&&r.text){ html=r.text; break; }
  }
  if(!html){
    // נסה JSON API
    const r=await proxyFetch(`http://${ip}/api/v1/system/info`,'GET');
    if(r&&!r.error&&r.status===200&&r.text) html=r.text;
  }

  if(html){
    for(const pat of db.fw_extract){
      const m=html.match(pat);
      if(m){ foundVer=(m[1]||m[0]).trim(); break; }
    }
    // חיפוש גנרי אם לא נמצא
    if(!foundVer){
      const m=html.match(/(\d+\.\d+\.\d+\.\d+\.\d+)/);
      if(m) foundVer=m[1];
    }
  }

  const latestVer=db.versions.find(v=>v.good)?.version||'?';
  const isOk = foundVer && foundVer===latestVer;
  const isUnknown = !foundVer;

  if(isUnknown){
    autoLog('לא זוהתה גרסה — ייתכן שהיחידה לא מגיבה', 'warn');
    resultDiv.innerHTML=`<div style="background:#162d44;border-radius:10px;padding:12px;direction:rtl">
      <div style="font-weight:700;color:#f9c846;font-size:14px">⚠️ לא ניתן לקרוא גרסה</div>
      <div style="font-size:12px;color:#9bbdd8;margin-top:4px">ודא שה-IP נכון ושהיחידה מחוברת</div>
      <div style="font-size:12px;color:#53bdeb;margin-top:4px">גרסה עדכנית: <b>${latestVer}</b></div>
    </div>`;
    return;
  }

  autoLog(`Firmware AirTies: ${foundVer}`, isOk?'ok':'warn');
  if(!isOk) autoLog(`גרסה עדכנית: ${latestVer} — שקול עדכון`, 'action');

  resultDiv.innerHTML=`<div style="background:${isOk?'#0d3a22':'#3d2a0a'};border-radius:10px;padding:12px;direction:rtl">
    <div style="font-size:16px;font-weight:700;color:${isOk?'#3dba6f':'#f9c846'}">${isOk?'✅':'⚠️'} ${db.name}</div>
    <div style="font-size:13px;color:#eaf4ff;margin-top:6px">גרסה נוכחית: <b style="color:#f9c846">${foundVer}</b></div>
    <div style="font-size:12px;color:${isOk?'#9bbdd8':'#f9c846'};margin-top:4px">
      ${isOk?'Firmware עדכני ✅':`⚠️ גרסה עדכנית: ${latestVer}`}
    </div>
  </div>`;
}

/* ══ WiFi / MESH functions ══ */
function checkMeshSignal(){
  const val=parseFloat(document.getElementById('mesh-signal').value);
  const div=document.getElementById('mesh-signal-result');
  if(isNaN(val)||val<0){div.innerHTML='';return;}
  const ok=val>=55;
  div.innerHTML=`<div style="background:${ok?'#1a3d22':'#3d1a1a'};border-radius:10px;padding:12px;direction:rtl;text-align:center">
    <div style="font-size:28px;font-weight:700;color:${ok?'#3dba6f':'#f15c6e'}">${val}%</div>
    <div style="font-size:14px;color:${ok?'#3dba6f':'#f15c6e'};font-weight:600;margin-top:4px">
      ${ok?'✅ Signal Level תקין — מיקום מצוין':'❌ Signal Level נמוך — קרב ליחידה לנתב'}
    </div>
    <div style="background:${ok?'#0d3a22':'#3d0d0d'};border-radius:8px;padding:6px 10px;margin-top:8px;font-size:12px;color:${ok?'#9bbdd8':'#f9c846'}">
      ${ok?'המשך התקנה — מיקום AirTies תקין':'הזז את יחידת ה-Super Wi-Fi קרוב יותר לנתב'}
    </div>
  </div>`;
}

function toggleAdbGuide(){
  const g=document.getElementById('adb-guide');
  g.style.display=g.style.display==='none'?'block':'none';
}

function showWifiMode(mode){
  const wps=document.getElementById('wifi-wps-guide');
  const ap=document.getElementById('wifi-ap-guide');
  const btnWps=document.getElementById('wifi-mode-wps');
  const btnAp=document.getElementById('wifi-mode-ap');
  if(!wps||!ap||!btnWps||!btnAp) return;
  if(mode==='wps'){
    wps.style.display='block'; ap.style.display='none';
    btnWps.style.borderColor='#3dba6f'; btnWps.style.background='#1a3d22'; btnWps.style.color='#3dba6f'; btnWps.style.fontWeight='700';
    btnAp.style.borderColor='#284461'; btnAp.style.background='#162d44'; btnAp.style.color='#6b94b8'; btnAp.style.fontWeight='400';
  } else {
    ap.style.display='block'; wps.style.display='none';
    btnAp.style.borderColor='#53bdeb'; btnAp.style.background='#1a3d4a'; btnAp.style.color='#53bdeb'; btnAp.style.fontWeight='700';
    btnWps.style.borderColor='#284461'; btnWps.style.background='#162d44'; btnWps.style.color='#6b94b8'; btnWps.style.fontWeight='400';
  }
}

function toggleWpsMultiUnit(){
  const g=document.getElementById('wps-multi-unit');
  if(!g) return;
  g.style.display=g.style.display==='none'?'block':'none';
}

function toggleMeshTab(name){
  const g=document.getElementById('mesh-tab-'+name);
  if(!g) return;
  g.style.display=g.style.display==='none'?'block':'none';
}

function toggleMeshPairGuide(){
  const g=document.getElementById('mesh-pair-guide');
  if(!g) return;
  g.style.display=g.style.display==='none'?'block':'none';
}

/* ADMIN_PHONE_JS placeholder — filled server-side */

/* ══════════════════════════════════════════════
   לשונית בדיקות — MOCA / Firmware / Settings
   ══════════════════════════════════════════════ */
let _mocaNodes=[];
let _logoOk=null;
let _cardStore={};          // call_id → card data
let _diagCard=null;         // פקעה פעילה בלשונית בדיקות

/* זיהוי דגם נתב מנתוני פקעה */
function detectRouterFromCard(c){
  const txt=[(c.existing||''),(c.planned||''),(c.infra||''),(c.technology||'')].join(' ');
  if(/5670|AIO.?ULTRA/i.test(txt))          return 'aio_ultra';
  if(/5657|AIO\b/i.test(txt))               return 'aio';
  if(/5698|STAR.?PRO|HEITS/i.test(txt))     return 'star_plus_heits';
  if(/5674|STAR.?PLUS|STAR\+/i.test(txt))   return 'star_plus';
  if(/5359|B35/i.test(txt))                 return 'b35';
  if(/VV58|VVS|ADB/i.test(txt))             return 'adb_vvs822';
  if(/ST6840|SwissCom/i.test(txt))          return 'swisscom';
  if(/Air4930|4930/i.test(txt))             return 'air4930';
  if(/Air4920|4920/i.test(txt))             return 'air4920';
  // ברירת מחדל לפי תשתית
  if(/סיב|Fiber|FTTH|FB|NV|IBC/i.test(txt)) return 'aio';
  if(/נחושת|Copper|VDSL|BN/i.test(txt))     return 'b35';
  return null;
}

/* פתיחת לשונית בדיקות לפקעה ספציפית */
function openDiagForCard(callId){
  _diagCard=_cardStore[callId]||null;
  switchTab('diag');
}

/* סנכרון שדות הבדיקות מנתוני הפקעה */
function syncDiagFromCard(c){
  if(!c) return;
  const routerKey=detectRouterFromCard(c);
  const routerIP='10.100.102.1';

  // מלא כל שדות IP
  ['auto-router-ip','fw-router-ip','fw-scan-ip'].forEach(id=>{
    const el=document.getElementById(id);
    if(el) el.value=routerIP;
  });

  // בחר דגם אם זוהה
  if(routerKey){
    const sel=document.getElementById('fw-router-model');
    if(sel) sel.value=routerKey;
  }

  // בניית באנר
  const banner=document.getElementById('diag-card-banner');
  if(!banner) return;
  const db=routerKey?ROUTER_FW_DB[routerKey]:null;
  const creds=db?`<div style="background:#0d1b2a;border-radius:8px;padding:8px 10px;margin-top:8px;font-size:12px;direction:rtl;line-height:1.9">
    <span style="color:#9b7fff;font-weight:600">כניסה לנתב: </span>
    <span style="color:#eaf4ff">${db.login.user} / <b>${db.login.pass}</b></span>
  </div>`:'';
  banner.style.display='block';
  banner.innerHTML=`
  <div style="background:linear-gradient(135deg,#0d2a1a,#0d1b2a);border:1.5px solid #3dba6f;border-radius:14px;padding:14px;margin-bottom:14px;direction:rtl">
    <div style="font-size:15px;font-weight:700;color:#3dba6f;margin-bottom:6px">📋 פקעה ${c.call_id||''}</div>
    <div style="font-size:13px;color:#eaf4ff">👤 ${c.name||''}</div>
    <div style="font-size:13px;color:#9bbdd8">📍 ${c.address||''}</div>
    ${c.infra?`<div style="font-size:12px;color:#53bdeb;margin-top:2px">🌐 ${c.infra}${c.technology?' · '+c.technology:''}</div>`:''}
    ${db?`<div style="font-size:13px;color:#f9c846;margin-top:6px">🔧 נתב: <b>${db.name}</b></div>`:'<div style="font-size:12px;color:#f9c846;margin-top:4px">⚠️ דגם נתב לא זוהה — בחר ידנית</div>'}
    ${c.existing?`<div style="font-size:12px;color:#6b94b8">ציוד קיים: ${c.existing}</div>`:''}
    ${creds}
    <div style="display:flex;gap:8px;margin-top:10px">
      <a href="http://${routerIP}" target="_blank"
        style="flex:1;background:#1a3d22;border:1px solid #3dba6f;color:#3dba6f;border-radius:8px;padding:9px 6px;font-size:13px;font-weight:700;text-align:center;text-decoration:none">
        🔗 פתח נתב
      </a>
      <button onclick="runRouterFwCheckSingle()"
        style="flex:1;background:#3a1a5c;border:1px solid #9b7fff;color:#eaf4ff;border-radius:8px;padding:9px 6px;font-size:13px;font-weight:700;cursor:pointer">
        💾 בדוק Firmware
      </button>
      <button onclick="runRouterScan()"
        style="flex:1;background:#1a2a4a;border:1px solid #53bdeb;color:#53bdeb;border-radius:8px;padding:9px 6px;font-size:13px;font-weight:700;cursor:pointer">
        🔍 סרוק
      </button>
    </div>
    <button onclick="showAllDiagSections()" style="margin-top:8px;width:100%;background:transparent;border:1px solid #284461;color:#6b94b8;border-radius:8px;padding:7px;font-size:12px;cursor:pointer;direction:rtl">
      👁 הצג את כל הסקציות
    </button>
  </div>`;

  // סנן סקציות לפי ציוד הפקעה
  filterDiagSections(detectDiagCategories(c));
}

/* זיהוי קטגוריות רלוונטיות מנתוני הפקעה */
function detectDiagCategories(c){
  const cats=new Set();
  const txt=[(c.existing||''),(c.planned||''),(c.infra||''),(c.technology||'')].join(' ');

  // MOCA — תמיד רלוונטי (כמעט לכל התקנה)
  cats.add('moca');

  // DSL / נחושת
  if(/נחושת|Copper|VDSL|DSL|BN|B35|5359/i.test(txt)) cats.add('dsl');

  // WiFi / AirTies
  if(/AirTies|Air49|4920|4930|MESH|WiFi|Wifi|אחיד|אקסס/i.test(txt)) cats.add('wifi');

  // סיב / Fiber
  if(/סיב|Fiber|FTTH|AIO|5657|5670|5698|STAR|FB|NV|IBC/i.test(txt)) cats.add('fiber');

  // Firmware — תמיד רלוונטי
  cats.add('fw');

  return cats;
}

/* הסתר/הצג סקציות לפי קבוצה */
function filterDiagSections(cats){
  const map={
    'diag-sec-moca': cats.has('moca'),
    'diag-sec-dsl':  cats.has('dsl'),
    'diag-sec-wifi': cats.has('wifi'),
    'diag-sec-fiber':cats.has('fiber'),
    'diag-sec-fw':   cats.has('fw')
  };
  Object.entries(map).forEach(([id,show])=>{
    const el=document.getElementById(id);
    if(el) el.style.display=show?'':'none';
  });
}

/* הצג את כל הסקציות (כשאין פקעה פעילה) */
function showAllDiagSections(){
  ['diag-sec-moca','diag-sec-dsl','diag-sec-wifi','diag-sec-fiber','diag-sec-fw'].forEach(id=>{
    const el=document.getElementById(id);
    if(el) el.style.display='';
  });
}

function loadDiagPanel(){
  const panel=document.getElementById('diag-panel');
  panel.dataset.loaded='1';
  panel.innerHTML=`
<div id="diag-card-banner" style="display:none"></div>
<div style="padding:4px 0 10px;font-size:18px;font-weight:700;color:#eaf4ff">🔧 כלי בדיקה לשטח</div>

<!-- ══ אוטומציה — בדיקה ותיקון אוטומטי ══ -->
<div style="background:linear-gradient(135deg,#1a3d22,#0d2a4a);border-radius:14px;padding:16px;margin-bottom:14px;border:1px solid #3dba6f;direction:rtl">
  <div style="font-size:16px;font-weight:700;color:#3dba6f;margin-bottom:6px">🤖 אוטומציה — בדיקה ותיקון</div>
  <div style="font-size:12px;color:#9bbdd8;margin-bottom:14px">מתחבר אוטומטית למכשירים ← בודק ← מתקן ← מדווח</div>

  <!-- הגדרת IP לסריקה -->
  <div style="display:grid;gap:8px;margin-bottom:12px">
    <div class="setting-row">
      <div class="setting-label" style="color:#3dba6f">IP נתב</div>
      <input id="auto-router-ip" value="10.100.102.1" class="setting-input" style="max-width:160px">
    </div>
    <div class="setting-row">
      <div class="setting-label" style="color:#53bdeb">IP מוקה</div>
      <input id="auto-moca-ip" placeholder="לדוג' 10.100.102.5" class="setting-input" style="max-width:160px">
    </div>
    <div class="setting-row">
      <div class="setting-label" style="color:#f9c846">IP AirTies</div>
      <input id="auto-airties-ip" placeholder="לדוג' 10.100.102.7" class="setting-input" style="max-width:160px">
    </div>
  </div>

  <!-- כפתורי אוטומציה -->
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:10px">
    <button onclick="runMocaAuto()" style="background:#1a3d22;border:1px solid #3dba6f;color:#3dba6f;border-radius:10px;padding:10px 6px;font-size:13px;font-weight:700;cursor:pointer">
      🔧 תקן מוקה
    </button>
    <button onclick="runAirtiesSync()" style="background:#1a3d4a;border:1px solid #53bdeb;color:#53bdeb;border-radius:10px;padding:10px 6px;font-size:13px;font-weight:700;cursor:pointer">
      📡 סנכרן AirTies
    </button>
    <button onclick="runMocaFwCheck()" style="background:#3d2a0a;border:1px solid #f9c846;color:#f9c846;border-radius:10px;padding:10px 6px;font-size:13px;font-weight:700;cursor:pointer">
      ⬆️ עדכן Firmware REV
    </button>
    <button onclick="runRouterScan()" style="background:#1c2a3d;border:1px solid #9b7fff;color:#9b7fff;border-radius:10px;padding:10px 6px;font-size:13px;font-weight:700;cursor:pointer">
      🔍 סרוק נתב
    </button>
  </div>
  <button onclick="runFullAuto()" style="width:100%;background:linear-gradient(135deg,#1a5c3a,#1a3d5c);border:2px solid #3dba6f;color:#eaf4ff;border-radius:12px;padding:13px;font-size:15px;font-weight:700;cursor:pointer">
    ⚡ הרץ בדיקה מלאה — תקן הכל
  </button>

  <!-- לוג אוטומציה -->
  <div id="auto-log" style="margin-top:12px;background:#050e1a;border-radius:10px;padding:10px;min-height:60px;max-height:260px;overflow-y:auto;font-family:monospace;font-size:12px;direction:ltr"></div>
  <!-- סיכום -->
  <div id="auto-summary" style="margin-top:8px"></div>
</div>

<div id="diag-sec-moca">
<!-- גישה לנתב ומוקה -->
<div class="diag-section">
  <div class="diag-title">🌐 גישה לנתב / מוקה</div>
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:8px">
    <a href="http://10.100.102.1" target="_blank" class="diag-btn-primary">
      🔗 פתח נתב<div style="font-size:11px;opacity:0.8;margin-top:2px">10.100.102.1</div>
    </a>
    <div style="background:#0d1b2a;border-radius:10px;padding:10px;font-size:12px;direction:rtl;line-height:1.9">
      <div style="color:#53bdeb;font-weight:600">נתב:</div>
      <div style="color:#eaf4ff">User-home / <b>Home013</b></div>
      <div style="color:#eaf4ff">Admin / <b>lz5jbzto</b></div>
      <div style="color:#53bdeb;font-weight:600;margin-top:4px">MOCA:</div>
      <div style="color:#eaf4ff">admin / <b>maxlinear</b></div>
      <div style="color:#eaf4ff">admin / <b>entropic</b></div>
    </div>
  </div>
</div>

<!-- גרסת Firmware -->
<div class="diag-section">
  <div class="diag-title">💾 גרסת Firmware MOCA</div>
  <div style="display:grid;gap:8px;margin:10px 0">
    <div class="fw-row fw-bad">
      <span class="fw-badge fw-rev">REV (89)</span>
      <span style="color:#f07080;font-weight:600">❌ דרוש עדכון</span>
      <button onclick="toggleUpdateGuide()" class="fw-update-btn">📖 הוראות</button>
    </div>
    <div class="fw-row fw-good">
      <span class="fw-badge fw-vsco">VSCO (94)</span>
      <span style="color:#3dba6f;font-weight:600">✅ תקין</span>
    </div>
    <div class="fw-row fw-good">
      <span class="fw-badge fw-mmxl">MMXL (2.12.8)</span>
      <span style="color:#3dba6f;font-weight:600">✅ תקין</span>
    </div>
  </div>
  <div id="update-guide" style="display:none;margin-top:8px">
    <div style="background:#0d1b2a;border-radius:10px;padding:14px;font-size:13px;line-height:2">
      <div style="font-weight:700;color:#f9c846;margin-bottom:8px">📋 שדרוג REV → VSCO</div>
      <div class="step-item" style="margin-bottom:5px">1️⃣ התחבר לכתובת ה-IP של המוקה</div>
      <div class="step-item" style="margin-bottom:5px">2️⃣ כניסה: <b>admin / maxlinear</b> או <b>admin / entropic</b></div>
      <div class="step-item" style="margin-bottom:5px">3️⃣ עבור ל: <b>Advanced ← Upgrade</b></div>
      <div class="step-item" style="margin-bottom:5px">4️⃣ בחר קובץ: <b>עדכון מוקה.bin</b></div>
      <div class="step-item" style="margin-bottom:5px">5️⃣ לחץ <b>Upload</b> — המתן ~3 דקות</div>
      <div class="step-item">6️⃣ המוקה יתאפס — ודא שגרסה עודכנה ל-VSCO</div>
      <div style="background:#3d2a0a;border-radius:8px;padding:8px 12px;margin-top:10px;color:#f9c846;font-size:12px">
        ⚠️ אל תנתק חשמל במהלך העדכון!
      </div>
    </div>
  </div>
</div>

<!-- בדיקת PHY Rates MOCA -->
<div class="diag-section">
  <div class="diag-title">📡 PHY Rates MOCA</div>
  <div style="font-size:12px;color:#8696a0;margin:4px 0 10px;line-height:1.7">
    VSCO ↔ VSCO: <b style="color:#3dba6f">&gt; 600 Mbps ✅</b><br>
    MMXL ↔ MMXL: <b style="color:#3dba6f">&gt; 1300 Mbps ✅</b><br>
    VSCO ↔ MMXL: <b style="color:#f9c846">~600 Mbps ⚠️ רשת מעורבת</b>
  </div>
  <div style="display:flex;gap:8px;margin-bottom:10px;align-items:center;flex-wrap:wrap">
    <input id="moca-node-name" placeholder="שם / IP"
      style="flex:1;min-width:90px;background:#0d1b2a;border:1px solid #284461;border-radius:8px;padding:8px;color:#eaf4ff;font-size:13px;direction:rtl">
    <select id="moca-node-fw"
      style="background:#0d1b2a;border:1px solid #284461;border-radius:8px;padding:8px;color:#eaf4ff;font-size:13px">
      <option value="vsco">VSCO (94)</option>
      <option value="mmxl">MMXL (2.12.8)</option>
      <option value="rev">REV (89) ❌</option>
    </select>
    <button onclick="addMocaNode()" class="diag-btn-add">+ הוסף</button>
  </div>
  <div id="moca-nodes-list" style="margin-bottom:10px;min-height:24px"></div>
  <div id="moca-matrix-container"></div>
  <div id="moca-summary" style="margin-top:8px"></div>
</div>

<!-- בדיקת הגדרות MOCA -->
<div class="diag-section">
  <div class="diag-title">⚙️ הגדרות MOCA</div>
  <div style="display:grid;gap:10px;margin-top:10px">
    <div class="setting-row">
      <div class="setting-label">LOF (ערוץ יעד)</div>
      <input id="lof-val" type="number" class="setting-input" placeholder="1450">
      <span style="color:#6b94b8;font-size:12px">MHz</span>
      <span id="lof-status" style="font-size:16px"></span>
    </div>
    <div class="setting-row">
      <div class="setting-label">Tx Power</div>
      <input id="txpow-val" type="number" class="setting-input" placeholder="8">
      <span style="color:#6b94b8;font-size:12px">dBm</span>
      <span id="txpow-status" style="font-size:16px"></span>
    </div>
    <div class="setting-row">
      <div class="setting-label">Beacon Power Level</div>
      <input id="beacon-val" type="number" class="setting-input" placeholder="10">
      <span style="color:#6b94b8;font-size:12px"></span>
      <span id="beacon-status" style="font-size:16px"></span>
    </div>
    <div class="setting-row">
      <div class="setting-label">לוגו סלקום</div>
      <button onclick="setLogoStatus(true)" id="logo-yes-btn" class="logo-btn-active">✅ קיים</button>
      <button onclick="setLogoStatus(false)" id="logo-no-btn" class="logo-btn">❌ חסר</button>
      <span id="logo-status" style="font-size:16px"></span>
    </div>
  </div>
  <button onclick="checkMocaSettings()" class="diag-btn-check" style="margin-top:14px">🔍 בדוק הגדרות</button>
  <div id="settings-result" style="margin-top:10px"></div>
</div>

</div><!-- /diag-sec-moca -->

<div id="diag-sec-dsl">
<!-- ══ תשתית נחושת DSL/VDSL ══ -->
<div style="padding:4px 0 10px;font-size:18px;font-weight:700;color:#f9c846;margin-top:8px">📡 תשתית נחושת DSL/VDSL</div>

<!-- גישה לנתב נחושת -->
<div class="diag-section">
  <div class="diag-title">🌐 גישה לנתב נחושת</div>
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:8px">
    <a href="http://10.100.102.1" target="_blank" class="diag-btn-primary">
      🔗 פתח נתב<div style="font-size:11px;opacity:0.8;margin-top:2px">10.100.102.1</div>
    </a>
    <div style="background:#0d1b2a;border-radius:10px;padding:10px;font-size:12px;direction:rtl;line-height:1.9">
      <div style="color:#f9c846;font-weight:600">לפני משיכת קובץ:</div>
      <div style="color:#eaf4ff">net013 / <b>lz5jbzto</b></div>
      <div style="color:#f9c846;font-weight:600;margin-top:4px">אחרי משיכת קובץ:</div>
      <div style="color:#eaf4ff">ADMIN (מהאפליקציה)</div>
    </div>
  </div>
  <div style="margin-top:8px;background:#0d1b2a;border-radius:8px;padding:8px 12px;font-size:12px;direction:rtl;color:#9bbdd8">
    📌 דגמים: <b style="color:#eaf4ff">SagemCom 5359</b> (נחושת עד 200Mbps) | <b style="color:#eaf4ff">VVS822</b> | <b style="color:#eaf4ff">ADB / Swiscom</b>
  </div>
</div>

<!-- בדיקת סינכרון DSL -->
<div class="diag-section">
  <div class="diag-title">⚡ בדיקת סינכרון DSL</div>
  <div class="diag-subtitle">
    T15: הורדה ~18,938 | T100: 55K-102K | T200: 150K-200K Kbps
  </div>
  <div style="display:grid;gap:10px;margin-top:4px">
    <div class="setting-row">
      <div class="setting-label">Downstream Rate</div>
      <input id="dsl-ds" type="number" class="setting-input" placeholder="Kbps" oninput="detectDslCable()">
      <span style="color:#6b94b8;font-size:12px">Kbps</span>
    </div>
    <div class="setting-row">
      <div class="setting-label">Upstream Rate</div>
      <input id="dsl-us" type="number" class="setting-input" placeholder="Kbps" oninput="detectDslCable()">
      <span style="color:#6b94b8;font-size:12px">Kbps</span>
    </div>
  </div>
  <div id="dsl-cable-result" style="margin-top:10px"></div>
</div>

<!-- Noise Margin & Attenuation -->
<div class="diag-section">
  <div class="diag-title">📶 איכות קו DSL</div>
  <div class="diag-subtitle">
    Statistics → Downstream / Upstream Measured Values
  </div>
  <div style="display:grid;gap:10px;margin-top:4px">
    <div class="setting-row">
      <div class="setting-label">Noise Margin DS</div>
      <input id="snr-ds" type="number" step="0.1" class="setting-input" placeholder="dB">
      <span style="color:#6b94b8;font-size:12px">dB</span>
    </div>
    <div class="setting-row">
      <div class="setting-label">Noise Margin US</div>
      <input id="snr-us" type="number" step="0.1" class="setting-input" placeholder="dB">
      <span style="color:#6b94b8;font-size:12px">dB</span>
    </div>
    <div class="setting-row">
      <div class="setting-label">Line Attenuation DS</div>
      <input id="att-ds" type="number" step="0.1" class="setting-input" placeholder="dB">
      <span style="color:#6b94b8;font-size:12px">dB</span>
    </div>
  </div>
  <button onclick="checkDslQuality()" class="diag-btn-check" style="margin-top:12px">🔍 בדוק איכות קו</button>
  <div id="dsl-quality-result" style="margin-top:8px"></div>
</div>

<!-- Advanced Statistics - שגיאות -->
<div class="diag-section">
  <div class="diag-title">🔬 Advanced Statistics — שגיאות</div>
  <div class="diag-subtitle">
    Settings ← Network Connections ← DSL Line ← Advanced Statistics
  </div>
  <div style="display:grid;gap:8px;margin-top:4px">
    <div class="setting-row">
      <div class="setting-label">FEC errors</div>
      <input id="err-fec" type="number" class="setting-input" placeholder="DS">
      <span style="color:#6b94b8;font-size:11px">DS</span>
      <input id="err-fec-us" type="number" class="setting-input" placeholder="US">
      <span style="color:#6b94b8;font-size:11px">US</span>
    </div>
    <div class="setting-row">
      <div class="setting-label">CRC errors</div>
      <input id="err-crc" type="number" class="setting-input" placeholder="DS">
      <span style="color:#6b94b8;font-size:11px">DS</span>
      <input id="err-crc-us" type="number" class="setting-input" placeholder="US">
      <span style="color:#6b94b8;font-size:11px">US</span>
    </div>
    <div class="setting-row">
      <div class="setting-label">ES (שניות שגיאה)</div>
      <input id="err-es" type="number" class="setting-input" placeholder="DS">
      <span style="color:#6b94b8;font-size:11px">DS</span>
      <input id="err-es-us" type="number" class="setting-input" placeholder="US">
      <span style="color:#6b94b8;font-size:11px">US</span>
    </div>
    <div class="setting-row">
      <div class="setting-label">SES (שניות חמורות)</div>
      <input id="err-ses" type="number" class="setting-input" placeholder="DS">
      <span style="color:#6b94b8;font-size:11px">DS</span>
      <input id="err-ses-us" type="number" class="setting-input" placeholder="US">
      <span style="color:#6b94b8;font-size:11px">US</span>
    </div>
    <div class="setting-row">
      <div class="setting-label">UAS (שניות ללא חיבור)</div>
      <input id="err-uas" type="number" class="setting-input" placeholder="DS">
      <span style="color:#6b94b8;font-size:11px">DS</span>
      <input id="err-uas-us" type="number" class="setting-input" placeholder="US">
      <span style="color:#6b94b8;font-size:11px">US</span>
    </div>
    <div class="setting-row">
      <div class="setting-label">LOS (אבדן סנכרון)</div>
      <input id="err-los" type="number" class="setting-input" placeholder="DS">
      <span style="color:#6b94b8;font-size:11px">DS</span>
      <input id="err-los-us" type="number" class="setting-input" placeholder="US">
      <span style="color:#6b94b8;font-size:11px">US</span>
    </div>
    <div class="setting-row">
      <div class="setting-label">LOM (שינוי SNR)</div>
      <input id="err-lom" type="number" class="setting-input" placeholder="DS">
      <span style="color:#6b94b8;font-size:11px">DS</span>
      <input id="err-lom-us" type="number" class="setting-input" placeholder="US">
      <span style="color:#6b94b8;font-size:11px">US</span>
    </div>
  </div>
  <button onclick="checkDslErrors()" class="diag-btn-check" style="margin-top:12px">🔍 בדוק שגיאות</button>
  <div id="dsl-errors-result" style="margin-top:8px"></div>
</div>

<!-- אין סנכרון - רשימת תיוג -->
<div class="diag-section">
  <div class="diag-title">🚫 אין סנכרון — רשימת בדיקות</div>
  <div class="diag-subtitle">בצע לפי הסדר לפני ניתוק עבודה</div>
  <div id="no-sync-checklist" style="background:#0d1b2a;border-radius:10px;margin-top:6px;overflow:hidden"></div>
  <button onclick="resetNoSyncList()" style="margin-top:10px;background:transparent;border:1px solid #284461;color:#6b94b8;border-radius:8px;padding:6px 14px;font-size:12px;cursor:pointer">↺ אפס רשימה</button>
</div>

</div><!-- /diag-sec-dsl -->

<div id="diag-sec-wifi">
<!-- ══ WiFi / AirTies MESH ══ -->
<div style="padding:4px 0 10px;font-size:18px;font-weight:700;color:#53bdeb;margin-top:8px">📶 WiFi / AirTies MESH</div>

<!-- גישה ליחידת AirTies -->
<div class="diag-section">
  <div class="diag-title">🌐 גישה ליחידת AirTies</div>
  <div class="diag-subtitle">
    כנס לנתב 10.100.102.1 → רשימת מכשירים → מצא IP של AirTies (לרוב 10.100.102.7) → פתח בדפדפן → Advanced Settings
  </div>
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:6px">
    <a href="http://10.100.102.1" target="_blank" class="diag-btn-primary">
      🔗 פתח נתב<div style="font-size:11px;opacity:0.8;margin-top:2px">מצא IP של AirTies</div>
    </a>
    <div style="background:#0d1b2a;border-radius:10px;padding:10px;font-size:12px;direction:rtl;line-height:1.9">
      <div style="color:#53bdeb;font-weight:600">כתובות נפוצות AirTies:</div>
      <div style="color:#eaf4ff">10.100.102.7 (Air4920)</div>
      <div style="color:#eaf4ff">10.100.102.3, .22 ...</div>
      <div style="color:#f9c846;font-size:11px;margin-top:4px">ניווט: Advanced Settings ← Wireless ← MESH</div>
    </div>
  </div>
</div>

<!-- AirTies Firmware Check -->
<div class="diag-section">
  <div class="diag-title">💾 Firmware יחידת AirTies</div>
  <div class="diag-subtitle">מתחבר אוטומטית ← קורא גרסה ← בודק עם הגרסה העדכנית</div>
  <div style="display:grid;gap:8px;margin-top:8px">
    <div class="setting-row">
      <div class="setting-label">IP AirTies</div>
      <input id="airties-fw-ip" placeholder="לדוג׳ 10.100.102.7" class="setting-input" style="max-width:160px">
    </div>
    <div class="setting-row">
      <div class="setting-label">דגם</div>
      <select id="airties-fw-model" style="flex:1;background:#0d1b2a;border:1px solid #284461;border-radius:8px;padding:8px 10px;color:#eaf4ff;font-size:13px;direction:rtl">
        <option value="air4920">Air4920IL</option>
        <option value="air4930">Air4930IL</option>
      </select>
    </div>
  </div>
  <button onclick="checkAirtiesFw()" class="diag-btn-check" style="margin-top:10px;background:linear-gradient(135deg,#1a2a4a,#1a3a5c);border-color:#53bdeb;color:#eaf4ff">📡 בדוק Firmware AirTies</button>
  <div id="airties-fw-result" style="margin-top:8px"></div>
  <!-- גרסאות ידועות -->
  <div style="margin-top:10px;background:#0d1b2a;border-radius:8px;padding:10px;font-size:12px;direction:rtl;line-height:2">
    <div style="color:#53bdeb;font-weight:600;margin-bottom:4px">גרסאות עדכניות:</div>
    <div style="color:#3dba6f">✅ Air4920IL — <b>3.112.14.4.1461</b></div>
    <div style="color:#3dba6f">✅ Air4930IL — <b>2.112.14.4.528</b></div>
  </div>
</div>

<!-- ADB Firmware Check -->
<div class="diag-section">
  <div class="diag-title">💾 Firmware נתב ADB</div>
  <div class="diag-subtitle">בדוק Firmware Version במסך הבית → Device Summary</div>
  <div style="display:grid;gap:8px;margin-top:6px">
    <div class="fw-row fw-bad">
      <span class="fw-badge fw-rev">גרסה ...0039 (39)</span>
      <span style="color:#f07080;font-weight:600">❌ Hard Reset נדרש</span>
      <button onclick="toggleAdbGuide()" class="fw-update-btn">📖 הוראות</button>
    </div>
    <div class="fw-row fw-good">
      <span class="fw-badge fw-vsco">גרסה אחרת</span>
      <span style="color:#3dba6f;font-weight:600">✅ תקין</span>
    </div>
  </div>
  <div id="adb-guide" style="display:none;margin-top:8px">
    <div style="background:#0d1b2a;border-radius:10px;padding:14px;font-size:13px;line-height:2">
      <div style="font-weight:700;color:#f9c846;margin-bottom:8px">🔄 Hard Reset לנתב ADB</div>
      <div class="step-item">1️⃣ לחץ Hard Reset על גבי הנתב</div>
      <div class="step-item" style="margin-top:4px">2️⃣ המתן ~5 דקות — הנתב יעלה מחדש עם הגרסה הרולינגית</div>
      <div class="step-item" style="margin-top:4px">3️⃣ ודא שגרסת ה-Firmware עודכנה</div>
      <div style="background:#3d2a0a;border-radius:8px;padding:8px 12px;margin-top:8px;color:#f9c846;font-size:12px">
        ⚠️ הנתב יאבד הגדרות — נדרש הגדרה מחדש!
      </div>
    </div>
  </div>
</div>

<!-- בדיקת MESH Signal Level -->
<div class="diag-section">
  <div class="diag-title">📡 MESH Signal Level</div>
  <div class="diag-subtitle">Advanced Settings → Wireless → MESH → בדוק Signal Level בטבלת MESH List</div>
  <div style="display:flex;align-items:center;gap:12px;margin-top:10px;flex-wrap:wrap">
    <div class="setting-row" style="flex:1;min-width:200px">
      <div class="setting-label">Signal Level</div>
      <input id="mesh-signal" type="number" min="0" max="100" class="setting-input" placeholder="%" oninput="checkMeshSignal()">
      <span style="color:#6b94b8;font-size:14px">%</span>
    </div>
  </div>
  <div id="mesh-signal-result" style="margin-top:10px"></div>
  <div style="margin-top:10px;background:#0d1b2a;border-radius:8px;padding:10px;font-size:12px;direction:rtl;line-height:1.8;color:#9bbdd8">
    <div style="color:#eaf4ff;font-weight:600;margin-bottom:4px">📌 פעולה לפי תוצאה:</div>
    <div>✅ <b>&gt; 55%</b> — מיקום תקין, המשך התקנה</div>
    <div>⚠️ <b>&lt; 55%</b> — קרב יחידת Super Wi-Fi לנתב</div>
    <div>📱 <b>התחברות להגדרות:</b> גש עם הטלפון ליחידה ← לחץ Home Page</div>
  </div>
</div>

<!-- בחר סוג נתב -->
<div class="diag-section">
  <div class="diag-title">🔀 בחר סוג נתב להתקנת WiFi</div>
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:10px">
    <button onclick="showWifiMode('wps')" id="wifi-mode-wps"
      style="padding:12px 8px;border-radius:10px;border:2px solid #3dba6f;background:#1a3d22;color:#3dba6f;font-weight:700;font-size:13px;cursor:pointer">
      📡 ADB 5820/5823<div style="font-size:11px;font-weight:400;margin-top:3px">שיוך WPS אלחוטי</div>
    </button>
    <button onclick="showWifiMode('ap')" id="wifi-mode-ap"
      style="padding:12px 8px;border-radius:10px;border:2px solid #284461;background:#162d44;color:#6b94b8;font-weight:400;font-size:13px;cursor:pointer">
      🔌 B35 / SagemCom / SwissCom<div style="font-size:11px;font-weight:400;margin-top:3px">Access Point דרך כבל</div>
    </button>
  </div>

  <!-- WPS Mode (ADB) -->
  <div id="wifi-wps-guide" style="margin-top:12px">
    <div style="background:#0d1b2a;border-radius:10px;padding:14px;font-size:13px;line-height:2;direction:rtl">
      <div style="font-weight:700;color:#3dba6f;margin-bottom:8px">🔧 שיוך WPS — ADB 5820/5823</div>
      <div style="margin-bottom:10px;padding:8px 10px;background:#162d44;border-radius:8px">
        <div style="color:#f9c846;font-weight:600;margin-bottom:4px">הכנה בנתב ADB (10.100.102.1):</div>
        <div class="step-item" style="margin-bottom:3px">1️⃣ נתב ← הגדרות אלחוטי</div>
        <div class="step-item" style="margin-bottom:3px">2️⃣ שם רשת + סיסמה זהים ל-2.4 ו-5GHz</div>
        <div class="step-item" style="margin-bottom:3px">3️⃣ Wi-Fi Protected Setup ← WPS Enabled: Yes ← Push Button: ✓ ← Apply</div>
      </div>
      <div style="padding:8px 10px;background:#162d44;border-radius:8px">
        <div style="color:#53bdeb;font-weight:600;margin-bottom:4px">שיוך AirTies (יחידה ראשית ← יחידה משנית):</div>
        <div class="step-item" style="margin-bottom:3px">4️⃣ הדלק AirTies ליד הנתב (עד 3 מטר) — המתן לנוריות 2.4+5GHz ירוק</div>
        <div class="step-item" style="margin-bottom:3px">5️⃣ לחץ WPS על נתב ADB בדיוק <b>3 שניות</b></div>
        <div class="step-item" style="margin-bottom:3px">6️⃣ לחץ WPS ביחידת AirTies — <b>3 שניות</b></div>
        <div class="step-item" style="margin-bottom:3px">7️⃣ AirTies תהבהב ← תיכבה ← המתן ~60 שניות</div>
        <div class="step-item">8️⃣ בדוק Signal Level ב-MESH ← בדוק WiFi בבית</div>
      </div>
      <div id="wps-multi-unit" style="margin-top:8px;padding:8px 10px;background:#162d44;border-radius:8px;display:none">
        <div style="color:#f9c846;font-weight:600;margin-bottom:4px">➕ שיוך יחידה משנית נוספת:</div>
        <div class="step-item" style="margin-bottom:3px">• המתן לסיום שיוך היחידה הראשית</div>
        <div class="step-item" style="margin-bottom:3px">• צימוד WPS עד 5 דקות — 5GHz תהבהב בשני המכשירים</div>
        <div class="step-item" style="margin-bottom:3px">• אם 5GHz ביחידה משנית לא מהבהבת אחרי 5 דקות — חזור לשלב</div>
        <div class="step-item" style="margin-bottom:3px">• אם לא עובד — בצע <b>Factory Reset</b> ליחידה המשנית</div>
      </div>
      <button onclick="toggleWpsMultiUnit()" style="margin-top:8px;background:transparent;border:1px solid #284461;color:#53bdeb;border-radius:8px;padding:6px 12px;font-size:12px;cursor:pointer">➕ שיוך יחידה שנייה/שלישית</button>
    </div>
  </div>

  <!-- AP Mode (B35 / SagemCom / SwissCom) -->
  <div id="wifi-ap-guide" style="display:none;margin-top:12px">
    <div style="background:#0d1b2a;border-radius:10px;padding:14px;font-size:13px;line-height:2;direction:rtl">
      <div style="font-weight:700;color:#53bdeb;margin-bottom:8px">🔌 הגדרת Access Point — B35 / SagemCom / SwissCom</div>
      <div style="margin-bottom:10px;padding:8px 10px;background:#162d44;border-radius:8px">
        <div style="color:#f9c846;font-weight:600;margin-bottom:4px">1️⃣ חיבור פיזי:</div>
        <div class="step-item" style="margin-bottom:3px">• חבר כבל מנתב ← יחידת AirTies ראשית (פורט LAN)</div>
        <div class="step-item">• SagemCom: Home Network → מצא IP של Air4920</div>
        <div class="step-item">• B35: Home Network → LAN Eth → מצא IP של Air4920</div>
        <div class="step-item">• SwissCom: Home Network → LAN → מצא IP</div>
      </div>
      <div style="margin-bottom:10px;padding:8px 10px;background:#162d44;border-radius:8px">
        <div style="color:#53bdeb;font-weight:600;margin-bottom:4px">2️⃣ הגדרת AirTies כ-Access Point:</div>
        <div class="step-item" style="margin-bottom:3px">• פתח דפדפן ← הכנס IP של AirTies</div>
        <div class="step-item" style="margin-bottom:3px">• Quick Setup ← לחץ לשונית <b>Access Point</b> (לא Repeater!)</div>
        <div class="step-item" style="margin-bottom:3px">• Network Name = SSID בנתב | Password = סיסמה בנתב</div>
        <div class="step-item">• לחץ <b>Save</b></div>
      </div>
      <div style="padding:8px 10px;background:#162d44;border-radius:8px">
        <div style="color:#3dba6f;font-weight:600;margin-bottom:4px">3️⃣ שיוך יחידה משנית (WPS בין 2 AirTies):</div>
        <div class="step-item" style="margin-bottom:3px">• הנח יחידה משנית עד 3 מטר מהיחידה הראשית</div>
        <div class="step-item" style="margin-bottom:3px">• לחץ WPS על יחידה ראשית — <b>2 שניות</b></div>
        <div class="step-item" style="margin-bottom:3px">• לחץ WPS על יחידה משנית — <b>2 שניות</b></div>
        <div class="step-item" style="margin-bottom:3px">• המתן ~5 דקות — נורית 5GHz תדלוק רציף (ירוק)</div>
        <div class="step-item" style="margin-bottom:3px">• אם לא הצליח: Factory Reset ← נסה שנית</div>
        <div class="step-item">• לאחר שיוך: הורד אפליקציית AirTies לכיול מיקום</div>
      </div>
      <div style="background:#3d2a0a;border-radius:8px;padding:8px 12px;margin-top:8px;color:#f9c846;font-size:12px">
        ⚠️ AirTies חייב להיות ב-Access Point — לא Repeater!<br>
        חיבור <b>כבל בלבד</b> מהנתב ליחידה הראשית
      </div>
    </div>
  </div>
</div>

</div><!-- /diag-sec-wifi -->

<div id="diag-sec-fiber">
<!-- ══ נתבי סיב — נוריות ובדיקות ══ -->
<div style="padding:4px 0 10px;font-size:18px;font-weight:700;color:#9b7fff;margin-top:8px">🔦 נתבי סיב — נוריות ובדיקות</div>

<!-- טבלת דגמי נתבים -->
<div class="diag-section">
  <div class="diag-title">📦 דגמי נתב סיב</div>
  <div style="display:grid;gap:8px;margin-top:8px">
    <div style="background:#0d1b2a;border-radius:8px;padding:10px;direction:rtl;font-size:13px">
      <div style="color:#9b7fff;font-weight:700">SagemCom F@ST 5657 — סופרפייבר</div>
      <div style="color:#9bbdd8;font-size:12px;margin-top:2px">FTTH + נתב משולב | ניתן להתקנה בכל מקום ללא POP</div>
    </div>
    <div style="background:#0d1b2a;border-radius:8px;padding:10px;direction:rtl;font-size:13px">
      <div style="color:#9b7fff;font-weight:700">SagemCom F@ST 5670 AIO ULTRA</div>
      <div style="color:#9bbdd8;font-size:12px;margin-top:2px">FTTH + משולב | WiFi 6 (AX) | עד 5.2 Gbps | פורט LAN ייעודי | תמיכה מפייא — קו 1</div>
    </div>
    <div style="background:#0d1b2a;border-radius:8px;padding:10px;direction:rtl;font-size:13px">
      <div style="color:#9b7fff;font-weight:700">SagemCom F@ST 266V2 MESH ULTRA Wi-Fi</div>
      <div style="color:#9bbdd8;font-size:12px;margin-top:2px">WiFi 6 (JAX) מלא | שיפור קליטה 15% | שיפור מהירות אלחוטי | משמש כיחידת ULTRA ראשית</div>
    </div>
  </div>
</div>

<!-- נוריות נתב סיב -->
<div class="diag-section">
  <div class="diag-title">💡 נוריות נתב סיב (5657 / 5670)</div>
  <div style="display:grid;gap:8px;margin-top:10px">
    <div style="background:#0d1b2a;border-radius:10px;overflow:hidden">
      <div style="display:flex;justify-content:space-between;padding:9px 12px;border-bottom:1px solid #1c3450;direction:rtl">
        <span style="color:#9b7fff;font-weight:600;min-width:80px">POWER</span>
        <span style="color:#eaf4ff">ירוקה / לבנה קבועה</span>
        <span style="color:#3dba6f">✅ מחובר חשמל</span>
      </div>
      <div style="display:flex;justify-content:space-between;padding:9px 12px;border-bottom:1px solid #1c3450;direction:rtl">
        <span style="color:#3dba6f;font-weight:600;min-width:80px">PON 🟢</span>
        <span style="color:#eaf4ff">קבועה</span>
        <span style="color:#3dba6f">✅ סיב מחובר — סנכרון תקין</span>
      </div>
      <div style="display:flex;justify-content:space-between;padding:9px 12px;border-bottom:1px solid #1c3450;direction:rtl">
        <span style="color:#f9c846;font-weight:600;min-width:80px">PON 🔆</span>
        <span style="color:#eaf4ff">מהבהבת</span>
        <span style="color:#f9c846">⚠️ ממתין לסנכרון — המתן לסיום הגדרה במערכת</span>
      </div>
      <div style="display:flex;justify-content:space-between;padding:9px 12px;border-bottom:1px solid #1c3450;direction:rtl">
        <span style="color:#f15c6e;font-weight:600;min-width:80px">PON ⭕</span>
        <span style="color:#eaf4ff">כבויה</span>
        <span style="color:#f15c6e">❌ חוסר סנכרון סיב — בדוק ארון ראשי</span>
      </div>
      <div style="display:flex;justify-content:space-between;padding:9px 12px;border-bottom:1px solid #1c3450;direction:rtl">
        <span style="color:#3dba6f;font-weight:600;min-width:80px">INTERNET 🟢</span>
        <span style="color:#eaf4ff">ירוקה</span>
        <span style="color:#3dba6f">✅ מחובר לאינטרנט</span>
      </div>
      <div style="display:flex;justify-content:space-between;padding:9px 12px;direction:rtl">
        <span style="color:#f15c6e;font-weight:600;min-width:80px">INTERNET 🔴</span>
        <span style="color:#eaf4ff">אדומה</span>
        <span style="color:#f15c6e">❌ בעיה בחיבור אינטרנט</span>
      </div>
    </div>
    <div style="background:#3d2a0a;border-radius:8px;padding:8px 12px;font-size:12px;color:#f9c846;direction:rtl">
      ⚠️ PON מהבהבת — המתן לסיום הגדרה במערכת. לא ניתן להקטב ארון ראשי בינתיים.
    </div>
  </div>
</div>

<!-- נוריות יחידת MESH (266V2 ULTRA) -->
<div class="diag-section">
  <div class="diag-title">💡 נוריות יחידת MESH (F@ST 266V2 ULTRA)</div>
  <div style="display:grid;gap:8px;margin-top:10px">
    <div style="background:#0d1b2a;border-radius:10px;overflow:hidden">
      <!-- POWER -->
      <div style="padding:6px 12px;background:#1c3450;font-size:12px;font-weight:700;color:#9b7fff;direction:rtl">⚡ POWER</div>
      <div style="display:flex;justify-content:space-between;padding:8px 12px;border-bottom:1px solid #1c3450;font-size:13px;direction:rtl">
        <span style="display:flex;align-items:center;gap:6px"><span style="width:12px;height:12px;border-radius:50%;background:#fff;display:inline-block"></span>לבן קבוע</span>
        <span style="color:#3dba6f">✅ MESH פעיל</span>
      </div>
      <div style="display:flex;justify-content:space-between;padding:8px 12px;border-bottom:1px solid #1c3450;font-size:13px;direction:rtl">
        <span style="display:flex;align-items:center;gap:6px"><span style="width:12px;height:12px;border-radius:50%;background:#aaa;display:inline-block"></span>לבן מהבהב</span>
        <span style="color:#f9c846">⏳ אתחול / Booting / שדרוג</span>
      </div>
      <div style="display:flex;justify-content:space-between;padding:8px 12px;border-bottom:2px solid #1c3450;font-size:13px;direction:rtl">
        <span style="display:flex;align-items:center;gap:6px"><span style="width:12px;height:12px;border-radius:50%;background:#333;border:1px solid #555;display:inline-block"></span>כבוי</span>
        <span style="color:#f15c6e">❌ MESH כבוי</span>
      </div>
      <!-- SIGNAL -->
      <div style="padding:6px 12px;background:#1c3450;font-size:12px;font-weight:700;color:#53bdeb;direction:rtl">📶 SIGNAL (איכות חיבור MESH)</div>
      <div style="display:flex;justify-content:space-between;padding:8px 12px;border-bottom:1px solid #1c3450;font-size:13px;direction:rtl">
        <span style="display:flex;align-items:center;gap:6px"><span style="width:12px;height:12px;border-radius:50%;background:#3dba6f;display:inline-block"></span>ירוק קבוע</span>
        <span style="color:#3dba6f">✅ Best Link Quality — מיקום מצוין</span>
      </div>
      <div style="display:flex;justify-content:space-between;padding:8px 12px;border-bottom:1px solid #1c3450;font-size:13px;direction:rtl">
        <span style="display:flex;align-items:center;gap:6px"><span style="width:12px;height:12px;border-radius:50%;background:#f9a825;display:inline-block"></span>כתום קבוע</span>
        <span style="color:#f9c846">⚠️ Good Link Quality — מיקום סביר</span>
      </div>
      <div style="display:flex;justify-content:space-between;padding:8px 12px;border-bottom:2px solid #1c3450;font-size:13px;direction:rtl">
        <span style="display:flex;align-items:center;gap:6px"><span style="width:12px;height:12px;border-radius:50%;background:#f15c6e;display:inline-block"></span>אדום קבוע</span>
        <span style="color:#f15c6e">❌ Poor Link Quality — הזז יחידה!</span>
      </div>
      <!-- WiFi -->
      <div style="padding:6px 12px;background:#1c3450;font-size:12px;font-weight:700;color:#3dba6f;direction:rtl">📡 WiFi</div>
      <div style="display:flex;justify-content:space-between;padding:8px 12px;border-bottom:1px solid #1c3450;font-size:13px;direction:rtl">
        <span>לבן קבוע</span><span style="color:#3dba6f">✅ WiFi ביחידה פעיל</span>
      </div>
      <div style="display:flex;justify-content:space-between;padding:8px 12px;border-bottom:1px solid #1c3450;font-size:13px;direction:rtl">
        <span>לבן מהבהב</span><span style="color:#f9c846">⏳ תהליך WPS פעיל</span>
      </div>
      <div style="display:flex;justify-content:space-between;padding:8px 12px;font-size:13px;direction:rtl">
        <span>כבוי</span><span style="color:#f15c6e">❌ WiFi כבוי</span>
      </div>
    </div>
  </div>
</div>

<!-- לשוניות MESH בנתב -->
<div class="diag-section">
  <div class="diag-title">🗂️ לשוניות MESH בנתב — מה לבדוק</div>
  <div style="display:grid;gap:8px;margin-top:8px">
    <button onclick="toggleMeshTab('overview')" style="background:#162d44;border:none;border-radius:8px;padding:10px 12px;color:#eaf4ff;font-size:13px;font-weight:600;cursor:pointer;text-align:right;direction:rtl">
      📍 Overview — מפת יחידות MESH ▼
    </button>
    <div id="mesh-tab-overview" style="display:none;background:#0d1b2a;border-radius:8px;padding:10px;font-size:13px;direction:rtl;line-height:1.9">
      <div style="color:#9bbdd8">• מציגה מפה ויזואלית של כל יחידות ה-MESH</div>
      <div style="color:#9bbdd8">• יחידה עליונה מחוברת <b>בכבל</b> לנתב ← מחוברת ל-S22 ב-5GHz</div>
      <div style="color:#9bbdd8">• יחידה תחתונה מחוברת <b>אלחוטית</b> לעליונה ב-2.4GHz</div>
      <div style="color:#f9c846;margin-top:4px">← ניתן לראות גם ציוד לקוח (טלפונים/טאבלטים)</div>
    </div>
    <button onclick="toggleMeshTab('extenders')" style="background:#162d44;border:none;border-radius:8px;padding:10px 12px;color:#eaf4ff;font-size:13px;font-weight:600;cursor:pointer;text-align:right;direction:rtl">
      📡 Extenders — פרטי יחידות MESH ▼
    </button>
    <div id="mesh-tab-extenders" style="display:none;background:#0d1b2a;border-radius:8px;padding:10px;font-size:13px;direction:rtl;line-height:1.9">
      <div style="color:#9bbdd8">• <b>UpTime</b> — זמן חיבור חשמלי של היחידה</div>
      <div style="color:#9bbdd8">• <b>Address/IP</b> — כתובת IP לוקאלית</div>
      <div style="color:#9bbdd8">• <b>Connection Type</b> — תצורת חיבור (קווי / אלחוטי)</div>
      <div style="color:#9bbdd8">• <b>Role</b> — תפקיד (gateway / extender)</div>
      <div style="color:#9bbdd8">• <b>Operating Software</b> — גרסת ליבה (firmware)</div>
    </div>
    <button onclick="toggleMeshTab('devices')" style="background:#162d44;border:none;border-radius:8px;padding:10px 12px;color:#eaf4ff;font-size:13px;font-weight:600;cursor:pointer;text-align:right;direction:rtl">
      📱 Devices — מכשירים מחוברים ▼
    </button>
    <div id="mesh-tab-devices" style="display:none;background:#0d1b2a;border-radius:8px;padding:10px;font-size:13px;direction:rtl;line-height:1.9">
      <div style="color:#9bbdd8">• <b>Name</b> — שם ההתקן</div>
      <div style="color:#9bbdd8">• <b>MAC Address</b> — כתובת MAC</div>
      <div style="color:#9bbdd8">• <b>IP Address</b> — כתובת IP לוקאלית</div>
      <div style="color:#9bbdd8">• <b>Band</b> — תדר חיבור (2.4GHz / 5GHz)</div>
      <div style="color:#9bbdd8">• <b>Signal Strength (RSSI)</b> — עוצמת האות:</div>
      <div style="margin-top:4px;background:#162d44;border-radius:8px;padding:8px">
        <div style="color:#3dba6f">• מעל -70 dBm = 🟢 חיבור מעולה</div>
        <div style="color:#f9c846">• -70 עד -80 dBm = 🟡 טוב</div>
        <div style="color:#f07080">• -80 עד -90 dBm = 🟠 חלש</div>
        <div style="color:#f15c6e">• מתחת -90 dBm = 🔴 גרוע — שקול הזזת יחידה</div>
      </div>
    </div>
  </div>
  <!-- שיוך יחידת MESH לנתב ULTRA -->
  <div style="margin-top:10px">
    <button onclick="toggleMeshPairGuide()" class="diag-btn-check" style="background:#4a2a7a">📋 הוראות שיוך יחידת MESH לנתב ULTRA</button>
    <div id="mesh-pair-guide" style="display:none;margin-top:8px;background:#0d1b2a;border-radius:10px;padding:14px;font-size:13px;direction:rtl;line-height:2">
      <div style="font-weight:700;color:#9b7fff;margin-bottom:8px">🔧 שיטה 1 — שיוך קווי לנתב ULTRA</div>
      <div class="step-item" style="margin-bottom:4px">1️⃣ חבר כבל רשת בין נתב ULTRA ← יחידת MESH</div>
      <div class="step-item" style="margin-bottom:4px">2️⃣ המתן עד 5 דקות — נוריות יתאמו לנוריות חיווי</div>
      <div class="step-item" style="margin-bottom:4px">3️⃣ ניתן לנתק כבל רשת ולהזיז יחידה למיקום הרצוי</div>
      <div class="step-item">4️⃣ ודא קישוריות/RSSI ← נתב MESH יתואם</div>
      <div style="margin-top:10px;font-weight:700;color:#9b7fff;margin-bottom:8px">🔧 שיטה 2 — שיוך WPS לנתב ULTRA</div>
      <div class="step-item" style="margin-bottom:4px">1️⃣ לחץ WPS על נתב ULTRA למשך <b>5 שניות</b></div>
      <div class="step-item" style="margin-bottom:4px">2️⃣ לחץ WPS על יחידת MESH (דופן שמאלי) <b>5 שניות</b></div>
      <div class="step-item" style="margin-bottom:4px">3️⃣ המתן עד 5 דקות — נוריות יתאמו</div>
      <div class="step-item" style="margin-bottom:4px">4️⃣ הזז יחידה למיקום הרצוי ← ודא קישוריות/RSSI</div>
      <div style="background:#3d2a0a;border-radius:8px;padding:8px 12px;margin-top:8px;color:#f9c846;font-size:12px">
        ⚠️ לכיול מיקום וקליטה — השתמש במילון המונחים וב-MESH overview
      </div>
    </div>
  </div>
</div>

</div><!-- /diag-sec-fiber -->

<div id="diag-sec-fw">
<!-- ══ בדיקת Firmware נתבים ══ -->
<div style="padding:4px 0 10px;font-size:18px;font-weight:700;color:#9b7fff;margin-top:8px">💾 בדיקת Firmware נתבים</div>

<!-- בדיקה אוטומטית לנתב בודד -->
<div class="diag-section">
  <div class="diag-title">🔍 בדוק Firmware — אוטומטי</div>
  <div style="font-size:12px;color:#9bbdd8;margin-bottom:10px">מתחבר לנתב → קורא גרסה → מדווח תקין / דרוש עדכון</div>
  <div style="display:grid;gap:8px;margin-top:4px">
    <div class="setting-row">
      <div class="setting-label">IP נתב</div>
      <input id="fw-router-ip" value="10.100.102.1" class="setting-input" style="max-width:160px">
    </div>
    <div class="setting-row">
      <div class="setting-label">דגם</div>
      <select id="fw-router-model" style="flex:1;background:#0d1b2a;border:1px solid #284461;border-radius:8px;padding:8px 10px;color:#eaf4ff;font-size:13px;direction:rtl">
        <option value="b35">נתב B35 (SagemCom 5359)</option>
        <option value="aio">SagemCom AIO (5657)</option>
        <option value="aio_ultra">SagemCom AIO ULTRA (5670)</option>
        <option value="star_plus">STAR PLUS (5674)</option>
        <option value="star_plus_heits">STAR PLUS HEITS (5698)</option>
        <option value="adb_vvs822">ADB VVS822/5823</option>
        <option value="swisscom">SwissCom (ST6840)</option>
      </select>
    </div>
  </div>
  <button onclick="runRouterFwCheckSingle()" class="diag-btn-check" style="margin-top:12px;background:linear-gradient(135deg,#3a1a5c,#1a2a5c);border-color:#9b7fff;color:#eaf4ff">🔍 בדוק Firmware</button>
  <div id="fw-single-result" style="margin-top:8px"></div>
</div>

<!-- סרוק כל נתבי הרשת -->
<div class="diag-section">
  <div class="diag-title">🔎 סרוק נתבים ברשת — כל הדגמים</div>
  <div style="font-size:12px;color:#9bbdd8;margin-bottom:8px">בדיקת כל הדגמים ב-IP נתב אחד (מחפש בכל הנתיבים)</div>
  <div class="setting-row">
    <div class="setting-label">IP לסריקה</div>
    <input id="fw-scan-ip" value="10.100.102.1" class="setting-input" style="max-width:160px">
  </div>
  <button onclick="runRouterFwScan()" class="diag-btn-check" style="margin-top:10px;background:linear-gradient(135deg,#1a3d5c,#1a2a4a);border-color:#53bdeb;color:#eaf4ff">🔎 סרוק כל הדגמים</button>
  <div id="fw-scan-result" style="margin-top:8px"></div>
</div>

<!-- טבלת גרסאות firmware ידועות -->
<div class="diag-section">
  <div class="diag-title">📋 גרסאות Firmware ידועות לפי דגם</div>
  <div style="font-size:12px;color:#9bbdd8;margin-bottom:8px">לעדכון גרסאות — פנה למנהל</div>
  <div id="fw-versions-table" style="margin-top:8px"></div>
</div>
</div><!-- /diag-sec-fw -->
`;
  renderMocaNodes();
  renderMocaMatrix();
  renderNoSyncChecklist();
  renderFwVersionsTable();
  if(_diagCard) syncDiagFromCard(_diagCard);
  else showAllDiagSections();
}

function toggleUpdateGuide(){
  const g=document.getElementById('update-guide');
  g.style.display=g.style.display==='none'?'block':'none';
}

function addMocaNode(){
  const nm=document.getElementById('moca-node-name');
  const fw=document.getElementById('moca-node-fw');
  const name=nm.value.trim()||('מוקה '+(_mocaNodes.length+1));
  _mocaNodes.push({name,fw:fw.value});
  nm.value='';
  renderMocaNodes();
  renderMocaMatrix();
}

function removeMocaNode(i){
  _mocaNodes.splice(i,1);
  renderMocaNodes();
  renderMocaMatrix();
}

function renderMocaNodes(){
  const list=document.getElementById('moca-nodes-list');
  if(!list) return;
  if(_mocaNodes.length===0){
    list.innerHTML='<div style="color:#8696a0;font-size:12px;text-align:center">הוסף מכשירי MOCA לבדיקה</div>';
    return;
  }
  const fwLabel={vsco:'VSCO✅',mmxl:'MMXL✅',rev:'REV❌'};
  const fwBg={vsco:'#1a4d30',mmxl:'#1a3d5c',rev:'#4d1a1a'};
  list.innerHTML=_mocaNodes.map((n,i)=>`
    <span style="display:inline-flex;align-items:center;gap:5px;background:${fwBg[n.fw]};border-radius:20px;padding:4px 10px;margin:2px;font-size:13px;color:#eaf4ff">
      ${n.name} <span style="opacity:0.7;font-size:11px">${fwLabel[n.fw]}</span>
      <button onclick="removeMocaNode(${i})" style="background:none;border:none;color:#f07080;cursor:pointer;font-size:14px;padding:0 0 0 4px;line-height:1">✕</button>
    </span>`).join('');
}

function getMocaThresholdInfo(fw1,fw2){
  if(fw1==='rev'||fw2==='rev') return {thr:0,type:'rev'};
  if(fw1==='mmxl'&&fw2==='mmxl') return {thr:1300,type:'mmxl'};
  if(fw1==='vsco'&&fw2==='vsco') return {thr:600,type:'vsco'};
  return {thr:600,type:'mixed'};
}

function renderMocaMatrix(){
  const container=document.getElementById('moca-matrix-container');
  const summary=document.getElementById('moca-summary');
  if(!container) return;
  if(_mocaNodes.length<2){
    container.innerHTML='<div style="color:#8696a0;font-size:12px;text-align:center;padding:8px">נדרשים לפחות 2 מכשירי MOCA</div>';
    if(summary) summary.innerHTML='';
    return;
  }
  const n=_mocaNodes.length;
  let h='<div style="overflow-x:auto"><table style="border-collapse:collapse;width:100%;font-size:12px;direction:ltr"><tr>';
  h+='<th style="background:#0d1b2a;padding:5px;color:#6b94b8;min-width:60px"></th>';
  _mocaNodes.forEach(nd=>{ h+=`<th style="background:#0d1b2a;padding:5px;color:#eaf4ff;text-align:center;min-width:80px">${nd.name}</th>`; });
  h+='</tr>';
  for(let i=0;i<n;i++){
    h+=`<tr><th style="background:#0d1b2a;padding:5px;color:#eaf4ff;text-align:right;white-space:nowrap">${_mocaNodes[i].name}</th>`;
    for(let j=0;j<n;j++){
      if(i===j){ h+='<td style="background:#162d44;padding:4px;text-align:center;color:#6b94b8">—</td>'; }
      else {
        h+=`<td style="padding:2px;background:#0d1b2a">
          <input id="mc-${i}-${j}" type="number" placeholder="Mbps"
            style="width:70px;background:#162d44;border:1px solid #284461;border-radius:6px;padding:5px 3px;color:#eaf4ff;font-size:12px;text-align:center"
            oninput="onMocaCellInput(${i},${j})">
        </td>`;
      }
    }
    h+='</tr>';
  }
  h+='</table></div>';
  container.innerHTML=h;
  if(summary) summary.innerHTML='';
}

function onMocaCellInput(i,j){
  const inp=document.getElementById('mc-'+i+'-'+j);
  if(!inp) return;
  const val=parseFloat(inp.value);
  const info=getMocaThresholdInfo(_mocaNodes[i].fw,_mocaNodes[j].fw);
  if(!isNaN(val)&&val>0){
    if(info.type==='rev'){ inp.style.background='#3d1a1a'; inp.style.color='#f07080'; }
    else if(info.type==='mixed'){
      inp.style.background=val>=info.thr?'#2d3a1a':'#3d1a1a';
      inp.style.color=val>=info.thr?'#f9c846':'#f07080';
    } else {
      inp.style.background=val>=info.thr?'#1a3d22':'#3d1a1a';
      inp.style.color=val>=info.thr?'#3dba6f':'#f07080';
    }
  } else { inp.style.background='#162d44'; inp.style.color='#eaf4ff'; }
  updateMocaSummary();
}

function updateMocaSummary(){
  const summary=document.getElementById('moca-summary');
  if(!summary||_mocaNodes.length<2) return;
  const n=_mocaNodes.length;
  const problems=[];
  let checked=0;
  for(let i=0;i<n;i++){
    for(let j=i+1;j<n;j++){
      const inp=document.getElementById('mc-'+i+'-'+j);
      if(!inp||!inp.value) continue;
      const val=parseFloat(inp.value);
      if(isNaN(val)||val<=0) continue;
      checked++;
      const info=getMocaThresholdInfo(_mocaNodes[i].fw,_mocaNodes[j].fw);
      if(info.type==='rev') problems.push(`${_mocaNodes[i].name} ↔ ${_mocaNodes[j].name}: REV — דרוש עדכון Firmware`);
      else if(info.type==='mixed') problems.push(`${_mocaNodes[i].name} ↔ ${_mocaNodes[j].name}: רשת מעורבת VSCO+MMXL ⚠️ (${val} Mbps)`);
      else if(val<info.thr) problems.push(`${_mocaNodes[i].name} ↔ ${_mocaNodes[j].name}: ${val} Mbps < ${info.thr} ❌`);
    }
  }
  if(checked===0){summary.innerHTML='';return;}
  if(problems.length===0){
    summary.innerHTML='<div style="background:#1a3d22;border-radius:8px;padding:10px;color:#3dba6f;font-size:13px;text-align:center">✅ כל קווי MOCA תקינים!</div>';
  } else {
    summary.innerHTML=`<div style="background:#3d1a1a;border-radius:8px;padding:10px;color:#f15c6e;font-size:13px;direction:rtl">
      <div style="font-weight:700;margin-bottom:6px">⚠️ בעיות:</div>
      ${problems.map(p=>'<div>• '+p+'</div>').join('')}
    </div>`;
  }
}

function setLogoStatus(ok){
  _logoOk=ok;
  const y=document.getElementById('logo-yes-btn');
  const nn=document.getElementById('logo-no-btn');
  if(!y||!nn) return;
  if(ok){ y.className='logo-btn-active'; nn.className='logo-btn'; }
  else { y.className='logo-btn'; nn.className='logo-btn-active'; }
}

function checkMocaSettings(){
  const lof=parseFloat(document.getElementById('lof-val').value);
  const txpow=parseFloat(document.getElementById('txpow-val').value);
  const beacon=parseFloat(document.getElementById('beacon-val').value);
  const checks=[
    {id:'lof',label:'LOF',val:lof,exp:1450,unit:'MHz'},
    {id:'txpow',label:'Tx Power',val:txpow,exp:8,unit:'dBm'},
    {id:'beacon',label:'Beacon Power Level',val:beacon,exp:10,unit:''},
  ];
  let rows='',allOk=true;
  checks.forEach(c=>{
    let ok,display;
    if(isNaN(c.val)){ok=null;display='לא הוזן';}
    else{ok=(c.val===c.exp);display=c.val+c.unit;}
    if(ok===false) allOk=false;
    const col=ok===true?'#3dba6f':ok===false?'#f15c6e':'#f9c846';
    const ico=ok===true?'✅':ok===false?'❌':'⚠️';
    const hint=ok===false?` (צריך: ${c.exp}${c.unit})`:'';
    rows+=`<div style="display:flex;justify-content:space-between;align-items:center;padding:8px 10px;border-bottom:1px solid #1c3450;font-size:13px;direction:rtl">
      <span style="color:#9bbdd8">${c.label}</span>
      <span style="color:${col}">${ico} ${display}${hint}</span>
    </div>`;
  });
  // logo
  let logoOk,logoDisplay;
  if(_logoOk===null){logoOk=null;logoDisplay='לא סומן';}
  else{logoOk=_logoOk;logoDisplay=_logoOk?'קיים':'חסר';}
  if(logoOk===false) allOk=false;
  const lCol=logoOk===true?'#3dba6f':logoOk===false?'#f15c6e':'#f9c846';
  const lIco=logoOk===true?'✅':logoOk===false?'❌':'⚠️';
  rows+=`<div style="display:flex;justify-content:space-between;align-items:center;padding:8px 10px;font-size:13px;direction:rtl">
    <span style="color:#9bbdd8">לוגו סלקום</span>
    <span style="color:${lCol}">${lIco} ${logoDisplay}</span>
  </div>`;
  const result=document.getElementById('settings-result');
  result.innerHTML=`<div style="background:#0d1b2a;border-radius:10px;overflow:hidden">${rows}</div>`
    +(allOk?'<div style="color:#3dba6f;text-align:center;margin-top:8px;font-size:14px;padding:4px">✅ כל ההגדרות תקינות!</div>':'');
}

// ── Pull-to-Refresh (נייד בלבד) ──────────────────────────────
(function(){
  if(!/Mobi|Android|iPhone|iPad/i.test(navigator.userAgent)) return;
  let _ptStartY=0, _ptEl=null, _ptIndicator=null, _ptTriggered=false;
  const THRESHOLD=70;

  function _getIndicator(){
    if(!_ptIndicator){
      _ptIndicator=document.createElement('div');
      _ptIndicator.style.cssText='position:fixed;top:0;left:0;right:0;z-index:99999;background:#25d366;color:#fff;text-align:center;font-size:14px;font-weight:700;padding:10px;transform:translateY(-100%);transition:transform 0.2s;pointer-events:none;direction:rtl';
      _ptIndicator.textContent='🔄 שחרר לרענון...';
      document.body.appendChild(_ptIndicator);
    }
    return _ptIndicator;
  }

  document.addEventListener('touchstart',function(e){
    const scr=document.getElementById('msgs')||document.scrollingElement||document.documentElement;
    const touchY=e.touches[0].clientY;
    // PTR רק אם: כבר בראש הדף + האצבע התחילה ב-100px העליונים של המסך
    if((scr.scrollTop||window.scrollY||0)===0 && touchY<=100){
      _ptStartY=touchY;
      _ptEl=scr;
      _ptTriggered=false;
    } else {
      _ptStartY=0;
    }
  },{passive:true});

  document.addEventListener('touchmove',function(e){
    if(!_ptStartY) return;
    const dy=e.touches[0].clientY-_ptStartY;
    if(dy<=0){_ptStartY=0;return;}
    const ind=_getIndicator();
    const pct=Math.min(dy/THRESHOLD,1);
    ind.style.transform=`translateY(${-100+pct*100}%)`;
    ind.textContent=pct>=1?'🔄 שחרר לרענון...':'⬇️ משוך לרענון';
    _ptTriggered=(pct>=1);
  },{passive:true});

  document.addEventListener('touchend',function(){
    if(!_ptStartY) return;
    _ptStartY=0;
    const ind=_getIndicator();
    ind.style.transform='translateY(-100%)';
    if(_ptTriggered){
      ind.textContent='🔄 מרענן...';
      ind.style.transform='translateY(0)';
      setTimeout(()=>location.reload(),300);
    }
    _ptTriggered=false;
  },{passive:true});
})();
</script>
<script>const ADMIN_PHONE_JS='__ADMIN_PHONE__';</script>

<!-- כפתור צ'אט צף -->
<button id="chat-float-btn" onclick="openChatModal()" style="display:none;position:fixed;bottom:76px;left:16px;width:50px;height:50px;border-radius:50%;background:#25d366;border:none;color:#fff;font-size:22px;cursor:pointer;z-index:900;box-shadow:0 3px 12px rgba(0,0,0,.4);align-items:center;justify-content:center">
  💬
  <span id="chat-unread-badge" style="display:none;position:absolute;top:-3px;right:-3px;background:#f15c6e;color:#fff;border-radius:50%;min-width:18px;height:18px;font-size:10px;font-weight:700;align-items:center;justify-content:center;pointer-events:none"></span>
</button>

<!-- מודל צ'אט -->
<div id="chat-modal" style="display:none;position:fixed;inset:0;z-index:3000;background:#0d1b2a;flex-direction:column;direction:rtl">
  <div id="chat-modal-header" style="display:flex;justify-content:space-between;align-items:center;padding:12px 14px;background:#d0e4f7;border-bottom:1px solid #284461;flex-shrink:0"></div>
  <div id="chat-modal-body" style="flex:1;overflow-y:auto;padding:12px 14px"></div>
  <!-- input bar — נסתר כשמוצגת רשימת שיחות -->
  <div id="chat-input-bar" style="display:flex;gap:8px;padding:10px 12px;background:#d0e4f7;border-top:1px solid #284461;flex-shrink:0">
    <input id="chat-inp" type="text" placeholder="כתוב הודעה..." onkeypress="if(event.key==='Enter')sendChatMessage()"
      style="flex:1;background:#223d58;border:1px solid #284461;border-radius:10px;padding:10px 12px;color:#ddeeff;font-size:15px;outline:none;direction:rtl">
    <button onclick="sendChatMessage()" style="background:#25d366;border:none;border-radius:10px;padding:10px 16px;color:#fff;font-size:20px;cursor:pointer">➤</button>
  </div>
</div>

<!-- מודל חודשים להרשאה -->
<div id="perm-modal" style="display:none;position:fixed;inset:0;z-index:5000;background:rgba(0,0,0,.75);align-items:center;justify-content:center;direction:rtl">
  <div style="background:#d0e4f7;border-radius:16px;padding:22px 20px;width:90%;max-width:310px;display:flex;flex-direction:column;gap:14px">
    <div style="font-weight:700;font-size:15px;color:#ddeeff">⏳ משך ההרשאה</div>
    <div id="perm-modal-label" style="font-size:13px;color:#6b94b8"></div>
    <select id="perm-months-sel" style="background:#223d58;border:1px solid #284461;border-radius:10px;padding:11px 12px;color:#ddeeff;font-size:15px;direction:rtl">
      <option value="0.25">שבוע (7 ימים)</option>
      <option value="0.5">שבועיים (14 ימים)</option>
      <option value="1" selected>חודש אחד</option>
      <option value="2">2 חודשים</option>
      <option value="3">3 חודשים</option>
      <option value="6">6 חודשים</option>
      <option value="12">שנה (12 חודשים)</option>
      <option value="0">ללא הגבלה</option>
    </select>
    <div style="display:flex;gap:10px">
      <button onclick="closePermModal(true)" style="flex:1;background:#3dba6f;color:#fff;border:none;border-radius:10px;padding:12px;font-size:15px;font-weight:700;cursor:pointer">אשר</button>
      <button onclick="closePermModal(false)" style="flex:1;background:#284461;color:#ddeeff;border:none;border-radius:10px;padding:12px;font-size:15px;cursor:pointer">ביטול</button>
    </div>
  </div>
</div>

<!-- מודל הענקת פרמיום -->
<div id="grant-modal" style="display:none;position:fixed;inset:0;z-index:4000;background:rgba(0,0,0,.7);align-items:center;justify-content:center;direction:rtl">
  <div style="background:#d0e4f7;border-radius:16px;padding:24px 20px;width:90%;max-width:320px;display:flex;flex-direction:column;gap:14px">
    <div style="font-weight:700;font-size:16px;color:#ddeeff">⭐ הענקת גישת פרמיום</div>
    <div style="font-size:14px;color:#6b94b8">בחר לכמה חודשים להעניק גישה:</div>
    <select id="grant-months" style="background:#223d58;border:1px solid #284461;border-radius:10px;padding:11px 12px;color:#ddeeff;font-size:15px;direction:rtl">
      <option value="1">חודש אחד</option>
      <option value="2">2 חודשים</option>
      <option value="3">3 חודשים</option>
      <option value="4">4 חודשים</option>
      <option value="5">5 חודשים</option>
      <option value="6">6 חודשים</option>
    </select>
    <div style="display:flex;gap:10px">
      <button id="grant-btn" onclick="doGrantPremium()" style="flex:1;background:#3dba6f;color:#fff;border:none;border-radius:10px;padding:12px;font-size:15px;font-weight:700;cursor:pointer">הענק</button>
      <button onclick="closeGrantModal()" style="flex:1;background:#284461;color:#ddeeff;border:none;border-radius:10px;padding:12px;font-size:15px;cursor:pointer">ביטול</button>
    </div>
  </div>
</div>
<!-- ══ MAP MODAL ══ -->
<div id="map-modal" style="display:none;position:fixed;inset:0;z-index:9999;background:#0d1b2a;flex-direction:column">
  <div style="display:flex;justify-content:space-between;align-items:center;padding:12px 16px;background:#0d1b2a;border-bottom:1px solid #1c3450;flex-shrink:0">
    <div style="font-weight:700;color:#eaf4ff;font-size:16px">🗺️ מפת ביקורים</div>
    <div style="display:flex;gap:10px;align-items:center">
      <div id="map-status" style="font-size:12px;color:#6b94b8;direction:rtl"></div>
      <button onclick="closeMap()" style="background:transparent;border:1px solid #284461;color:#6b94b8;border-radius:8px;padding:6px 14px;font-size:14px;cursor:pointer">✕ סגור</button>
    </div>
  </div>
  <div style="display:flex;gap:8px;padding:8px 12px;background:#0d1b2a;flex-shrink:0;flex-wrap:wrap">
    <div style="display:flex;align-items:center;gap:4px;font-size:12px;color:#ddeeff"><span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:#3dba6f;border:2px solid #fff"></span>המיקום שלי</div>
    <div style="display:flex;align-items:center;gap:4px;font-size:12px;color:#ddeeff"><span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:#53bdeb;border:2px solid #fff"></span>ביקור רגיל</div>
    <div style="display:flex;align-items:center;gap:4px;font-size:12px;color:#ddeeff"><span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:#f9c846;border:2px solid #fff"></span>VIP</div>
    <div style="display:flex;align-items:center;gap:4px;font-size:12px;color:#ddeeff"><span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:#f15c6e;border:2px solid #fff"></span>לא הושלם</div>
  </div>
  <div id="map-container" style="flex:1;width:100%"></div>
</div>
</body>
</html>"""

# אתחול DB בהפעלה — יוצר את כל הטבלאות כולל users
try:
    init_db().close()
except Exception as _e:
    print(f"DB init warning: {_e}")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    print(f"רץ על http://localhost:{port}")
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
