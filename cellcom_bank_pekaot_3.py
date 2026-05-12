"""
cellcom_bank_pekaot_3.py — שליפת בנק פקעות (רשימת משימות) מ-API של סלקום
==========================================================================

תיאור:
------
סקריפט Python עצמאי ששולף את רשימת המשימות הפתוחות (פקעות) של הטכנאי
ישירות מ-API הפנימי של סלקום (אותו API שמשמש את אפליקציית HomeTech).

הפונקציונליות:
--------------
- מבצע GET ל-endpoint של בנק פקעות
- לכל פקעה מושך פרטים מלאים (כתובת, שם לקוח, טלפון, ציוד מתוכנן, תשתית)
- מגלה ביקורים חוזרים (חשד לחזורת) לפי היסטוריה ב-DB
- מייצא ל-Excel עם עמודות מוגדרות ועיצוב

שימוש:
------
  python cellcom_bank_pekaot_3.py

דרישות:
-------
  - קובץ token.txt עם Bearer token תקף (מהרשמה ב-chat_server.py)
  - גישה לרשת האינטרנט
"""

import requests
import os
import json
import sys
import time
from datetime import date, datetime
from collections import defaultdict
import urllib3
urllib3.disable_warnings()

# ============================================================
# הגדרות
# ============================================================
PHONE_DEVICE_ID  = "FF210C6E-5313-4961-846D-229DF3FAC0FC"
DEVICE_MODEL     = "ios_Apple_iPhone 16_SysVer_26.3.1_appVer_23.03.26.1P"
OUTPUT_DIR       = r"C:\Users\Avihu\Documents\cellcom_reports"
TOKEN_FILE       = os.path.join(OUTPUT_DIR, "token.txt")
TECHNICIAN_NAME  = "אביהו עזאני"   # שם הטכנאי כפי שמופיע בהיסטוריית ביקורים

BASE_HEADERS = {
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "he-IL,he;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Content-Type": "application/json",
    "format": "application/json",
    "phonedeviceid": PHONE_DEVICE_ID,
    "devicemodel": DEVICE_MODEL,
    "User-Agent": "HomeTechAppClient/23.03.26 CFNetwork/3860.400.51 Darwin/25.3.0",
}

INFRA_MAP = {
    "FB": "סיבים IBC",
    "BF": "סיבים בזק",
    "BN": "נחושת בזק",
    "HO": "HOT",
    "NV": "סיבים NV",
    "IB": "IBC",
}

# ============================================================
# טוקן
# ============================================================
def get_token():
    if not os.path.exists(TOKEN_FILE):
        print("לא נמצא טוקן. הרץ קודם את cellcom_fetcher.py")
        exit(1)
    with open(TOKEN_FILE, "r") as f:
        token = f.read().strip()
    if not token:
        print("קובץ הטוקן ריק.")
        exit(1)
    print("טוקן נטען.")
    return token

# ============================================================
# שליפת בנק פקעות
# ============================================================
def fetch_potential_tasks(token):
    url = "https://tech-api.cellcom.co.il/api/technician/authorize/callActivities/getPotentialTasks"
    headers = {**BASE_HEADERS, "Authorization": f"Bearer {token}"}
    body = {"DeviceModel": DEVICE_MODEL, "PhoneDeviceId": PHONE_DEVICE_ID}
    try:
        response = requests.post(url, headers=headers, json=body, verify=False, timeout=15)
        if response.status_code == 401:
            print("הטוקן פג. הרץ קודם את cellcom_fetcher.py")
            exit(1)
        data = response.json()
        if data.get("Header", {}).get("ReturnCode") != "00":
            return []
        response_data = data.get("Body", {}).get("ResponseData", "{}")
        if isinstance(response_data, str):
            response_data = json.loads(response_data)
        return response_data.get("Tasks", {}).get("Task", [])
    except Exception as e:
        print(f"שגיאה: {e}")
        return []

# ============================================================
# שליפת פרטי ציוד
# ============================================================
def fetch_call_details(token, call_id, customer_id, source_system):
    url = "https://tech-api.cellcom.co.il/api/technician/authorize/callDetails/getCallDetails"
    headers = {**BASE_HEADERS, "Authorization": f"Bearer {token}"}
    body = {
        "DeviceModel": DEVICE_MODEL,
        "CallId": call_id,
        "CrmCustomerId": customer_id,
        "SourceSystem": source_system,
        "IsRefresh": False,
        "PhoneDeviceId": PHONE_DEVICE_ID
    }
    try:
        response = requests.post(url, headers=headers, json=body, verify=False, timeout=15)
        if response.status_code != 200:
            return {}
        data = response.json()
        if data.get("Header", {}).get("ReturnCode") != "00":
            return {}
        return data.get("Body", {})
    except:
        return {}

def fetch_visit_history(token, ban, user_id, user_type="JET"):
    url = "https://tech-api.cellcom.co.il/api/technician/authorize/TechnicianVisitsHistory/GetTechnicianVisitsHistory"
    headers = {**BASE_HEADERS, "Authorization": f"Bearer {token}"}
    body = {
        "DeviceModel": DEVICE_MODEL,
        "PhoneDeviceId": PHONE_DEVICE_ID,
        "ban": ban,
        "userId": user_id,
        "userType": user_type,
    }
    try:
        response = requests.post(url, headers=headers, json=body, verify=False, timeout=15)
        if response.status_code != 200:
            return []
        data = response.json()
        if data.get("Header", {}).get("ReturnCode") != "00":
            return []
        return data.get("Body", {}).get("visitsHistory", [])
    except:
        return []

def parse_visit_history(visits):
    lines = []
    for v in visits:
        date  = v.get("dateEnd", "")
        time_ = v.get("timeEnd", "")
        vtype = v.get("visitType", "")
        status = v.get("status", "")
        tech  = v.get("technicianName", "")
        reasons = " / ".join(
            i.get("visitCloseReason", "") or ""
            for i in v.get("visitInfo", [])
            if i.get("visitCloseReason")
        )
        parts = [p for p in [f"{date} {time_}".strip(), vtype, status, tech, reasons] if p]
        lines.append(" | ".join(parts))
    return "\n".join(lines)

def check_last_visit_warning(visits, task_type):
    if "תקלה" not in (task_type or ""):
        return False
    cutoff = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    from datetime import timedelta
    cutoff -= timedelta(days=30)
    recent = []
    for v in visits:
        raw = v.get("fullDateTimeTo", "") or v.get("dateEnd", "")
        try:
            dt = datetime.strptime(raw[:16], "%d/%m/%Y %H:%M") if " " in raw else datetime.strptime(raw, "%d/%m/%Y")
        except:
            continue
        if dt >= cutoff:
            recent.append((dt, v.get("technicianName", "")))
    if not recent:
        return False
    recent.sort(key=lambda x: x[0], reverse=True)
    return recent[0][1] == TECHNICIAN_NAME

def parse_equipment(body):
    infra_code = ""
    infra_name = ""
    planned_equipment = []
    existing_equipment = []
    tv_equipment = []
    technology = ""

    spm = body.get("supplyProductServicesModel", {})
    general = spm.get("generalInfo", {})
    technology = general.get("technologyType", "")

    services = spm.get("services", [])
    for service in services:
        stype = service.get("serviceType", "")
        if stype == "INTERNET":
            infra_model = service.get("infrastructureModel", {})
            if infra_model:
                infra_code = infra_model.get("infraComOperator", "") or ""
                infra_name = INFRA_MAP.get(infra_code, infra_code)
        elif stype == "TV":
            for kit in service.get("equipmentList", []):
                for item in kit.get("serialItems", []):
                    if item.get("productStatus") == "Installed":
                        desc = item.get("productDescription", "")
                        serial = item.get("serialNumber", "")
                        if desc:
                            tv_equipment.append(f"{desc} ({serial})" if serial else desc)

    for item in spm.get("allVisitSerialEquipment", []):
        status = item.get("productStatus", "")
        desc = item.get("productDescription", "") or item.get("originProductDescription", "")
        actions = item.get("availableEquipmentActions", [])
        if "Install" in actions or status in ["SupplyProcess", "Add"]:
            if desc:
                planned_equipment.append(desc)
        elif status == "Installed":
            if desc:
                existing_equipment.append(desc)

    return {
        "infra_code": infra_code,
        "infra_name": infra_name,
        "technology": technology,
        "planned_equipment": " | ".join(planned_equipment),
        "existing_equipment": " | ".join(existing_equipment),
        "tv_equipment": " | ".join(tv_equipment),
    }

# ============================================================
# ייצוא לאקסל
# ============================================================
def export_to_excel(tasks, equipment_data, district_filter=""):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        os.system("pip install openpyxl")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    today = date.today()
    district_suffix = f"_{district_filter}" if district_filter else ""
    base = f"bank_pekaot{district_suffix}_{today.strftime('%Y-%m-%d')}"
    filepath = os.path.join(OUTPUT_DIR, f"{base}.xlsx")
    counter = 1
    while True:
        try:
            open(filepath, 'a').close()
            break
        except PermissionError:
            filepath = os.path.join(OUTPUT_DIR, f"{base}_{counter}.xlsx")
            counter += 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "בנק פקעות"
    ws.sheet_view.rightToLeft = True

    headers = [
        "מס' קריאה", "מס' לקוח", "שם לקוח", "טלפון",
        "סוג משימה", "תשתית", "טכנולוגיה",
        "ציוד להתקנה", "ציוד קיים", "ציוד TV",
        "כתובת", "עיר", "מחוז", "תאריך", "שעת התחלה", "שעת סיום",
        "הערות", "היסטוריית ביקורים"
    ]
    col_widths = [12, 14, 22, 14, 28, 14, 12, 30, 25, 20, 22, 12, 14, 13, 11, 11, 35, 50]

    BLUE        = "4472C4"
    WHITE       = "FFFFFF"
    HEADER_FNT  = Font(color=WHITE, bold=True, size=11)
    HEADER_FILL = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    CENTER      = Alignment(horizontal="center", vertical="center")
    WRAP_RIGHT  = Alignment(horizontal="right", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FNT
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 22

    def parse_date(t):
        try:
            return datetime.fromisoformat(t.get("start_date", "").replace("+03:00", ""))
        except:
            return datetime.min

    tasks_sorted = sorted(tasks, key=parse_date)

    for row_idx, task in enumerate(tasks_sorted, 2):
        try:
            start_dt  = datetime.fromisoformat(task.get("start_date", "").replace("+03:00", ""))
            end_dt    = datetime.fromisoformat(task.get("end_date", "").replace("+03:00", ""))
            date_str  = start_dt.strftime("%d/%m/%Y")
            start_str = start_dt.strftime("%H:%M")
            end_str   = end_dt.strftime("%H:%M")
        except:
            date_str = start_str = end_str = ""

        street  = task.get("street", "")
        home_no = task.get("home_no", "")
        apt_no  = task.get("apartment_no", "")
        address = f"{street} {home_no}" + (f" דירה {apt_no}" if apt_no and apt_no != "1" else "")

        call_id = task.get("call_id", "")
        eq = equipment_data.get(call_id, {})

        row_data = [
            call_id,
            task.get("customer_id", ""),
            task.get("contact_name", ""),
            task.get("contact_phone", ""),
            task.get("task_type", ""),
            eq.get("infra_name", ""),
            eq.get("technology", ""),
            eq.get("planned_equipment", ""),
            eq.get("existing_equipment", ""),
            eq.get("tv_equipment", ""),
            address,
            task.get("city", ""),
            task.get("district", ""),
            date_str,
            start_str,
            end_str,
            task.get("comment_text", "") or "",
            eq.get("visit_history", ""),
        ]

        if eq.get("last_visit_warning"):
            row_fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
            warn_font = Font(bold=True, color="CC0000")
        else:
            fill_color = "EEF4FF" if row_idx % 2 == 0 else "FFFFFF"
            row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            warn_font = None

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill = row_fill
            cell.alignment = WRAP_RIGHT
            if warn_font and col == 1:
                cell.font = warn_font

    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    summary_row = len(tasks_sorted) + 3
    ws.cell(row=summary_row, column=1, value=f"סה\"כ פקעות: {len(tasks_sorted)}").font = Font(bold=True, size=12)

    infra_counts = defaultdict(int)
    for eq in equipment_data.values():
        name = eq.get("infra_name", "") or "לא ידוע"
        infra_counts[name] += 1

    ws.cell(row=summary_row + 2, column=1, value="פירוט לפי תשתית:").font = Font(bold=True)
    for i, (infra, cnt) in enumerate(sorted(infra_counts.items(), key=lambda x: -x[1])):
        ws.cell(row=summary_row + 3 + i, column=1, value=infra)
        ws.cell(row=summary_row + 3 + i, column=2, value=cnt)

    wb.save(filepath)
    print(f"אקסל נשמר: {filepath}")
    return filepath

# ============================================================
# הרצה ראשית
# ============================================================
if __name__ == "__main__":
    # פרמטר מחוז אופציונלי
    district_filter = sys.argv[1] if len(sys.argv) > 1 else ""

    token = get_token()

    print("שולף בנק פקעות...")
    all_tasks = fetch_potential_tasks(token)

    if not all_tasks:
        print("לא נמצאו פקעות.")
        exit(0)

    # סינון לפי מחוז
    if district_filter:
        tasks = [t for t in all_tasks if district_filter in t.get("district", "")]
        print(f"סינון לפי מחוז '{district_filter}': {len(tasks)} מתוך {len(all_tasks)} פקעות")
    else:
        tasks = all_tasks
        print(f"נמצאו {len(tasks)} פקעות (כל המחוזות)")

    if not tasks:
        print(f"לא נמצאו פקעות במחוז '{district_filter}'.")
        exit(0)

    print("שולף פרטי ציוד והיסטוריית ביקורים...")
    equipment_data = {}
    for i, task in enumerate(tasks, 1):
        call_id     = task.get("call_id", "")
        customer_id = task.get("customer_id", "")
        source      = task.get("system_source", "JET")
        ban         = task.get("ban", customer_id)
        user_id     = task.get("user_id", customer_id)
        print(f"  [{i}/{len(tasks)}] קריאה {call_id}...")
        body = fetch_call_details(token, call_id, customer_id, source)
        eq = parse_equipment(body) if body else {}
        visits = fetch_visit_history(token, ban, user_id, source)
        eq["visit_history"] = parse_visit_history(visits)
        eq["last_visit_warning"] = check_last_visit_warning(visits, task.get("task_type", ""))
        equipment_data[call_id] = eq
        time.sleep(0.3)

    print("בונה אקסל...")
    excel_path = export_to_excel(tasks, equipment_data, district_filter)
    print("הושלם!")
    os.startfile(excel_path)
